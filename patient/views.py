import json
import logging
from datetime import datetime, date, timedelta
import openpyxl
from django.views import View
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from django.http import HttpResponse
from datetime import date, timedelta
from django.db.models import Q
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.core.exceptions import ObjectDoesNotExist, ValidationError, PermissionDenied
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.serializers.json import DjangoJSONEncoder
from django.db import transaction
from django.db.models import Q, Sum, Count
from django.http import HttpResponse, JsonResponse, Http404
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.views.generic import TemplateView, CreateView, UpdateView, DeleteView, ListView, DetailView

from admin_site.models import SiteInfoModel
from patient.models import PatientModel, PatientSettingModel, PatientWalletModel, RegistrationFeeModel, \
    RegistrationPaymentModel, ConsultationDocument, RegistrationReportTemplate, ConsultationReportTemplate
from patient.forms import PatientForm, PatientEditForm, PatientSettingForm, PatientSearchForm, PatientWalletTopUpForm, \
    RegistrationFeeForm, ConsultationDocumentForm, BiodataReportFilterForm, RegistrationReportTemplateForm, \
    ConsultationReportTemplateForm
from admin_site.utility import state_list

# Set up logging
logger = logging.getLogger(__name__)


class FlashFormErrorsMixin:
    """
    Mixin for CreateView/UpdateView to flash form errors and redirect safely.
    Use before SuccessMessageMixin in MRO so messages appear before redirect.
    """
    def form_invalid(self, form):
        try:
            for field, errors in form.errors.items():
                label = form.fields.get(field).label if form.fields.get(field) else field
                for error in errors:
                    messages.error(self.request, f"{label}: {error}")
        except Exception:
            logger.exception("Error while processing form_invalid errors.")
            messages.error(self.request, "There was an error processing the form. Please try again.")
        return redirect(self.get_success_url())


class RegistrationFeeCreateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, CreateView
):
    model = RegistrationFeeModel
    permission_required = 'patient.add_registrationfeemodel'
    form_class = RegistrationFeeForm
    template_name = 'patient/registration_fee/index.html'
    success_message = 'Registration Fee Successfully Created'

    def form_valid(self, form):
        # Attach the logged-in user as created_by
        form.instance.created_by = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('registration_fee_index')

    def dispatch(self, request, *args, **kwargs):
        # Redirect GET back to index (only handle POST)
        if request.method == 'GET':
            return redirect(reverse('registration_fee_index'))
        return super().dispatch(request, *args, **kwargs)


class RegistrationFeeListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = RegistrationFeeModel
    permission_required = 'patient.view_registrationfeemodel'
    template_name = 'patient/registration_fee/index.html'
    context_object_name = "registration_fee_list"

    def get_queryset(self):
        return RegistrationFeeModel.objects.all().order_by('patient_type', 'title')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = RegistrationFeeForm()
        return context


class RegistrationFeeUpdateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, UpdateView
):
    model = RegistrationFeeModel
    permission_required = 'patient.add_registrationfeemodel'
    form_class = RegistrationFeeForm
    template_name = 'patient/registration_fee/index.html'
    success_message = 'Registration Fee Successfully Updated'

    def form_valid(self, form):
        # Attach the logged-in user as updated_by
        form.instance.updated_by = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('registration_fee_index')

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET':
            return redirect(reverse('registration_fee_index'))
        return super().dispatch(request, *args, **kwargs)


class RegistrationFeeDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = RegistrationFeeModel
    permission_required = 'patient.view_registrationfeemodel'
    template_name = 'patient/registration_fee/detail.html'
    context_object_name = "registration_fee"


class RegistrationFeeDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = RegistrationFeeModel
    permission_required = 'patient.delete_registrationfeemodel'
    template_name = 'patient/registration_fee/delete.html'
    context_object_name = "registration_fee"
    success_message = 'Registration Fee Successfully Deleted'

    def get_success_url(self):
        return reverse('registration_fee_index')


class PatientSettingCreateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, CreateView):
    """Create patient settings - only if none exist"""
    model = PatientSettingModel
    form_class = PatientSettingForm
    permission_required = 'patient.change_patientsettingmodel'
    success_message = 'Patient settings created successfully'
    template_name = 'patient/setting/create.html'

    def dispatch(self, request, *args, **kwargs):
        try:
            setting = PatientSettingModel.objects.first()
            if setting:
                messages.info(request, 'Patient settings already exist. Redirecting to edit.')
                return redirect(reverse('patient_setting_edit', kwargs={'pk': setting.pk}))
            return super().dispatch(request, *args, **kwargs)
        except Exception as e:
            logger.error(f"Error in PatientSettingCreateView dispatch: {str(e)}")
            messages.error(request, 'An error occurred while accessing patient settings.')
            return redirect('dashboard')  # Adjust to your dashboard URL

    def get_success_url(self):
        return reverse('patient_setting_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        try:
            with transaction.atomic():
                return super().form_valid(form)
        except ValidationError as e:
            messages.error(self.request, f'Validation error: {str(e)}')
            return self.form_invalid(form)
        except Exception as e:
            logger.error(f"Error creating patient settings: {str(e)}")
            messages.error(self.request, 'An unexpected error occurred. Please try again.')
            return self.form_invalid(form)


class PatientSettingDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """View patient settings details"""
    model = PatientSettingModel
    permission_required = 'patient.view_patientsettingmodel'
    template_name = 'patient/setting/detail.html'
    context_object_name = "patient_setting"

    def dispatch(self, request, *args, **kwargs):
        try:
            setting = PatientSettingModel.objects.first()
            if not setting:
                messages.warning(request, 'No patient settings found. Please create settings first.')
                return redirect(reverse('patient_setting_create'))

            # Ensure we're always looking at the single settings record
            if self.kwargs.get('pk') != setting.id:
                return redirect(reverse('patient_setting_detail', kwargs={'pk': setting.pk}))

            return super().dispatch(request, *args, **kwargs)
        except Exception as e:
            logger.error(f"Error in PatientSettingDetailView dispatch: {str(e)}")
            messages.error(request, 'An error occurred while accessing patient settings.')
            return redirect('admin_dashboard')


class PatientSettingUpdateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, UpdateView):
    """Update patient settings"""
    model = PatientSettingModel
    form_class = PatientSettingForm
    permission_required = 'patient.change_patientsettingmodel'
    success_message = 'Patient settings updated successfully'
    template_name = 'patient/setting/create.html'

    def get_success_url(self):
        return reverse('patient_setting_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        try:
            with transaction.atomic():
                return super().form_valid(form)
        except ValidationError as e:
            messages.error(self.request, f'Validation error: {str(e)}')
            return self.form_invalid(form)
        except Exception as e:
            logger.error(f"Error updating patient settings: {str(e)}")
            messages.error(self.request, 'An unexpected error occurred. Please try again.')
            return self.form_invalid(form)


class PatientCreateView(
    LoginRequiredMixin,
    PermissionRequiredMixin,
    SuccessMessageMixin,
    CreateView
):
    """
    Create new patient — registration MUST originate from a pending RegistrationPaymentModel
    Expects pay_id as a URL path parameter: /patients/register/<int:pay_id>/
    """
    model = PatientModel
    permission_required = 'patient.add_patientmodel'
    form_class = PatientForm
    template_name = 'patient/patient/create.html'
    success_message = 'Patient registered successfully'

    def dispatch(self, request, *args, **kwargs):
        """
        Validate environment & payment before allowing access to the form.
        """
        try:
            # Ensure patient settings exist
            patient_setting = PatientSettingModel.objects.first()
            if not patient_setting:
                messages.error(request, 'Patient settings not configured. Please contact administrator.')
                return redirect('patient_setting_create')

            # pay_id must be provided as a path parameter (kwargs)
            pay_id = kwargs.get('pay_id')
            if not pay_id:
                messages.warning(request, 'Registration must originate from a finance payment. Select a pending payment to register.')
                return redirect('pending_patient_index')

            # Validate the payment exists and is pending
            try:
                payment = RegistrationPaymentModel.objects.get(pk=pay_id)
            except RegistrationPaymentModel.DoesNotExist:
                messages.error(request, 'Invalid payment ID.')
                return redirect('pending_patient_index')

            if payment.registration_status != 'pending':
                messages.error(request, 'This payment has already been used for registration or is not pending.')
                return redirect('pending_patient_index')

            if payment.status == 'reverted':
                messages.error(request, 'This payment has been reverted.')
                return redirect('pending_patient_index')

            # Attach to request for reuse in other methods to avoid requery
            request._registration_payment = payment
            return super().dispatch(request, *args, **kwargs)

        except Exception as e:
            logger.error(f"Error in PatientCreateView.dispatch: {e}", exc_info=True)
            messages.error(request, 'An unexpected error occurred.')
            return redirect('dashboard')

    def get_initial(self):
        """
        Optionally set initial form values if you want to prefill fields in the form itself.
        We'll keep get_context_data for name-part prefilling, but this is available if useful.
        """
        initial = super().get_initial()
        # Example: initial['card_number'] = generate_card_number()  # if needed
        return initial

    def form_valid(self, form):
        try:
            with transaction.atomic():
                # set creator
                form.instance.created_by = self.request.user

                # attach the registration payment BEFORE saving
                payment = getattr(self.request, "_registration_payment", None)
                if not payment:
                    messages.error(self.request, "Missing registration payment.")
                    return self.form_invalid(form)

                form.instance.registration_payment = payment

                # Save patient (sets self.object)
                response = super().form_valid(form)

                # Mark payment completed
                payment.registration_status = 'completed'
                payment.save()

                # Create patient wallet
                PatientWalletModel.objects.get_or_create(patient=self.object)

                logger.info(
                    f"Patient {getattr(self.object, 'card_number', self.object.pk)} created by {self.request.user}")
                return response

        except ValidationError as e:
            messages.error(self.request, f'Validation error: {e}')
            return self.form_invalid(form)
        except Exception as e:
            logger.error(f"Error creating patient: {e}", exc_info=True)
            messages.error(self.request, 'An unexpected error occurred during registration.')
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse('patient_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        """
        Prefill first/middle/last for template display using payment.full_name.
        Return pay_id, patient settings, and old card number (if available) for template consumption.
        """
        context = super().get_context_data(**kwargs)
        pay_id = self.kwargs.get('pay_id')
        first_name = middle_name = last_name = old_card_number = ''

        if pay_id:
            try:
                payment = getattr(self.request, "_registration_payment", None) or get_object_or_404(
                    RegistrationPaymentModel, pk=pay_id)

                # split full name into parts
                full_name = (payment.full_name or '').strip()
                if full_name:
                    parts = full_name.split()
                    if len(parts) == 1:
                        first_name = parts[0]
                    elif len(parts) == 2:
                        first_name, last_name = parts
                    elif len(parts) >= 3:
                        first_name, middle_name = parts[0], parts[1]
                        last_name = ' '.join(parts[2:])

                # old card number (if provided on the payment)
                old_card_number = (getattr(payment, 'old_card_number', '') or '').strip()

            except Exception as e:
                logger.warning(f"Error extracting name or card from payment {pay_id}: {e}", exc_info=True)

        context.update({
            'pay_id': pay_id,
            'state_list': state_list,  # keep your existing state data source
            'first_name': first_name,
            'middle_name': middle_name,
            'last_name': last_name,
            'old_card_number': old_card_number,
            'patient_setting': PatientSettingModel.objects.first()
        })
        return context


class PatientListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """List all patients with search and pagination"""
    model = PatientModel
    permission_required = 'patient.view_patientmodel'
    template_name = 'patient/patient/index.html'
    context_object_name = "patients"
    paginate_by = 20

    def get_queryset(self):
        queryset = PatientModel.objects.select_related().filter(status='active')

        # Handle search
        search_form = PatientSearchForm(self.request.GET)
        if search_form.is_valid():
            search_query = search_form.cleaned_data.get('search_query')
            gender = search_form.cleaned_data.get('gender')
            blood_group = search_form.cleaned_data.get('blood_group')
            date_from = search_form.cleaned_data.get('date_from')
            date_to = search_form.cleaned_data.get('date_to')

            if search_query:
                queryset = queryset.filter(
                    Q(first_name__icontains=search_query) |
                    Q(middle_name__icontains=search_query) |
                    Q(last_name__icontains=search_query) |
                    Q(card_number__icontains=search_query) |
                    Q(mobile__icontains=search_query) |
                    Q(email__icontains=search_query)
                )

            if gender:
                queryset = queryset.filter(gender=gender)

            if blood_group:
                queryset = queryset.filter(blood_group=blood_group)

            if date_from:
                queryset = queryset.filter(registration_date__gte=date_from)

            if date_to:
                queryset = queryset.filter(registration_date__lte=date_to)

        return queryset.order_by('first_name', 'last_name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_form'] = PatientSearchForm(self.request.GET)
        context['patient_setting'] = PatientSettingModel.objects.first()
        context['total_patients'] = PatientModel.objects.filter(status='active').count()

        # Add pagination info for debugging
        if context['patients']:
            page_obj = context['page_obj']
            context['pagination_info'] = {
                'current_page': page_obj.number,
                'total_pages': page_obj.paginator.num_pages,
                'total_items': page_obj.paginator.count,
                'start_index': page_obj.start_index(),
                'end_index': page_obj.end_index(),
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
            }

        return context


class PatientDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """View patient details with related information"""
    model = PatientModel
    permission_required = 'patient.view_patientmodel'
    template_name = 'patient/patient/detail.html'
    context_object_name = "patient"

    def get_object(self, queryset=None):
        try:
            obj = super().get_object(queryset)
            if obj.status != 'active':
                raise Http404("Patient not found or inactive")
            return obj
        except PatientModel.DoesNotExist:
            raise Http404("Patient not found")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        patient = self.object

        # Get or create wallet
        wallet, created = PatientWalletModel.objects.get_or_create(patient=patient)

        consultation_documents = patient.consultation_documents.all()
        document_form = ConsultationDocumentForm()

        context.update({
            'patient_wallet': wallet,
            'patient_age': patient.age(),
            'wallet_topup_form': PatientWalletTopUpForm(),
            'recent_visits': [],  # Add when you have visit models
            'upcoming_appointments': [],  # Add when you have appointment models
            'consultation_documents': consultation_documents,
            'document_form': document_form,
        })

        return context


class PatientPaymentHistoryView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """
    Patient sector: shows processed registration payments
    """
    permission_required = 'patient.view_patientmodel'
    model = RegistrationPaymentModel
    template_name = 'patient/patient/payment_history.html'
    context_object_name = 'payment_list'
    paginate_by = 50  # optional

    def get_queryset(self):
        qs = RegistrationPaymentModel.objects.filter(registration_status='completed').order_by('-created_at')

        # Filter by date range
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')

        if start_date:
            qs = qs.filter(created_at__date__gte=start_date)
        else:
            # Default to today
            today = timezone.now().date()
            qs = qs.filter(created_at__date=today)

        if end_date:
            qs = qs.filter(created_at__date__lte=end_date)

        # Filter by transaction ID (search)
        transaction_id = self.request.GET.get('transaction_id')
        if transaction_id:
            qs = qs.filter(transaction_id__icontains=transaction_id)

        # Optional: filter by full_name search
        search_name = self.request.GET.get('search')
        if search_name:
            qs = qs.filter(full_name__icontains=search_name)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'start_date': self.request.GET.get('start_date', timezone.now().date()),
            'end_date': self.request.GET.get('end_date', ''),
            'transaction_id': self.request.GET.get('transaction_id', ''),
            'search': self.request.GET.get('search', ''),
        })
        return context


class PatientUpdateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, UpdateView):
    """Update patient information"""
    model = PatientModel
    permission_required = 'patient.change_patientmodel'
    form_class = PatientEditForm
    template_name = 'patient/patient/edit.html'
    success_message = 'Patient information updated successfully'

    def get_object(self, queryset=None):
        try:
            obj = super().get_object(queryset)
            if obj.status != 'active':
                raise Http404("Patient not found or inactive")
            return obj
        except PatientModel.DoesNotExist:
            raise Http404("Patient not found")

    def form_valid(self, form):
        try:
            with transaction.atomic():
                response = super().form_valid(form)
                logger.info(f"Patient {self.object.card_number} updated by {self.request.user}")
                return response
        except ValidationError as e:
            messages.error(self.request, f'Validation error: {str(e)}')
            return self.form_invalid(form)
        except Exception as e:
            logger.error(f"Error updating patient {self.object.pk}: {str(e)}")
            messages.error(self.request, 'An unexpected error occurred while updating patient.')
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse('patient_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'state_list': state_list,
            'patient': self.object,
            'patient_setting': PatientSettingModel.objects.first()
        })
        return context


class PatientDeleteView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, DeleteView):
    """Soft delete patient (set status to inactive)"""
    model = PatientModel
    permission_required = 'patient.delete_patientmodel'
    template_name = 'patient/patient/delete.html'
    context_object_name = "patient"
    success_message = 'Patient deactivated successfully'
    success_url = reverse_lazy('patient_index')

    def get_object(self, queryset=None):
        try:
            obj = super().get_object(queryset)
            if obj.status != 'active':
                raise Http404("Patient not found or already inactive")
            return obj
        except PatientModel.DoesNotExist:
            raise Http404("Patient not found")

    def delete(self, request, *args, **kwargs):
        """Soft delete - set status to inactive instead of actually deleting"""
        try:
            with transaction.atomic():
                self.object = self.get_object()
                self.object.status = 'inactive'
                self.object.save()

                logger.info(f"Patient {self.object.card_number} deactivated by {request.user}")
                messages.success(request, self.success_message)

                return redirect(self.success_url)
        except Exception as e:
            logger.error(f"Error deactivating patient {self.object.pk}: {str(e)}")
            messages.error(request, 'An error occurred while deactivating patient.')
            return redirect('patient_detail', pk=self.object.pk)


def get_patient_dashboard_context(request):
    """
    Get the context data for patient dashboard
    """
    # Get current date and time periods
    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)
    year_start = today.replace(month=1, day=1)

    # Basic patient statistics
    total_patients = PatientModel.objects.count()
    male_patients = PatientModel.objects.filter(gender='male').count()
    female_patients = PatientModel.objects.filter(gender='female').count()

    # Registration status statistics
    pending_registrations = RegistrationPaymentModel.objects.filter(
        registration_status='pending'
    ).count()

    # Old vs New patients (based on registration fee type)
    old_patients_total = RegistrationPaymentModel.objects.filter(
        registration_fee__patient_type='old'
    ).count()
    new_patients_total = RegistrationPaymentModel.objects.filter(
        registration_fee__patient_type='new'
    ).count()

    # Time-based statistics for old patients
    old_patients_today = RegistrationPaymentModel.objects.filter(
        registration_fee__patient_type='old',
        date=today
    ).count()

    old_patients_week = RegistrationPaymentModel.objects.filter(
        registration_fee__patient_type='old',
        date__gte=week_start
    ).count()

    old_patients_month = RegistrationPaymentModel.objects.filter(
        registration_fee__patient_type='old',
        date__gte=month_start
    ).count()

    # Time-based statistics for new patients
    new_patients_today = RegistrationPaymentModel.objects.filter(
        registration_fee__patient_type='new',
        date=today
    ).count()

    new_patients_week = RegistrationPaymentModel.objects.filter(
        registration_fee__patient_type='new',
        date__gte=week_start
    ).count()

    new_patients_month = RegistrationPaymentModel.objects.filter(
        registration_fee__patient_type='new',
        date__gte=month_start
    ).count()

    # Calculate growth percentages (comparing this month to last month)
    last_month_start = (month_start - timedelta(days=1)).replace(day=1)
    last_month_end = month_start - timedelta(days=1)

    old_patients_last_month = RegistrationPaymentModel.objects.filter(
        registration_fee__patient_type='old',
        date__gte=last_month_start,
        date__lte=last_month_end
    ).count()

    new_patients_last_month = RegistrationPaymentModel.objects.filter(
        registration_fee__patient_type='new',
        date__gte=last_month_start,
        date__lte=last_month_end
    ).count()

    # Calculate growth percentages
    old_patients_growth = calculate_growth_percentage(old_patients_month, old_patients_last_month)
    new_patients_growth = calculate_growth_percentage(new_patients_month, new_patients_last_month)

    # Age group statistics
    age_groups = calculate_age_groups()

    # Recent registrations for chart data (last 7 days)
    registration_chart_data = get_registration_chart_data()

    # Monthly registration trends (last 12 months)
    monthly_trends = get_monthly_trends()

    # Gender distribution for pie chart
    gender_distribution = [
        {'name': 'Male', 'value': male_patients},
        {'name': 'Female', 'value': female_patients}
    ]

    # Revenue statistics
    total_revenue = RegistrationPaymentModel.objects.filter(
        registration_status='completed'
    ).aggregate(Sum('amount'))['amount__sum'] or 0

    revenue_today = RegistrationPaymentModel.objects.filter(
        registration_status='completed',
        date=today
    ).aggregate(Sum('amount'))['amount__sum'] or 0

    revenue_month = RegistrationPaymentModel.objects.filter(
        registration_status='completed',
        date__gte=month_start
    ).aggregate(Sum('amount'))['amount__sum'] or 0

    return {
        'total_patients': total_patients,
        'male_patients': male_patients,
        'female_patients': female_patients,
        'pending_registrations': pending_registrations,
        'old_patients_total': old_patients_total,
        'new_patients_total': new_patients_total,
        'old_patients_today': old_patients_today,
        'old_patients_week': old_patients_week,
        'old_patients_month': old_patients_month,
        'new_patients_today': new_patients_today,
        'new_patients_week': new_patients_week,
        'new_patients_month': new_patients_month,
        'old_patients_growth': old_patients_growth,
        'new_patients_growth': new_patients_growth,
        'age_groups': age_groups,
        'registration_chart_data': json.dumps(registration_chart_data),
        'monthly_trends': json.dumps(monthly_trends),
        'gender_distribution': json.dumps(gender_distribution),
        'total_revenue': total_revenue,
        'revenue_today': revenue_today,
        'revenue_month': revenue_month,
    }


@login_required
@permission_required('patient.can_view_patient_dashboard', raise_exception=True)
def patient_dashboard(request):
    """
    Comprehensive patient dashboard with statistics and analytics
    """
    context = get_patient_dashboard_context(request)
    return render(request, 'patient/patient/dashboard.html', context)


@login_required
@permission_required('patient.can_view_patient_dashboard', raise_exception=True)
def patient_dashboard_print(request):
    """
    Printable version of patient dashboard statistics
    """
    # Get the same context data as the main dashboard
    context = get_patient_dashboard_context(request)

    # Add print-specific data
    context.update({
        'print_date': timezone.now(),
        'generated_by': request.user,
    })

    return render(request, 'patient/patient/dashboard_print.html', context)


def calculate_growth_percentage(current, previous):
    """Calculate growth percentage between two periods"""
    if previous == 0:
        return 100 if current > 0 else 0
    return round(((current - previous) / previous) * 100, 1)


def calculate_age_groups():
    """Calculate patient count by age groups"""
    from datetime import date

    age_groups = {
        '0-12': 0,
        '13-17': 0,
        '18-40': 0,
        '41-65': 0,
        '66+': 0
    }

    patients_with_dob = PatientModel.objects.exclude(date_of_birth__isnull=True)

    for patient in patients_with_dob:
        age = patient.age()
        if isinstance(age, int):
            if 0 <= age <= 12:
                age_groups['0-12'] += 1
            elif 13 <= age <= 17:
                age_groups['13-17'] += 1
            elif 18 <= age <= 40:
                age_groups['18-40'] += 1
            elif 41 <= age <= 65:
                age_groups['41-65'] += 1
            elif age >= 66:
                age_groups['66+'] += 1

    return [{'name': k, 'value': v} for k, v in age_groups.items()]


def get_registration_chart_data():
    """Get registration data for the last 7 days"""
    today = timezone.now().date()
    data = []

    for i in range(6, -1, -1):  # Last 7 days
        date = today - timedelta(days=i)
        new_count = RegistrationPaymentModel.objects.filter(
            registration_fee__patient_type='new',
            date=date
        ).count()
        old_count = RegistrationPaymentModel.objects.filter(
            registration_fee__patient_type='old',
            date=date
        ).count()

        data.append({
            'date': date.strftime('%Y-%m-%d'),
            'new_patients': new_count,
            'old_patients': old_count
        })

    return data


def get_monthly_trends():
    """Get monthly registration trends for the last 12 months"""
    today = timezone.now().date()
    data = []

    for i in range(11, -1, -1):  # Last 12 months
        # Calculate month start
        if today.month - i > 0:
            month = today.month - i
            year = today.year
        else:
            month = 12 + (today.month - i)
            year = today.year - 1

        month_start = date(year, month, 1)
        if month == 12:
            month_end = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = date(year, month + 1, 1) - timedelta(days=1)

        new_count = RegistrationPaymentModel.objects.filter(
            registration_fee__patient_type='new',
            date__gte=month_start,
            date__lte=month_end
        ).count()

        old_count = RegistrationPaymentModel.objects.filter(
            registration_fee__patient_type='old',
            date__gte=month_start,
            date__lte=month_end
        ).count()

        data.append({
            'month': month_start.strftime('%b %Y'),
            'new_patients': new_count,
            'old_patients': old_count,
            'total': new_count + old_count
        })

    return data


class PatientPendingListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """List pending patient registrations (paid but not registered)"""
    model = RegistrationPaymentModel
    permission_required = 'patient.add_patientmodel'
    template_name = 'patient/patient/pending_index.html'
    context_object_name = "pending_payment_list"
    paginate_by = 20

    def get_queryset(self):
        return RegistrationPaymentModel.objects.filter(
            registration_status='pending', status='confirmed'
        ).order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_pending'] = self.get_queryset().count()
        return context


# AJAX and Utility Views
@login_required
@require_http_methods(["GET"])
def get_patient_with_card(request):
    """Get patient details by card number (AJAX endpoint)"""
    card_number = request.GET.get('card_number', '').strip()
    payment_type = request.GET.get('payment_type', '')

    if not card_number:
        return JsonResponse({
            'success': False,
            'error': 'Card number is required'
        })

    try:
        patient = PatientModel.objects.get(card_number=card_number, status='active')

        # Return JSON response for AJAX calls
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'patient': {
                    'id': patient.pk,
                    'full_name': str(patient),
                    'card_number': patient.card_number,
                    'mobile': patient.mobile,
                    'gender': patient.gender,
                    'age': patient.age(),
                    'blood_group': patient.blood_group,
                }
            })

        # Return template for non-AJAX calls
        context = {
            'patient': patient,
            'payment_type': payment_type,
        }
        return render(request, 'patient/patient/get_detail_from_card.html', context)

    except PatientModel.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': 'Patient not found with this card number'
            })

        context = {
            'patient': None,
            'payment_type': payment_type,
            'error': 'Patient not found with this card number'
        }
        return render(request, 'patient/patient/get_detail_from_card.html', context)

    except Exception as e:
        logger.error(f"Error retrieving patient with card {card_number}: {str(e)}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': 'An error occurred while searching for patient'
            })

        context = {
            'patient': None,
            'payment_type': payment_type,
            'error': 'An error occurred while searching for patient'
        }
        return render(request, 'patient/patient/get_detail_from_card.html', context)


@login_required
@require_http_methods(["POST"])
def patient_wallet_topup(request, patient_id):
    """Add funds to patient wallet"""
    try:
        patient = get_object_or_404(PatientModel, pk=patient_id, status='active')
        form = PatientWalletTopUpForm(request.POST)

        if form.is_valid():
            with transaction.atomic():
                amount = form.cleaned_data['amount']
                payment_method = form.cleaned_data['payment_method']
                reference = form.cleaned_data.get('reference', '')

                # Get or create wallet
                wallet, created = PatientWalletModel.objects.get_or_create(patient=patient)
                wallet.add_funds(amount)

                # Log the transaction (you might want to create a transaction model)
                logger.info(f"Wallet top-up: {amount} added to patient {patient.card_number} by {request.user}")

                messages.success(request, f'Successfully added ₦{amount:,.2f} to patient wallet.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')

    except Exception as e:
        logger.error(f"Error topping up wallet for patient {patient_id}: {str(e)}")
        messages.error(request, 'An error occurred while adding funds to wallet.')

    return redirect('patient_detail', pk=patient_id)


@login_required
def patient_statistics(request):
    """Get patient statistics for dashboard"""
    try:
        stats = {
            'total_patients': PatientModel.objects.filter(status='active').count(),
            'new_this_month': PatientModel.objects.filter(
                status='active',
                registration_date__month=timezone.now().month,
                registration_date__year=timezone.now().year
            ).count(),
            'pending_registrations': RegistrationPaymentModel.objects.filter(
                registration_status='pending'
            ).count(),
            'total_wallet_balance': PatientWalletModel.objects.aggregate(
                total=Sum('amount')
            )['total'] or 0,
        }

        return JsonResponse({
            'success': True,
            'statistics': stats
        })

    except Exception as e:
        logger.error(f"Error getting patient statistics: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Error retrieving statistics'
        })


@login_required
@permission_required('patient.add_patientmodel', raise_exception=True)
def upload_consultation_document(request, patient_id):
    """Upload consultation document via AJAX"""
    patient = get_object_or_404(PatientModel, id=patient_id)

    if request.method == 'POST':
        form = ConsultationDocumentForm(request.POST, request.FILES)
        if form.is_valid():
            document = form.save(commit=False)
            document.patient = patient
            document.uploaded_by = request.user
            document.save()

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Document uploaded successfully',
                    'document': {
                        'id': document.id,
                        'filename': document.filename,
                        'title': document.title,
                        'uploaded_at': document.uploaded_at.strftime('%b %d, %Y %H:%M'),
                        'uploaded_by': document.uploaded_by.get_full_name() or document.uploaded_by.username,
                        'url': document.document.url
                    }
                })
            else:
                messages.success(request, 'Document uploaded successfully')
                return redirect('patient_detail', patient_id=patient_id)
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                })

    return redirect('patient_detail', patient_id=patient_id)


@login_required
@permission_required('patient.add_patientmodel', raise_exception=True)
def delete_consultation_document(request, document_id):
    """Delete consultation document via AJAX"""
    document = get_object_or_404(ConsultationDocument, id=document_id)
    patient_id = document.patient.id

    if request.method == 'POST':
        document.document.delete()  # Delete file from storage
        document.delete()  # Delete from database

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': 'Document deleted successfully'
            })
        else:
            messages.success(request, 'Document deleted successfully')

    return redirect('patient_detail', patient_id=patient_id)


# ========================= REPORTS HUB =========================
class ReportsHubView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Central hub for all patient reports"""
    template_name = 'patient/reports/hub.html'
    permission_required = 'patient.view_patientmodel'


# ========================= BIO DATA REPORT =========================
class BiodataReportView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """Bio-data report with advanced filtering"""
    model = PatientModel
    template_name = 'patient/reports/biodata.html'
    context_object_name = 'patients'
    permission_required = 'patient.view_patientmodel'

    def get_paginate_by(self, queryset):
        return int(self.request.GET.get('per_page', 100))

    def get_queryset(self):
        queryset = PatientModel.objects.filter(status='active').order_by('-created_at')

        # Name search
        name_search = self.request.GET.get('name_search', '').strip()
        if name_search:
            queryset = queryset.filter(
                Q(first_name__icontains=name_search) |
                Q(middle_name__icontains=name_search) |
                Q(last_name__icontains=name_search)
            )

        # State filter
        state = self.request.GET.get('state', '').strip()
        if state:
            queryset = queryset.filter(state__icontains=state)

        # LGA filter
        lga = self.request.GET.get('lga', '').strip()
        if lga:
            queryset = queryset.filter(lga__icontains=lga)

        # Age range filter
        age_min = self.request.GET.get('age_min')
        age_max = self.request.GET.get('age_max')

        if age_min or age_max:
            today = date.today()

            if age_max:
                max_birth_date = date(today.year - int(age_max), today.month, today.day)
                queryset = queryset.filter(date_of_birth__gte=max_birth_date)

            if age_min:
                min_birth_date = date(today.year - int(age_min) - 1, today.month, today.day)
                queryset = queryset.filter(date_of_birth__lte=min_birth_date)

        # Address search
        address_search = self.request.GET.get('address_search', '').strip()
        if address_search:
            queryset = queryset.filter(address__icontains=address_search)

        # Gender filter
        gender = self.request.GET.get('gender', '').strip()
        if gender:
            queryset = queryset.filter(gender=gender)

        # Marital status filter
        marital_status = self.request.GET.get('marital_status', '').strip()
        if marital_status:
            queryset = queryset.filter(marital_status=marital_status)

        # Religion filter
        religion = self.request.GET.get('religion', '').strip()
        if religion:
            queryset = queryset.filter(religion=religion)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = BiodataReportFilterForm(self.request.GET)
        context['total_count'] = self.get_queryset().count()
        return context


class BiodataReportExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Export bio-data report to Excel"""
    permission_required = 'patient.view_patientmodel'

    def get(self, request, *args, **kwargs):
        # Get filtered queryset (reuse same logic as BiodataReportView)
        queryset = PatientModel.objects.filter(status='active').order_by('-created_at')

        # Apply all filters (same as BiodataReportView)
        name_search = request.GET.get('name_search', '').strip()
        if name_search:
            queryset = queryset.filter(
                Q(first_name__icontains=name_search) |
                Q(middle_name__icontains=name_search) |
                Q(last_name__icontains=name_search)
            )

        state = request.GET.get('state', '').strip()
        if state:
            queryset = queryset.filter(state__icontains=state)

        lga = request.GET.get('lga', '').strip()
        if lga:
            queryset = queryset.filter(lga__icontains=lga)

        age_min = request.GET.get('age_min')
        age_max = request.GET.get('age_max')
        if age_min or age_max:
            today = date.today()
            if age_max:
                max_birth_date = date(today.year - int(age_max), today.month, today.day)
                queryset = queryset.filter(date_of_birth__gte=max_birth_date)
            if age_min:
                min_birth_date = date(today.year - int(age_min) - 1, today.month, today.day)
                queryset = queryset.filter(date_of_birth__lte=min_birth_date)

        address_search = request.GET.get('address_search', '').strip()
        if address_search:
            queryset = queryset.filter(address__icontains=address_search)

        gender = request.GET.get('gender', '').strip()
        if gender:
            queryset = queryset.filter(gender=gender)

        marital_status = request.GET.get('marital_status', '').strip()
        if marital_status:
            queryset = queryset.filter(marital_status=marital_status)

        religion = request.GET.get('religion', '').strip()
        if religion:
            queryset = queryset.filter(religion=religion)

        # Get selected fields (default fields always included)
        export_fields = request.GET.getlist('export_fields')
        if not export_fields:
            export_fields = ['full_name', 'card_number', 'mobile', 'address']

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Patient Bio Data"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Build headers
        headers = []
        if 'full_name' in export_fields:
            headers.append('Full Name')
        if 'card_number' in export_fields:
            headers.append('Card Number')
        if 'mobile' in export_fields:
            headers.append('Phone')
        if 'email' in export_fields:
            headers.append('Email')
        if 'address' in export_fields:
            headers.append('Address')
        if 'gender' in export_fields:
            headers.append('Gender')
        if 'age' in export_fields:
            headers.append('Age')
        if 'marital_status' in export_fields:
            headers.append('Marital Status')
        if 'religion' in export_fields:
            headers.append('Religion')
        if 'state' in export_fields:
            headers.append('State')
        if 'lga' in export_fields:
            headers.append('LGA')

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Write data
        row = 2
        for patient in queryset:
            col = 1

            if 'full_name' in export_fields:
                ws.cell(row=row, column=col, value=str(patient))
                col += 1
            if 'card_number' in export_fields:
                ws.cell(row=row, column=col, value=patient.card_number)
                col += 1
            if 'mobile' in export_fields:
                ws.cell(row=row, column=col, value=patient.mobile or '')
                col += 1
            if 'email' in export_fields:
                ws.cell(row=row, column=col, value=patient.email or '')
                col += 1
            if 'address' in export_fields:
                ws.cell(row=row, column=col, value=patient.address or '')
                col += 1
            if 'gender' in export_fields:
                ws.cell(row=row, column=col, value=patient.get_gender_display() if patient.gender else '')
                col += 1
            if 'age' in export_fields:
                ws.cell(row=row, column=col, value=patient.age() or '')
                col += 1
            if 'marital_status' in export_fields:
                ws.cell(row=row, column=col,
                        value=patient.get_marital_status_display() if patient.marital_status else '')
                col += 1
            if 'religion' in export_fields:
                ws.cell(row=row, column=col, value=patient.get_religion_display() if patient.religion else '')
                col += 1
            if 'state' in export_fields:
                ws.cell(row=row, column=col, value=patient.state or '')
                col += 1
            if 'lga' in export_fields:
                ws.cell(row=row, column=col, value=patient.lga or '')
                col += 1

            row += 1

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"patient_biodata_{date.today().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response


# ========================= REGISTRATION REPORT =========================
class RegistrationReportView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """Registration report with templates"""
    model = RegistrationReportTemplate
    template_name = 'patient/reports/registration.html'
    context_object_name = 'templates'
    permission_required = 'patient.view_patientmodel'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get date range
        from_date = self.request.GET.get('from_date')
        to_date = self.request.GET.get('to_date')

        # Default to current month
        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = date.fromisoformat(from_date)

        if not to_date:
            # Last day of current month
            today = date.today()
            if today.month == 12:
                to_date = today.replace(day=31)
            else:
                to_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        else:
            to_date = date.fromisoformat(to_date)

        context['from_date'] = from_date
        context['to_date'] = to_date

        # Calculate counts for each template
        template_data = []
        total_patients = 0

        for template in self.get_queryset():
            count = template.get_patient_count(from_date, to_date)
            template_data.append({
                'template': template,
                'count': count
            })
            total_patients += count

        context['template_data'] = template_data
        context['total_patients'] = total_patients

        return context


class RegistrationTemplateCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create registration report template"""
    model = RegistrationReportTemplate
    form_class = RegistrationReportTemplateForm
    template_name = 'patient/reports/registration_template_form.html'
    permission_required = 'patient.add_patientmodel'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, f'Template "{form.instance.title}" created successfully.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('registration_report')


class RegistrationTemplateUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update registration report template"""
    model = RegistrationReportTemplate
    form_class = RegistrationReportTemplateForm
    template_name = 'patient/reports/registration_template_form.html'
    permission_required = 'patient.add_patientmodel'

    def form_valid(self, form):
        messages.success(self.request, f'Template "{form.instance.title}" updated successfully.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('registration_report')


class RegistrationTemplateDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete registration report template"""
    model = RegistrationReportTemplate
    template_name = 'patient/reports/registration_template_delete.html'
    permission_required = 'patient.add_patientmodel'

    def get_success_url(self):
        messages.success(self.request, 'Template deleted successfully.')
        return reverse('registration_report')


class RegistrationReportExportPDFView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Export registration report to PDF"""
    permission_required = 'patient.view_patientmodel'

    def get(self, request, *args, **kwargs):
        # Get date range
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')

        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = date.fromisoformat(from_date)

        if not to_date:
            today = date.today()
            if today.month == 12:
                to_date = today.replace(day=31)
            else:
                to_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        else:
            to_date = date.fromisoformat(to_date)

        # Get templates and counts
        templates = RegistrationReportTemplate.objects.all()
        template_data = []
        total_patients = 0

        for template in templates:
            count = template.get_patient_count(from_date, to_date)
            template_data.append({
                'title': template.title,
                'count': count
            })
            total_patients += count

        # Get hospital info
        site_info = SiteInfoModel.objects.first()

        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30,
                                topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()

        title_style = styles['Heading1']
        title_style.alignment = 1  # Center

        # Header
        if site_info:
            elements.append(Paragraph(site_info.name, title_style))

        elements.append(Spacer(1, 0.3 * inch))
        elements.append(Paragraph("REGISTRATION REPORT", title_style))
        elements.append(Paragraph(
            f"Period: {from_date.strftime('%B %d, %Y')} - {to_date.strftime('%B %d, %Y')}",
            styles['Normal']
        ))
        elements.append(Spacer(1, 0.4 * inch))

        # Build table
        table_data = [['S/N', 'Template Title', 'Total Patients']]

        for idx, data in enumerate(template_data, 1):
            table_data.append([
                str(idx),
                data['title'],
                str(data['count'])
            ])

        # Total row
        table_data.append(['', 'TOTAL REGISTRATIONS', str(total_patients)])

        # Create table
        table = Table(table_data, colWidths=[0.6 * inch, 4 * inch, 1.5 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f0f0f0')]),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f0f0f0')),
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#4472C4')),
        ]))

        elements.append(table)
        doc.build(elements)
        buffer.seek(0)

        # Response
        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f"registration_report_{from_date.strftime('%Y%m%d')}_{to_date.strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ========================= CONSULTATION REPORT TEMPLATES =========================
class ConsultationTemplateListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """List consultation report templates with counts"""
    model = ConsultationReportTemplate
    template_name = 'patient/reports/consultation_templates.html'
    context_object_name = 'templates'
    permission_required = 'patient.view_patientmodel'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get date range
        from_date = self.request.GET.get('from_date')
        to_date = self.request.GET.get('to_date')

        # Default to current month
        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = date.fromisoformat(from_date)

        if not to_date:
            today = date.today()
            if today.month == 12:
                to_date = today.replace(day=31)
            else:
                to_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        else:
            to_date = date.fromisoformat(to_date)

        context['from_date'] = from_date
        context['to_date'] = to_date

        # Calculate counts for each template
        template_data = []
        total_cases = 0

        for template in self.get_queryset():
            count = template.get_consultation_count(from_date, to_date)
            template_data.append({
                'template': template,
                'count': count
            })
            total_cases += count

        context['template_data'] = template_data
        context['total_cases'] = total_cases

        return context


class ConsultationTemplateCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create consultation report template"""
    model = ConsultationReportTemplate
    form_class = ConsultationReportTemplateForm
    template_name = 'patient/reports/consultation_template_form.html'
    permission_required = 'patient.add_patientmodel'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, f'Template "{form.instance.title}" created successfully.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('consultation_templates')


class ConsultationTemplateUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update consultation report template"""
    model = ConsultationReportTemplate
    form_class = ConsultationReportTemplateForm
    template_name = 'patient/reports/consultation_template_form.html'
    permission_required = 'patient.add_patientmodel'

    def form_valid(self, form):
        messages.success(self.request, f'Template "{form.instance.title}" updated successfully.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('consultation_templates')


class ConsultationTemplateDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete consultation report template"""
    model = ConsultationReportTemplate
    template_name = 'patient/reports/consultation_template_delete.html'
    permission_required = 'patient.add_patientmodel'

    def get_success_url(self):
        messages.success(self.request, 'Template deleted successfully.')
        return reverse('consultation_templates')


class ConsultationTemplateExportPDFView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Export consultation template report to PDF"""
    permission_required = 'patient.view_patientmodel'

    def get(self, request, *args, **kwargs):
        # Get date range
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')

        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = date.fromisoformat(from_date)

        if not to_date:
            today = date.today()
            if today.month == 12:
                to_date = today.replace(day=31)
            else:
                to_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        else:
            to_date = date.fromisoformat(to_date)

        # Get templates and counts
        templates = ConsultationReportTemplate.objects.all()
        template_data = []
        total_cases = 0

        for template in templates:
            count = template.get_consultation_count(from_date, to_date)
            template_data.append({
                'title': template.title,
                'count': count
            })
            total_cases += count

        # Get hospital info
        site_info = SiteInfoModel.objects.first()

        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30,
                                topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()

        title_style = styles['Heading1']
        title_style.alignment = 1  # Center

        # Header
        if site_info:
            elements.append(Paragraph(site_info.name, title_style))

        elements.append(Spacer(1, 0.3 * inch))
        elements.append(Paragraph("CONSULTATION DIAGNOSIS REPORT", title_style))
        elements.append(Paragraph(
            f"Period: {from_date.strftime('%B %d, %Y')} - {to_date.strftime('%B %d, %Y')}",
            styles['Normal']
        ))
        elements.append(Spacer(1, 0.4 * inch))

        # Build table
        table_data = [['S/N', 'Template Title', 'Total Cases']]

        for idx, data in enumerate(template_data, 1):
            table_data.append([
                str(idx),
                data['title'],
                str(data['count'])
            ])

        # Total row
        table_data.append(['', 'TOTAL CASES', str(total_cases)])

        # Create table
        table = Table(table_data, colWidths=[0.6 * inch, 4 * inch, 1.5 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f0f0f0')]),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f0f0f0')),
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#4472C4')),
        ]))

        elements.append(table)
        doc.build(elements)
        buffer.seek(0)

        # Response
        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f"consultation_diagnosis_report_{from_date.strftime('%Y%m%d')}_{to_date.strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
