# views.py
import calendar
from datetime import timedelta, date, datetime
from decimal import Decimal, ROUND_HALF_UP

from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.models import User
from django.contrib.messages.views import SuccessMessageMixin
from django.core import serializers
from django.db import transaction
from django import forms
from django.db.models.functions import TruncMonth
from django.forms import modelformset_factory
from django.shortcuts import redirect, get_object_or_404, render
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.timezone import now
from django.views.decorators.http import require_http_methods
from django.views.generic import CreateView, ListView, DetailView, UpdateView, DeleteView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.db.models import Q, Sum, Avg, F, Count, ExpressionWrapper, DecimalField, Prefetch
from django.contrib import messages
from django.http import HttpResponse, JsonResponse, HttpResponseBadRequest
from django.template.loader import render_to_string
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from admin_site.models import SiteInfoModel
from consultation.models import ConsultationFeeModel, SpecializationModel, PatientQueueModel
from finance.forms import FinanceSettingForm, ExpenseCategoryForm, \
    SalaryStructureForm, StaffBankDetailForm, IncomeForm, IncomeCategoryForm, \
    ExpenseForm, PaysheetRowForm, MoneyRemittanceForm, OtherPaymentServiceForm, OtherPaymentForm
from finance.models import PatientTransactionModel, FinanceSettingModel, ExpenseCategory, SalaryStructure, \
    StaffBankDetail, SalaryRecord, Income, IncomeCategory, Expense, MoneyRemittance, OtherPaymentService, \
    WalletWithdrawalRecord
from human_resource.models import DepartmentModel, StaffModel
from human_resource.views import FlashFormErrorsMixin
from inpatient.models import Admission, Surgery
from insurance.models import PatientInsuranceModel
from laboratory.models import LabTestOrderModel
from patient.forms import RegistrationPaymentForm
from patient.models import RegistrationPaymentModel, RegistrationFeeModel, PatientModel, PatientWalletModel

import json
from django.core.serializers.json import DjangoJSONEncoder
import uuid

from pharmacy.models import DrugOrderModel
from scan.models import ScanOrderModel
from service.models import PatientServiceTransaction


class RegistrationPaymentCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = RegistrationPaymentModel
    form_class = RegistrationPaymentForm
    template_name = "finance/registration_payment/create.html"
    permission_required = "finance.add_patienttransactionmodel"

    def get_success_url(self):
        return reverse('registration_payment_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        # Generate transaction ID
        if not form.instance.transaction_id:
            form.instance.transaction_id = f"REG{uuid.uuid4().hex[:8].upper()}"

        # Calculate total amount
        total = form.instance.registration_fee.amount if form.instance.registration_fee else 0
        if form.instance.consultation_fee:
            form.instance.consultation_paid = True
            form.instance.consultation_amount = form.instance.consultation_fee.amount
            total += form.instance.consultation_amount

        form.instance.amount = total
        messages.success(self.request, f"Payment recorded successfully! Transaction ID: {form.instance.transaction_id}")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Fetch registration fees
        reg_fees = RegistrationFeeModel.objects.all().values('id', 'title', 'amount', 'patient_type')
        context['registration_fees'] = json.dumps(list(reg_fees), cls=DjangoJSONEncoder)

        # Consultation fees for the select
        context['consultation_fees'] = ConsultationFeeModel.objects.filter(patient_category='regular')
        return context


class RegistrationPaymentDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = RegistrationPaymentModel
    template_name = "finance/registration_payment/detail.html"
    context_object_name = "payment"
    permission_required = "finance.view_patienttransactionmodel"


class RegistrationPaymentListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = RegistrationPaymentModel
    template_name = "finance/registration_payment/index.html"
    context_object_name = "payments"
    permission_required = "finance.view_patienttransactionmodel"

    def get_queryset(self):
        qs = RegistrationPaymentModel.objects.select_related('registration_fee', 'consultation_fee', 'created_by')

        # Default: only today's payments
        start_date = self.request.GET.get("start_date")
        end_date = self.request.GET.get("end_date")
        search = self.request.GET.get("search")

        if not start_date and not end_date:
            today = now().date()
            qs = qs.filter(date=today)

        if start_date and end_date:
            qs = qs.filter(date__range=[start_date, end_date])

        if search:
            qs = qs.filter(Q(full_name__icontains=search) | Q(old_card_number__icontains=search) | Q(
                transaction_id__icontains=search))

        return qs.order_by("-created_at")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Determine timeframe for title
        start_date = self.request.GET.get("start_date")
        end_date = self.request.GET.get("end_date")

        if start_date and end_date:
            if start_date == end_date:
                context['timeframe'] = f"for {start_date}"
            else:
                context['timeframe'] = f"from {start_date} to {end_date}"
        else:
            context['timeframe'] = "for Today"

        # Calculate total amount
        payments = context['payments']
        context['total_amount'] = sum(payment.amount for payment in payments)

        return context


@login_required
def print_receipt(request, pk):
    """Generate printable POS receipt"""
    payment = get_object_or_404(RegistrationPaymentModel, pk=pk)

    # Check permission
    if not request.user.has_perm('patient.add_patienttransactionmodel'):
        messages.error(request, "You don't have permission to view this receipt.")
        return redirect('registration_payment_index')

    context = {
        'payment': payment,
        'site_info': SiteInfoModel.objects.first(),
        'current_user': request.user,
        'print_time': now(),
    }

    html = render_to_string('finance/registration_payment/receipt.html', context)

    # Return as HTML for printing
    response = HttpResponse(html)
    response['Content-Type'] = 'text/html'
    return response


@login_required
@permission_required('finance.add_patienttransactionmodel', raise_exception=True)
def patient_wallet_funding(request):
    """Initial page for patient wallet funding"""
    context = {
        'finance_setting': FinanceSettingModel.objects.first()
    }
    try:
        # optional amount to prefill
        if 'amount' in request.GET:
            try:
                context['amount'] = float(request.GET.get('amount'))
            except (TypeError, ValueError):
                context['amount'] = None

        # optional redirect target (kept as string)
        if 'redirect' in request.GET:
            context['redirect'] = request.GET.get('redirect')

        # optional patient_id: load patient and pass to template
        if 'patient_id' in request.GET:
            patient_id = request.GET.get('patient_id')
            # get_object_or_404 will raise 404 if invalid
            patient = get_object_or_404(PatientModel, pk=patient_id)
            context['patient'] = patient

    except Exception:
        # keep old behavior if anything goes wrong
        pass

    return render(request, 'finance/wallet/funding.html', context)


def calculate_insurance_amount(base_amount, coverage_percentage):
    """Calculate patient's portion after insurance"""
    if coverage_percentage and coverage_percentage > 0:
        covered_amount = base_amount * (coverage_percentage / 100)
        # Ensure result is rounded correctly and not negative
        return max(Decimal('0.00'), base_amount - covered_amount).quantize(Decimal('0.01'))
    return base_amount


@login_required
def verify_patient_ajax(request):
    """Verify patient by card number and return wallet details with pending payments"""
    from insurance.claim_helpers import get_orders_with_claim_info, get_pending_claim_for_order

    card_number = request.GET.get('card_number', '').strip()

    if not card_number:
        return JsonResponse({'error': 'Card number required'}, status=400)

    try:
        # Look up patient by card number
        patient = PatientModel.objects.get(card_number__iexact=card_number)

        # Get or create wallet
        wallet, created = PatientWalletModel.objects.get_or_create(
            patient=patient,
            defaults={'amount': Decimal('0.00')}
        )

        # Calculate pending payments for last 30 days
        thirty_days_ago = timezone.now() - timedelta(days=30)

        # --- QUERIES ---

        # 1. Get pending Service/Item Transactions
        pending_services_items = PatientServiceTransaction.objects.filter(
            patient=patient,
            status='pending_payment',
            created_at__gte=thirty_days_ago
        ).select_related(
            'service', 'service__category',
            'service_item', 'service_item__category'
        )

        # 2. Get pending drug orders
        pending_drugs = DrugOrderModel.objects.filter(
            patient=patient,
            status__in=['pending'],
            ordered_at__gte=thirty_days_ago
        ).select_related('drug')

        # 3. Get pending lab orders
        pending_labs = LabTestOrderModel.objects.filter(
            patient=patient,
            status__in=['pending'],
            ordered_at__gte=thirty_days_ago
        ).select_related('template')

        # 4. Get pending scan orders
        pending_scans = ScanOrderModel.objects.filter(
            patient=patient,
            status__in=['pending'],
            ordered_at__gte=thirty_days_ago
        ).select_related('template')

        # --- INSURANCE CHECK ---
        active_insurance = None
        if hasattr(patient, 'insurance_policies'):
            active_insurance = patient.insurance_policies.filter(
                is_active=True,
                valid_to__gte=date.today()
            ).select_related('hmo', 'coverage_plan').first()

        # --- PROCESSING WITH CLAIM-BASED LOGIC ---

        # 1. Process Service/Item Orders
        service_items_list = []
        service_total = Decimal('0.00')

        for transaction in pending_services_items:
            # Determine the name and category
            category_obj = None
            if transaction.service:
                name = transaction.service.name
                category_obj = transaction.service.category
            elif transaction.service_item:
                name = transaction.service_item.name
                category_obj = transaction.service_item.category
            else:
                continue

            base_amount = transaction.total_amount

            # Get claim info
            pending_claim = get_pending_claim_for_order(transaction)

            # For services, use base amount (no auto claims yet in your system)
            patient_amount = base_amount

            service_items_list.append({
                'id': transaction.id,
                'name': f"{category_obj.name} - {name}",
                'quantity': transaction.quantity,
                'base_amount': float(base_amount),
                'patient_amount': float(patient_amount),
                'covered_amount': 0.00,
                'status': transaction.status,
                'ordered_date': transaction.created_at.strftime('%Y-%m-%d'),
                'has_approved_claim': False,
                'claim_number': None,
                'claim_status': None,
                'has_pending_claim': pending_claim['has_pending_claim'],
                'pending_claim_number': pending_claim['claim_number'],
                'insurance_applied': False
            })
            service_total += patient_amount

        # 2. Process drug orders with claim-based logic
        drug_results = get_orders_with_claim_info(pending_drugs, 'drug')
        drug_items = []
        drug_total = Decimal('0.00')

        for result in drug_results:
            order = result['order']
            drug_items.append({
                'id': order.id,
                'name': f"{order.drug.__str__()} (x{float(order.quantity_ordered)})",
                'quantity': float(order.quantity_ordered),
                'base_amount': float(result['base_amount']),
                'patient_amount': float(result['patient_amount']),
                'covered_amount': float(result['covered_amount']),
                'order_number': order.order_number,
                'status': order.status,
                'ordered_date': order.ordered_at.strftime('%Y-%m-%d'),
                'has_approved_claim': result['has_approved_claim'],
                'claim_number': result['claim_number'],
                'claim_status': result['claim_status'],
                'has_pending_claim': result['has_pending_claim'],
                'pending_claim_number': result['pending_claim_number'],
                'insurance_applied': result['has_approved_claim']
            })
            drug_total += result['patient_amount']

        # 3. Process lab orders with claim-based logic
        lab_results = get_orders_with_claim_info(pending_labs, 'lab')
        lab_items = []
        lab_total = Decimal('0.00')

        for result in lab_results:
            order = result['order']
            lab_items.append({
                'id': order.id,
                'name': order.template.name,
                'base_amount': float(result['base_amount']),
                'patient_amount': float(result['patient_amount']),
                'covered_amount': float(result['covered_amount']),
                'order_number': order.order_number,
                'status': order.status,
                'ordered_date': order.ordered_at.strftime('%Y-%m-%d'),
                'has_approved_claim': result['has_approved_claim'],
                'claim_number': result['claim_number'],
                'claim_status': result['claim_status'],
                'has_pending_claim': result['has_pending_claim'],
                'pending_claim_number': result['pending_claim_number'],
                'insurance_applied': result['has_approved_claim']
            })
            lab_total += result['patient_amount']

        # 4. Process scan orders with claim-based logic
        scan_results = get_orders_with_claim_info(pending_scans, 'scan')
        scan_items = []
        scan_total = Decimal('0.00')

        for result in scan_results:
            order = result['order']
            scan_items.append({
                'id': order.id,
                'name': order.template.name,
                'base_amount': float(result['base_amount']),
                'patient_amount': float(result['patient_amount']),
                'covered_amount': float(result['covered_amount']),
                'order_number': order.order_number,
                'status': order.status,
                'ordered_date': order.ordered_at.strftime('%Y-%m-%d'),
                'has_approved_claim': result['has_approved_claim'],
                'claim_number': result['claim_number'],
                'claim_status': result['claim_status'],
                'has_pending_claim': result['has_pending_claim'],
                'pending_claim_number': result['pending_claim_number'],
                'insurance_applied': result['has_approved_claim']
            })
            scan_total += result['patient_amount']

        # Calculate grand total
        grand_total = drug_total + lab_total + scan_total + service_total

        # Active Admissions and Surgeries
        active_admissions = Admission.objects.filter(
            patient=patient,
            status='active'
        )
        admission_count = active_admissions.count()

        active_surgeries = Surgery.objects.filter(
            patient=patient,
            status__in=['scheduled', 'in_progress']
        )
        surgery_count = active_surgeries.count()

        return JsonResponse({
            'success': True,
            'patient': {
                'id': patient.id,
                'full_name': str(patient),
                'card_number': patient.card_number,
                'phone': getattr(patient, 'mobile', ''),
                'email': getattr(patient, 'email', ''),
                'age': patient.age() if hasattr(patient, 'age') and callable(patient.age) else '',
                'gender': getattr(patient, 'gender', ''),
            },
            'wallet': {
                'balance': float(wallet.amount),
                'formatted_balance': f'₦{wallet.amount:,.2f}'
            },
            'insurance': {
                'has_insurance': bool(active_insurance),
                'hmo_name': active_insurance.hmo.name if active_insurance else None,
                'plan_name': active_insurance.coverage_plan.name if active_insurance else None,
                'policy_number': active_insurance.policy_number if active_insurance else None,
            },
            'pending_payments': {
                'drugs': {
                    'items': drug_items,
                    'total': float(drug_total),
                    'count': len(drug_items)
                },
                'labs': {
                    'items': lab_items,
                    'total': float(lab_total),
                    'count': len(lab_items)
                },
                'scans': {
                    'items': scan_items,
                    'total': float(scan_total),
                    'count': len(scan_items)
                },
                'services': {
                    'items': service_items_list,
                    'total': float(service_total),
                    'count': len(service_items_list)
                },
                'admissions': {'count': admission_count, 'total': 0},
                'surgeries': {'count': surgery_count, 'total': 0},
                'grand_total': float(grand_total),
                'formatted_grand_total': f'₦{grand_total:,.2f}'
            }
        })

    except PatientModel.DoesNotExist:
        return JsonResponse({
            'error': 'Patient not found with this card number'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'error': f'Error verifying patient: {type(e).__name__}: {str(e)}'
        }, status=500)


@login_required
@permission_required('finance.add_patienttransactionmodel', raise_exception=True)
@require_http_methods(["POST"])
def process_wallet_funding(request):
    """Process wallet funding and optional payments"""
    try:
        # Get form data
        patient_id = request.POST.get('patient_id')
        funding_amount = Decimal(request.POST.get('funding_amount', '0'))
        payment_method = request.POST.get('payment_method', 'cash')
        action_type = request.POST.get('action_type', 'fund_only')  # 'fund_only' or 'fund_and_pay'

        # Validate inputs
        if not patient_id:
            return JsonResponse({'error': 'Patient ID required'}, status=400)

        if funding_amount <= 0:
            return JsonResponse({'error': 'Funding amount must be greater than 0'}, status=400)

        patient = get_object_or_404(PatientModel, id=patient_id)
        wallet, created = PatientWalletModel.objects.get_or_create(
            patient=patient,
            defaults={'amount': Decimal('0.00')}
        )

        patient_old_balance = wallet.amount

        with transaction.atomic():
            # Add funds to wallet
            wallet.amount += funding_amount
            wallet.save()

            patient_new_balance = wallet.amount

            PatientTransactionModel.objects.create(
                patient=patient,
                transaction_type='wallet_funding',
                transaction_direction='in',
                amount=funding_amount,
                old_balance=patient_old_balance,
                new_balance=patient_new_balance,
                payment_method=payment_method,
                received_by=request.user,
                status='completed',
                date=timezone.now().date()
            )

            if action_type == 'fund_only':
                messages.success(request, f'Successfully funded wallet with ₦{funding_amount:,.2f}')
                return JsonResponse({
                    'success': True,
                    'message': f'Wallet funded with ₦{funding_amount:,.2f}',
                    'new_balance': float(wallet.amount),
                    'redirect_url': reverse('patient_wallet_dashboard', args=[patient.id])
                })

            elif action_type == 'fund_and_pay':
                payment_type = request.POST.get('payment_type')

                if not payment_type:
                    return JsonResponse({
                        'error': 'Payment type and items selection required'
                    }, status=400)

                # Redirect to appropriate payment page
                redirect_urls = {
                    'consultation': reverse('finance_consultation_patient_payment', args=[patient.id]),
                    'lab': reverse('finance_laboratory_patient_payment', args=[patient.id]),
                    'scan': reverse('finance_scan_patient_payment', args=[patient.id]),
                    'drug': reverse('finance_pharmacy_patient_payment', args=[patient.id]),
                    'inpatient': reverse('finance_admission_funding', args=[patient.id]),
                    'service': reverse('finance_service_patient_payment', args=[patient.id]),
                }

                redirect_url = redirect_urls.get(payment_type)
                if not redirect_url:
                    return JsonResponse({'error': 'Invalid payment type'}, status=400)

                messages.success(request, f'Wallet funded with ₦{funding_amount:,.2f}. Proceeding to payment...')
                return JsonResponse({
                    'success': True,
                    'message': f'Wallet funded. Redirecting to {payment_type} payment...',
                    'new_balance': float(wallet.amount),
                    'redirect_url': redirect_url
                })

    except PatientModel.DoesNotExist:
        return JsonResponse({'error': 'Patient not found'}, status=404)
    except ValueError:
        return JsonResponse({'error': 'Invalid funding amount'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Error processing funding: {str(e)}'}, status=500)


@login_required
def calculate_payment_total_ajax(request):
    """Calculate total based on selected items - CLAIM-BASED VERSION"""
    from insurance.claim_helpers import calculate_patient_amount_with_claim

    try:
        patient_id = request.GET.get('patient_id')
        selected_drugs = request.GET.getlist('drugs[]')
        selected_labs = request.GET.getlist('labs[]')
        selected_scans = request.GET.getlist('scans[]')

        if not patient_id:
            return JsonResponse({'error': 'Patient ID required'}, status=400)

        patient = get_object_or_404(PatientModel, id=patient_id)

        # Get active insurance for display
        active_insurance = None
        if hasattr(patient, 'insurance_policies'):
            active_insurance = patient.insurance_policies.filter(
                is_active=True,
                valid_to__gte=date.today()
            ).select_related('coverage_plan').first()

        total_amount = Decimal('0.00')
        items_breakdown = {
            'drugs': [],
            'labs': [],
            'scans': []
        }

        # Calculate drug totals using claims
        if selected_drugs:
            drug_orders = DrugOrderModel.objects.filter(
                id__in=selected_drugs,
                patient=patient
            ).select_related('drug')

            for order in drug_orders:
                base_amount = (
                    order.drug.selling_price * Decimal(str(order.quantity_ordered))
                    if hasattr(order, 'drug') and hasattr(order.drug, 'selling_price')
                    else Decimal('0.00')
                )

                # Use claim-based calculation
                claim_info = calculate_patient_amount_with_claim(order, base_amount)
                patient_amount = claim_info['patient_amount']

                total_amount += patient_amount
                items_breakdown['drugs'].append({
                    'id': order.id,
                    'base_amount': float(base_amount),
                    'patient_amount': float(patient_amount),
                    'covered_amount': float(claim_info['covered_amount']),
                    'has_claim': claim_info['has_approved_claim'],
                    'claim_number': claim_info['claim_number']
                })

        # Calculate lab totals using claims
        if selected_labs:
            lab_orders = LabTestOrderModel.objects.filter(
                id__in=selected_labs,
                patient=patient
            ).select_related('template')

            for order in lab_orders:
                base_amount = order.amount_charged or order.template.price

                # Use claim-based calculation
                claim_info = calculate_patient_amount_with_claim(order, base_amount)
                patient_amount = claim_info['patient_amount']

                total_amount += patient_amount
                items_breakdown['labs'].append({
                    'id': order.id,
                    'base_amount': float(base_amount),
                    'patient_amount': float(patient_amount),
                    'covered_amount': float(claim_info['covered_amount']),
                    'has_claim': claim_info['has_approved_claim'],
                    'claim_number': claim_info['claim_number']
                })

        # Calculate scan totals using claims
        if selected_scans:
            scan_orders = ScanOrderModel.objects.filter(
                id__in=selected_scans,
                patient=patient
            ).select_related('template')

            for order in scan_orders:
                base_amount = order.amount_charged or order.template.price

                # Use claim-based calculation
                claim_info = calculate_patient_amount_with_claim(order, base_amount)
                patient_amount = claim_info['patient_amount']

                total_amount += patient_amount
                items_breakdown['scans'].append({
                    'id': order.id,
                    'base_amount': float(base_amount),
                    'patient_amount': float(patient_amount),
                    'covered_amount': float(claim_info['covered_amount']),
                    'has_claim': claim_info['has_approved_claim'],
                    'claim_number': claim_info['claim_number']
                })

        # Calculate summary statistics
        total_claims_applied = sum(
            1 for item_list in items_breakdown.values()
            for item in item_list
            if item['has_claim']
        )

        total_covered = sum(
            item['covered_amount']
            for item_list in items_breakdown.values()
            for item in item_list
        )

        return JsonResponse({
            'success': True,
            'total_amount': float(total_amount),
            'formatted_total': f'₦{total_amount:,.2f}',
            'insurance_info': {
                'has_insurance': bool(active_insurance),
                'claims_applied': total_claims_applied,
                'total_covered': float(total_covered),
                'formatted_covered': f'₦{total_covered:,.2f}'
            },
            'breakdown': items_breakdown
        })

    except PatientModel.DoesNotExist:
        return JsonResponse({'error': 'Patient not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error calculating total: {str(e)}'}, status=500)


@login_required
@permission_required('finance.view_patienttransactionmodel', raise_exception=True)
def patient_wallet_dashboard(request, patient_id):
    """Patient wallet dashboard showing balance and recent transactions"""
    patient = get_object_or_404(PatientModel, id=patient_id)
    wallet, created = PatientWalletModel.objects.get_or_create(
        patient=patient,
        defaults={'amount': Decimal('0.00')}
    )

    # Filter transactions for the patient
    patient_transactions = PatientTransactionModel.objects.filter(patient=patient)

    # Get the current month and year
    current_month = timezone.now().month
    current_year = timezone.now().year

    # Aggregate funded (IN) and spent (OUT) amounts for the current month
    monthly_summary = patient_transactions.filter(
        created_at__year=current_year,
        created_at__month=current_month
    ).aggregate(
        funded_amount=Sum('amount', filter=Q(transaction_direction='in')),
        spent_amount=Sum('amount', filter=Q(transaction_direction='out'))
    )

    # Use 0 if the sum is None (no transactions found)
    funded_amount = monthly_summary.get('funded_amount', Decimal('0.00')) or Decimal('0.00')
    spent_amount = monthly_summary.get('spent_amount', Decimal('0.00')) or Decimal('0.00')

    # Get recent transactions (e.g., last 20) for the patient
    recent_transactions = patient_transactions.order_by('-created_at')[:20]

    context = {
        'patient': patient,
        'wallet': wallet,
        'wallet_balance': wallet.amount,
        'recent_transactions': recent_transactions,
        'funded_amount': funded_amount,
        'spent_amount': spent_amount,
    }

    return render(request, 'finance/wallet/dashboard.html', context)


@login_required
@permission_required('finance.add_patienttransactionmodel', raise_exception=True)
@require_http_methods(["POST"])
def process_wallet_payment(request):
    """Process payment from wallet for selected services"""
    try:
        patient_id = request.POST.get('patient_id')
        payment_type = request.POST.get('payment_type')
        selected_items = request.POST.getlist('selected_items[]')

        if not all([patient_id, payment_type, selected_items]):
            return JsonResponse({
                'error': 'Missing required payment information'
            }, status=400)

        patient = get_object_or_404(PatientModel, id=patient_id)
        wallet = get_object_or_404(PatientWalletModel, patient=patient)

        # Get active insurance
        active_insurance = patient.insurance_policies.filter(
            is_active=True,
            valid_to__gte=date.today()
        ).select_related('coverage_plan').first()

        def calculate_patient_amount(base_amount, coverage_percentage):
            if coverage_percentage and coverage_percentage > 0:
                covered_amount = base_amount * (coverage_percentage / 100)
                return base_amount - covered_amount
            return base_amount

        total_to_pay = Decimal('0.00')
        payment_items = []

        with transaction.atomic():
            # Process based on payment type
            if payment_type == 'drug':
                orders = DrugOrderModel.objects.filter(
                    id__in=selected_items,
                    patient=patient,
                    status='pending'
                ).select_related('drug')

                for order in orders:
                    base_amount = order.amount_charged or order.drug.selling_price

                    if active_insurance and active_insurance.coverage_plan.is_drug_covered(order.drug):
                        patient_amount = calculate_patient_amount(
                            base_amount,
                            active_insurance.coverage_plan.drug_coverage_percentage
                        )
                    else:
                        patient_amount = base_amount

                    total_to_pay += patient_amount
                    payment_items.append({
                        'order': order,
                        'amount': patient_amount
                    })

            elif payment_type == 'lab':
                orders = LabTestOrderModel.objects.filter(
                    id__in=selected_items,
                    patient=patient,
                    status='pending'
                ).select_related('template')

                for order in orders:
                    base_amount = order.amount_charged or order.template.price

                    if active_insurance and active_insurance.coverage_plan.is_lab_covered(order.template):
                        patient_amount = calculate_patient_amount(
                            base_amount,
                            active_insurance.coverage_plan.lab_coverage_percentage
                        )
                    else:
                        patient_amount = base_amount

                    total_to_pay += patient_amount
                    payment_items.append({
                        'order': order,
                        'amount': patient_amount
                    })

            elif payment_type == 'scan':
                orders = ScanOrderModel.objects.filter(
                    id__in=selected_items,
                    patient=patient,
                    status='pending'
                ).select_related('template')

                for order in orders:
                    base_amount = order.amount_charged or order.template.price

                    if active_insurance and active_insurance.coverage_plan.is_radiology_covered(order.template):
                        patient_amount = calculate_patient_amount(
                            base_amount,
                            active_insurance.coverage_plan.radiology_coverage_percentage
                        )
                    else:
                        patient_amount = base_amount

                    total_to_pay += patient_amount
                    payment_items.append({
                        'order': order,
                        'amount': patient_amount
                    })

            # Check if wallet has sufficient funds
            if wallet.amount < total_to_pay:
                return JsonResponse({
                    'error': f'Insufficient wallet balance. Available: ₦{wallet.amount:,.2f}, Required: ₦{total_to_pay:,.2f}'
                }, status=400)

            # Process payment
            wallet.amount -= total_to_pay
            wallet.save()

            # Update order statuses
            for item in payment_items:
                order = item['order']


                if payment_type in ['lab', 'scan']:
                    order.status = 'paid'
                elif payment_type == 'drug':
                    order.status = 'paid'

                order.save()

            # Create payment record (if you have a payment model)
            # PaymentModel.objects.create(
            #     patient=patient,
            #     payment_type=payment_type,
            #     amount=total_to_pay,
            #     payment_method='wallet',
            #     processed_by=request.user,
            #     items_paid=selected_items
            # )

        messages.success(
            request,
            f'Payment of ₦{total_to_pay:,.2f} processed successfully from wallet. '
            f'New balance: ₦{wallet.amount:,.2f}'
        )

        return JsonResponse({
            'success': True,
            'message': f'Payment of ₦{total_to_pay:,.2f} processed successfully',
            'new_balance': float(wallet.amount),
            'formatted_balance': f'₦{wallet.amount:,.2f}',
            'redirect_url': reverse('wallet:patient_dashboard', args=[patient.id])
        })

    except Exception as e:
        return JsonResponse({
            'error': f'Error processing payment: {str(e)}'
        }, status=500)


THIRTY_DAYS = 30


def _to_decimal(v):
    if isinstance(v, Decimal):
        return v
    try:
        return Decimal(str(v))
    except Exception:
        return Decimal('0.00')


def _quantize_money(d: Decimal) -> Decimal:
    return d.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)


@login_required
@permission_required('finance.add_patienttransactionmodel', raise_exception=True)
def finance_payment_select(request):
    """
    Initial payment selection page.
    The patient is verified by AJAX (finance_verify_patient_ajax) and the page
    will show radio options (consultation, drug, lab, scan) with counts & totals.
    """
    return render(request, 'finance/payment/select.html', {})


@login_required
def finance_verify_patient_ajax(request):
    """
    AJAX endpoint to verify patient and return pending payments summary
    UPDATED: Uses claim-based insurance calculation
    """
    from insurance.claim_helpers import get_orders_with_claim_info

    card_number = request.GET.get('card_number', '').strip()
    if not card_number:
        return JsonResponse({'success': False, 'error': 'Card number is required'})

    try:
        # Find patient by card number
        patient = PatientModel.objects.get(card_number=card_number)
    except PatientModel.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Patient not found with this card number'})

    # Get or create wallet
    try:
        wallet = patient.wallet
        has_wallet = True
        balance = float(wallet.amount)
        formatted_balance = f'₦{wallet.amount:,.2f}'
    except:
        from patient.models import PatientWalletModel
        wallet = PatientWalletModel.objects.create(patient=patient, amount=Decimal('0.00'))
        has_wallet = True
        balance = 0.00
        formatted_balance = '₦0.00'

    # Get active insurance
    active_insurance = None
    try:
        policies_qs = patient.insurance_policies.all()
    except:
        try:
            policies_qs = patient.insurancepolicy_set.all()
        except:
            policies_qs = PatientInsuranceModel.objects.none()

    active_insurance = policies_qs.filter(
        is_active=True,
        valid_to__gte=timezone.now().date()
    ).select_related('hmo', 'coverage_plan').first()

    # Calculate pending payments for last 30 days
    thirty_days_ago = timezone.now() - timedelta(days=THIRTY_DAYS)

    # Get pending orders
    pending_drugs = DrugOrderModel.objects.filter(
        patient=patient,
        status__in=['pending'],
        ordered_at__gte=thirty_days_ago
    ).select_related('drug')

    pending_labs = LabTestOrderModel.objects.filter(
        patient=patient,
        status__in=['pending', 'unpaid', 'awaiting_payment'],
        ordered_at__gte=thirty_days_ago
    ).select_related('template')

    pending_scans = ScanOrderModel.objects.filter(
        patient=patient,
        status__in=['pending', 'unpaid', 'awaiting_payment'],
        ordered_at__gte=thirty_days_ago
    ).select_related('template')

    # Process with claim-based logic
    drug_results = get_orders_with_claim_info(pending_drugs, 'drug')
    drug_total = sum(r['patient_amount'] for r in drug_results)
    drug_count = len(drug_results)

    lab_results = get_orders_with_claim_info(pending_labs, 'lab')
    lab_total = sum(r['patient_amount'] for r in lab_results)
    lab_count = len(lab_results)

    scan_results = get_orders_with_claim_info(pending_scans, 'scan')
    scan_total = sum(r['patient_amount'] for r in scan_results)
    scan_count = len(scan_results)

    grand_total = drug_total + lab_total + scan_total

    return JsonResponse({
        'success': True,
        'patient': {
            'id': patient.id,
            'full_name': f'{patient.first_name} {patient.last_name}',
            'card_number': patient.card_number,
            'phone': getattr(patient, 'phone', ''),
            'age': getattr(patient, 'age', None),
            'patient_id': patient.card_number
        },
        'wallet': {
            'has_wallet': has_wallet,
            'balance': balance,
            'formatted_balance': formatted_balance
        },
        'insurance': {
            'has_insurance': active_insurance is not None,
            'hmo_name': active_insurance.hmo.name if active_insurance else None,
            'plan_name': active_insurance.coverage_plan.name if active_insurance else None
        },
        'pending_payments': {
            'drugs': {'count': drug_count, 'total': float(drug_total)},
            'labs': {'count': lab_count, 'total': float(lab_total)},
            'scans': {'count': scan_count, 'total': float(scan_total)},
            'grand_total': float(grand_total)
        }
    })


class OtherPaymentView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    template_name = 'finance/payment/other_payment.html'
    permission_required = 'finance.view_patienttransactionmodel'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # We pass the empty form for the other fields (like amount)
        context['form'] = OtherPaymentForm()
        # We also pass the list of all active services directly
        context['services'] = OtherPaymentService.objects.filter(is_active=True)
        return context


@login_required
@permission_required('finance.add_patienttransactionmodel', raise_exception=True)
def process_other_payment_ajax(request):
    if request.method == 'POST':
        patient_id = request.POST.get('patient_id')
        patient = get_object_or_404(PatientModel, pk=patient_id)

        form = OtherPaymentForm(request.POST)

        if form.is_valid():
            # ... (successful payment logic remains the same) ...
            service = form.cleaned_data['other_service']
            amount = form.cleaned_data['amount']

            if service.is_fixed_amount and service.default_amount != amount:
                return JsonResponse({'success': False, 'error': 'The amount for this service is fixed.'}, status=400)

            wallet = get_object_or_404(PatientWalletModel, patient=patient)
            if wallet.amount < amount:
                return JsonResponse({'success': False, 'error': 'Insufficient wallet balance.'}, status=400)

            try:
                with transaction.atomic():
                    old_balance = wallet.amount
                    wallet.deduct_funds(amount)
                    new_transaction = PatientTransactionModel.objects.create(
                        patient=patient,
                        transaction_type='other_payment',
                        transaction_direction='out',
                        other_service=service,
                        amount=amount,
                        old_balance=old_balance,
                        new_balance=wallet.amount,
                        date=timezone.now().date(),
                        received_by=request.user,
                        payment_method='wallet',
                        status='completed'
                    )
                return JsonResponse({
                    'success': True,
                    'message': 'Payment successful!',
                    'redirect_url': reverse('transaction_detail', kwargs={'transaction_id': new_transaction.pk})
                })
            except Exception as e:
                return JsonResponse({'success': False, 'error': f'An error occurred: {str(e)}'}, status=400)

        else:
            # THIS IS THE CHANGE: Return a structured error response
            return JsonResponse({
                'success': False,
                'error': 'Please correct the errors below.',
                'errors': form.errors
            }, status=400)

    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=400)


class AdmissionSurgeryFundingView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    template_name = 'finance/payment/admission_surgery.html'
    permission_required = 'finance.add_patienttransactionmodel'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        patient = get_object_or_404(PatientModel, pk=self.kwargs.get('patient_id'))
        context['patient'] = patient
        context['wallet'], _ = PatientWalletModel.objects.get_or_create(patient=patient)

        funding_targets = []

        paid_statuses_drug = ['paid', 'partially_dispensed', 'dispensed']
        paid_statuses_lab_scan = ['paid', 'collected', 'processing', 'completed']

        # --- 1. Get and process active admissions ---
        active_admissions = Admission.objects.filter(patient=patient, status='active')
        for admission in active_admissions:
            drug_orders = admission.drug_orders.all()
            lab_orders = admission.lab_test_orders.all()
            scan_orders = admission.scan_orders.all()
            other_services = admission.service_drug_orders.all()

            # --- THIS IS THE FIX ---
            # Apply ExpressionWrapper to the drug cost calculation
            drug_costs = drug_orders.aggregate(
                total=Sum(
                    ExpressionWrapper(F('quantity_ordered') * F('drug__selling_price'), output_field=DecimalField()))
            )['total'] or 0

            lab_costs = lab_orders.aggregate(total=Sum('amount_charged'))['total'] or 0
            scan_costs = scan_orders.aggregate(total=Sum('amount_charged'))['total'] or 0
            other_services_costs = other_services.aggregate(total=Sum('total_amount'))['total'] or 0

            base_fee = (admission.admission_fee_charged or 0) + (admission.bed_fee_charged or 0)
            total_bill = base_fee + drug_costs + lab_costs + scan_costs + other_services_costs

            total_paid = PatientTransactionModel.objects.filter(
                admission=admission, status='completed',
            ).aggregate(total=Sum('amount'))['total'] or 0

            # The calculation was also missing here
            total_paid += sum(order.drug.selling_price * Decimal(order.quantity_ordered) for order in drug_orders if
                              order.status in paid_statuses_drug)
            total_paid += sum(order.amount_charged for order in lab_orders if order.status in paid_statuses_lab_scan)
            total_paid += sum(order.amount_charged for order in scan_orders if order.status in paid_statuses_lab_scan)
            total_paid += sum(service.total_amount for service in other_services if
                              service.status in ['paid', 'fully_dispensed', 'partially_dispensed'])

            balance = total_paid - total_bill

            funding_targets.append({
                'record': admission, 'type': 'Admission', 'identifier': admission.admission_number,
                'base_fee': base_fee,
                'drugs': {'total': drug_costs},
                'labs': {'total': lab_costs},
                'scans': {'total': scan_costs},
                'others': {'total': other_services_costs},
                'total_bill': total_bill, 'total_paid': total_paid,
                'balance': balance, 'abs_balance': abs(balance)
            })

        # --- 2. Get and process active surgeries (with the same fix) ---
        active_surgeries = Surgery.objects.filter(patient=patient, status__in=['scheduled', 'in_progress'])
        for surgery in active_surgeries:
            drug_orders = surgery.drug_orders.all()
            lab_orders = surgery.lab_test_orders.all()
            scan_orders = surgery.scan_orders.all()
            other_services = surgery.service_drug_orders.all()

            # --- APPLY THE SAME FIX HERE ---
            drug_costs = drug_orders.aggregate(
                total=Sum(
                    ExpressionWrapper(F('quantity_ordered') * F('drug__selling_price'), output_field=DecimalField()))
            )['total'] or 0

            lab_costs = lab_orders.aggregate(total=Sum('amount_charged'))['total'] or 0
            scan_costs = scan_orders.aggregate(total=Sum('amount_charged'))['total'] or 0
            other_services_costs = other_services.aggregate(total=Sum('total_amount'))['total'] or 0

            base_fee = surgery.total_surgery_cost
            total_bill = base_fee + drug_costs + lab_costs + scan_costs + other_services_costs

            total_paid = PatientTransactionModel.objects.filter(
                surgery=surgery, status='completed',
            ).aggregate(total=Sum('amount'))['total'] or 0

            total_paid += sum(order.drug.selling_price * Decimal(order.quantity_ordered) for order in drug_orders if
                              order.status in paid_statuses_drug)
            total_paid += sum(order.amount_charged for order in lab_orders if order.status in paid_statuses_lab_scan)
            total_paid += sum(order.amount_charged for order in scan_orders if order.status in paid_statuses_lab_scan)
            total_paid += sum(service.total_amount for service in other_services if
                              service.status in ['paid', 'fully_dispensed', 'partially_dispensed'])

            balance = total_paid - total_bill

            funding_targets.append({
                'record': surgery, 'type': 'Surgery', 'identifier': surgery.surgery_number,
                'base_fee': base_fee,
                'drugs': {'total': drug_costs},
                'labs': {'total': lab_costs},
                'scans': {'total': scan_costs},
                'others': {'total': other_services_costs},
                'total_bill': total_bill, 'total_paid': total_paid,
                'balance': balance, 'abs_balance': abs(balance)
            })

        context['funding_targets'] = funding_targets
        return context


@login_required
@permission_required('finance.add_patienttransactionmodel', raise_exception=True)
def ajax_process_admission_funding(request):
    if request.method == 'POST':
        try:
            patient_id = request.POST.get('patient_id')
            admission_id = request.POST.get('admission_id')
            surgery_id = request.POST.get('surgery_id')
            amount = Decimal(request.POST.get('amount', '0'))

            if amount <= 0:
                return JsonResponse({'success': False, 'error': 'Amount must be a positive number.'}, status=400)

            patient = get_object_or_404(PatientModel, pk=patient_id)
            wallet = get_object_or_404(PatientWalletModel, patient=patient)

            # Check for sufficient wallet balance
            if wallet.amount < amount:
                return JsonResponse({'success': False, 'error': 'Insufficient wallet balance.'}, status=400)

            admission = get_object_or_404(Admission, pk=admission_id) if admission_id else None
            surgery = get_object_or_404(Surgery, pk=surgery_id) if surgery_id else None

            with transaction.atomic():
                old_balance = wallet.amount

                # Deduct funds from the wallet
                wallet.amount -= amount
                wallet.save()

                new_balance = wallet.amount

                PatientTransactionModel.objects.create(
                    patient=patient,
                    transaction_type='admission_payment' if admission else 'surgery_payment',
                    transaction_direction='out',  # Direction is now 'out'
                    admission=admission,
                    surgery=surgery,
                    amount=amount,
                    old_balance=old_balance,  # Record old wallet balance
                    new_balance=new_balance,  # Record new wallet balance
                    date=timezone.now().date(),
                    received_by=request.user,
                    payment_method='wallet',  # Payment is from the wallet
                    status='completed'
                )

            record_identifier = admission.admission_number if admission else surgery.surgery_number
            success_message = f'Successfully paid ₦{amount:,.2f} for {record_identifier} from wallet.'

            return JsonResponse({
                'success': True,
                'message': success_message
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)

    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=400)


@login_required
def get_consultation_fees_ajax(request):
    """
    AJAX endpoint to get consultation fees for a specialization and patient
    """
    specialization_id = request.GET.get('specialization_id')
    patient_id = request.GET.get('patient_id')

    if not specialization_id or not patient_id:
        return JsonResponse({'success': False, 'error': 'Missing required parameters'})

    try:
        patient = PatientModel.objects.get(id=patient_id)
        specialization = SpecializationModel.objects.get(id=specialization_id)
    except (PatientModel.DoesNotExist, SpecializationModel.DoesNotExist):
        return JsonResponse({'success': False, 'error': 'Patient or specialization not found'})

    # Check if patient has active insurance
    try:
        policies_qs = patient.insurance_policies.all()
    except:
        try:
            policies_qs = patient.insurancepolicy_set.all()
        except:
            policies_qs = PatientInsuranceModel.objects.none()

    active_insurance = policies_qs.filter(
        is_active=True,
        valid_to__gte=timezone.now().date()
    ).select_related('hmo', 'coverage_plan').first()

    # Get appropriate fee
    try:
        if active_insurance and active_insurance.coverage_plan.consultation_covered:
            # Try to get insurance-specific fee first
            fee = ConsultationFeeModel.objects.filter(
                specialization=specialization,
                patient_category='insurance',
                insurance=active_insurance.coverage_plan,
                is_active=True
            ).first()

            if not fee:
                # Fall back to regular insurance fee
                fee = ConsultationFeeModel.objects.filter(
                    specialization=specialization,
                    patient_category='insurance',
                    insurance__isnull=True,
                    is_active=True
                ).first()
        else:
            fee = None

        if not fee:
            # Get regular patient fee
            fee = ConsultationFeeModel.objects.filter(
                specialization=specialization,
                patient_category='regular',
                is_active=True
            ).first()

        if not fee:
            return JsonResponse({'success': False, 'error': 'No fee structure found for this specialization'})

        # Calculate patient amount after insurance
        base_amount = fee.amount
        patient_amount = base_amount

        if (active_insurance and
                active_insurance.coverage_plan.consultation_covered and
                fee.patient_category == 'insurance'):
            pct = _to_decimal(active_insurance.coverage_plan.consultation_coverage_percentage)
            covered = (base_amount * (pct / Decimal('100')))
            patient_amount = base_amount - covered

        return JsonResponse({
            'success': True,
            'fee': {
                'id': fee.id,
                'amount': float(patient_amount),
                'base_amount': float(base_amount),
                'formatted_amount': f'₦{patient_amount:,.2f}',
                'patient_category': fee.patient_category,
                'duration_minutes': 30,  # Default or from fee model
                'coverage_applied': patient_amount < base_amount
            }
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error retrieving fees: {str(e)}'})


@login_required
@permission_required('finance.add_patienttransactionmodel', raise_exception=True)
@require_http_methods(["GET", "POST"])
def finance_consultation_patient_payment(request, patient_id):
    """
    Consultation payment - handles both verification and payment processing
    """
    patient = get_object_or_404(PatientModel, id=patient_id)

    if request.method == 'GET':
        # ... (GET request logic is unchanged)
        specializations = SpecializationModel.objects.all().order_by('name')
        return render(request, 'finance/payment/consultation.html', {
            'patient': patient,
            'specializations': specializations
        })

    # --- POST: process consultation payment ---
    fee_id = request.POST.get('fee_structure') or request.POST.get('fee_structure_id')
    amount_paid = _to_decimal(request.POST.get('amount_paid') or request.POST.get('amount_due') or '0')

    if amount_paid <= 0:
        return JsonResponse({'success': False, 'error': 'Invalid payment amount.'}, status=400)

    try:
        # --- THIS IS THE FIX ---
        # 1. Fetch the ConsultationFeeModel instance using the fee_id
        fee_structure = get_object_or_404(ConsultationFeeModel, id=fee_id)

        with transaction.atomic():
            # ... (wallet logic is unchanged)
            wallet = get_object_or_404(PatientWalletModel, patient=patient)
            if wallet.amount < amount_paid:
                return JsonResponse({'success': False, 'error': 'Insufficient wallet balance.'}, status=400)

            old_balance = wallet.amount
            wallet.amount -= amount_paid
            wallet.save()

            # 2. Link the fee_structure to the new transaction
            payment = PatientTransactionModel.objects.create(
                patient=patient,
                transaction_type='consultation_payment',
                transaction_direction='out',
                fee_structure=fee_structure,  # Link the fee structure here
                amount=amount_paid,
                old_balance=old_balance,
                new_balance=wallet.amount,
                date=timezone.now().date(),
                received_by=request.user,
                payment_method='wallet',
                status='completed'
            )

            # 3. Create the queue entry
            PatientQueueModel.objects.create(
                patient=patient,
                payment=payment,
                specialization=fee_structure.specialization, # Use the specialization from the fee object
                status='waiting_vitals'
            )

        return JsonResponse({
            'success': True,
            'message': 'Consultation payment successful.',
            'new_balance': float(wallet.amount),
            'formatted_new_balance': f'₦{wallet.amount:,.2f}',
            'redirect_url': reverse('transaction_detail', kwargs={'transaction_id': payment.id})
        })

    except ConsultationFeeModel.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Invalid fee structure selected.'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error processing payment: {str(e)}'}, status=500)


@login_required
@permission_required('finance.add_patienttransactionmodel', raise_exception=True)
@require_http_methods(["GET", "POST"])
def finance_pharmacy_patient_payment(request, patient_id):
    """
    Pharmacy (drug) payment UI + processing - CLAIM-BASED VERSION
    - GET: render list of pending drug orders with claim status
    - POST: process selected orders using approved claims
    """
    from insurance.claim_helpers import get_orders_with_claim_info, calculate_patient_amount_with_claim

    patient = get_object_or_404(PatientModel, id=patient_id)
    thirty_days_ago = timezone.now() - timedelta(days=THIRTY_DAYS)

    if request.method == 'GET':
        pending_drugs = DrugOrderModel.objects.filter(
            patient=patient,
            status__in=['pending'],
            ordered_at__gte=thirty_days_ago
        ).select_related('drug')

        # Get active insurance for display
        active_insurance = None
        try:
            policies_qs = patient.insurance_policies.all()
        except Exception:
            try:
                policies_qs = patient.insurancepolicy_set.all()
            except:
                policies_qs = PatientInsuranceModel.objects.none()

        active_insurance = policies_qs.filter(
            is_active=True,
            valid_to__gte=timezone.now().date()
        ).select_related('hmo', 'coverage_plan').first()

        # Process with claim-based logic
        drug_results = get_orders_with_claim_info(pending_drugs, 'drug')

        items = []
        total = Decimal('0.00')

        for result in drug_results:
            order = result['order']
            items.append({
                'id': order.id,
                'order_number': getattr(order, 'order_number', ''),
                'name': f"{order.drug.__str__()}",
                'quantity': getattr(order, 'quantity_ordered', 1),
                'base_amount': float(result['base_amount']),
                'patient_amount': float(result['patient_amount']),
                'covered_amount': float(result['covered_amount']),
                'formatted_patient_amount': f'₦{result["patient_amount"]:,.2f}',
                'status': order.status,
                'ordered_date': getattr(order, 'ordered_at').date().isoformat() if getattr(order, 'ordered_at',
                                                                                           None) else '',
                'has_approved_claim': result['has_approved_claim'],
                'claim_number': result['claim_number'],
                'claim_status': result['claim_status'],
                'has_pending_claim': result['has_pending_claim'],
                'pending_claim_number': result['pending_claim_number'],
                'insurance_covered': result['has_approved_claim']
            })
            total += result['patient_amount']

        pending_claims_count = sum(1 for item in items if item.get('has_pending_claim'))

        context = {
            'patient': patient,
            'items': items,
            'total': _quantize_money(total),
            'insurance': active_insurance,
            'pending_claims_count': pending_claims_count
        }
        return render(request, 'finance/payment/pharmacy.html', context)

    # POST - process payment with claim validation
    selected_ids = request.POST.getlist('selected_items[]') or request.POST.getlist('selected_items')
    if not selected_ids:
        return JsonResponse({'success': False, 'error': 'No items selected for payment.'}, status=400)

    try:
        selected_orders = list(
            DrugOrderModel.objects.filter(
                id__in=selected_ids,
                patient=patient,
                status='pending'
            ).select_related('drug')
        )

        if not selected_orders:
            return JsonResponse({'success': False, 'error': 'Selected orders not found.'}, status=404)

        total_amount = Decimal('0.00')
        order_details = []

        # Calculate amounts using claim-based logic
        for order in selected_orders:
            if getattr(order, 'ordered_at', timezone.now()) < thirty_days_ago:
                return JsonResponse({
                    'success': False,
                    'error': 'One or more selected orders are older than 30 days and cannot be paid here.'
                }, status=400)

            base_amount = (
                order.drug.selling_price * Decimal(str(order.quantity_ordered))
                if hasattr(order, 'drug') and hasattr(order.drug, 'selling_price')
                else Decimal('0.00')
            )

            # Use claim-based calculation
            claim_info = calculate_patient_amount_with_claim(order, base_amount)
            patient_amount = claim_info['patient_amount']

            total_amount += _quantize_money(patient_amount)
            order_details.append({
                'order': order,
                'base_amount': base_amount,
                'patient_amount': patient_amount,
                'covered_amount': claim_info['covered_amount'],
                'has_claim': claim_info['has_approved_claim'],
                'claim_number': claim_info['claim_number']
            })

        # Process wallet deduction
        with transaction.atomic():
            wallet_qs = PatientWalletModel.objects.select_for_update().filter(patient=patient)
            if wallet_qs.exists():
                wallet = wallet_qs.first()
            else:
                wallet = PatientWalletModel.objects.create(patient=patient, amount=Decimal('0.00'))

            if wallet.amount < total_amount:
                shortfall = _quantize_money(total_amount - wallet.amount)
                return JsonResponse({
                    'success': False,
                    'error': 'Insufficient wallet balance.',
                    'shortfall': float(shortfall),
                    'formatted_shortfall': f'₦{shortfall:,.2f}'
                }, status=400)

            old_balance = wallet.amount
            wallet.amount = _quantize_money(wallet.amount - total_amount)
            wallet.save()

            # Mark orders as paid
            for detail in order_details:
                order = detail['order']
                order.status = 'paid'

                order.save(update_fields=['status'])

            # Create transaction record
            try:
                payment = PatientTransactionModel.objects.create(
                    patient=patient,
                    transaction_type='drug_payment',
                    transaction_direction='out',
                    amount=total_amount,
                    old_balance=old_balance,
                    new_balance=wallet.amount,
                    date=timezone.now().date(),
                    received_by=request.user,
                    payment_method='wallet',
                    status='completed'
                )
            except Exception:
                payment = None

        # Build success message
        claims_count = sum(1 for d in order_details if d['has_claim'])
        message = f'Payment successful. ₦{total_amount:,.2f} deducted from wallet.'
        if claims_count > 0:
            message += f' Insurance claims applied to {claims_count} order(s).'

        return JsonResponse({
            'success': True,
            'message': message,
            'new_balance': float(wallet.amount),
            'formatted_new_balance': f'₦{wallet.amount:,.2f}',
            'redirect_url': reverse('transaction_detail', kwargs={'transaction_id': payment.pk}) if payment else reverse(
                'finance:payment_select'),
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error processing payment: {str(e)}'}, status=500)


# Re-define insurance calculation helper (using the provided logic)
def _calculate_insurance_amount(base_amount, coverage_percentage):
    """Calculate patient's portion after insurance, ensuring correct rounding."""
    base_amount = _to_decimal(base_amount)
    coverage_percentage = _to_decimal(coverage_percentage)
    if coverage_percentage and coverage_percentage > Decimal('0'):
        covered = base_amount * (coverage_percentage / Decimal('100'))
        return _quantize_money(base_amount - covered)
    return _quantize_money(base_amount)


# Re-define insurance query helper (using the provided logic)
def _get_active_insurance(patient):
    """Attempts to find the active insurance policy for the patient."""
    try:
        policies_qs = patient.insurance_policies.all()
    except Exception:
        try:
            policies_qs = patient.insurancepolicy_set.all()
        except:
            # Assuming PatientInsuranceModel is the correct model if lookups fail
            policies_qs = PatientInsuranceModel.objects.none()

    return policies_qs.filter(
        is_active=True,
        valid_to__gte=timezone.now().date()
    ).select_related('hmo', 'coverage_plan').first()


@login_required
@permission_required('finance.add_patienttransactionmodel', raise_exception=True)
@require_http_methods(["GET", "POST"])
def finance_service_patient_payment(request, patient_id):
    """
    Service/Item payment UI + processing:
    - GET: render list of pending PatientServiceTransaction orders (last 30 days)
    - POST: process selected transactions, deduct wallet, mark paid
    """
    patient = get_object_or_404(PatientModel, id=patient_id)
    thirty_days_ago = timezone.now() - timedelta(days=THIRTY_DAYS)

    # Pre-fetch the insurance once
    active_insurance = _get_active_insurance(patient)

    if request.method == 'GET':
        # Get pending transactions (pending_payment status is assumed from your model)
        pending_transactions = PatientServiceTransaction.objects.filter(
            patient=patient,
            status='pending_payment',
            created_at__gte=thirty_days_ago
        ).select_related(
            'service', 'service__category',
            'service_item', 'service_item__category'
        )

        items = []
        total = Decimal('0.00')

        for t in pending_transactions:
            # Determine item name, base amount, and category for insurance logic
            if t.service:
                name = f"Service: {t.service.name}"
                category_obj = t.service.category
            elif t.service_item:
                name = f"Item: {t.service_item.name}"
                category_obj = t.service_item.category
            else:
                continue  # Skip malformed transaction

            base_amount = t.total_amount
            patient_amount = base_amount

            # Apply insurance logic based on category
            if active_insurance and category_obj:
                coverage_plan = active_insurance.coverage_plan

                # IMPORTANT: Assuming CoveragePlan has these methods/attributes for Service Categories
                if hasattr(coverage_plan, 'is_service_category_covered'):
                    try:
                        if coverage_plan.is_service_category_covered(category_obj):
                            pct = _to_decimal(getattr(coverage_plan, 'service_coverage_percentage', 0))
                            patient_amount = _calculate_insurance_amount(base_amount, pct)
                    except Exception:
                        patient_amount = base_amount

            patient_amount = _quantize_money(patient_amount)

            items.append({
                'id': t.id,
                'name': name,
                'quantity': t.quantity,
                'unit_price': float(t.unit_price),
                'discount': float(t.discount),
                'base_amount': float(base_amount),
                'patient_amount': float(patient_amount),
                'formatted_patient_amount': f'₦{patient_amount:,.2f}',
                'status': t.status,
                'ordered_date': t.created_at.date().isoformat(),
                'insurance_covered': patient_amount < base_amount
            })
            total += patient_amount

        context = {
            'patient': patient,
            'items': items,
            'total': _quantize_money(total),
            'insurance': active_insurance
        }
        return render(request, 'finance/payment/service.html', context)

    # POST - process payment
    selected_ids = request.POST.getlist('selected_items[]') or request.POST.getlist('selected_items')
    if not selected_ids:
        return JsonResponse({'success': False, 'error': 'No items selected for payment.'}, status=400)

    try:
        # Re-fetch the selected transactions and compute final sum (server-side authority)
        selected_transactions = list(
            PatientServiceTransaction.objects.filter(
                id__in=selected_ids,
                patient=patient,
                status='pending_payment'
            ).select_related('service', 'service__category', 'service_item', 'service_item__category')
        )

        if not selected_transactions:
            return JsonResponse({'success': False, 'error': 'Selected transactions not found or already paid.'},
                                status=404)

        total_amount = Decimal('0.00')

        # Compute patient portion for each
        for t in selected_transactions:
            if t.created_at < thirty_days_ago:
                return JsonResponse({'success': False,
                                     'error': 'One or more selected items are older than 30 days and cannot be paid here.'},
                                    status=400)

            # Determine category for insurance
            category_obj = t.service.category if t.service else t.service_item.category
            base_amount = t.total_amount
            patient_amount = base_amount

            # Re-apply insurance logic
            if active_insurance and category_obj:
                coverage_plan = active_insurance.coverage_plan
                if hasattr(coverage_plan, 'is_service_category_covered'):
                    try:
                        if coverage_plan.is_service_category_covered(category_obj):
                            pct = _to_decimal(getattr(coverage_plan, 'service_coverage_percentage', 0))
                            patient_amount = _calculate_insurance_amount(base_amount, pct)
                    except Exception:
                        patient_amount = base_amount

            total_amount += _quantize_money(patient_amount)

        # process wallet deduction inside a transaction with row lock
        with transaction.atomic():
            wallet_qs = PatientWalletModel.objects.select_for_update().filter(patient=patient)
            if wallet_qs.exists():
                wallet = wallet_qs.first()
            else:
                # create if missing
                wallet = PatientWalletModel.objects.create(patient=patient, amount=Decimal('0.00'))

            if wallet.amount < total_amount:
                shortfall = _quantize_money(total_amount - wallet.amount)
                return JsonResponse({
                    'success': False,
                    'error': 'Insufficient wallet balance.',
                    'shortfall': float(shortfall),
                    'formatted_shortfall': f'₦{shortfall:,.2f}'
                }, status=400)

            # deduct
            old_balance = wallet.amount
            wallet.amount = _quantize_money(wallet.amount - total_amount)
            wallet.save()

            # mark transactions as paid
            for t in selected_transactions:
                # Note: Setting status to 'paid' may trigger stock deduction in the transaction's save method
                # based on your PatientServiceTransaction.save() logic.
                t.status = 'paid'
                t.amount_paid = t.total_amount  # Mark as fully paid
                t.save(update_fields=['status', 'amount_paid'])

            # Create transaction record
            payment = PatientTransactionModel.objects.create(
                patient=patient,
                # Decide if 'service' or 'item' is more appropriate, or use a new type
                transaction_type='service',
                transaction_direction='out',
                amount=total_amount,
                old_balance=old_balance,
                new_balance=wallet.amount,
                date=timezone.now().date(),
                received_by=request.user,
                payment_method='wallet',
                status='completed',
                # You might want to link one of the transactions, or none
                # Linking is complex for multiple items, but we link the first one for reference
                service=selected_transactions[0] if selected_transactions else None
            )

        # success
        return JsonResponse({
            'success': True,
            'message': f'Payment successful. {total_amount:,.2f} deducted from wallet.',
            'new_balance': float(wallet.amount),
            'formatted_new_balance': f'₦{wallet.amount:,.2f}',
            'redirect_url': reverse('transaction_detail', kwargs={'transaction_id': payment.pk}),
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error processing payment: {str(e)}'}, status=500)


# Lab and Scan views remain the same as in your original code
@login_required
@permission_required('finance.add_patienttransactionmodel', raise_exception=True)
@require_http_methods(["GET", "POST"])
def finance_generic_order_payment(request, patient_id, order_type):
    """
    Generic handler used for lab/scan payments - CLAIM-BASED VERSION
    order_type: 'lab' or 'scan'
    """
    from insurance.claim_helpers import get_orders_with_claim_info, calculate_patient_amount_with_claim

    patient = get_object_or_404(PatientModel, id=patient_id)
    thirty_days_ago = timezone.now() - timedelta(days=THIRTY_DAYS)

    if order_type == 'lab':
        Model = LabTestOrderModel
        template = 'finance/payment/lab.html'
        transaction_type = 'lab_payment'
        order_type_key = 'lab'
    elif order_type == 'scan':
        Model = ScanOrderModel
        template = 'finance/payment/scan.html'
        transaction_type = 'scan_payment'
        order_type_key = 'scan'
    else:
        return HttpResponseBadRequest('Invalid order type')

    if request.method == 'GET':
        pending = Model.objects.filter(
            patient=patient,
            status__in=['pending', 'unpaid', 'awaiting_payment'],
            ordered_at__gte=thirty_days_ago
        ).select_related('template')

        # Get active insurance for display
        active_insurance = None
        try:
            policies_qs = patient.insurance_policies.all()
        except Exception:
            try:
                policies_qs = patient.insurancepolicy_set.all()
            except:
                policies_qs = PatientInsuranceModel.objects.none()

        active_insurance = policies_qs.filter(
            is_active=True,
            valid_to__gte=timezone.now().date()
        ).select_related('hmo', 'coverage_plan').first()

        # Process with claim-based logic
        order_results = get_orders_with_claim_info(pending, order_type_key)

        items = []
        total = Decimal('0.00')

        for result in order_results:
            order = result['order']
            items.append({
                'id': order.id,
                'order_number': getattr(order, 'order_number', ''),
                'name': getattr(order.template, 'name', ''),
                'base_amount': float(result['base_amount']),
                'patient_amount': float(result['patient_amount']),
                'covered_amount': float(result['covered_amount']),
                'formatted_patient_amount': f'₦{result["patient_amount"]:,.2f}',
                'status': order.status,
                'ordered_date': getattr(order, 'ordered_at').date().isoformat() if getattr(order, 'ordered_at',
                                                                                           None) else '',
                'has_approved_claim': result['has_approved_claim'],
                'claim_number': result['claim_number'],
                'claim_status': result['claim_status'],
                'has_pending_claim': result['has_pending_claim'],
                'pending_claim_number': result['pending_claim_number'],
                'insurance_covered': result['has_approved_claim']
            })
            total += result['patient_amount']

        pending_claims_count = sum(1 for item in items if item.get('has_pending_claim'))

        context = {
            'patient': patient,
            'items': items,
            'total': _quantize_money(total),
            'order_type': order_type,
            'insurance': active_insurance,
            'pending_claims_count': pending_claims_count
        }
        return render(request, template, context)

    # POST - process payment with claim validation
    selected_ids = request.POST.getlist('selected_items[]') or request.POST.getlist('selected_items')
    if not selected_ids:
        return JsonResponse({'success': False, 'error': 'No items selected for payment.'}, status=400)

    try:
        selected_orders = list(
            Model.objects.filter(
                id__in=selected_ids,
                patient=patient,
                status__in=['pending', 'unpaid', 'awaiting_payment']
            ).select_related('template')
        )

        if not selected_orders:
            return JsonResponse({'success': False, 'error': 'Selected orders not found.'}, status=404)

        total_amount = Decimal('0.00')
        order_details = []

        # Calculate amounts using claim-based logic
        for order in selected_orders:
            if getattr(order, 'ordered_at', timezone.now()) < thirty_days_ago:
                return JsonResponse({
                    'success': False,
                    'error': 'One or more selected orders are older than 30 days and cannot be paid here.'
                }, status=400)

            base_amount = _to_decimal(
                getattr(order, 'amount_charged', None) or getattr(order.template, 'price', 0)
            )

            # Use claim-based calculation
            claim_info = calculate_patient_amount_with_claim(order, base_amount)
            patient_amount = claim_info['patient_amount']

            total_amount += _quantize_money(patient_amount)
            order_details.append({
                'order': order,
                'base_amount': base_amount,
                'patient_amount': patient_amount,
                'covered_amount': claim_info['covered_amount'],
                'has_claim': claim_info['has_approved_claim'],
                'claim_number': claim_info['claim_number']
            })

        # Process wallet deduction
        with transaction.atomic():
            wallet_qs = PatientWalletModel.objects.select_for_update().filter(patient=patient)
            if wallet_qs.exists():
                wallet = wallet_qs.first()
            else:
                wallet = PatientWalletModel.objects.create(patient=patient, amount=Decimal('0.00'))

            if wallet.amount < total_amount:
                shortfall = _quantize_money(total_amount - wallet.amount)
                return JsonResponse({
                    'success': False,
                    'error': 'Insufficient wallet balance.',
                    'shortfall': float(shortfall),
                    'formatted_shortfall': f'₦{shortfall:,.2f}'
                }, status=400)

            old_balance = wallet.amount
            wallet.amount = _quantize_money(wallet.amount - total_amount)
            wallet.save()

            # Mark orders as paid
            for detail in order_details:
                order = detail['order']
                order.status = 'paid'
                order.save(update_fields=['status'])

            # Create transaction record
            try:
                payment = PatientTransactionModel.objects.create(
                    patient=patient,
                    transaction_type=transaction_type,
                    transaction_direction='out',
                    amount=total_amount,
                    old_balance=old_balance,
                    new_balance=wallet.amount,
                    date=timezone.now().date(),
                    received_by=request.user,
                    payment_method='wallet',
                    status='completed'
                )
            except Exception:
                payment = None

        # Build success message
        claims_count = sum(1 for d in order_details if d['has_claim'])
        message = f'Payment successful. ₦{total_amount:,.2f} deducted from wallet.'
        if claims_count > 0:
            message += f' Insurance claims applied to {claims_count} order(s).'

        return JsonResponse({
            'success': True,
            'message': message,
            'new_balance': float(wallet.amount),
            'formatted_new_balance': f'₦{wallet.amount:,.2f}',
            'redirect_url': reverse('transaction_detail', kwargs={'transaction_id': payment.pk}) if payment else reverse(
                'finance:payment_select'),
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error processing payment: {str(e)}'}, status=500)


# convenience wrappers (keep these as-is)
@login_required
def finance_laboratory_patient_payment(request, patient_id):
    return finance_generic_order_payment(request, patient_id, 'lab')


@login_required
def finance_scan_patient_payment(request, patient_id):
    return finance_generic_order_payment(request, patient_id, 'scan')


class PatientTransactionListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = PatientTransactionModel
    template_name = "finance/payment/index.html"
    context_object_name = "transactions"
    permission_required = "finance.view_patienttransactionmodel"
    paginate_by = 20

    def get(self, request, *args, **kwargs):
        """Override get method to handle Excel download requests"""
        if request.GET.get('download') == 'excel':
            return self.download_excel()
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        """
        Filters the queryset based on GET parameters from the URL.
        """
        # Eager load related models to prevent N+1 query issues
        qs = PatientTransactionModel.objects.select_related('patient', 'received_by')

        # Get filter parameters from the request
        start_date = self.request.GET.get("start_date")
        end_date = self.request.GET.get("end_date")
        search = self.request.GET.get("search")
        transaction_type = self.request.GET.get("transaction_type")
        transaction_direction = self.request.GET.get("transaction_direction")

        # Default to today's transactions if no date range is provided
        if not start_date and not end_date:
            today = now().date()
            qs = qs.filter(date=today)
        elif start_date and end_date:
            qs = qs.filter(date__range=[start_date, end_date])

        # Apply search filter for patient name or transaction ID
        if search:
            qs = qs.filter(
                Q(patient__full_name__icontains=search) |
                Q(transaction_id__icontains=search)
            )

        # Apply choice filters
        if transaction_type:
            qs = qs.filter(transaction_type=transaction_type)
        if transaction_direction:
            qs = qs.filter(transaction_direction=transaction_direction)

        return qs.order_by("-created_at")

    def get_context_data(self, **kwargs):
        """
        Adds summary data and filter choices to the template context.
        """
        context = super().get_context_data(**kwargs)
        queryset = context['object_list']  # Use the paginated queryset from the context

        # Determine timeframe for the page title
        start_date = self.request.GET.get("start_date")
        end_date = self.request.GET.get("end_date")

        if start_date and end_date:
            context['timeframe'] = f"from {start_date} to {end_date}"
        else:
            context['timeframe'] = "for Today"

        # Calculate total inflow and outflow using database aggregation for efficiency
        totals = queryset.aggregate(
            total_inflow=Sum('amount', filter=Q(transaction_direction='in')),
            total_outflow=Sum('amount', filter=Q(transaction_direction='out'))
        )
        context['total_inflow'] = totals['total_inflow'] or Decimal('0.00')
        context['total_outflow'] = totals['total_outflow'] or Decimal('0.00')

        # Pass transaction type choices to the template for the filter dropdown
        context['transaction_types'] = PatientTransactionModel.TRANSACTION_TYPE

        return context

    def download_excel(self):
        """
        Generate and download Excel file with the same filtered transactions
        """
        # Get all transactions with the same filters (no pagination for Excel)
        transactions = self.get_queryset()

        # Create Excel workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Patient Transactions"

        # Define styles for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center")

        # Define column headers
        headers = [
            "S/N", "Patient Name", "Transaction ID", "Transaction Type",
            "Direction", "Amount (₦)", "Old Balance (₦)", "New Balance (₦)",
            "Date", "Time", "Status", "Received By"
        ]

        # Write headers to first row
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment

        # Write transaction data
        for row_num, transaction in enumerate(transactions, 2):
            worksheet.cell(row=row_num, column=1, value=row_num - 1)  # Serial Number
            worksheet.cell(row=row_num, column=2, value=str(transaction.patient) if transaction.patient else "N/A")
            worksheet.cell(row=row_num, column=3, value=transaction.transaction_id)
            worksheet.cell(row=row_num, column=4, value=transaction.get_transaction_type_display())
            worksheet.cell(row=row_num, column=5,
                           value="Credit" if transaction.transaction_direction == 'in' else "Debit")
            worksheet.cell(row=row_num, column=6, value=float(transaction.amount))
            worksheet.cell(row=row_num, column=7, value=float(transaction.old_balance))
            worksheet.cell(row=row_num, column=8, value=float(transaction.new_balance))
            worksheet.cell(row=row_num, column=9, value=transaction.date.strftime('%d %b %Y'))
            worksheet.cell(row=row_num, column=10, value=transaction.created_at.strftime('%H:%M'))
            worksheet.cell(row=row_num, column=11, value=transaction.status.title())
            worksheet.cell(row=row_num, column=12,
                           value=str(transaction.received_by) if transaction.received_by else "System")

        # Add summary section if there are transactions
        if transactions:
            summary_row = len(transactions) + 3

            # Calculate summary totals
            total_inflow = sum(float(t.amount) for t in transactions if t.transaction_direction == 'in')
            total_outflow = sum(float(t.amount) for t in transactions if t.transaction_direction == 'out')

            # Add summary data
            worksheet.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=12)
            worksheet.cell(row=summary_row + 1, column=1, value="Total Transactions:")
            worksheet.cell(row=summary_row + 1, column=2, value=len(transactions))
            worksheet.cell(row=summary_row + 2, column=1, value="Total Inflow (Credit):")
            worksheet.cell(row=summary_row + 2, column=2, value=total_inflow)
            worksheet.cell(row=summary_row + 3, column=1, value="Total Outflow (Debit):")
            worksheet.cell(row=summary_row + 3, column=2, value=total_outflow)
            worksheet.cell(row=summary_row + 4, column=1, value="Net Amount:")
            worksheet.cell(row=summary_row + 4, column=2, value=total_inflow - total_outflow)

        # Auto-adjust column widths
        column_widths = [8, 25, 20, 18, 12, 15, 18, 18, 12, 10, 12, 20]
        for col_num, width in enumerate(column_widths, 1):
            column_letter = openpyxl.utils.get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = width

        # Generate filename based on applied filters
        start_date = self.request.GET.get("start_date")
        end_date = self.request.GET.get("end_date")

        if start_date and end_date:
            filename = f"Patient_Transactions_{start_date}_to_{end_date}.xlsx"
        else:
            today = now().date().strftime('%Y-%m-%d')
            filename = f"Patient_Transactions_{today}.xlsx"

        # Prepare HTTP response for file download
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Save workbook to response
        workbook.save(response)
        return response


class PatientTransactionDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = PatientTransactionModel
    template_name = "finance/payment/detail.html"
    context_object_name = "transaction"
    permission_required = "finance.view_patienttransactionmodel"

    def get_queryset(self):
        """
        Overrides the default queryset to optimize database queries.

        By using `select_related`, we fetch the related 'patient' and 'received_by'
        objects in the same database query as the transaction itself. This prevents
        additional database hits when the template accesses `{{ transaction.patient }}`
        or `{{ transaction.received_by }}`.
        """
        return super().get_queryset().select_related('patient', 'received_by')


def get_finance_setting_instance():
    """
    Helper function to load the singleton FinanceSettingModel instance.
    This uses the .load() classmethod defined on the model.
    """
    return FinanceSettingModel.objects.first()


# -------------------------
# Finance Setting Views
# -------------------------
class FinanceSettingCreateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, CreateView):
    model = FinanceSettingModel
    form_class = FinanceSettingForm
    permission_required = 'finance.add_financesettingmodel'
    success_message = 'Finance Setting Created Successfully'
    template_name = 'finance/setting/create.html'

    def dispatch(self, request, *args, **kwargs):
        """
        If a setting object already exists, redirect to the edit page.
        This enforces the singleton pattern.
        """
        setting = get_finance_setting_instance()
        # The .load() method creates a default if none exists, so we check if its pk is not None.
        # If it has a pk other than the default 1, or has been saved before, it exists.
        # A simpler check might just be to see if we can find an object with pk=1.
        if FinanceSettingModel.objects.filter(pk=1).exists():
            return redirect('finance_setting_edit', pk=1)
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self):
        return reverse('finance_setting_detail', kwargs={'pk': self.object.pk})


class FinanceSettingDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = FinanceSettingModel
    permission_required = 'finance.view_financesettingmodel'
    template_name = 'finance/setting/detail.html'
    context_object_name = "finance_setting"

    def get_object(self, queryset=None):
        """
        Always return the single, global instance of the finance settings.
        """
        return get_finance_setting_instance()


class FinanceSettingUpdateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, UpdateView):
    model = FinanceSettingModel
    form_class = FinanceSettingForm
    permission_required = 'finance.add_financesettingmodel'
    success_message = 'Finance Setting Updated Successfully'
    template_name = 'finance/setting/create.html' # Reuse the create template for editing

    def get_object(self, queryset=None):
        """
        Always return the single, global instance to be edited.
        """
        return get_finance_setting_instance()

    def get_success_url(self):
        return reverse('finance_setting_detail', kwargs={'pk': self.object.pk})


# -------------------------
# Expense Category Views (Simple - Same HTML Pattern)
# -------------------------
class ExpenseCategoryCreateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, CreateView
):
    model = ExpenseCategory
    permission_required = 'finance.add_expensecategory'
    form_class = ExpenseCategoryForm
    template_name = 'finance/expense_category/index.html'
    success_message = 'Expense Category Successfully Created'

    def get_success_url(self):
        return reverse('expense_category_index')

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET':
            return redirect(reverse('expense_category_index'))
        return super().dispatch(request, *args, **kwargs)


class ExpenseCategoryListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = ExpenseCategory
    permission_required = 'finance.view_expensecategory'
    template_name = 'finance/expense_category/index.html'
    context_object_name = "category_list"

    def get_queryset(self):
        return ExpenseCategory.objects.all().order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = ExpenseCategoryForm()
        return context


class ExpenseCategoryUpdateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, UpdateView
):
    model = ExpenseCategory
    permission_required = 'finance.change_expensecategory'
    form_class = ExpenseCategoryForm
    template_name = 'finance/expense_category/index.html'
    success_message = 'Expense Category Successfully Updated'

    def get_success_url(self):
        return reverse('expense_category_index')

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET':
            return redirect(reverse('expense_category_index'))
        return super().dispatch(request, *args, **kwargs)


class ExpenseCategoryDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = ExpenseCategory
    permission_required = 'finance.delete_expensecategory'
    template_name = 'finance/expense_category/delete.html'
    context_object_name = "category"
    success_message = 'Expense Category Successfully Deleted'

    def get_success_url(self):
        return reverse('expense_category_index')


# -------------------------
# Expense Views
# -------------------------
class ExpenseListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Expense
    permission_required = 'finance.view_expense'
    template_name = 'finance/expense/index.html'
    context_object_name = "expense_list"
    paginate_by = 20

    def get_queryset(self):
        queryset = Expense.objects.select_related(
            'category', 'department', 'paid_by'
        ).order_by('-date')

        # Filter functionality
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category_id=category)

        department = self.request.GET.get('department')
        if department:
            queryset = queryset.filter(department_id=department)

        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(expense_number__icontains=search) |
                Q(title__icontains=search)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = ExpenseCategory.objects.all()
        context['departments'] = DepartmentModel.objects.all()
        context['total_amount'] = self.get_queryset().aggregate(Sum('amount'))['amount__sum'] or 0
        return context


class ExpenseCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Expense
    permission_required = 'finance.add_expense'
    form_class = ExpenseForm
    template_name = 'finance/expense/create.html'
    success_message = 'Expense Successfully Created'

    def get_success_url(self):
        return reverse('expense_detail', kwargs={'pk': self.object.pk})


class ExpenseUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Expense
    permission_required = 'finance.change_expense'
    form_class = ExpenseForm
    template_name = 'finance/expense/edit.html'
    success_message = 'Expense Successfully Updated'

    def get_success_url(self):
        return reverse('expense_detail', kwargs={'pk': self.object.pk})


class ExpenseDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Expense
    permission_required = 'finance.view_expense'
    template_name = 'finance/expense/detail.html'
    context_object_name = "expense"


# -------------------------
# Income Category Views (Simple - Same HTML Pattern)
# -------------------------
class IncomeCategoryCreateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, CreateView
):
    model = IncomeCategory
    permission_required = 'finance.add_incomecategory'
    form_class = IncomeCategoryForm
    template_name = 'finance/income_category/index.html'
    success_message = 'Income Category Successfully Created'

    def get_success_url(self):
        return reverse('income_category_index')

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET':
            return redirect(reverse('income_category_index'))
        return super().dispatch(request, *args, **kwargs)


class IncomeCategoryListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = IncomeCategory
    permission_required = 'finance.view_incomecategory'
    template_name = 'finance/income_category/index.html'
    context_object_name = "category_list"

    def get_queryset(self):
        return IncomeCategory.objects.all().order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = IncomeCategoryForm()
        return context


class IncomeCategoryUpdateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, UpdateView
):
    model = IncomeCategory
    permission_required = 'finance.change_incomecategory'
    form_class = IncomeCategoryForm
    template_name = 'finance/income_category/index.html'
    success_message = 'Income Category Successfully Updated'

    def get_success_url(self):
        return reverse('income_category_index')

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET':
            return redirect(reverse('income_category_index'))
        return super().dispatch(request, *args, **kwargs)


class IncomeCategoryDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = IncomeCategory
    permission_required = 'finance.delete_incomecategory'
    template_name = 'finance/income_category/delete.html'
    context_object_name = "category"
    success_message = 'Income Category Successfully Deleted'

    def get_success_url(self):
        return reverse('income_category_index')


# -------------------------
# Income Views
# -------------------------
class IncomeListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Income
    permission_required = 'finance.view_income'
    template_name = 'finance/income/index.html'
    context_object_name = "income_list"
    paginate_by = 20

    def get_queryset(self):
        queryset = Income.objects.select_related(
            'category', 'department', 'received_by'
        ).order_by('-date')

        # Filter functionality
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category_id=category)

        department = self.request.GET.get('department')
        if department:
            queryset = queryset.filter(department_id=department)

        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(income_number__icontains=search) |
                Q(title__icontains=search) |
                Q(source__icontains=search)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = IncomeCategory.objects.all()
        context['departments'] = DepartmentModel.objects.all()
        context['total_amount'] = self.get_queryset().aggregate(Sum('amount'))['amount__sum'] or 0
        return context


class IncomeCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Income
    permission_required = 'finance.add_income'
    form_class = IncomeForm
    template_name = 'finance/income/create.html'
    success_message = 'Income Successfully Created'

    def get_success_url(self):
        return reverse('income_detail', kwargs={'pk': self.object.pk})


class IncomeUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Income
    permission_required = 'finance.change_income'
    form_class = IncomeForm
    template_name = 'finance/income/edit.html'
    success_message = 'Income Successfully Updated'

    def get_success_url(self):
        return reverse('income_detail', kwargs={'pk': self.object.pk})


class IncomeDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Income
    permission_required = 'finance.view_income'
    template_name = 'finance/income/detail.html'
    context_object_name = "income"


class OtherPaymentServiceListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = OtherPaymentService
    permission_required = 'service.view_otherpaymentservice'
    template_name = 'finance/other_payment/index.html'
    context_object_name = "service_list"
    paginate_by = 20

    def get_queryset(self):
        queryset = OtherPaymentService.objects.select_related('created_by').order_by('name')

        # Filter and search functionality
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category=category)

        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(description__icontains=search)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Pass category choices to the template for the filter dropdown
        context['categories'] = OtherPaymentService._meta.get_field('category').choices
        return context


class OtherPaymentServiceCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = OtherPaymentService
    permission_required = 'service.add_otherpaymentservice'
    form_class = OtherPaymentServiceForm
    template_name = 'finance/other_payment/create.html'

    def form_valid(self, form):
        """Set the creator of the service to the current user."""
        form.instance.created_by = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        """Redirect to the detail page of the newly created service."""
        # You might want to add a success message here as well
        # messages.success(self.request, 'Service created successfully.')
        return reverse('other_payment_service_detail', kwargs={'pk': self.object.pk})


class OtherPaymentServiceDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = OtherPaymentService
    permission_required = 'service.view_otherpaymentservice'
    template_name = 'finance/other_payment/detail.html'
    context_object_name = "service"


class OtherPaymentServiceUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = OtherPaymentService
    permission_required = 'service.add_otherpaymentservice'
    form_class = OtherPaymentServiceForm
    template_name = 'finance/other_payment/edit.html'
    context_object_name = "service"

    def get_success_url(self):
        """Redirect to the detail page of the updated service."""
        # messages.success(self.request, 'Service updated successfully.')
        return reverse('other_payment_service_detail', kwargs={'pk': self.object.pk})


class OtherPaymentServiceDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = OtherPaymentService
    permission_required = 'service.add_otherpaymentservice'
    template_name = 'finance/other_payment/delete.html'
    context_object_name = "service"

    def get_success_url(self):
        """Redirect back to the list view after deletion."""
        # messages.success(self.request, 'Service deleted successfully.')
        return reverse_lazy('other_payment_service_list')


# -------------------------
# Staff Bank Detail Views (Simple - Same HTML Pattern)
# -------------------------
class StaffBankDetailCreateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, CreateView
):
    model = StaffBankDetail
    permission_required = 'finance.add_staffbankdetail'
    form_class = StaffBankDetailForm
    template_name = 'finance/staff_bank/index.html'
    success_message = 'Bank Detail Successfully Created'

    def get_success_url(self):
        return reverse('staff_bank_detail_index')

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET':
            return redirect(reverse('staff_bank_detail_index'))
        return super().dispatch(request, *args, **kwargs)


class StaffBankDetailListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = StaffBankDetail
    permission_required = 'finance.view_staffbankdetail'
    template_name = 'finance/staff_bank/index.html'
    context_object_name = "bank_detail_list"

    def get_queryset(self):
        return StaffBankDetail.objects.select_related('staff').order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = StaffBankDetailForm()
        return context


class StaffBankDetailUpdateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, UpdateView
):
    model = StaffBankDetail
    permission_required = 'finance.change_staffbankdetail'
    form_class = StaffBankDetailForm
    template_name = 'finance/staff_bank/index.html'
    success_message = 'Bank Detail Successfully Updated'

    def get_success_url(self):
        return reverse('staff_bank_detail_index')

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET':
            return redirect(reverse('staff_bank_detail_index'))
        return super().dispatch(request, *args, **kwargs)


class StaffBankDetailDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = StaffBankDetail
    permission_required = 'finance.delete_staffbankdetail'
    template_name = 'finance/staff_bank/delete.html'
    context_object_name = "bank_detail"
    success_message = 'Bank Detail Successfully Deleted'

    def get_success_url(self):
        return reverse('staff_bank_detail_index')


# -------------------------
# Salary Structure Views
# -------------------------
class SalaryStructureListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = SalaryStructure
    permission_required = 'finance.view_salarystructure'
    template_name = 'finance/salary_structure/index.html'
    context_object_name = "salary_structure_list"

    def get_queryset(self):
        return SalaryStructure.objects.select_related('staff').order_by('-created_at')


class SalaryStructureCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = SalaryStructure
    permission_required = 'finance.add_salarystructure'
    form_class = SalaryStructureForm
    template_name = 'finance/salary_structure/create.html'
    success_message = 'Salary Structure Successfully Created'

    def get_success_url(self):
        return reverse('salary_structure_detail', kwargs={'pk': self.object.pk})


class SalaryStructureUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = SalaryStructure
    permission_required = 'finance.change_salarystructure'
    form_class = SalaryStructureForm
    template_name = 'finance/salary_structure/edit.html'
    success_message = 'Salary Structure Successfully Updated'

    def get_success_url(self):
        return reverse('salary_structure_detail', kwargs={'pk': self.object.pk})


class SalaryStructureDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = SalaryStructure
    permission_required = 'finance.view_salarystructure'
    template_name = 'finance/salary_structure/detail.html'
    context_object_name = "salary_structure"


# -------------------------
# Salary Record Views
# -------------------------
@login_required
@permission_required('finance.add_salaryrecord')
def process_payroll_view(request):
    """
    An interactive view to manage the payroll for a specific month and year.
    Lists all staff with active salary structures and allows for inline editing.
    """
    # 1. Determine the period to process from GET parameters or use current month/year
    current_year = datetime.now().year
    current_month = datetime.now().month

    try:
        year = int(request.GET.get('year', current_year))
        month = int(request.GET.get('month', current_month))
    except (ValueError, TypeError):
        year = current_year
        month = current_month

    # 2. Get all staff who should be on the paysheet (i.e., have an active structure)
    staff_with_structures = StaffModel.objects.filter(salary_structure__is_active=True).select_related(
        'salary_structure')

    # 3. For each staff member, ensure a SalaryRecord exists for the period.
    # This automatically creates missing records, replacing the need for bulk generation.
    for staff in staff_with_structures:
        structure = staff.salary_structure
        record, created = SalaryRecord.objects.get_or_create(
            staff=staff,
            year=year,
            month=month,
            # 'defaults' are only used if a new record is being created
            defaults={
                'basic_salary': structure.basic_salary,
                'housing_allowance': structure.housing_allowance,
                'transport_allowance': structure.transport_allowance,
                'medical_allowance': structure.medical_allowance,
                'other_allowances': structure.other_allowances,
                'tax_amount': structure.tax_amount,
                'pension_amount': structure.pension_amount,
            }
        )

    # 4. Create a Formset. This is a collection of forms for our editable table.
    queryset = SalaryRecord.objects.filter(year=year, month=month, staff__in=staff_with_structures).select_related(
        'staff')
    PaysheetFormSet = modelformset_factory(SalaryRecord, form=PaysheetRowForm, extra=0)

    if request.method == 'POST':
        formset = PaysheetFormSet(request.POST, queryset=queryset)
        if formset.is_valid():
            # Save all the inline changes (bonus, deductions, notes, etc.)
            formset.save()

            # Handle the bulk "Mark as Paid" action
            paid_ids = request.POST.getlist('mark_as_paid')
            if paid_ids:
                paid_records = SalaryRecord.objects.filter(id__in=paid_ids)
                for record in paid_records:
                    if not record.is_paid:  # Only update if not already paid
                        record.is_paid = True
                        # If amount_paid is still 0, assume full payment on bulk mark
                        if record.amount_paid == 0:
                            record.amount_paid = record.net_salary
                        record.paid_date = date.today()
                        record.paid_by = request.user
                        record.save()

            messages.success(request, 'Paysheet saved successfully!')
            # Redirect back to the same page with the same filters
            return redirect(reverse('process_payroll') + f'?year={year}&month={month}')
        else:
            messages.error(request, 'Please correct the errors below. Invalid data was submitted.')

    else:
        # For a GET request, just display the formset with the current data
        formset = PaysheetFormSet(queryset=queryset)

    context = {
        'formset': formset,
        'year': year,
        'month': month,
        'years': range(2020, datetime.now().year + 2),
        'months': [(i, datetime(2000, i, 1).strftime('%B')) for i in range(1, 13)],
    }
    return render(request, 'finance/salary_record/process_payroll.html', context)


@login_required
@permission_required('finance.view_salaryrecord')
def export_payroll_to_excel(request, year, month):
    """
    Generates an Excel file with a detailed breakdown of the payroll for a given month and year.
    """
    # 1. Fetch the relevant salary records
    queryset = SalaryRecord.objects.filter(year=year, month=month).select_related('staff')

    # 2. Create an in-memory Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    month_name = datetime(2000, month, 1).strftime('%B')
    worksheet.title = f'Payroll_{year}_{month_name}'

    # 3. Define the detailed header row
    headers = [
        'Staff ID', 'Full Name', 'Basic Salary', 'Housing', 'Transport', 'Medical',
        'Other Allowances', 'Bonus', 'Gross Salary', 'Tax (PAYE)', 'Pension',
        'Other Deductions', 'Total Deductions', 'Net Salary', 'Amount Paid', 'Status', 'Notes'
    ]
    for col_num, header_title in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header_title)
        cell.font = Font(bold=True)

    # 4. Write data rows for each salary record
    for row_num, record in enumerate(queryset, 2):
        worksheet.cell(row=row_num, column=1, value=record.staff.staff_id)
        worksheet.cell(row=row_num, column=2, value=record.staff.__str__())
        worksheet.cell(row=row_num, column=3, value=record.basic_salary)
        worksheet.cell(row=row_num, column=4, value=record.housing_allowance)
        worksheet.cell(row=row_num, column=5, value=record.transport_allowance)
        worksheet.cell(row=row_num, column=6, value=record.medical_allowance)
        worksheet.cell(row=row_num, column=7, value=record.other_allowances)
        worksheet.cell(row=row_num, column=8, value=record.bonus)
        worksheet.cell(row=row_num, column=9, value=record.gross_salary)
        worksheet.cell(row=row_num, column=10, value=record.tax_amount)
        worksheet.cell(row=row_num, column=11, value=record.pension_amount)
        worksheet.cell(row=row_num, column=12, value=record.other_deductions)
        worksheet.cell(row=row_num, column=13, value=record.total_deductions)
        worksheet.cell(row=row_num, column=14, value=record.net_salary)
        worksheet.cell(row=row_num, column=15, value=record.amount_paid)
        worksheet.cell(row=row_num, column=16, value=record.payment_status)
        worksheet.cell(row=row_num, column=17, value=record.notes)

    # 5. Create the HttpResponse object with the correct headers
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="payroll_{year}_{month_name}.xlsx"'

    # Save the workbook to the response
    workbook.save(response)

    return response


# -------------------------
# AJAX Helper Views
# -------------------------
@login_required
def get_staff_salary_structure_ajax(request, staff_id):
    """Get salary structure details for a staff member"""
    try:
        staff = User.objects.get(id=staff_id)
        salary_structure = SalaryStructure.objects.get(staff=staff, is_active=True)

        return JsonResponse({
            'success': True,
            'data': {
                'basic_salary': str(salary_structure.basic_salary),
                'housing_allowance': str(salary_structure.housing_allowance),
                'transport_allowance': str(salary_structure.transport_allowance),
                'medical_allowance': str(salary_structure.medical_allowance),
                'other_allowances': str(salary_structure.other_allowances),
                'gross_salary': str(salary_structure.gross_salary),
                'tax_amount': str(salary_structure.tax_amount),
                'pension_amount': str(salary_structure.pension_amount),
                'net_salary': str(salary_structure.net_salary),
            }
        })
    except (User.DoesNotExist, SalaryStructure.DoesNotExist):
        return JsonResponse({
            'success': False,
            'error': 'No active salary structure found for this staff member'
        })


@login_required
@permission_required('finance.view_salaryrecord')
def payroll_dashboard_view(request):
    """
    Provides data for a dashboard with payroll statistics and visualizations.
    Calculates net salary at the database level to fix the FieldError.
    """
    # Define the net salary calculation using F() objects for database-level arithmetic
    net_salary_expression = (
            F('basic_salary') + F('housing_allowance') + F('transport_allowance') +
            F('medical_allowance') + F('other_allowances') + F('bonus') -
            (F('tax_amount') + F('pension_amount') + F('other_deductions'))
    )

    # Get the current period
    today = datetime.now()
    current_year = today.year
    current_month = today.month

    # Get last month's period for comparison
    last_month_date = today.replace(day=1) - timedelta(days=1)
    last_month = last_month_date.month
    last_year = last_month_date.year

    # --- 1. KPI Cards Data ---
    current_month_payroll = SalaryRecord.objects.filter(year=current_year, month=current_month)
    last_month_payroll = SalaryRecord.objects.filter(year=last_year, month=last_month)

    total_payroll_current = current_month_payroll.aggregate(total=Sum(net_salary_expression))['total'] or 0
    total_payroll_last = last_month_payroll.aggregate(total=Sum(net_salary_expression))['total'] or 0

    staff_paid_count = current_month_payroll.count()
    average_net_salary = current_month_payroll.aggregate(avg=Avg(net_salary_expression))['avg'] or 0

    # Calculate percentage change
    if total_payroll_last > 0:
        percent_change = ((total_payroll_current - total_payroll_last) / total_payroll_last) * 100
    else:
        percent_change = 100 if total_payroll_current > 0 else 0

    # --- 2. Charts Data ---

    # Chart 1: Payroll Cost by Department (Bar Chart)
    dept_payroll = SalaryRecord.objects.filter(year=current_year, month=current_month) \
        .values('staff__department__name') \
        .annotate(total_cost=Sum(net_salary_expression)) \
        .order_by('-total_cost')

    dept_payroll_data = [
        {'name': item['staff__department__name'] or 'Unassigned', 'value': float(item['total_cost'])}
        for item in dept_payroll if item['total_cost']
    ]

    # Chart 2: Salary Trend (Line Chart - Last 12 Months)
    twelve_months_ago = today - timedelta(days=365)
    salary_trend = SalaryRecord.objects.filter(paid_date__gte=twelve_months_ago) \
        .annotate(month_year=TruncMonth('paid_date')) \
        .values('month_year') \
        .annotate(total_net=Sum(net_salary_expression)) \
        .order_by('month_year')

    salary_trend_data = [
        {'month': item['month_year'].strftime('%b %Y'), 'total': float(item['total_net'])}
        for item in salary_trend if item['month_year'] and item['total_net']
    ]

    # Chart 3: Average Salary by Position (Bar Chart)
    position_payroll = SalaryRecord.objects.filter(year=current_year, month=current_month) \
        .values('staff__position__name') \
        .annotate(avg_salary=Avg(net_salary_expression)) \
        .order_by('-avg_salary')

    position_payroll_data = [
        {'name': item['staff__position__name'] or 'Unassigned', 'value': float(item['avg_salary'])}
        for item in position_payroll if item['avg_salary']
    ]

    context = {
        # KPI Cards
        'total_payroll_current': total_payroll_current,
        'staff_paid_count': staff_paid_count,
        'average_net_salary': average_net_salary,
        'percent_change': percent_change,

        # Chart Data (passed as JSON)
        'dept_payroll_data': json.dumps(dept_payroll_data),
        'salary_trend_data': json.dumps(salary_trend_data),
        'position_payroll_data': json.dumps(position_payroll_data),
    }

    return render(request, 'finance/salary_record/dashboard.html', context)


# =================================================================================
# NEW & UPDATED MONEY REMITTANCE VIEWS
# =================================================================================

@login_required
@permission_required('finance.view_moneyremittance')
def remittance_dashboard_view(request):
    """
    Shows a dashboard of staff members and the total unremitted funds they are holding,
    as well as overall statistics.
    """
    # Find all staff who have ever received wallet funding payments
    today = date.today()

    staff_with_transactions = User.objects.filter(
        # Use Q objects to define the conditions in the related model:
        Q(
            patienttransactionmodel__transaction_type='wallet_funding',
            patienttransactionmodel__remittance__isnull=True
        ) |
        Q(
            patienttransactionmodel__transaction_type='consultation_payment',
            patienttransactionmodel__date=today,
            patienttransactionmodel__remittance__isnull=True
        )
    ).distinct()

    # Calculate individual balances
    staff_balances = []
    total_outstanding = Decimal('0.00')

    for staff in staff_with_transactions:
        unremitted_txns = PatientTransactionModel.objects.filter(
            received_by=staff,
            remittance__isnull=True
        ).filter(
            # Use Q objects to create the required OR condition:
            # (wallet_funding, any date) OR (consultation_payment, today's date)
            Q(transaction_type='wallet_funding') |
            Q(
                transaction_type='consultation_payment',
                date=today  # Filter on the 'date' DateField for today only
            )
        )
        cash_total = unremitted_txns.filter(payment_method__iexact='cash').aggregate(total=Sum('amount'))['total'] or 0

        if cash_total > 0:
            total_outstanding += cash_total
            staff_balances.append({
                'staff': staff,
                'cash_outstanding': cash_total,
            })

    # Calculate overall remitted totals
    total_remitted = MoneyRemittance.objects.filter(status='APPROVED').aggregate(total=Sum('amount_remitted_cash'))[
                         'total'] or 0

    context = {
        'staff_balances': sorted(staff_balances, key=lambda x: x['cash_outstanding'], reverse=True),
        'total_outstanding': total_outstanding,
        'total_remitted': total_remitted,
    }
    return render(request, 'finance/remittance/dashboard.html', context)


@login_required
@permission_required('finance.add_moneyremittance')
def create_remittance_view(request):
    """
    Allows an admin/accountant to record a remittance from a staff member.
    The staff dropdown only shows users with unremitted funds.
    """
    today = timezone.now().date()
    # Find users who have unremitted cash transactions to populate the dropdown
    owing_staff_ids = PatientTransactionModel.objects.filter(
        remittance__isnull=True,
        payment_method__iexact='cash'
    ).filter(
        # Use Q objects to create the required OR condition:
        # (wallet_funding, any date) OR (consultation_payment, today's date)
        Q(transaction_type='wallet_funding') |
        Q(
            transaction_type='consultation_payment',
            date=today  # Filter on the 'date' DateField for today only
        )
    ).values_list('received_by_id', flat=True).distinct()

    owing_staff = User.objects.filter(id__in=owing_staff_ids)

    if request.method == 'POST':
        form = MoneyRemittanceForm(request.POST)
        # --- THIS IS THE FIX ---
        # We must set the queryset for the field before validation runs.
        # This ensures the submitted choice is considered valid.
        form.fields['remitted_by'].queryset = owing_staff

        if form.is_valid():
            remitted_by_user = form.cleaned_data['remitted_by']

            transactions_to_remit = PatientTransactionModel.objects.filter(
                received_by=remitted_by_user,
                transaction_type='wallet_funding',
                remittance__isnull=True
            )

            with transaction.atomic():
                if not transactions_to_remit.exists():
                    messages.error(request, "This staff member has no pending funds to remit.")
                    return redirect('remittance_create')

                cash_expected = \
                transactions_to_remit.filter(payment_method__iexact='cash').aggregate(total=Sum('amount'))['total'] or 0
                transfer_expected = \
                transactions_to_remit.filter(payment_method__iexact='transfer').aggregate(total=Sum('amount'))[
                    'total'] or 0

                remittance = form.save(commit=False)
                remittance.total_cash_expected = cash_expected
                remittance.total_transfer_expected = transfer_expected
                remittance.save()

                transactions_to_remit.update(remittance=remittance)

                messages.success(request,
                                 f"Remittance batch {remittance.remittance_id} created for {remitted_by_user.get_full_name()} and is awaiting approval.")
                return redirect('remittance_list')
    else:
        form = MoneyRemittanceForm()
        # Limit the 'remitted_by' dropdown to only staff who owe money
        form.fields['remitted_by'].queryset = owing_staff

    context = {
        'form': form,
    }
    return render(request, 'finance/remittance/create.html', context)


@login_required
def get_staff_remittance_details_ajax(request):
    """
    AJAX endpoint to get the unremitted totals for a selected staff member.
    """
    staff_id = request.GET.get('staff_id')
    if not staff_id:
        return JsonResponse({'success': False, 'error': 'Staff ID is required.'})

    try:
        staff = User.objects.get(pk=staff_id)
        today = timezone.now().date()

        unremitted_txns = PatientTransactionModel.objects.filter(
            received_by=staff,
            remittance__isnull=True
        ).filter(
            # Use Q objects to create the required OR condition:
            # (wallet_funding, any date) OR (consultation_payment, today's date)
            Q(transaction_type='wallet_funding') |
            Q(
                transaction_type='consultation_payment',
                date=today  # Filter on the 'date' DateField for today only
            )
        )
        cash_expected = unremitted_txns.filter(payment_method__iexact='cash').aggregate(total=Sum('amount'))[
                            'total'] or 0
        transfer_expected = unremitted_txns.filter(payment_method__iexact='transfer').aggregate(total=Sum('amount'))[
                                'total'] or 0

        return JsonResponse({
            'success': True,
            'cash_expected': f'{cash_expected:,.2f}',
            'transfer_expected': f'{transfer_expected:,.2f}',
            'total_expected': f'{cash_expected + transfer_expected:,.2f}',
        })
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Staff not found.'})


class RemittanceListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = MoneyRemittance
    permission_required = 'finance.view_moneyremittance'
    template_name = 'finance/remittance/list.html'
    context_object_name = "remittance_list"
    paginate_by = 20

    def get_queryset(self):
        queryset = MoneyRemittance.objects.select_related('remitted_by', 'approved_by')

        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        staff_id = self.request.GET.get('staff_id')

        if start_date and end_date:
            queryset = queryset.filter(created_at__date__range=[start_date, end_date])

        if staff_id:
            queryset = queryset.filter(remitted_by_id=staff_id)

        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Provide a list of staff who have ever received funds for the filter dropdown
        staff_ids = PatientTransactionModel.objects.filter(transaction_type='wallet_funding').values_list(
            'received_by_id', flat=True).distinct()
        context['staff_list'] = User.objects.filter(id__in=staff_ids)
        return context


class RemittanceDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = MoneyRemittance
    permission_required = 'finance.change_moneyremittance'  # Permission to approve
    template_name = 'finance/remittance/detail.html'
    context_object_name = "remittance"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['transactions'] = self.object.transactions.select_related('patient').order_by('-created_at')
        return context

    def post(self, request, *args, **kwargs):
        remittance = self.get_object()
        action = request.POST.get('action')

        if action == 'approve' and remittance.status == 'PENDING':
            remittance.status = 'APPROVED'
            remittance.approved_by = request.user
            remittance.approved_at = timezone.now()
            remittance.save()
            messages.success(request, f"Remittance {remittance.remittance_id} has been approved.")

        elif action == 'query' and remittance.status == 'PENDING':
            remittance.status = 'DISCREPANCY'
            remittance.approved_by = request.user  # The user who noted the discrepancy
            remittance.approved_at = timezone.now()
            remittance.save()
            messages.warning(request, f"Discrepancy noted for remittance {remittance.remittance_id}.")

        return redirect('remittance_detail', pk=remittance.pk)


# Add this new view to your views.py file
@login_required
@permission_required('finance.view_moneyremittance')
def staff_remittance_detail_view(request, staff_id):
    staff = get_object_or_404(User, pk=staff_id)

    unremitted_transactions = PatientTransactionModel.objects.filter(
        received_by=staff,
        transaction_type='wallet_funding',
        remittance__isnull=True
    ).select_related('patient').order_by('-created_at')

    # Calculate totals for each payment method
    summary = unremitted_transactions.aggregate(
        cash_total=Sum('amount', filter=Q(payment_method__iexact='cash')),
        transfer_total=Sum('amount', filter=Q(payment_method__iexact='transfer')),
        card_total=Sum('amount', filter=Q(payment_method__iexact='card'))
    )

    cash = summary.get('cash_total') or 0
    transfer = summary.get('transfer_total') or 0
    card = summary.get('card_total') or 0

    context = {
        'staff_member': staff,
        'transactions': unremitted_transactions,
        'cash_total': cash,
        'transfer_total': transfer,
        'card_total': card,
        'grand_total': cash + transfer + card, # The overall total
    }
    return render(request, 'finance/remittance/staff_detail.html', context)


@login_required
@permission_required('finance.view_patienttransactionmodel', raise_exception=True)
def finance_dashboard(request):
    """Finance dashboard showing patient transaction analytics and wallet data"""

    today = timezone.now().date()
    current_month_start = today.replace(day=1)
    last_month_start = (current_month_start - timedelta(days=1)).replace(day=1)
    last_month_end = current_month_start - timedelta(days=1)
    thirty_days_ago = today - timedelta(days=30)

    # Basic transaction metrics
    total_transactions = PatientTransactionModel.objects.filter(status='completed')

    # Revenue calculations (all inflow transactions)
    revenue_transactions = total_transactions.filter(transaction_direction='in')
    total_revenue = revenue_transactions.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')

    # Current month revenue
    revenue_month = revenue_transactions.filter(
        date__gte=current_month_start
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    # Last month revenue for growth calculation
    revenue_last_month = revenue_transactions.filter(
        date__range=[last_month_start, last_month_end]
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    # Calculate revenue growth
    if revenue_last_month > 0:
        revenue_growth = float(((revenue_month - revenue_last_month) / revenue_last_month) * 100)
        revenue_growth = round(revenue_growth, 1)
    else:
        revenue_growth = 0 if revenue_month == 0 else 100

    # Today's metrics
    revenue_today = revenue_transactions.filter(
        date=today
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    transactions_today_count = total_transactions.filter(date=today).count()

    # Total wallet balances (current patient wallet funds)
    total_wallet_balance = PatientWalletModel.objects.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')

    # Outflow calculations (all outflow transactions - payments made by patients)
    outflow_transactions = total_transactions.filter(transaction_direction='out')
    total_outflow = outflow_transactions.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')

    outflow_today = outflow_transactions.filter(
        date=today
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    # Net revenue (inflow - outflow)
    net_revenue = total_revenue - total_outflow

    # Pending actions
    pending_transactions = PatientTransactionModel.objects.filter(status='pending').count()
    pending_remittances = MoneyRemittance.objects.filter(status='PENDING').count()
    discrepancy_remittances = MoneyRemittance.objects.filter(status='DISCREPANCY').count()

    # Daily revenue data for chart (last 30 days)
    daily_revenue_data = []
    for i in range(30):
        date_point = today - timedelta(days=29 - i)
        daily_revenue = revenue_transactions.filter(
            date=date_point
        ).aggregate(total=Sum('amount'))['total'] or 0

        daily_revenue_data.append({
            'date': date_point.strftime('%Y-%m-%d'),
            'revenue': float(daily_revenue)
        })

    # Monthly comparison data (last 12 months)
    monthly_comparison_data = []
    for i in range(12):
        # Calculate the first day of the month for i months ago
        target_date = today.replace(day=1) - timedelta(days=i * 30)
        # Adjust to actual first day of that month
        month_start = target_date.replace(day=1)
        # Get last day of the month
        last_day = calendar.monthrange(month_start.year, month_start.month)[1]
        month_end = month_start.replace(day=last_day)

        monthly_revenue = revenue_transactions.filter(
            date__range=[month_start, month_end]
        ).aggregate(total=Sum('amount'))['total'] or 0

        monthly_outflow = outflow_transactions.filter(
            date__range=[month_start, month_end]
        ).aggregate(total=Sum('amount'))['total'] or 0

        monthly_comparison_data.insert(0, {
            'month': month_start.strftime('%b %Y'),
            'revenue': float(monthly_revenue),
            'outflow': float(monthly_outflow),
            'net': float(monthly_revenue - monthly_outflow)
        })

    # Transaction types analysis (current month)
    transaction_types = total_transactions.filter(
        date__gte=current_month_start
    ).values('transaction_type').annotate(
        count=Count('id'),
        total=Sum('amount')
    ).order_by('-total')[:10]

    # Payment methods analysis (current month)
    payment_methods = total_transactions.filter(
        date__gte=current_month_start,
        payment_method__isnull=False
    ).exclude(payment_method='').values('payment_method').annotate(
        count=Count('id'),
        total=Sum('amount')
    ).order_by('-total')[:10]

    # Recent large transactions (≥ ₦10,000)
    recent_large_transactions = total_transactions.filter(
        amount__gte=Decimal('10000.00')
    ).select_related('patient')[:20]

    # Wallet analytics
    active_wallets_count = PatientWalletModel.objects.filter(amount__gt=0).count()
    average_wallet_balance = PatientWalletModel.objects.filter(
        amount__gt=0
    ).aggregate(avg=Avg('amount'))['avg'] or Decimal('0.00')

    # Top wallet balances
    top_wallet_holders = PatientWalletModel.objects.filter(
        amount__gt=0
    ).select_related('patient').order_by('-amount')[:10]

    # Service popularity (from other payment services)
    popular_services = total_transactions.filter(
        other_service__isnull=False,
        date__gte=current_month_start
    ).values('other_service__name', 'other_service__category').annotate(
        count=Count('id'),
        total=Sum('amount')
    ).order_by('-count')[:10]

    context = {
        # Main metrics
        'total_revenue': total_revenue,
        'revenue_month': revenue_month,
        'revenue_growth': revenue_growth,
        'revenue_today': revenue_today,
        'total_outflow': total_outflow,
        'outflow_today': outflow_today,
        'net_revenue': net_revenue,
        'transactions_today_count': transactions_today_count,

        # Wallet metrics
        'total_wallet_balance': total_wallet_balance,
        'active_wallets_count': active_wallets_count,
        'average_wallet_balance': average_wallet_balance,
        'top_wallet_holders': top_wallet_holders,

        # Pending actions
        'pending_transactions': pending_transactions,
        'pending_remittances': pending_remittances,
        'discrepancy_remittances': discrepancy_remittances,

        # Analytics data
        'transaction_types': transaction_types,
        'payment_methods': payment_methods,
        'recent_large_transactions': recent_large_transactions,
        'popular_services': popular_services,

        # Chart data (JSON serialized)
        'daily_revenue_data': json.dumps(daily_revenue_data),
        'monthly_comparison_data': json.dumps(monthly_comparison_data),

        # Additional context
        'current_month': current_month_start.strftime('%B %Y'),
        'today': today,
    }

    return render(request, 'finance/dashboard.html', context)


@login_required
def finance_dashboard_print(request):
    """Printable version of the finance dashboard"""

    # Get simplified data for printing
    today = timezone.now().date()
    start_of_month = today.replace(day=1)

    # Basic statistics
    total_transactions = PatientTransactionModel.objects.filter(status='completed')
    revenue_transactions = total_transactions.filter(transaction_direction='in')

    total_revenue = revenue_transactions.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    revenue_month = revenue_transactions.filter(date__gte=start_of_month).aggregate(total=Sum('amount'))[
                        'total'] or Decimal('0.00')

    total_expenses = Expense.objects.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    expenses_month = Expense.objects.filter(date__gte=start_of_month).aggregate(total=Sum('amount'))[
                         'total'] or Decimal('0.00')

    total_income = Income.objects.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    income_month = Income.objects.filter(date__gte=start_of_month).aggregate(total=Sum('amount'))['total'] or Decimal(
        '0.00')

    # Top categories
    top_expense_categories = ExpenseCategory.objects.annotate(
        total_amount=Sum('expense__amount')
    ).filter(total_amount__gt=0).order_by('-total_amount')[:5]

    top_income_categories = IncomeCategory.objects.annotate(
        total_amount=Sum('income__amount')
    ).filter(total_amount__gt=0).order_by('-total_amount')[:5]

    context = {
        'total_revenue': total_revenue,
        'revenue_month': revenue_month,
        'total_expenses': total_expenses,
        'expenses_month': expenses_month,
        'total_income': total_income,
        'income_month': income_month,
        'net_profit': (total_revenue + total_income) - total_expenses,
        'net_profit_month': (revenue_month + income_month) - expenses_month,
        'top_expense_categories': top_expense_categories,
        'top_income_categories': top_income_categories,
        'current_date': today,
        'total_transactions_count': total_transactions.count(),
    }

    return render(request, 'finance/dashboard_print.html', context)


@login_required
@permission_required('finance.add_patientrefundmodel', raise_exception=True)
@require_http_methods(["GET", "POST"])
def finance_wallet_withdrawal(request, patient_id):
    """Handles patient wallet withdrawal."""
    patient = get_object_or_404(PatientModel, id=patient_id)
    wallet = get_object_or_404(PatientWalletModel, patient=patient)

    if request.method == 'POST':
        try:
            amount_to_withdraw = _to_decimal(request.POST.get('amount'))
            notes = request.POST.get('notes', '')

            if amount_to_withdraw <= Decimal('0.00'):
                return JsonResponse({'success': False, 'error': 'Amount must be positive.'}, status=400)

            with transaction.atomic():
                # Re-fetch and lock wallet
                wallet_qs = PatientWalletModel.objects.select_for_update().filter(patient=patient)
                wallet = wallet_qs.first()

                if wallet.amount < amount_to_withdraw:
                    shortfall = _quantize_money(amount_to_withdraw - wallet.amount)
                    return JsonResponse({
                        'success': False,
                        'error': 'Insufficient wallet balance for withdrawal.',
                        'shortfall': float(shortfall),
                        'formatted_shortfall': f'₦{shortfall:,.2f}'
                    }, status=400)

                # Deduct
                old_balance = wallet.amount
                wallet.amount = _quantize_money(wallet.amount - amount_to_withdraw)
                wallet.save()

                # 1. Create OUT transaction (Wallet Deduction)
                transaction_out = PatientTransactionModel.objects.create(
                    patient=patient,
                    transaction_type='wallet_withdrawal',
                    transaction_direction='out',
                    amount=amount_to_withdraw,
                    old_balance=old_balance,
                    new_balance=wallet.amount,
                    date=timezone.now().date(),
                    received_by=request.user,
                    payment_method='cash',  # Assuming cash
                    status='completed'
                )

                # 2. Create Withdrawal Record
                WalletWithdrawalRecord.objects.create(
                    patient=patient,
                    transaction=transaction_out,
                    amount=amount_to_withdraw,
                    withdrawn_by=request.user,
                    notes=notes
                )

                return JsonResponse({
                    'success': True,
                    'message': f"Successfully withdrew ₦{amount_to_withdraw:,.2f}.",
                    'new_balance': float(wallet.amount),
                    'redirect_url': reverse('patient_wallet_dashboard', args=[patient.id])
                })

        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Error processing withdrawal: {str(e)}'}, status=500)

    context = {
        'patient': patient,
        'wallet_balance': wallet.amount
    }
    return render(request, 'finance/wallet/withdrawal.html', context)


REFUNDABLE_STATUSES = ['paid', 'partially_dispensed']


@login_required
@permission_required('finance.add_patientrefundmodel', raise_exception=True)
@require_http_methods(["GET", "POST"])
def finance_process_refund(request, patient_id):
    """
    Handles refunds for paid items/services that have not been fully consumed/rendered.
    """
    patient = get_object_or_404(PatientModel, id=patient_id)

    # Helper to get name from order/transaction (Moved here for scope)
    def get_item_name(item):
        if hasattr(item,
                   'drug'): return f"{getattr(item.drug, 'brand_name', 'Drug')} ({getattr(item.drug, 'generic_name', 'N/A')})"
        if hasattr(item, 'template'): return item.template.name
        if hasattr(item, 'service') and item.service: return item.service.name
        if hasattr(item, 'service_item') and item.service_item: return item.service_item.name
        return "Unknown Item"

    # --- GET LOGIC (No changes needed here, only in POST) ---
    if request.method == 'GET':
        # ... (GET logic remains the same) ...
        one_year_ago = timezone.now() - timedelta(days=365)

        # Drug Orders: Paid but not fully dispensed
        refundable_drugs = DrugOrderModel.objects.filter(
            patient=patient,
            status__in=REFUNDABLE_STATUSES,
            quantity_dispensed__lt=F('quantity_ordered'),
            ordered_at__gte=one_year_ago
        ).select_related('drug')

        # Lab/Scan Orders: Paid but awaiting results/processing ('paid' or 'collected')
        refundable_labs = LabTestOrderModel.objects.filter(
            patient=patient,
            status__in=['paid', 'collected'],
            ordered_at__gte=one_year_ago
        ).select_related('template')

        refundable_scans = ScanOrderModel.objects.filter(
            patient=patient,
            status__in=['paid', 'collected'],
            ordered_at__gte=one_year_ago
        ).select_related('template')

        # Services/Items: Paid but not fully dispensed (items) or not yet completed (services)
        refundable_services = PatientServiceTransaction.objects.filter(
            patient=patient,
            status__in=REFUNDABLE_STATUSES,
            amount_paid__gt=Decimal('0.00'),
            created_at__gte=one_year_ago
        ).select_related('service', 'service_item')

        items_to_display = []

        # The actual item list compilation is done here...
        # We re-include the items list construction to avoid leaving it as a comment block
        for o in refundable_drugs:
            remaining_qty = o.quantity_ordered - o.quantity_dispensed
            items_to_display.append({
                'ref': f'drug-{o.id}',
                'type': 'Drug Order',
                'name': get_item_name(o),
                'status': o.get_status_display(),
                'notes': f"Remaining Qty: {remaining_qty}",
                'amount': 0.00
            })

        for o in refundable_labs:
            items_to_display.append({
                'ref': f'lab-{o.id}',
                'type': 'Lab Test',
                'name': get_item_name(o),
                'status': o.get_status_display(),
                'notes': f"Total Charged: ₦{o.amount_charged:,.2f}",
                'amount': float(o.amount_charged)
            })

        for o in refundable_scans:
            items_to_display.append({
                'ref': f'scan-{o.id}',
                'type': 'Scan Order',
                'name': get_item_name(o),
                'status': o.get_status_display(),
                'notes': f"Total Charged: ₦{o.amount_charged:,.2f}",
                'amount': float(o.amount_charged)
            })

        for t in refundable_services:
            items_to_display.append({
                'ref': f'service-{t.id}',
                'type': 'Service/Item',
                'name': get_item_name(t),
                'status': t.get_status_display(),
                'notes': f"Amount Paid: ₦{t.amount_paid:,.2f}",
                'amount': float(t.amount_paid)
            })

        context = {
            'patient': patient,
            'items': items_to_display
        }
        return render(request, 'finance/wallet/refund.html', context)

    # 2. POST LOGIC: Handles the actual refund transaction
    if request.method == 'POST':
        try:
            selected_items = request.POST.getlist('selected_items[]')

            if not selected_items:
                return JsonResponse({'success': False, 'error': 'No items selected for refund.'}, status=400)

            refund_total = Decimal('0.00')
            items_to_update = []

            with transaction.atomic():
                wallet_qs = PatientWalletModel.objects.select_for_update().filter(patient=patient)
                wallet = wallet_qs.first()

                if not wallet:
                    return JsonResponse({'success': False, 'error': 'Wallet not found.'}, status=404)

                # --- CALCULATE REFUND & UPDATE ITEM STATUSES ---
                for item_ref in selected_items:
                    try:
                        item_type, item_id = item_ref.split('-')
                        item_id = int(item_id)

                        refund_amount = Decimal('0.00')
                        item = None

                        if item_type == 'drug':
                            item = DrugOrderModel.objects.get(pk=item_id, patient=patient,
                                                              status__in=REFUNDABLE_STATUSES)

                            # FIX: Ensure all operands are Decimal
                            remaining_qty = _to_decimal(item.quantity_ordered - item.quantity_dispensed)

                            if remaining_qty <= Decimal('0.00'):
                                continue

                            unit_net_price = _to_decimal(item.drug.selling_price)

                            if unit_net_price <= Decimal('0.00'):
                                continue

                            refund_amount = remaining_qty * unit_net_price  # Multiplication is now safe

                            item.status = 'cancelled'
                            item.quantity_dispensed = item.quantity_ordered

                        elif item_type == 'lab':
                            item = LabTestOrderModel.objects.get(pk=item_id, patient=patient,
                                                                 status__in=['paid', 'collected'])
                            refund_amount = _to_decimal(item.amount_charged)
                            if refund_amount <= Decimal('0.00'): continue
                            item.status = 'cancelled'

                        elif item_type == 'scan':
                            item = ScanOrderModel.objects.get(pk=item_id, patient=patient,
                                                              status__in=['paid', 'collected'])
                            refund_amount = _to_decimal(item.amount_charged)
                            if refund_amount <= Decimal('0.00'): continue
                            item.status = 'cancelled'

                        elif item_type == 'service':
                            item = PatientServiceTransaction.objects.get(pk=item_id, patient=patient,
                                                                         status__in=REFUNDABLE_STATUSES)

                            refund_amount = _to_decimal(item.amount_paid)  # Defensive use of _to_decimal
                            if refund_amount <= Decimal('0.00'): continue

                            item.status = 'cancelled'
                            item.amount_paid = Decimal('0.00')

                            # Finalize
                        if item and refund_amount > Decimal('0.00'):
                            items_to_update.append(item)
                            refund_total += _quantize_money(refund_amount)

                    except Exception as e:
                        # Log the error but continue processing other items
                        print(f"Refund item processing error for {item_ref}: {e}")
                        continue

                        # --- FAILURE POINT CHECK ---
                if refund_total <= Decimal('0.00'):
                    return JsonResponse(
                        {'success': False, 'error': 'No valid refund amount calculated. Please check selected items.'},
                        status=400)

                # ... (rest of the transaction logic: update wallet, create transaction, save items)
                old_balance = wallet.amount
                wallet.amount += refund_total
                wallet.save()

                # Create IN transaction (Refund)
                PatientTransactionModel.objects.create(
                    patient=patient,
                    transaction_type='refund_to_wallet',
                    transaction_direction='in',
                    amount=refund_total,
                    old_balance=old_balance,
                    new_balance=wallet.amount,
                    date=timezone.now().date(),
                    received_by=request.user,
                    payment_method='wallet_return',
                    status='completed'
                )

                # Save all updated items
                for item in items_to_update:
                    item.save()

                return JsonResponse({
                    'success': True,
                    'message': f"Successfully refunded ₦{refund_total:,.2f} to wallet.",
                    'new_balance': float(wallet.amount),
                    'redirect_url': reverse('patient_wallet_dashboard', args=[patient.id])
                })

        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Error processing refund: {type(e).__name__}: {str(e)}'},
                                status=500)

# ----------------------------------------------------
# 3. VIEW: Wallet History
# ----------------------------------------------------

@login_required
@permission_required('finance.add_patienttransactionmodel', raise_exception=True)
def finance_wallet_history(request, patient_id):
    """Displays a detailed history of all wallet-related transactions."""
    patient = get_object_or_404(PatientModel, id=patient_id)

    # Get all transactions, ordered by date
    history = PatientTransactionModel.objects.filter(
        patient=patient
    ).select_related('received_by').order_by('-created_at')

    context = {
        'patient': patient,
        'history': history
    }
    return render(request, 'finance/wallet/history.html', context)


# ----------------------------------------------------
# 4. VIEW: Central Entry Point
# ----------------------------------------------------

@login_required
def finance_wallet_tools_entry(request):
    """Central entry point for wallet tools (Verification -> Options)."""
    # This view just renders the HTML containing the verification form.
    return render(request, 'finance/wallet/tools_entry.html')


@login_required
@permission_required('finance.add_patienttransactionmodel', raise_exception=True)
@require_http_methods(["POST"])
def process_direct_payment(request):
    """
    Process direct payment for selected items with optional wallet usage.
    This replaces the old fund-then-pay flow.
    """
    try:
        # Get form data
        patient_id = request.POST.get('patient_id')
        use_wallet = request.POST.get('use_wallet') == 'true'
        payment_method = request.POST.get('payment_method', 'cash')
        direct_amount = Decimal(request.POST.get('direct_amount', '0'))

        # Get selected items (sent as arrays from frontend)
        selected_items = request.POST.getlist('selected_items[]')
        selected_types = request.POST.getlist('selected_types[]')

        # Validate inputs
        if not patient_id:
            return JsonResponse({'error': 'Patient ID required'}, status=400)

        if not selected_items or not selected_types:
            return JsonResponse({'error': 'No items selected for payment'}, status=400)

        if len(selected_items) != len(selected_types):
            return JsonResponse({'error': 'Items and types mismatch'}, status=400)

        # Get patient and wallet
        patient = get_object_or_404(PatientModel, id=patient_id)
        wallet, created = PatientWalletModel.objects.get_or_create(
            patient=patient,
            defaults={'amount': Decimal('0.00')}
        )

        # Group items by type for processing
        items_by_type = {
            'drugs': [],
            'labs': [],
            'scans': [],
            'services': []
        }

        for item_id, item_type in zip(selected_items, selected_types):
            if item_type in items_by_type:
                items_by_type[item_type].append(item_id)

        # Fetch all orders
        drug_orders = DrugOrderModel.objects.filter(
            id__in=items_by_type['drugs'],
            patient=patient,
            status='pending'
        ).select_related('drug')

        lab_orders = LabTestOrderModel.objects.filter(
            id__in=items_by_type['labs'],
            patient=patient,
            status='pending'
        ).select_related('template')

        scan_orders = ScanOrderModel.objects.filter(
            id__in=items_by_type['scans'],
            patient=patient,
            status='pending'
        ).select_related('template')

        service_transactions = PatientServiceTransaction.objects.filter(
            id__in=items_by_type['services'],
            patient=patient,
            status='pending_payment'
        ).select_related('service', 'service_item')

        # Calculate total amount needed
        total_amount = Decimal('0.00')

        for order in drug_orders:
            total_amount += order.total_amount

        for order in lab_orders:
            # Use amount_charged for lab orders
            total_amount += order.amount_charged or Decimal('0.00')

        for order in scan_orders:
            # Use amount_charged for scan orders (assuming same structure)
            total_amount += order.amount_charged or Decimal('0.00')

        for trans in service_transactions:
            total_amount += trans.total_amount

        # Calculate wallet usage
        wallet_amount_to_use = Decimal('0.00')
        if use_wallet and wallet.amount > 0:
            wallet_amount_to_use = min(wallet.amount, total_amount)

        remaining_amount = total_amount - wallet_amount_to_use

        # Validate direct payment amount matches remaining
        if abs(direct_amount - remaining_amount) > Decimal('0.01'):
            return JsonResponse({
                'error': f'Direct payment amount (₦{direct_amount}) does not match required amount (₦{remaining_amount})'
            }, status=400)

        # Start atomic transaction
        with transaction.atomic():
            old_wallet_balance = wallet.amount
            parent_transaction = None
            wallet_withdrawal_transaction = None

            # Step 1: Deduct from wallet if applicable
            if wallet_amount_to_use > 0:
                wallet.amount -= wallet_amount_to_use
                wallet.save()

                wallet_withdrawal_transaction = PatientTransactionModel.objects.create(
                    patient=patient,
                    transaction_type='wallet_withdrawal',
                    transaction_direction='out',
                    amount=wallet_amount_to_use,
                    old_balance=old_wallet_balance,
                    new_balance=wallet.amount,
                    payment_method='wallet',
                    received_by=request.user,
                    status='completed',
                    date=timezone.now().date(),
                    wallet_amount_used=wallet_amount_to_use,
                    direct_payment_amount=Decimal('0.00')
                )

            # Step 2: Create parent transaction for direct payment (if any)
            if direct_amount > 0:
                parent_transaction = PatientTransactionModel.objects.create(
                    patient=patient,
                    transaction_type='direct_payment',
                    transaction_direction='in',
                    amount=direct_amount,
                    old_balance=wallet.amount,
                    new_balance=wallet.amount,  # Wallet unchanged by direct payment
                    payment_method=payment_method,
                    received_by=request.user,
                    status='completed',
                    date=timezone.now().date(),
                    wallet_amount_used=Decimal('0.00'),
                    direct_payment_amount=direct_amount
                )

            # Step 3: Create child transactions and update orders
            child_transactions_created = []

            # Process drug orders
            for order in drug_orders:
                order_amount = order.total_amount
                wallet_portion = min(wallet_amount_to_use, order_amount)
                direct_portion = order_amount - wallet_portion
                wallet_amount_to_use -= wallet_portion

                child_trans = PatientTransactionModel.objects.create(
                    patient=patient,
                    transaction_type='drug_payment',
                    transaction_direction='out',
                    amount=order_amount,
                    old_balance=wallet.amount,
                    new_balance=wallet.amount,
                    payment_method=payment_method if direct_portion > 0 else 'wallet',
                    received_by=request.user,
                    status='completed',
                    date=timezone.now().date(),
                    parent_transaction=parent_transaction if direct_portion > 0 else wallet_withdrawal_transaction,
                    wallet_amount_used=wallet_portion,
                    direct_payment_amount=direct_portion,
                    drug_order=order
                )

                # Update order status
                order.status = 'paid'
                order.save()

                child_transactions_created.append(child_trans)

            # Process lab orders
            for order in lab_orders:
                order_amount = order.amount_charged or Decimal('0.00')

                wallet_portion = min(wallet_amount_to_use, order_amount)
                direct_portion = order_amount - wallet_portion
                wallet_amount_to_use -= wallet_portion

                child_trans = PatientTransactionModel.objects.create(
                    patient=patient,
                    transaction_type='lab_payment',
                    transaction_direction='out',
                    amount=order_amount,
                    old_balance=wallet.amount,
                    new_balance=wallet.amount,
                    payment_method=payment_method if direct_portion > 0 else 'wallet',
                    received_by=request.user,
                    status='completed',
                    date=timezone.now().date(),
                    parent_transaction=parent_transaction if direct_portion > 0 else wallet_withdrawal_transaction,
                    wallet_amount_used=wallet_portion,
                    direct_payment_amount=direct_portion,
                    lab_structure=order
                )

                order.status = 'paid'
                order.save()

                child_transactions_created.append(child_trans)

            # Process scan orders
            for order in scan_orders:
                order_amount = order.amount_charged or Decimal('0.00')

                wallet_portion = min(wallet_amount_to_use, order_amount)
                direct_portion = order_amount - wallet_portion
                wallet_amount_to_use -= wallet_portion

                child_trans = PatientTransactionModel.objects.create(
                    patient=patient,
                    transaction_type='scan_payment',
                    transaction_direction='out',
                    amount=order_amount,
                    old_balance=wallet.amount,
                    new_balance=wallet.amount,
                    payment_method=payment_method if direct_portion > 0 else 'wallet',
                    received_by=request.user,
                    status='completed',
                    date=timezone.now().date(),
                    parent_transaction=parent_transaction if direct_portion > 0 else wallet_withdrawal_transaction,
                    wallet_amount_used=wallet_portion,
                    direct_payment_amount=direct_portion,
                    scan_order=order
                )

                order.status = 'paid'
                order.save()

                child_transactions_created.append(child_trans)

            # Process service transactions
            for svc_trans in service_transactions:
                order_amount = svc_trans.total_amount
                wallet_portion = min(wallet_amount_to_use, order_amount)
                direct_portion = order_amount - wallet_portion
                wallet_amount_to_use -= wallet_portion

                child_trans = PatientTransactionModel.objects.create(
                    patient=patient,
                    transaction_type='service',
                    transaction_direction='out',
                    amount=order_amount,
                    old_balance=wallet.amount,
                    new_balance=wallet.amount,
                    payment_method=payment_method if direct_portion > 0 else 'wallet',
                    received_by=request.user,
                    status='completed',
                    date=timezone.now().date(),
                    parent_transaction=parent_transaction if direct_portion > 0 else wallet_withdrawal_transaction,
                    wallet_amount_used=wallet_portion,
                    direct_payment_amount=direct_portion,
                    service=svc_trans
                )

                svc_trans.status = 'completed'
                svc_trans.save()

                child_transactions_created.append(child_trans)

            # Determine which transaction to show in detail
            detail_transaction_id = parent_transaction.id if parent_transaction else wallet_withdrawal_transaction.id

            return JsonResponse({
                'success': True,
                'message': f'Payment successful! ₦{total_amount:,.2f} paid for {len(child_transactions_created)} item(s)',
                'total_paid': float(total_amount),
                'wallet_used': float(wallet_amount_to_use + (old_wallet_balance - wallet.amount)),
                'direct_paid': float(direct_amount),
                'new_wallet_balance': float(wallet.amount),
                'items_paid': len(child_transactions_created),
                'redirect_url': reverse('transaction_detail', args=[detail_transaction_id])
            })

    except PatientModel.DoesNotExist:
        return JsonResponse({'error': 'Patient not found'}, status=404)
    except ValueError as e:
        return JsonResponse({'error': f'Invalid amount: {str(e)}'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Error processing payment: {str(e)}'}, status=500)


@login_required
def transaction_detail(request, transaction_id):
    """
    Display transaction details with support for:
    - Standalone transactions (old style)
    - Parent transactions (with multiple child items)
    - Child transactions (part of a parent payment)
    """
    transaction = get_object_or_404(
        PatientTransactionModel.objects.select_related(
            'patient',
            'received_by',
            'received_by__user_staff_profile',
            'received_by__user_staff_profile__staff',
            'parent_transaction',
            'lab_structure',
            'lab_structure__template',
            'service',
            'service__service',
            'service__service__category',
            'service__service_item',
            'service__service_item__category',
            'other_service',
            'fee_structure',
            'admission',
            'surgery',
        ).prefetch_related(
            Prefetch(
                'child_transactions',
                queryset=PatientTransactionModel.objects.select_related(
                    'lab_structure',
                    'lab_structure__template',
                    'service',
                    'service__service',
                    'service__service__category',
                    'service__service_item',
                    'service__service_item__category',
                ).order_by('transaction_type', 'created_at')
            ),
        ),
        id=transaction_id
    )

    # Determine transaction structure
    is_parent = transaction.is_parent_transaction
    is_child = transaction.is_child_transaction
    is_standalone = transaction.is_standalone_transaction

    # Group child transactions by type for easier template rendering
    child_transactions_by_type = None
    if is_parent:
        child_transactions_by_type = {
            'drugs': [],
            'labs': [],
            'scans': [],
            'services': []
        }

        for child in transaction.child_transactions.all():
            if child.transaction_type == 'drug_payment':
                child_transactions_by_type['drugs'].append(child)
            elif child.transaction_type == 'lab_payment':
                child_transactions_by_type['labs'].append(child)
            elif child.transaction_type == 'scan_payment':
                child_transactions_by_type['scans'].append(child)
            elif child.transaction_type == 'service':
                child_transactions_by_type['services'].append(child)

    context = {
        'transaction': transaction,
        'is_parent': is_parent,
        'is_child': is_child,
        'is_standalone': is_standalone,
        'child_transactions_by_type': child_transactions_by_type,
    }

    return render(request, 'finance/payment/detail.html', context)


@login_required
def transaction_list(request):
    """
    List all transactions with support for filtering
    Shows item count badge for parent transactions
    EXCLUDES child transactions (they're part of parent payments)
    """
    # Only show top-level transactions (no parent)
    # This includes:
    # 1. Direct payment (parent of multi-item payments)
    # 2. Wallet withdrawal (parent of wallet-only multi-item payments)
    # 3. Standalone transactions (old style single payments)
    transactions = PatientTransactionModel.objects.filter(
        parent_transaction__isnull=True  # Only show top-level transactions
    ).select_related(
        'patient',
        'received_by',
        'received_by__user_staff_profile',
        'received_by__user_staff_profile__staff',
    ).prefetch_related(
        'child_transactions'
    ).order_by('-created_at')

    # Apply filters if any
    patient_id = request.GET.get('patient')
    if patient_id:
        transactions = transactions.filter(patient_id=patient_id)

    transaction_type = request.GET.get('transaction_type')
    if transaction_type:
        transactions = transactions.filter(transaction_type=transaction_type)

    date_from = request.GET.get('date_from')
    if date_from:
        transactions = transactions.filter(date__gte=date_from)

    date_to = request.GET.get('date_to')
    if date_to:
        transactions = transactions.filter(date__lte=date_to)

    direction = request.GET.get('transaction_direction')
    if direction:
        transactions = transactions.filter(transaction_direction=direction)

    # Calculate totals for displayed transactions
    total_inflow = sum(
        t.amount for t in transactions if t.transaction_direction == 'in'
    )
    total_outflow = sum(
        t.amount for t in transactions if t.transaction_direction == 'out'
    )

    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(transactions, 50)  # 50 transactions per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Get transaction type choices for filter dropdown
    transaction_types = PatientTransactionModel.TRANSACTION_TYPE

    context = {
        'transactions': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'paginator': paginator,
        'total_inflow': total_inflow,
        'total_outflow': total_outflow,
        'transaction_types': transaction_types,
        'timeframe': 'All Time',
    }

    return render(request, 'finance/transaction/index.html', context)


@login_required
def print_receipt(request, transaction_id):
    """
    Generate a printable receipt for a transaction.
    Shows ALL items regardless of count.
    """
    transaction = get_object_or_404(
        PatientTransactionModel.objects.select_related(
            'patient',
            'received_by',
            'received_by__user_staff_profile',
            'received_by__user_staff_profile__staff',
            'parent_transaction',
            'lab_structure',
            'lab_structure__template',
            'service',
            'service__service',
            'service__service__category',
            'service__service_item',
            'service__service_item__category',
            'other_service',
            'fee_structure',
            'admission',
            'surgery',
        ).prefetch_related(
            Prefetch(
                'child_transactions',
                queryset=PatientTransactionModel.objects.select_related(
                    'lab_structure',
                    'lab_structure__template',
                    'service',
                    'service__service',
                    'service__service__category',
                    'service__service_item',
                    'service__service_item__category',
                ).order_by('transaction_type', 'created_at')
            ),
        ),
        id=transaction_id
    )

    # Get hospital/facility info (adjust based on your models)
    site_info = SiteInfoModel.objects.first()
    hospital_info = {
        'name': site_info.name.title() or 'Hospital Name',
        'address': site_info.address.title() or 'Hospital Address',
        'phone': site_info.mobile_1 or 'Phone Number',
        'email': site_info.email.lower() or 'Hospital Email',
    }

    context = {
        'transaction': transaction,
        'hospital_info': hospital_info,
        'is_parent': transaction.is_parent_transaction,
        'is_child': transaction.is_child_transaction,
        'is_standalone': transaction.is_standalone_transaction,
    }

    return render(request, 'finance/payment/print.html', context)


# OPTIONAL: Create a simple wallet-only funding page
@login_required
def wallet_funding_only_page(request):
    """
    Simple page for wallet funding only (no payment processing)
    """
    patient_id = request.GET.get('patient')
    patient = None

    if patient_id:
        patient = get_object_or_404(PatientModel, id=patient_id)

    context = {
        'patient': patient,
    }

    return render(request, 'finance/wallet_funding_only.html', context)


from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.views.generic import TemplateView
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.db import transaction
from django.db.models import Sum, F, DecimalField, ExpressionWrapper
from django.utils import timezone
from django.urls import reverse
from decimal import Decimal
from datetime import date, timedelta, time
import logging

logger = logging.getLogger(__name__)


class UnifiedPaymentView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Unified payment page for Consultation, Admission, Surgery, and Other Payments"""
    template_name = 'finance/payment/unified_payment.html'
    permission_required = 'finance.add_patienttransactionmodel'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['specializations'] = SpecializationModel.objects.order_by('name')
        context['other_services'] = OtherPaymentService.objects.filter(is_active=True).order_by('category', 'name')
        return context


@login_required
def ajax_get_admission_surgery_details(request):
    """Get details for active admissions or surgeries"""
    patient_id = request.GET.get('patient_id')
    record_type = request.GET.get('type')  # 'admission' or 'surgery'

    if not patient_id or not record_type:
        return JsonResponse({'success': False, 'error': 'Missing parameters'}, status=400)

    patient = get_object_or_404(PatientModel, pk=patient_id)

    paid_statuses_drug = ['paid', 'partially_dispensed', 'dispensed']
    paid_statuses_lab_scan = ['paid', 'collected', 'processing', 'completed']

    records = []

    if record_type == 'admission':
        active_records = Admission.objects.filter(patient=patient, status='active')

        for admission in active_records:
            drug_orders = admission.drug_orders.all()
            lab_orders = admission.lab_test_orders.all()
            scan_orders = admission.scan_orders.all()
            other_services = admission.service_drug_orders.all()

            drug_costs = drug_orders.aggregate(
                total=Sum(ExpressionWrapper(
                    F('quantity_ordered') * F('drug__selling_price'),
                    output_field=DecimalField()
                ))
            )['total'] or 0

            lab_costs = lab_orders.aggregate(total=Sum('amount_charged'))['total'] or 0
            scan_costs = scan_orders.aggregate(total=Sum('amount_charged'))['total'] or 0
            other_services_costs = other_services.aggregate(total=Sum('total_amount'))['total'] or 0

            base_fee = (admission.admission_fee_charged or 0) + (admission.bed_fee_charged or 0)
            total_bill = base_fee + drug_costs + lab_costs + scan_costs + other_services_costs

            total_paid = PatientTransactionModel.objects.filter(
                admission=admission,
                status='completed',
            ).aggregate(total=Sum('amount'))['total'] or 0

            total_paid += sum(
                order.drug.selling_price * Decimal(order.quantity_ordered)
                for order in drug_orders if order.status in paid_statuses_drug
            )
            total_paid += sum(order.amount_charged for order in lab_orders if order.status in paid_statuses_lab_scan)
            total_paid += sum(order.amount_charged for order in scan_orders if order.status in paid_statuses_lab_scan)
            total_paid += sum(
                service.total_amount for service in other_services
                if service.status in ['paid', 'fully_dispensed', 'partially_dispensed']
            )

            balance = total_paid - total_bill

            records.append({
                'id': admission.id,
                'identifier': admission.admission_number,
                'base_fee': float(base_fee),
                'drug_costs': float(drug_costs),
                'lab_costs': float(lab_costs),
                'scan_costs': float(scan_costs),
                'other_costs': float(other_services_costs),
                'total_bill': float(total_bill),
                'total_paid': float(total_paid),
                'balance': float(balance),
                'abs_balance': float(abs(balance)),
                'status': admission.status
            })

    elif record_type == 'surgery':
        active_records = Surgery.objects.filter(patient=patient, status__in=['scheduled', 'in_progress'])

        for surgery in active_records:
            drug_orders = surgery.drug_orders.all()
            lab_orders = surgery.lab_test_orders.all()
            scan_orders = surgery.scan_orders.all()
            other_services = surgery.service_drug_orders.all()

            drug_costs = drug_orders.aggregate(
                total=Sum(ExpressionWrapper(
                    F('quantity_ordered') * F('drug__selling_price'),
                    output_field=DecimalField()
                ))
            )['total'] or 0

            lab_costs = lab_orders.aggregate(total=Sum('amount_charged'))['total'] or 0
            scan_costs = scan_orders.aggregate(total=Sum('amount_charged'))['total'] or 0
            other_services_costs = other_services.aggregate(total=Sum('total_amount'))['total'] or 0

            base_fee = surgery.total_surgery_cost
            total_bill = base_fee + drug_costs + lab_costs + scan_costs + other_services_costs

            total_paid = PatientTransactionModel.objects.filter(
                surgery=surgery,
                status='completed',
            ).aggregate(total=Sum('amount'))['total'] or 0

            total_paid += sum(
                order.drug.selling_price * Decimal(order.quantity_ordered)
                for order in drug_orders if order.status in paid_statuses_drug
            )
            total_paid += sum(order.amount_charged for order in lab_orders if order.status in paid_statuses_lab_scan)
            total_paid += sum(order.amount_charged for order in scan_orders if order.status in paid_statuses_lab_scan)
            total_paid += sum(
                service.total_amount for service in other_services
                if service.status in ['paid', 'fully_dispensed', 'partially_dispensed']
            )

            balance = total_paid - total_bill

            records.append({
                'id': surgery.id,
                'identifier': surgery.surgery_number,
                'base_fee': float(base_fee),
                'drug_costs': float(drug_costs),
                'lab_costs': float(lab_costs),
                'scan_costs': float(scan_costs),
                'other_costs': float(other_services_costs),
                'total_bill': float(total_bill),
                'total_paid': float(total_paid),
                'balance': float(balance),
                'abs_balance': float(abs(balance)),
                'status': surgery.status
            })

    return JsonResponse({
        'success': True,
        'records': records,
        'count': len(records)
    })


@login_required
@permission_required('finance.add_patienttransactionmodel', raise_exception=True)
def ajax_process_consultation_payment(request):
    """Process consultation payment with direct payment or wallet"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=400)

    try:
        patient_id = request.POST.get('patient_id')
        specialization_id = request.POST.get('specialization_id')
        fee_structure_id = request.POST.get('fee_structure_id')
        payment_method = request.POST.get('payment_method', 'cash')
        use_wallet = request.POST.get('use_wallet', 'false').lower() == 'true'

        patient = get_object_or_404(PatientModel, pk=patient_id)
        specialization = get_object_or_404(SpecializationModel, pk=specialization_id)
        fee_structure = get_object_or_404(ConsultationFeeModel, pk=fee_structure_id)
        wallet = get_object_or_404(PatientWalletModel, patient=patient)

        total_amount = fee_structure.amount
        wallet_used = Decimal('0.00')
        direct_payment = total_amount

        # Calculate wallet usage
        if use_wallet and wallet.amount > 0:
            wallet_used = min(wallet.amount, total_amount)
            direct_payment = total_amount - wallet_used

        with transaction.atomic():
            old_balance = wallet.amount

            # Deduct from wallet if applicable
            if wallet_used > 0:
                wallet.amount -= wallet_used
                wallet.save()

            new_balance = wallet.amount

            # Calculate validity
            validity_days = fee_structure.validity_in_days
            today = date.today()
            cutoff_time = time(20, 0)
            start_date_for_validity = today if timezone.now().time() < cutoff_time else today + timedelta(days=1)
            valid_till = start_date_for_validity + timedelta(days=validity_days - 1)

            # Create consultation payment transaction
            consult_payment = PatientTransactionModel.objects.create(
                patient=patient,
                fee_structure=fee_structure,
                transaction_type='consultation_payment',
                transaction_direction='out',
                amount=total_amount,
                wallet_amount_used=wallet_used,
                direct_payment_amount=direct_payment,
                date=today,
                payment_method=payment_method,
                status='completed',
                received_by=request.user,
                old_balance=old_balance,
                new_balance=new_balance,
                valid_till=valid_till
            )

            # Add patient to queue
            PatientQueueModel.objects.create(
                patient=patient,
                payment=consult_payment,
                specialization=specialization,
                status='waiting_vitals'
            )

            logger.info(f"Consultation payment {consult_payment.transaction_id} created for patient {patient.pk}")

        return JsonResponse({
            'success': True,
            'message': f'Payment successful! Patient added to {specialization.name} queue.',
            'redirect_url': reverse('transaction_detail', kwargs={'transaction_id': consult_payment.pk})
        })

    except Exception as e:
        logger.error(f"Error processing consultation payment: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@permission_required('finance.add_patienttransactionmodel', raise_exception=True)
def ajax_process_admission_funding(request):
    """Process admission/surgery funding with direct payment or wallet"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=400)

    try:
        patient_id = request.POST.get('patient_id')
        admission_id = request.POST.get('admission_id')
        surgery_id = request.POST.get('surgery_id')
        amount = Decimal(request.POST.get('amount', '0'))
        payment_method = request.POST.get('payment_method', 'cash')
        use_wallet = request.POST.get('use_wallet', 'false').lower() == 'true'

        if amount <= 0:
            return JsonResponse({'success': False, 'error': 'Amount must be a positive number.'}, status=400)

        patient = get_object_or_404(PatientModel, pk=patient_id)
        wallet = get_object_or_404(PatientWalletModel, patient=patient)

        admission = get_object_or_404(Admission, pk=admission_id) if admission_id else None
        surgery = get_object_or_404(Surgery, pk=surgery_id) if surgery_id else None

        wallet_used = Decimal('0.00')
        direct_payment = amount

        # Calculate wallet usage
        if use_wallet and wallet.amount > 0:
            wallet_used = min(wallet.amount, amount)
            direct_payment = amount - wallet_used

        with transaction.atomic():
            old_balance = wallet.amount

            # Deduct from wallet if applicable
            if wallet_used > 0:
                wallet.amount -= wallet_used
                wallet.save()

            new_balance = wallet.amount

            payment_transaction = PatientTransactionModel.objects.create(
                patient=patient,
                transaction_type='admission_payment' if admission else 'surgery_payment',
                transaction_direction='out',
                admission=admission,
                surgery=surgery,
                amount=amount,
                wallet_amount_used=wallet_used,
                direct_payment_amount=direct_payment,
                old_balance=old_balance,
                new_balance=new_balance,
                date=timezone.now().date(),
                received_by=request.user,
                payment_method=payment_method,
                status='completed'
            )

        record_identifier = admission.admission_number if admission else surgery.surgery_number
        success_message = f'Successfully paid ₦{amount:,.2f} for {record_identifier}.'

        return JsonResponse({
            'success': True,
            'message': success_message,
            'redirect_url': reverse('transaction_detail', kwargs={'transaction_id': payment_transaction.pk})
        })

    except Exception as e:
        logger.error(f"Error processing admission/surgery funding: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@permission_required('finance.add_patienttransactionmodel', raise_exception=True)
def ajax_process_other_payment(request):
    """Process other payment with direct payment or wallet"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=400)

    try:
        patient_id = request.POST.get('patient_id')
        service_id = request.POST.get('other_service')
        amount = Decimal(request.POST.get('amount', '0'))
        payment_method = request.POST.get('payment_method', 'cash')
        use_wallet = request.POST.get('use_wallet', 'false').lower() == 'true'

        patient = get_object_or_404(PatientModel, pk=patient_id)
        service = get_object_or_404(OtherPaymentService, pk=service_id)
        wallet = get_object_or_404(PatientWalletModel, patient=patient)

        # Validate fixed amount
        if service.is_fixed_amount and service.default_amount != amount:
            return JsonResponse({'success': False, 'error': 'The amount for this service is fixed.'}, status=400)

        if amount <= 0:
            return JsonResponse({'success': False, 'error': 'Amount must be greater than zero.'}, status=400)

        wallet_used = Decimal('0.00')
        direct_payment = amount

        # Calculate wallet usage
        if use_wallet and wallet.amount > 0:
            wallet_used = min(wallet.amount, amount)
            direct_payment = amount - wallet_used

        with transaction.atomic():
            old_balance = wallet.amount

            # Deduct from wallet if applicable
            if wallet_used > 0:
                wallet.amount -= wallet_used
                wallet.save()

            new_balance = wallet.amount

            new_transaction = PatientTransactionModel.objects.create(
                patient=patient,
                transaction_type='other_payment',
                transaction_direction='out',
                other_service=service,
                amount=amount,
                wallet_amount_used=wallet_used,
                direct_payment_amount=direct_payment,
                old_balance=old_balance,
                new_balance=new_balance,
                date=timezone.now().date(),
                received_by=request.user,
                payment_method=payment_method,
                status='completed'
            )

        return JsonResponse({
            'success': True,
            'message': f'Payment of ₦{amount:,.2f} for {service.name} successful!',
            'redirect_url': reverse('transaction_detail', kwargs={'transaction_id': new_transaction.pk})
        })

    except Exception as e:
        logger.error(f"Error processing other payment: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@permission_required('finance.add_patienttransactionmodel', raise_exception=True)
def ajax_reuse_consultation_payment(request):
    """Add patient to queue using existing valid payment"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=400)

    try:
        patient_id = request.POST.get('patient_id')
        payment_id = request.POST.get('payment_id')
        specialization_id = request.POST.get('specialization_id')

        patient = get_object_or_404(PatientModel, pk=patient_id)
        payment = get_object_or_404(PatientTransactionModel, pk=payment_id)
        specialization = get_object_or_404(SpecializationModel, pk=specialization_id)

        with transaction.atomic():
            PatientQueueModel.objects.create(
                patient=patient,
                payment=payment,
                specialization=specialization,
                status='waiting_vitals'
            )

        return JsonResponse({
            'success': True,
            'message': f'Patient added to {specialization.name} queue successfully!',
            'redirect_url': reverse('finance_unified_payment')
        })

    except Exception as e:
        logger.error(f"Error reusing consultation payment: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=400)