import datetime
from datetime import timedelta
import json
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, DecimalField, F
from django.views import View
from django.views.decorators.http import require_http_methods
from django.db.models import Sum, Count, Q
from django.http import HttpResponse
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from decimal import Decimal
from django.contrib.messages.views import SuccessMessageMixin, messages
from django.views.generic import TemplateView
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.utils import timezone
from admin_site.forms import SiteInfoForm
from admin_site.models import SiteInfoModel, ActivityLogModel
from consultation.models import SpecializationModel, ConsultationSessionModel
from finance.models import PatientTransactionModel
from human_resource.models import StaffProfileModel
from inpatient.models import Surgery
from laboratory.models import LabSettingModel, LabTestOrderModel
from patient.models import PatientModel, RegistrationPaymentModel
from patient.views import calculate_growth_percentage, calculate_age_groups, get_monthly_trends, \
    get_registration_chart_data

from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.urls import reverse
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.views.decorators.cache import never_cache
import logging

from pharmacy.models import DrugOrderModel
from scan.models import ScanOrderModel, ScanSettingModel
from service.models import ServiceCategory, PatientServiceTransaction

logger = logging.getLogger(__name__)


def create_activity_log(user, log, category, sub_category, keywords=""):
    ActivityLogModel.objects.create(
        user=user,
        log=log,
        category=category,
        sub_category=sub_category,
        keywords=keywords
    )


def get_patient_dashboard_context(request):
    """
    Get the context data for patient dashboard
    """
    # Get current date and time periods
    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)
    year_start = today.replace(month=1, day=1)

    # Basic patient statistics
    total_patients = PatientModel.objects.count()
    male_patients = PatientModel.objects.filter(gender='male').count()
    female_patients = PatientModel.objects.filter(gender='female').count()

    # Registration status statistics
    pending_registrations = RegistrationPaymentModel.objects.filter(
        registration_status='pending', status='confirmed',
    ).count()

    # Old vs New patients (based on registration fee type)
    old_patients_total = RegistrationPaymentModel.objects.filter(
        registration_fee__patient_type='old', status='confirmed',
    ).count()
    new_patients_total = RegistrationPaymentModel.objects.filter(
        registration_fee__patient_type='new', status='confirmed',
    ).count()

    # Time-based statistics for old patients
    old_patients_today = RegistrationPaymentModel.objects.filter(
        registration_fee__patient_type='old', status='confirmed',
        date=today
    ).count()

    old_patients_week = RegistrationPaymentModel.objects.filter(
        registration_fee__patient_type='old', status='confirmed',
        date__gte=week_start
    ).count()

    old_patients_month = RegistrationPaymentModel.objects.filter(
        registration_fee__patient_type='old', status='confirmed',
        date__gte=month_start
    ).count()

    # Time-based statistics for new patients
    new_patients_today = RegistrationPaymentModel.objects.filter(
        registration_fee__patient_type='new', status='confirmed',
        date=today
    ).count()

    new_patients_week = RegistrationPaymentModel.objects.filter(
        registration_fee__patient_type='new', status='confirmed',
        date__gte=week_start
    ).count()

    new_patients_month = RegistrationPaymentModel.objects.filter(
        registration_fee__patient_type='new', status='confirmed',
        date__gte=month_start
    ).count()

    # Calculate growth percentages (comparing this month to last month)
    last_month_start = (month_start - timedelta(days=1)).replace(day=1)
    last_month_end = month_start - timedelta(days=1)

    old_patients_last_month = RegistrationPaymentModel.objects.filter(
        registration_fee__patient_type='old', status='confirmed',
        date__gte=last_month_start,
        date__lte=last_month_end
    ).count()

    new_patients_last_month = RegistrationPaymentModel.objects.filter(
        registration_fee__patient_type='new', status='confirmed',
        date__gte=last_month_start,
        date__lte=last_month_end
    ).count()

    # Calculate growth percentages
    old_patients_growth = calculate_growth_percentage(old_patients_month, old_patients_last_month)
    new_patients_growth = calculate_growth_percentage(new_patients_month, new_patients_last_month)

    # Age group statistics
    age_groups = calculate_age_groups()

    # Recent registrations for chart data (last 7 days)
    registration_chart_data = get_registration_chart_data()

    # Monthly registration trends (last 12 months)
    monthly_trends = get_monthly_trends()

    # Gender distribution for pie chart
    gender_distribution = [
        {'name': 'Male', 'value': male_patients},
        {'name': 'Female', 'value': female_patients}
    ]

    # Revenue statistics
    total_revenue = RegistrationPaymentModel.objects.filter(
        registration_status='completed', status='confirmed',
    ).aggregate(Sum('amount'))['amount__sum'] or 0

    revenue_today = RegistrationPaymentModel.objects.filter(
        registration_status='completed', status='confirmed',
        date=today
    ).aggregate(Sum('amount'))['amount__sum'] or 0

    revenue_month = RegistrationPaymentModel.objects.filter(
        registration_status='completed', status='confirmed',
        date__gte=month_start
    ).aggregate(Sum('amount'))['amount__sum'] or 0

    return {
        'total_patients': total_patients,
        'male_patients': male_patients,
        'female_patients': female_patients,
        'pending_registrations': pending_registrations,
        'old_patients_total': old_patients_total,
        'new_patients_total': new_patients_total,
        'old_patients_today': old_patients_today,
        'old_patients_week': old_patients_week,
        'old_patients_month': old_patients_month,
        'new_patients_today': new_patients_today,
        'new_patients_week': new_patients_week,
        'new_patients_month': new_patients_month,
        'old_patients_growth': old_patients_growth,
        'new_patients_growth': new_patients_growth,
        'age_groups': age_groups,
        'registration_chart_data': json.dumps(registration_chart_data),
        'monthly_trends': json.dumps(monthly_trends),
        'gender_distribution': json.dumps(gender_distribution),
        'total_revenue': total_revenue,
        'revenue_today': revenue_today,
        'revenue_month': revenue_month,
    }


@login_required
def dashboard(request):
    """
    Comprehensive patient dashboard with statistics and analytics
    """
    context = get_patient_dashboard_context(request)
    return render(request, 'admin_site/dashboard.html', context)


class SiteInfoCreateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, CreateView):
    model = SiteInfoModel
    form_class = SiteInfoForm
    permission_required = 'admin_site.add_siteinfomodel'
    success_message = 'Site Information Created Successfully'
    template_name = 'admin_site/site_info/create.html'

    def dispatch(self, *args, **kwargs):
        site_info = SiteInfoModel.objects.first()
        if not site_info:
            return super().dispatch(*args, **kwargs)
        else:
            return redirect(reverse('site_info_edit', kwargs={'pk': site_info.pk}))

    def get_success_url(self):
        return reverse('site_info_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['mode'] = 'create'
        return context


class SiteInfoDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = SiteInfoModel
    permission_required = 'admin_site.view_siteinfomodel'
    template_name = 'admin_site/site_info/detail.html'
    context_object_name = "site_info"

    def dispatch(self, request, *args, **kwargs):
        site_info = SiteInfoModel.objects.first()
        if site_info:
            if str(self.kwargs.get('pk')) != str(site_info.id):
                return redirect(reverse('site_info_detail', kwargs={'pk': site_info.pk}))
            return super().dispatch(request, *args, **kwargs)
        else:
            messages.info(request, "No site info found. Please create one first.")
            return redirect(reverse('site_info_create'))

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['mode'] = 'detail'
        return context


class SiteInfoUpdateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, UpdateView):
    model = SiteInfoModel
    form_class = SiteInfoForm
    permission_required = 'admin_site.change_siteinfomodel'
    success_message = 'Site Information Updated Successfully'
    template_name = 'admin_site/site_info/create.html'

    def get_success_url(self):
        return reverse('site_info_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['site_info'] = self.object
        context['mode'] = 'update'
        return context


def user_sign_in_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']

        user = authenticate(request, username=username, password=password)
        if user is not None:
            intended_route = request.POST.get('next')
            if not intended_route:
                intended_route = request.GET.get('next')

            remember_me = request.POST.get('remember_me')
            if not remember_me:
                remember_me = request.GET.get('remember_me')

            if user.is_superuser:
                login(request, user)
                messages.success(request, 'welcome back {}'.format(user.username.title()))
                if remember_me:
                    request.session.set_expiry(3600 * 24 * 30)
                else:
                    request.session.set_expiry(0)
                if intended_route:
                    return redirect(intended_route)
                return redirect(reverse('admin_dashboard'))
            try:
                user_role = StaffProfileModel.objects.get(user=user)
            except StaffProfileModel.DoesNotExist:
                messages.error(request, 'Unknown Identity, Access Denied')
                return redirect(reverse('login'))

            if user_role.staff:
                login(request, user)
                messages.success(request, 'welcome back {}'.format(user_role.staff))
                if remember_me:
                    request.session.set_expiry(3600 * 24 * 30)
                else:
                    request.session.set_expiry(0)
                if intended_route:
                    return redirect(intended_route)
                return redirect(reverse('admin_dashboard'))

            else:
                messages.error(request, 'Unknown Identity, Access Denied')
                return redirect(reverse('login'))
        else:
            messages.error(request, 'Invalid Credentials')
            return redirect(reverse('login'))

    return render(request, 'admin_site/user/sign_in.html')


def user_sign_out_view(request):
    logout(request)
    return redirect(reverse('login'))


@login_required
@never_cache
@require_http_methods(["GET", "POST"])
def change_password_view(request):
    """
    View to handle password change for authenticated users.
    Validates current password and updates to new password.
    """

    if request.method == 'POST':
        # Get form data
        current_password = request.POST.get('current_password', '').strip()
        new_password1 = request.POST.get('new_password1', '').strip()
        new_password2 = request.POST.get('new_password2', '').strip()

        # Validation
        errors = []

        # Check if all fields are provided
        if not current_password:
            errors.append("Current password is required.")

        if not new_password1:
            errors.append("New password is required.")

        if not new_password2:
            errors.append("Password confirmation is required.")

        # Check if new passwords match
        if new_password1 and new_password2 and new_password1 != new_password2:
            errors.append("New passwords do not match.")

        # Check password length and complexity
        if new_password1 and len(new_password1) < 8:
            errors.append("New password must be at least 8 characters long.")

        # Check if new password is different from current
        if current_password and new_password1 and current_password == new_password1:
            errors.append("New password must be different from current password.")

        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, 'admin_site/user/change_password.html')

        # Verify current password
        user = authenticate(username=request.user.username, password=current_password)
        if user is None:
            messages.error(request, "Current password is incorrect.")
            logger.warning(
                f"Failed password change attempt for user {request.user.username} - incorrect current password")
            return render(request, 'admin_site/user/change_password.html')

        try:
            # Change password
            user.set_password(new_password1)
            user.save()

            # Keep user logged in after password change
            update_session_auth_hash(request, user)

            messages.success(request, "Your password has been successfully changed!")
            logger.info(f"Password successfully changed for user {request.user.username}")

            # Redirect to dashboard or profile page
            return redirect('admin_dashboard')  # Change this to your desired redirect URL

        except Exception as e:
            logger.exception(f"Error changing password for user {request.user.username}: {str(e)}")
            messages.error(request, "An error occurred while changing your password. Please try again.")
            return render(request, 'admin_site/user/change_password.html')

    # GET request - show the form
    return render(request, 'admin_site/user/change_password.html')


def custom_404_view(request, exception):
    """
    Handles Page Not Found errors (404).

    It checks if the URL path contains 'portal' to decide whether to show
    the admin error page or the public website error page.
    """
    # Get the path of the URL that could not be found
    path = request.path

    # Check if the path starts with '/portal/'. This is more specific
    # than just checking if 'portal' is in the path.
    if path.startswith('/portal/'):
        # If it's a portal URL, render the admin-themed 404 page
        template_name = 'admin_site/errors/404.html'
    else:
        # Otherwise, render the public website's 404 page
        template_name = 'website/errors/404.html'

    return render(request, template_name, status=404)


def custom_500_view(request):
    """
    Handles Server Errors (500).
    Django calls this view when there's a server-side crash.
    """
    return render(request, 'errors/500.html', status=500)


def custom_403_view(request, exception):
    """
    Handles Permission Denied errors (403).
    This is triggered when a user tries to access a resource they don't have permission for.
    """
    return render(request, 'admin_site/errors/403.html', status=403)


def custom_csrf_failure_view(request, reason=""):
    """
    Handles CSRF verification failures.
    This is a special case handled via settings.py.
    """
    return render(request, 'admin_site/errors/403_csrf.html', status=403)


class LabFinancialReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    template_name = 'admin_site/reports/lab_financial.html'
    permission_required = 'laboratory.view_financial_report'  # Uses the new permission

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # === Date range ===
        from_date = self.request.GET.get('from_date')
        to_date = self.request.GET.get('to_date')

        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()

        if not to_date:
            today = date.today()
            if today.month == 12:
                to_date = today.replace(day=31)
            else:
                to_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        else:
            to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()

        # === Title ===
        month_year = from_date.strftime('%B %Y')
        default_title = f'Laboratory Financial Report for {month_year}'
        report_title = self.request.GET.get('title', default_title)

        # === Query ===
        # We only care about orders that are not pending or cancelled
        orders = LabTestOrderModel.objects.filter(
            ordered_at__date__range=[from_date, to_date]
        ).exclude(status__in=['pending', 'cancelled'])

        # Get order by parameter
        order_by = self.request.GET.get('order_by', 'name')

        # === Breakdown Query ===
        test_breakdown = orders.values(
            'template__name',
            'template__id',
            'template__price'  # Get the CURRENT unit price
        ).annotate(
            total_orders=Count('id'),
            total_amount=Sum('amount_charged')  # SUM of what was ACTUALLY charged
        )

        # Apply ordering
        if order_by == 'total':
            test_breakdown = test_breakdown.order_by('-total_orders', 'template__name')
        elif order_by == 'amount':
            test_breakdown = test_breakdown.order_by('-total_amount', 'template__name')
        else:  # default to name
            test_breakdown = test_breakdown.order_by('template__name')

        # === Summary Statistics ===
        grand_total = orders.aggregate(
            total=Sum('amount_charged'),
            total_count=Count('id')
        )

        context.update({
            'from_date': from_date,
            'to_date': to_date,
            'report_title': report_title,
            'test_breakdown': test_breakdown,
            'order_by': order_by,

            # Summary stats
            'grand_total_amount': grand_total['total'] or Decimal('0.00'),
            'grand_total_orders': grand_total['total_count'] or 0,
        })

        return context


class LabFinancialReportExportExcelView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'laboratory.view_financial_report'

    def get(self, request, *args, **kwargs):
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        report_title = request.GET.get('title', 'Laboratory Financial Report')
        order_by = request.GET.get('order_by', 'name')

        # Date filtering
        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()
        if not to_date:
            today = date.today()
            if today.month == 12:
                to_date = today.replace(day=31)
            else:
                to_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        else:
            to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()

        # Get lab and site info
        lab_setting = LabSettingModel.objects.first()
        site_info = SiteInfoModel.objects.first()

        # Query data
        orders = LabTestOrderModel.objects.filter(
            ordered_at__date__range=[from_date, to_date]
        ).exclude(status__in=['pending', 'cancelled'])

        test_breakdown = orders.values(
            'template__name', 'template__id', 'template__price'
        ).annotate(
            total_orders=Count('id'),
            total_amount=Sum('amount_charged')
        )

        # Apply ordering
        if order_by == 'total':
            test_breakdown = test_breakdown.order_by('-total_orders', 'template__name')
        elif order_by == 'amount':
            test_breakdown = test_breakdown.order_by('-total_amount', 'template__name')
        else:
            test_breakdown = test_breakdown.order_by('template__name')

        # Totals
        grand_total = orders.aggregate(total=Sum('amount_charged'), total_count=Count('id'))

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Lab Financial Report"

        # Styles
        header_font = Font(bold=True, size=14);
        title_font = Font(bold=True, size=12)
        table_header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))
        total_font = Font(bold=True)

        row = 1
        # Header section (Same as your other report)
        if lab_setting and lab_setting.lab_name:
            ws.merge_cells(f'A{row}:E{row}');
            cell = ws[f'A{row}'];
            cell.value = lab_setting.lab_name;
            cell.font = header_font;
            cell.alignment = Alignment(horizontal='center');
            row += 1
        elif site_info:
            ws.merge_cells(f'A{row}:E{row}');
            cell = ws[f'A{row}'];
            cell.value = site_info.name;
            cell.font = header_font;
            cell.alignment = Alignment(horizontal='center');
            row += 1
        row += 1
        ws.merge_cells(f'A{row}:E{row}');
        cell = ws[f'A{row}'];
        cell.value = report_title;
        cell.font = title_font;
        cell.alignment = Alignment(horizontal='center');
        row += 1
        ws.merge_cells(f'A{row}:E{row}');
        cell = ws[f'A{row}'];
        cell.value = f"Period: {from_date.strftime('%B %d, %Y')} to {to_date.strftime('%B %d, %Y')}";
        cell.alignment = Alignment(horizontal='center');
        row += 2

        # Table headers
        headers = ['S/N', 'Test', 'Unit Cost', 'Total Orders', 'Total Amount']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col);
            cell.value = header;
            cell.font = table_header_font;
            cell.fill = header_fill;
            cell.alignment = Alignment(horizontal='center');
            cell.border = border
        row += 1

        # Data rows
        for idx, test in enumerate(test_breakdown, 1):
            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=2, value=test['template__name']).border = border
            cell_unit = ws.cell(row=row, column=3, value=test['template__price']);
            cell_unit.number_format = '#,##0.00';
            cell_unit.border = border
            ws.cell(row=row, column=4, value=test['total_orders']).border = border
            cell_total = ws.cell(row=row, column=5, value=test['total_amount']);
            cell_total.number_format = '#,##0.00';
            cell_total.border = border
            row += 1

        # Total Row
        ws.cell(row=row, column=2, value="GRAND TOTAL").font = total_font;
        ws.cell(row=row, column=2).border = border
        ws.cell(row=row, column=3).border = border  # Empty unit cost
        cell_grand_orders = ws.cell(row=row, column=4, value=grand_total['total_count']);
        cell_grand_orders.font = total_font;
        cell_grand_orders.border = border
        cell_grand_total = ws.cell(row=row, column=5, value=grand_total['total']);
        cell_grand_total.font = total_font;
        cell_grand_total.number_format = '#,##0.00';
        cell_grand_total.border = border

        # Adjust column widths
        ws.column_dimensions['A'].width = 8;
        ws.column_dimensions['B'].width = 40;
        ws.column_dimensions['C'].width = 15;
        ws.column_dimensions['D'].width = 15;
        ws.column_dimensions['E'].width = 18

        # Prepare response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"lab_financial_report_{from_date.strftime('%Y%m%d')}_{to_date.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class LabFinancialReportExportPDFView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'laboratory.view_financial_report'

    def get(self, request, *args, **kwargs):
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        report_title = request.GET.get('title', 'Laboratory Financial Report')
        order_by = request.GET.get('order_by', 'name')

        # Date filtering
        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()
        if not to_date:
            today = date.today()
            if today.month == 12:
                to_date = today.replace(day=31)
            else:
                to_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        else:
            to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()

        # Get lab and site info
        lab_setting = LabSettingModel.objects.first()
        site_info = SiteInfoModel.objects.first()

        # Query data
        orders = LabTestOrderModel.objects.filter(
            ordered_at__date__range=[from_date, to_date]
        ).exclude(status__in=['pending', 'cancelled'])

        test_breakdown = orders.values(
            'template__name', 'template__id', 'template__price'
        ).annotate(
            total_orders=Count('id'),
            total_amount=Sum('amount_charged')
        )

        if order_by == 'total':
            test_breakdown = test_breakdown.order_by('-total_orders', 'template__name')
        elif order_by == 'amount':
            test_breakdown = test_breakdown.order_by('-total_amount', 'template__name')
        else:
            test_breakdown = test_breakdown.order_by('template__name')

        # Totals
        grand_total = orders.aggregate(total=Sum('amount_charged'), total_count=Count('id'))

        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, alignment=TA_CENTER,
                                     spaceAfter=12)
        subtitle_style = ParagraphStyle('CustomSubtitle', parent=styles['Normal'], fontSize=11, alignment=TA_CENTER,
                                        spaceAfter=6)

        # Header
        if lab_setting and lab_setting.lab_name:
            elements.append(Paragraph(lab_setting.lab_name, title_style))
        elif site_info:
            elements.append(Paragraph(site_info.name, title_style))
        elements.append(Spacer(1, 0.3 * inch))
        elements.append(Paragraph(report_title, title_style))
        elements.append(
            Paragraph(f"Period: {from_date.strftime('%B %d, %Y')} to {to_date.strftime('%B %d, %Y')}", subtitle_style))
        elements.append(Spacer(1, 0.4 * inch))

        # Build table data
        headers = ['S/N', 'Test', 'Unit Cost', 'Total Orders', 'Total Amount']
        col_widths = [0.5 * inch, 3 * inch, 1.2 * inch, 1 * inch, 1.2 * inch]
        table_data = [headers]

        for idx, test in enumerate(test_breakdown, 1):
            row = [
                str(idx),
                test['template__name'],
                f"{test['template__price']:,.2f}",
                str(test['total_orders']),
                f"{test['total_amount']:,.2f}"
            ]
            table_data.append(row)

        # Total Row
        total_row = [
            '', 'GRAND TOTAL', '',
            str(grand_total['total_count']),
            f"{grand_total['total']:,.2f}"
        ]
        table_data.append(total_row)

        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # S/N
            ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),  # Numbers
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f0f0f0')]),
            # Total Row
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f0f0f0')),
            ('GRID', (0, -1), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)

        # --- Add this code for signatures ---

        elements.append(Spacer(1, 0.5 * inch))  # Add space before signatures

        # Define a style for the signature text
        sig_style = ParagraphStyle(
            'SigStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_LEFT,
        )

        # Create the text for each signature
        sig1_text = """
                    <br/><br/><br/>
                    _________________________<br/>
                    <b>Name:</b><br/>
                    <b>Title:</b>
                    """
        sig2_text = """
                    <br/><br/><br/>
                    _________________________<br/>
                    <b>Name:</b><br/>
                    <b>Title:</b>
                    """

        sig1 = Paragraph(sig1_text, sig_style)
        sig2 = Paragraph(sig2_text, sig_style)

        # Create a 1x2 table to hold them side-by-side
        # Adjust colWidths to change spacing
        signature_table = Table([[sig1, sig2]], colWidths=[3.5 * inch, 3.5 * inch])

        # Style the table to have no borders
        signature_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0, colors.white),
            ('VALIGN', (0, 0), (-1, -1), 'BOTTOM'),
        ]))

        elements.append(signature_table)

        # --- End of new code ---

        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f"lab_financial_report_{from_date.strftime('%Y%m%d')}_{to_date.strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class ScanFinancialReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    template_name = 'admin_site/reports/scan_financial.html'
    permission_required = 'scan.view_financial_report'  # Uses the new permission

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from_date = self.request.GET.get('from_date')
        to_date = self.request.GET.get('to_date')

        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()
        if not to_date:
            today = date.today()
            if today.month == 12:
                to_date = today.replace(day=31)
            else:
                to_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        else:
            to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()

        month_year = from_date.strftime('%B %Y')
        default_title = f'Scan Financial Report for {month_year}'
        report_title = self.request.GET.get('title', default_title)

        orders = ScanOrderModel.objects.filter(
            ordered_at__date__range=[from_date, to_date]
        ).exclude(status__in=['pending', 'cancelled'])

        order_by = self.request.GET.get('order_by', 'name')

        test_breakdown = orders.values(
            'template__name',
            'template__id',
            'template__price'  # Get the CURRENT unit price
        ).annotate(
            total_orders=Count('id'),
            total_amount=Sum('amount_charged')  # SUM of what was ACTUALLY charged
        )

        if order_by == 'total':
            test_breakdown = test_breakdown.order_by('-total_orders', 'template__name')
        elif order_by == 'amount':
            test_breakdown = test_breakdown.order_by('-total_amount', 'template__name')
        else:
            test_breakdown = test_breakdown.order_by('template__name')

        grand_total = orders.aggregate(
            total=Sum('amount_charged'),
            total_count=Count('id')
        )

        context.update({
            'from_date': from_date,
            'to_date': to_date,
            'report_title': report_title,
            'test_breakdown': test_breakdown,
            'order_by': order_by,
            'grand_total_amount': grand_total['total'] or Decimal('0.00'),
            'grand_total_orders': grand_total['total_count'] or 0,
        })
        return context


class ScanFinancialReportExportExcelView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'scan.view_financial_report'

    def get(self, request, *args, **kwargs):
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        report_title = request.GET.get('title', 'Scan Financial Report')
        order_by = request.GET.get('order_by', 'name')

        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()
        if not to_date:
            today = date.today()
            if today.month == 12:
                to_date = today.replace(day=31)
            else:
                to_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        else:
            to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()

        scan_setting = ScanSettingModel.objects.first()
        site_info = SiteInfoModel.objects.first()

        orders = ScanOrderModel.objects.filter(
            ordered_at__date__range=[from_date, to_date]
        ).exclude(status__in=['pending', 'cancelled'])

        test_breakdown = orders.values(
            'template__name', 'template__id', 'template__price'
        ).annotate(
            total_orders=Count('id'),
            total_amount=Sum('amount_charged')
        )

        if order_by == 'total':
            test_breakdown = test_breakdown.order_by('-total_orders', 'template__name')
        elif order_by == 'amount':
            test_breakdown = test_breakdown.order_by('-total_amount', 'template__name')
        else:
            test_breakdown = test_breakdown.order_by('template__name')

        grand_total = orders.aggregate(total=Sum('amount_charged'), total_count=Count('id'))

        wb = Workbook()
        ws = wb.active
        ws.title = "Scan Financial Report"

        header_font = Font(bold=True, size=14);
        title_font = Font(bold=True, size=12)
        table_header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))
        total_font = Font(bold=True)

        row = 1
        if scan_setting and scan_setting.scan_name:
            ws.merge_cells(f'A{row}:E{row}');
            cell = ws[f'A{row}'];
            cell.value = scan_setting.scan_name;
            cell.font = header_font;
            cell.alignment = Alignment(horizontal='center');
            row += 1
        elif site_info:
            ws.merge_cells(f'A{row}:E{row}');
            cell = ws[f'A{row}'];
            cell.value = site_info.name;
            cell.font = header_font;
            cell.alignment = Alignment(horizontal='center');
            row += 1
        row += 1
        ws.merge_cells(f'A{row}:E{row}');
        cell = ws[f'A{row}'];
        cell.value = report_title;
        cell.font = title_font;
        cell.alignment = Alignment(horizontal='center');
        row += 1
        ws.merge_cells(f'A{row}:E{row}');
        cell = ws[f'A{row}'];
        cell.value = f"Period: {from_date.strftime('%B %d, %Y')} to {to_date.strftime('%B %d, %Y')}";
        cell.alignment = Alignment(horizontal='center');
        row += 2

        headers = ['S/N', 'Scan', 'Unit Cost', 'Total Orders', 'Total Amount']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col);
            cell.value = header;
            cell.font = table_header_font;
            cell.fill = header_fill;
            cell.alignment = Alignment(horizontal='center');
            cell.border = border
        row += 1

        for idx, test in enumerate(test_breakdown, 1):
            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=2, value=test['template__name']).border = border
            cell_unit = ws.cell(row=row, column=3, value=test['template__price']);
            cell_unit.number_format = '#,##0.00';
            cell_unit.border = border
            ws.cell(row=row, column=4, value=test['total_orders']).border = border
            cell_total = ws.cell(row=row, column=5, value=test['total_amount']);
            cell_total.number_format = '#,##0.00';
            cell_total.border = border
            row += 1

        ws.cell(row=row, column=2, value="GRAND TOTAL").font = total_font;
        ws.cell(row=row, column=2).border = border
        ws.cell(row=row, column=3).border = border
        cell_grand_orders = ws.cell(row=row, column=4, value=grand_total['total_count']);
        cell_grand_orders.font = total_font;
        cell_grand_orders.border = border
        cell_grand_total = ws.cell(row=row, column=5, value=grand_total['total']);
        cell_grand_total.font = total_font;
        cell_grand_total.number_format = '#,##0.00';
        cell_grand_total.border = border

        ws.column_dimensions['A'].width = 8;
        ws.column_dimensions['B'].width = 40;
        ws.column_dimensions['C'].width = 15;
        ws.column_dimensions['D'].width = 15;
        ws.column_dimensions['E'].width = 18

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"scan_financial_report_{from_date.strftime('%Y%m%d')}_{to_date.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class ScanFinancialReportExportPDFView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'scan.view_financial_report'

    def get(self, request, *args, **kwargs):
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        report_title = request.GET.get('title', 'Scan Financial Report')
        order_by = request.GET.get('order_by', 'name')

        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()
        if not to_date:
            today = date.today()
            if today.month == 12:
                to_date = today.replace(day=31)
            else:
                to_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        else:
            to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()

        scan_setting = ScanSettingModel.objects.first()
        site_info = SiteInfoModel.objects.first()

        orders = ScanOrderModel.objects.filter(
            ordered_at__date__range=[from_date, to_date]
        ).exclude(status__in=['pending', 'cancelled'])

        test_breakdown = orders.values(
            'template__name', 'template__id', 'template__price'
        ).annotate(
            total_orders=Count('id'),
            total_amount=Sum('amount_charged')
        )

        if order_by == 'total':
            test_breakdown = test_breakdown.order_by('-total_orders', 'template__name')
        elif order_by == 'amount':
            test_breakdown = test_breakdown.order_by('-total_amount', 'template__name')
        else:
            test_breakdown = test_breakdown.order_by('template__name')

        grand_total = orders.aggregate(total=Sum('amount_charged'), total_count=Count('id'))

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, alignment=TA_CENTER,
                                     spaceAfter=12)
        subtitle_style = ParagraphStyle('CustomSubtitle', parent=styles['Normal'], fontSize=11, alignment=TA_CENTER,
                                        spaceAfter=6)

        if scan_setting and scan_setting.scan_name:
            elements.append(Paragraph(scan_setting.scan_name, title_style))
        elif site_info:
            elements.append(Paragraph(site_info.name, title_style))
        elements.append(Spacer(1, 0.3 * inch))
        elements.append(Paragraph(report_title, title_style))
        elements.append(
            Paragraph(f"Period: {from_date.strftime('%B %d, %Y')} to {to_date.strftime('%B %d, %Y')}", subtitle_style))
        elements.append(Spacer(1, 0.4 * inch))

        headers = ['S/N', 'Scan', 'Unit Cost', 'Total Orders', 'Total Amount']
        col_widths = [0.5 * inch, 3 * inch, 1.2 * inch, 1 * inch, 1.2 * inch]
        table_data = [headers]

        for idx, test in enumerate(test_breakdown, 1):
            row = [
                str(idx),
                test['template__name'],
                f"{test['template__price']:,.2f}",
                str(test['total_orders']),
                f"{test['total_amount']:,.2f}"
            ]
            table_data.append(row)

        total_row = [
            '', 'GRAND TOTAL', '',
            str(grand_total['total_count']),
            f"{grand_total['total']:,.2f}"
        ]
        table_data.append(total_row)

        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f0f0f0')]),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f0f0f0')),
            ('GRID', (0, -1), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)

        # --- Add Signatures ---
        elements.append(Spacer(1, 0.5 * inch))
        sig_style = ParagraphStyle('SigStyle', parent=styles['Normal'], fontSize=10, alignment=TA_LEFT)
        sig1_text = "<br/><br/><br/>_________________________<br/><b>Name:</b><br/><b>Title:</b>"
        sig2_text = "<br/><br/><br/>_________________________<br/><b>Name:</b><br/><b>Title:</b>"
        sig1 = Paragraph(sig1_text, sig_style)
        sig2 = Paragraph(sig2_text, sig_style)
        signature_table = Table([[sig1, sig2]], colWidths=[3.5 * inch, 3.5 * inch])
        signature_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0, colors.white),
            ('VALIGN', (0, 0), (-1, -1), 'BOTTOM'),
        ]))
        elements.append(signature_table)
        # --- End Signatures ---

        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f"scan_financial_report_{from_date.strftime('%Y%m%d')}_{to_date.strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class ConsultationReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    template_name = 'admin_site/reports/consultation_report.html'
    permission_required = 'consultation.view_consultation_report'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # === Date range ===
        from_date = self.request.GET.get('from_date')
        to_date = self.request.GET.get('to_date')

        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()

        if not to_date:
            today = date.today()
            if today.month == 12:
                to_date = today.replace(day=31)
            else:
                to_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        else:
            to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()

        # === Specialization Filter ===
        specialization_id = self.request.GET.get('specialization')
        selected_specialization = None

        if specialization_id and specialization_id != 'all':
            selected_specialization = SpecializationModel.objects.filter(id=specialization_id).first()

        # === Report Title ===
        month_year = from_date.strftime('%B %Y')
        if selected_specialization:
            default_title = f'Consultation Report for {selected_specialization.name} - {month_year}'
        else:
            default_title = f'Consultation Report (All Specializations) - {month_year}'

        report_title = self.request.GET.get('title', default_title)

        # === Signature Toggle ===
        show_signature = self.request.GET.get('show_signature', 'true') == 'true'

        # === Query Consultations ===
        consultations = ConsultationSessionModel.objects.filter(
            created_at__date__range=[from_date, to_date],
            status='completed'
        )

        if selected_specialization:
            consultations = consultations.filter(
                queue_entry__specialization=selected_specialization
            )

        # === 1. Diagnosis Breakdown ===
        # Get primary diagnoses
        primary_diagnosis_counts = consultations.filter(
            primary_diagnosis__isnull=False
        ).values(
            'primary_diagnosis__name'
        ).annotate(
            count=Count('id')
        ).order_by('-count')

        # Get secondary diagnoses
        secondary_diagnosis_counts = {}
        for consultation in consultations:
            for diagnosis in consultation.secondary_diagnoses.all():
                if diagnosis.name in secondary_diagnosis_counts:
                    secondary_diagnosis_counts[diagnosis.name] += 1
                else:
                    secondary_diagnosis_counts[diagnosis.name] = 1

        # Combine primary and secondary
        diagnosis_breakdown = {}
        for item in primary_diagnosis_counts:
            diagnosis_breakdown[item['primary_diagnosis__name']] = item['count']

        for name, count in secondary_diagnosis_counts.items():
            if name in diagnosis_breakdown:
                diagnosis_breakdown[name] += count
            else:
                diagnosis_breakdown[name] = count

        # Sort by count
        diagnosis_breakdown = sorted(
            diagnosis_breakdown.items(),
            key=lambda x: x[1],
            reverse=True
        )

        total_patients = consultations.count()

        # === 2. Financial Breakdown ===
        # Get consultation IDs for filtering
        consultation_ids = list(consultations.values_list('id', flat=True))

        # A. Drug Income
        drug_income = DrugOrderModel.objects.filter(
            consultation_id__in=consultation_ids,
            status__in=['paid', 'dispensed', 'partially_dispensed']
        ).aggregate(
            total=Sum(
                (F('quantity_ordered') * F('drug__selling_price')),
                output_field=DecimalField()
            )
        )['total'] or Decimal('0.00')

        print(drug_income)

        # B. Service/Item Income by Category
        service_income_breakdown = []

        # Get service categories linked to this specialization (or all if no filter)
        if selected_specialization:
            service_categories = ServiceCategory.objects.filter(
                specializations=selected_specialization,
                is_active=True
            )
        else:
            service_categories = ServiceCategory.objects.filter(is_active=True)

        for category in service_categories:
            category_income = PatientServiceTransaction.objects.filter(
                consultation_id__in=consultation_ids,
                status__in=['paid', 'fully_dispensed', 'partially_dispensed']
            ).filter(
                Q(service__category=category) | Q(service_item__category=category)
            ).aggregate(
                total=Sum('amount_paid')
            )['total'] or Decimal('0.00')

            if category_income > 0:
                service_income_breakdown.append({
                    'category': category.name,
                    'amount': category_income
                })

        # C. Surgery Income
        surgery_filter = Q(status='completed')
        if selected_specialization:
            surgery_filter &= Q(specialization=selected_specialization)

        surgery_income = Surgery.objects.filter(
            surgery_filter,
            created_at__date__range=[from_date, to_date]
        ).aggregate(
            total=Sum(
                F('custom_surgeon_fee') +
                F('custom_anesthesia_fee') +
                F('custom_facility_fee'),
                output_field=DecimalField()
            )
        )['total'] or Decimal('0.00')

        # Total income
        total_service_income = sum(item['amount'] for item in service_income_breakdown)
        total_income = drug_income + total_service_income + surgery_income

        # === All Specializations for Dropdown ===
        all_specializations = SpecializationModel.objects.all().order_by('name')

        context.update({
            'from_date': from_date,
            'to_date': to_date,
            'report_title': report_title,
            'show_signature': show_signature,

            # Filters
            'selected_specialization': selected_specialization,
            'all_specializations': all_specializations,

            # Diagnosis data
            'diagnosis_breakdown': diagnosis_breakdown,
            'total_patients': total_patients,

            # Financial data
            'drug_income': drug_income,
            'service_income_breakdown': service_income_breakdown,
            'surgery_income': surgery_income,
            'total_service_income': total_service_income,
            'total_income': total_income,
        })

        return context


class ConsultationReportExportExcelView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'consultation.view_consultation_report'

    def get(self, request, *args, **kwargs):
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        report_title = request.GET.get('title', 'Consultation Report')
        specialization_id = request.GET.get('specialization')
        show_signature = request.GET.get('show_signature', 'true') == 'true'

        # Date filtering
        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()
        if not to_date:
            today = date.today()
            if today.month == 12:
                to_date = today.replace(day=31)
            else:
                to_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        else:
            to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()

        # Specialization filter
        selected_specialization = None
        if specialization_id and specialization_id != 'all':
            selected_specialization = SpecializationModel.objects.filter(id=specialization_id).first()

        # Query consultations
        consultations = ConsultationSessionModel.objects.filter(
            created_at__date__range=[from_date, to_date],
            status='completed'
        )

        if selected_specialization:
            consultations = consultations.filter(
                queue_entry__specialization=selected_specialization
            )

        # Get diagnosis breakdown
        primary_diagnosis_counts = consultations.filter(
            primary_diagnosis__isnull=False
        ).values(
            'primary_diagnosis__name'
        ).annotate(
            count=Count('id')
        ).order_by('-count')

        secondary_diagnosis_counts = {}
        for consultation in consultations:
            for diagnosis in consultation.secondary_diagnoses.all():
                if diagnosis.name in secondary_diagnosis_counts:
                    secondary_diagnosis_counts[diagnosis.name] += 1
                else:
                    secondary_diagnosis_counts[diagnosis.name] = 1

        diagnosis_breakdown = {}
        for item in primary_diagnosis_counts:
            diagnosis_breakdown[item['primary_diagnosis__name']] = item['count']

        for name, count in secondary_diagnosis_counts.items():
            if name in diagnosis_breakdown:
                diagnosis_breakdown[name] += count
            else:
                diagnosis_breakdown[name] = count

        diagnosis_breakdown = sorted(
            diagnosis_breakdown.items(),
            key=lambda x: x[1],
            reverse=True
        )

        total_patients = consultations.count()

        # Get financial data
        consultation_ids = list(consultations.values_list('id', flat=True))

        # Drug income
        drug_income = DrugOrderModel.objects.filter(
            consultation_id__in=consultation_ids,
            status__in=['paid', 'dispensed', 'partially_dispensed']
        ).aggregate(
            total=Sum(
                (F('quantity_ordered') * F('drug__selling_price')),
                output_field=DecimalField()
            )
        )['total'] or Decimal('0.00')

        # Service income
        service_income_breakdown = []
        if selected_specialization:
            service_categories = ServiceCategory.objects.filter(
                specializations=selected_specialization,
                is_active=True
            )
        else:
            service_categories = ServiceCategory.objects.filter(is_active=True)

        for category in service_categories:
            category_income = PatientServiceTransaction.objects.filter(
                consultation_id__in=consultation_ids,
                status__in=['paid', 'fully_dispensed', 'partially_dispensed']
            ).filter(
                Q(service__category=category) | Q(service_item__category=category)
            ).aggregate(
                total=Sum('amount_paid')
            )['total'] or Decimal('0.00')

            if category_income > 0:
                service_income_breakdown.append({
                    'category': category.name,
                    'amount': category_income
                })

        # Surgery income
        surgery_filter = Q(status='completed')
        if selected_specialization:
            surgery_filter &= Q(specialization=selected_specialization)

        surgery_income = Surgery.objects.filter(
            surgery_filter,
            created_at__date__range=[from_date, to_date]
        ).aggregate(
            total=Sum(
                F('custom_surgeon_fee') +
                F('custom_anesthesia_fee') +
                F('custom_facility_fee'),
                output_field=DecimalField()
            )
        )['total'] or Decimal('0.00')

        total_service_income = sum(item['amount'] for item in service_income_breakdown)
        total_income = drug_income + total_service_income + surgery_income

        lab_setting = LabSettingModel.objects.first()
        site_info = SiteInfoModel.objects.first()

        wb = Workbook()
        ws = wb.active
        ws.title = "Consultation Report"

        # Styles
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True, size=12)
        section_font = Font(bold=True, size=11)
        table_header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        total_font = Font(bold=True)

        row = 1
        # Header section
        if lab_setting and lab_setting.lab_name:
            ws.merge_cells(f'A{row}:C{row}')
            cell = ws[f'A{row}']
            cell.value = lab_setting.lab_name
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            row += 1
        elif site_info:
            ws.merge_cells(f'A{row}:C{row}')
            cell = ws[f'A{row}']
            cell.value = site_info.name
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            row += 1

        row += 1
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = report_title
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center')
        row += 1

        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = f"Period: {from_date.strftime('%B %d, %Y')} to {to_date.strftime('%B %d, %Y')}"
        cell.alignment = Alignment(horizontal='center')
        row += 2

        # Section 1: Diagnosis table
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "1. Total Number of Patients"
        cell.font = section_font
        row += 1

        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "Conditions seen are as follows:"
        row += 1

        # Table headers
        headers = ['S/N', 'Condition Seen', 'Number']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = table_header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        row += 1

        # Data rows
        for idx, (diagnosis_name, count) in enumerate(diagnosis_breakdown, 1):
            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=2, value=diagnosis_name).border = border
            ws.cell(row=row, column=3, value=count).border = border
            row += 1

        # Total row
        ws.cell(row=row, column=2, value="TOTAL PATIENTS SEEN").font = total_font
        ws.cell(row=row, column=2).border = border
        cell_total = ws.cell(row=row, column=3, value=total_patients)
        cell_total.font = total_font
        cell_total.border = border
        row += 2

        # Section 2: Financial Summary
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "2. Financial Summary"
        cell.font = section_font
        row += 1

        # Drug income
        ws.cell(row=row, column=1, value="a. Total income from drugs")
        cell_amount = ws.cell(row=row, column=3, value=float(drug_income))
        cell_amount.number_format = '#,##0.00'
        row += 1

        # Service categories
        letter_idx = ord('b')
        for item in service_income_breakdown:
            ws.cell(row=row, column=1, value=f"{chr(letter_idx)}. Total income from {item['category'].lower()}")
            cell_amount = ws.cell(row=row, column=3, value=float(item['amount']))
            cell_amount.number_format = '#,##0.00'
            row += 1
            letter_idx += 1

        # Surgery income
        ws.cell(row=row, column=1, value=f"{chr(letter_idx)}. Total income from surgery")
        cell_amount = ws.cell(row=row, column=3, value=float(surgery_income))
        cell_amount.number_format = '#,##0.00'
        row += 1
        letter_idx += 1

        # Total income
        section_label = selected_specialization.name.lower() if selected_specialization else "all sections"
        ws.cell(row=row, column=1, value=f"{chr(letter_idx)}. Total income from {section_label}").font = total_font
        cell_amount = ws.cell(row=row, column=3, value=float(total_income))
        cell_amount.font = total_font
        cell_amount.number_format = '#,##0.00'
        row += 2

        # Signatures
        if show_signature:
            ws.merge_cells(f'A{row}:C{row}')
            cell = ws[f'A{row}']
            cell.value = "3. Approval Signatures"
            cell.font = section_font
            row += 2

            for title in ['Compiled by:', 'Clinic Incharge:', 'Supervisor:']:
                ws.cell(row=row, column=1, value=title).font = Font(bold=True)
                row += 1
                ws.cell(row=row, column=1, value="Name: _______________________")
                row += 1
                ws.cell(row=row, column=1, value="Date: _______________________")
                row += 2

        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 18

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"consultation_report_{from_date.strftime('%Y%m%d')}_{to_date.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class ConsultationReportExportPDFView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'consultation.view_consultation_report'

    def get(self, request, *args, **kwargs):
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        report_title = request.GET.get('title', 'Consultation Report')
        specialization_id = request.GET.get('specialization')
        show_signature = request.GET.get('show_signature', 'true') == 'true'

        # Date filtering (same as Excel)
        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()
        if not to_date:
            today = date.today()
            if today.month == 12:
                to_date = today.replace(day=31)
            else:
                to_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        else:
            to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()

        # Get data (reuse logic from Excel export)
        selected_specialization = None
        if specialization_id and specialization_id != 'all':
            selected_specialization = SpecializationModel.objects.filter(id=specialization_id).first()

        consultations = ConsultationSessionModel.objects.filter(
            created_at__date__range=[from_date, to_date],
            status='completed'
        )

        if selected_specialization:
            consultations = consultations.filter(
                queue_entry__specialization=selected_specialization
            )

        # Diagnosis breakdown
        primary_diagnosis_counts = consultations.filter(
            primary_diagnosis__isnull=False
        ).values(
            'primary_diagnosis__name'
        ).annotate(
            count=Count('id')
        ).order_by('-count')

        secondary_diagnosis_counts = {}
        for consultation in consultations:
            for diagnosis in consultation.secondary_diagnoses.all():
                if diagnosis.name in secondary_diagnosis_counts:
                    secondary_diagnosis_counts[diagnosis.name] += 1
                else:
                    secondary_diagnosis_counts[diagnosis.name] = 1

        diagnosis_breakdown = {}
        for item in primary_diagnosis_counts:
            diagnosis_breakdown[item['primary_diagnosis__name']] = item['count']

        for name, count in secondary_diagnosis_counts.items():
            if name in diagnosis_breakdown:
                diagnosis_breakdown[name] += count
            else:
                diagnosis_breakdown[name] = count

        diagnosis_breakdown = sorted(
            diagnosis_breakdown.items(),
            key=lambda x: x[1],
            reverse=True
        )

        total_patients = consultations.count()

        # Financial data
        consultation_ids = list(consultations.values_list('id', flat=True))

        drug_income = DrugOrderModel.objects.filter(
            consultation_id__in=consultation_ids,
            status__in=['paid', 'dispensed', 'partially_dispensed']
        ).aggregate(
            total=Sum(
                (F('quantity_ordered') * F('drug__selling_price')),
                output_field=DecimalField()
            )
        )['total'] or Decimal('0.00')

        service_income_breakdown = []
        if selected_specialization:
            service_categories = ServiceCategory.objects.filter(
                specializations=selected_specialization,
                is_active=True
            )
        else:
            service_categories = ServiceCategory.objects.filter(is_active=True)

        for category in service_categories:
            category_income = PatientServiceTransaction.objects.filter(
                consultation_id__in=consultation_ids,
                status__in=['paid', 'fully_dispensed', 'partially_dispensed']
            ).filter(
                Q(service__category=category) | Q(service_item__category=category)
            ).aggregate(
                total=Sum('amount_paid')
            )['total'] or Decimal('0.00')

            if category_income > 0:
                service_income_breakdown.append({
                    'category': category.name,
                    'amount': category_income
                })

        surgery_filter = Q(status='completed')
        if selected_specialization:
            surgery_filter &= Q(specialization=selected_specialization)

        surgery_income = Surgery.objects.filter(
            surgery_filter,
            created_at__date__range=[from_date, to_date]
        ).aggregate(
            total=Sum(
                F('custom_surgeon_fee') +
                F('custom_anesthesia_fee') +
                F('custom_facility_fee'),
                output_field=DecimalField()
            )
        )['total'] or Decimal('0.00')

        total_service_income = sum(item['amount'] for item in service_income_breakdown)
        total_income = drug_income + total_service_income + surgery_income

        lab_setting = LabSettingModel.objects.first()
        site_info = SiteInfoModel.objects.first()

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30,
                                topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                     fontSize=16, alignment=TA_CENTER, spaceAfter=12)
        subtitle_style = ParagraphStyle('CustomSubtitle', parent=styles['Normal'],
                                        fontSize=11, alignment=TA_CENTER, spaceAfter=6)
        section_style = ParagraphStyle('SectionTitle', parent=styles['Heading2'],
                                       fontSize=12, spaceAfter=8, spaceBefore=16)

        # Header
        if lab_setting and lab_setting.lab_name:
            elements.append(Paragraph(lab_setting.lab_name, title_style))
        elif site_info:
            elements.append(Paragraph(site_info.name, title_style))

        elements.append(Spacer(1, 0.3 * inch))
        elements.append(Paragraph(report_title, title_style))
        elements.append(
            Paragraph(
                f"Period: {from_date.strftime('%B %d, %Y')} to {to_date.strftime('%B %d, %Y')}",
                subtitle_style
            )
        )
        elements.append(Spacer(1, 0.4 * inch))

        # Section 1: Diagnosis Table
        elements.append(Paragraph("1. Total Number of Patients", section_style))
        elements.append(Paragraph("Conditions seen are as follows:", styles['Normal']))
        elements.append(Spacer(1, 0.2 * inch))

        # Build diagnosis table
        headers = ['S/N', 'Condition Seen', 'Number']
        col_widths = [0.6 * inch, 4 * inch, 1.2 * inch]
        table_data = [headers]

        for idx, (diagnosis_name, count) in enumerate(diagnosis_breakdown, 1):
            row = [str(idx), diagnosis_name, str(count)]
            table_data.append(row)

        # Total row
        total_row = ['', 'TOTAL PATIENTS SEEN', str(total_patients)]
        table_data.append(total_row)

        diagnosis_table = Table(table_data, colWidths=col_widths)
        diagnosis_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # S/N
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),  # Numbers
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f0f0f0')]),
            # Total row
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f0f0f0')),
        ]))
        elements.append(diagnosis_table)
        elements.append(Spacer(1, 0.4 * inch))

        # Section 2: Financial Summary
        elements.append(Paragraph("2. Financial Summary", section_style))
        elements.append(Spacer(1, 0.2 * inch))

        financial_data = []
        financial_data.append(['a. Total income from drugs', f"₦{drug_income:,.2f}"])

        letter_idx = ord('b')
        for item in service_income_breakdown:
            financial_data.append([
                f"{chr(letter_idx)}. Total income from {item['category'].lower()}",
                f"₦{item['amount']:,.2f}"
            ])
            letter_idx += 1

        financial_data.append([
            f"{chr(letter_idx)}. Total income from surgery",
            f"₦{surgery_income:,.2f}"
        ])
        letter_idx += 1

        section_label = selected_specialization.name.lower() if selected_specialization else "all sections"
        financial_data.append([
            f"{chr(letter_idx)}. Total income from {section_label}",
            f"₦{total_income:,.2f}"
        ])

        financial_table = Table(financial_data, colWidths=[4 * inch, 1.8 * inch])
        financial_table.setStyle(TableStyle([
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#4472C4')),
            ('LINEBELOW', (0, -1), (-1, -1), 2, colors.HexColor('#4472C4')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f0f0f0')),
        ]))
        elements.append(financial_table)

        # Section 3: Signatures
        if show_signature:
            elements.append(Spacer(1, 0.5 * inch))
            elements.append(Paragraph("3. Approval Signatures", section_style))

            sig_style = ParagraphStyle(
                'SigStyle',
                parent=styles['Normal'],
                fontSize=10,
                alignment=TA_LEFT,
            )

            signatures = [
                "Compiled by:",
                "Clinic Incharge:",
                "Supervisor:"
            ]

            for sig_title in signatures:
                sig_text = f"""
                    <br/>
                    <b>{sig_title}</b><br/>
                    <br/>
                    _________________________<br/>
                    <b>Name:</b> _______________________<br/>
                    <b>Date:</b> _______________________
                """
                elements.append(Paragraph(sig_text, sig_style))
                elements.append(Spacer(1, 0.3 * inch))

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f"consultation_report_{from_date.strftime('%Y%m%d')}_{to_date.strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


"""
Admin Site / Reports App Views - Complete Implementation
=========================================================
General Financial Report (Vertical Format)
Excel & PDF Exports
"""

# ============================================================================
# MAIN VIEW
# ============================================================================


class GeneralFinancialReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    template_name = 'admin_site/reports/general_financial_report.html'
    permission_required = 'finance.view_financial_reports'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Date range
        from_date = self.request.GET.get('from_date')
        to_date = self.request.GET.get('to_date')

        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()

        if not to_date:
            today = date.today()
            if today.month == 12:
                to_date = today.replace(day=31)
            else:
                to_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        else:
            to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()

        # Report title
        month_year = from_date.strftime('%B %Y')
        default_title = f'General Financial Report for {month_year}'
        report_title = self.request.GET.get('title', default_title)

        # Build vertical table data
        financial_data = []

        # 1. CARD (Registration)
        card_total = RegistrationPaymentModel.objects.filter(
            date__range=[from_date, to_date], status='confirmed'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        financial_data.append({
            'category': 'Card (Registration)',
            'amount': card_total
        })

        # 2. CONS (Consultation)
        cons_total = PatientTransactionModel.objects.filter(
            date__range=[from_date, to_date],
            transaction_type='consultation_payment',
            status='completed'
        ).aggregate(total=Sum('direct_payment_amount'))['total'] or Decimal('0.00')

        financial_data.append({
            'category': 'Consultation',
            'amount': cons_total
        })

        # 3. LAB
        lab_total = PatientTransactionModel.objects.filter(
            date__range=[from_date, to_date],
            transaction_type='lab_payment',
            status='completed'
        ).aggregate(total=Sum('direct_payment_amount'))['total'] or Decimal('0.00')

        financial_data.append({
            'category': 'Laboratory',
            'amount': lab_total
        })

        # 4. DRUGS
        drugs_total = PatientTransactionModel.objects.filter(
            date__range=[from_date, to_date],
            transaction_type='drug_payment',
            status='completed'
        ).aggregate(total=Sum('direct_payment_amount'))['total'] or Decimal('0.00')

        financial_data.append({
            'category': 'Drugs',
            'amount': drugs_total
        })

        # 5. SCAN
        scan_total = PatientTransactionModel.objects.filter(
            date__range=[from_date, to_date],
            transaction_type='scan_payment',
            status='completed'
        ).aggregate(total=Sum('direct_payment_amount'))['total'] or Decimal('0.00')

        financial_data.append({
            'category': 'Scan/Imaging',
            'amount': scan_total
        })

        # 6. Service Categories (Dynamic)
        service_categories = ServiceCategory.objects.filter(
            show_as_record_column=True,
            is_active=True
        ).order_by('name')

        for category in service_categories:
            category_total = PatientTransactionModel.objects.filter(
                date__range=[from_date, to_date],
                transaction_type__in=['service', 'item'],
                status='completed'
            ).filter(
                Q(service__service__category=category) |
                Q(service__service_item__category=category)
            ).aggregate(total=Sum('direct_payment_amount'))['total'] or Decimal('0.00')

            if category_total > 0:  # Only add if there are transactions
                financial_data.append({
                    'category': category.name,
                    'amount': category_total
                })

        # 7. SURGERY
        surgery_total = PatientTransactionModel.objects.filter(
            date__range=[from_date, to_date],
            transaction_type='surgery_payment',
            status='completed'
        ).aggregate(total=Sum('direct_payment_amount'))['total'] or Decimal('0.00')

        financial_data.append({
            'category': 'Surgery',
            'amount': surgery_total
        })

        # 8. Other Payments (Dynamic by service name)
        other_payments = PatientTransactionModel.objects.filter(
            date__range=[from_date, to_date],
            transaction_type='other_payment',
            status='completed'
        ).values('other_service__name').annotate(
            total=Sum('direct_payment_amount')
        ).order_by('other_service__name')

        for item in other_payments:
            if item['total']:
                financial_data.append({
                    'category': item['other_service__name'] or 'Other Payment',
                    'amount': item['total']
                })

        # Calculate grand total
        grand_total = sum(item['amount'] for item in financial_data)

        context.update({
            'from_date': from_date,
            'to_date': to_date,
            'report_title': report_title,
            'financial_data': financial_data,
            'grand_total': grand_total,
        })

        return context


# ============================================================================
# EXCEL EXPORT
# ============================================================================

class GeneralFinancialReportExcelView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'finance.view_financial_reports'

    def get(self, request, *args, **kwargs):
        # Get parameters
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        report_title = request.GET.get('title', '')

        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()

        if not to_date:
            today = date.today()
            if today.month == 12:
                to_date = today.replace(day=31)
            else:
                to_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        else:
            to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()

        if not report_title:
            month_year = from_date.strftime('%B %Y')
            report_title = f'General Financial Report for {month_year}'

        # Build financial data (same logic as view)
        financial_data = []

        # 1. CARD (Registration)
        card_total = RegistrationPaymentModel.objects.filter(
            date__range=[from_date, to_date], status='confirmed'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        financial_data.append(('Card (Registration)', card_total))

        # 2. CONS (Consultation)
        cons_total = PatientTransactionModel.objects.filter(
            date__range=[from_date, to_date],
            transaction_type='consultation_payment',
            status='completed'
        ).aggregate(total=Sum('direct_payment_amount'))['total'] or Decimal('0.00')

        financial_data.append(('Consultation', cons_total))

        # 3. LAB
        lab_total = PatientTransactionModel.objects.filter(
            date__range=[from_date, to_date],
            transaction_type='lab_payment',
            status='completed'
        ).aggregate(total=Sum('direct_payment_amount'))['total'] or Decimal('0.00')

        financial_data.append(('Laboratory', lab_total))

        # 4. DRUGS
        drugs_total = PatientTransactionModel.objects.filter(
            date__range=[from_date, to_date],
            transaction_type='drug_payment',
            status='completed'
        ).aggregate(total=Sum('direct_payment_amount'))['total'] or Decimal('0.00')

        financial_data.append(('Drugs', drugs_total))

        # 5. SCAN
        scan_total = PatientTransactionModel.objects.filter(
            date__range=[from_date, to_date],
            transaction_type='scan_payment',
            status='completed'
        ).aggregate(total=Sum('direct_payment_amount'))['total'] or Decimal('0.00')

        financial_data.append(('Scan/Imaging', scan_total))

        # 6. Service Categories (Dynamic)
        service_categories = ServiceCategory.objects.filter(
            show_as_record_column=True,
            is_active=True
        ).order_by('name')

        for category in service_categories:
            category_total = PatientTransactionModel.objects.filter(
                date__range=[from_date, to_date],
                transaction_type__in=['service', 'item'],
                status='completed'
            ).filter(
                Q(service__service__category=category) |
                Q(service__service_item__category=category)
            ).aggregate(total=Sum('direct_payment_amount'))['total'] or Decimal('0.00')

            if category_total > 0:
                financial_data.append((category.name, category_total))

        # 7. SURGERY
        surgery_total = PatientTransactionModel.objects.filter(
            date__range=[from_date, to_date],
            transaction_type='surgery_payment',
            status='completed'
        ).aggregate(total=Sum('direct_payment_amount'))['total'] or Decimal('0.00')

        financial_data.append(('Surgery', surgery_total))

        # 8. Other Payments (Dynamic)
        other_payments = PatientTransactionModel.objects.filter(
            date__range=[from_date, to_date],
            transaction_type='other_payment',
            status='completed'
        ).values('other_service__name').annotate(
            total=Sum('direct_payment_amount')
        ).order_by('other_service__name')

        for item in other_payments:
            if item['total']:
                financial_data.append((item['other_service__name'] or 'Other Payment', item['total']))

        # Calculate grand total
        grand_total = sum(amount for _, amount in financial_data)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Financial Report"

        # Styles
        header_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=11, color="FFFFFF")
        section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        total_font = Font(bold=True, size=12, color="FFFFFF")
        total_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        row = 1

        # Header
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = report_title
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        row += 1

        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = f"Period: {from_date.strftime('%B %d, %Y')} to {to_date.strftime('%B %d, %Y')}"
        cell.alignment = Alignment(horizontal='center')
        row += 2

        # Column headers
        ws[f'A{row}'] = "Category"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws[f'B{row}'] = "Amount"
        ws[f'B{row}'].font = section_font
        ws[f'B{row}'].fill = section_fill
        ws[f'B{row}'].alignment = Alignment(horizontal='right')
        row += 1

        # Data rows
        for category, amount in financial_data:
            ws.cell(row=row, column=1, value=category)
            ws.cell(row=row, column=2, value=float(amount))
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
            row += 1

        # Grand total
        ws.cell(row=row, column=1, value="GRAND TOTAL").font = total_font
        ws.cell(row=row, column=1).fill = total_fill
        ws.cell(row=row, column=2, value=float(grand_total)).font = total_font
        ws.cell(row=row, column=2).fill = total_fill
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')

        # Column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18

        # Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"financial_report_{from_date.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


# ============================================================================
# PDF EXPORT
# ============================================================================

class GeneralFinancialReportPDFView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'finance.view_financial_reports'

    def get(self, request, *args, **kwargs):
        # Get parameters
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        report_title = request.GET.get('title', '')

        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()

        if not to_date:
            today = date.today()
            if today.month == 12:
                to_date = today.replace(day=31)
            else:
                to_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        else:
            to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()

        if not report_title:
            month_year = from_date.strftime('%B %Y')
            report_title = f'General Financial Report for {month_year}'

        # Build financial data (same logic as view)
        financial_data = []

        # 1. CARD (Registration)
        card_total = RegistrationPaymentModel.objects.filter(
            date__range=[from_date, to_date], status='confirmed'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        financial_data.append(('Card (Registration)', card_total))

        # 2. CONS (Consultation)
        cons_total = PatientTransactionModel.objects.filter(
            date__range=[from_date, to_date],
            transaction_type='consultation_payment',
            status='completed'
        ).aggregate(total=Sum('direct_payment_amount'))['total'] or Decimal('0.00')

        financial_data.append(('Consultation', cons_total))

        # 3. LAB
        lab_total = PatientTransactionModel.objects.filter(
            date__range=[from_date, to_date],
            transaction_type='lab_payment',
            status='completed'
        ).aggregate(total=Sum('direct_payment_amount'))['total'] or Decimal('0.00')

        financial_data.append(('Laboratory', lab_total))

        # 4. DRUGS
        drugs_total = PatientTransactionModel.objects.filter(
            date__range=[from_date, to_date],
            transaction_type='drug_payment',
            status='completed'
        ).aggregate(total=Sum('direct_payment_amount'))['total'] or Decimal('0.00')

        financial_data.append(('Drugs', drugs_total))

        # 5. SCAN
        scan_total = PatientTransactionModel.objects.filter(
            date__range=[from_date, to_date],
            transaction_type='scan_payment',
            status='completed'
        ).aggregate(total=Sum('direct_payment_amount'))['total'] or Decimal('0.00')

        financial_data.append(('Scan/Imaging', scan_total))

        # 6. Service Categories (Dynamic)
        service_categories = ServiceCategory.objects.filter(
            show_as_record_column=True,
            is_active=True
        ).order_by('name')

        for category in service_categories:
            category_total = PatientTransactionModel.objects.filter(
                date__range=[from_date, to_date],
                transaction_type__in=['service', 'item'],
                status='completed'
            ).filter(
                Q(service__service__category=category) |
                Q(service__service_item__category=category)
            ).aggregate(total=Sum('direct_payment_amount'))['total'] or Decimal('0.00')

            if category_total > 0:
                financial_data.append((category.name, category_total))

        # 7. SURGERY
        surgery_total = PatientTransactionModel.objects.filter(
            date__range=[from_date, to_date],
            transaction_type='surgery_payment',
            status='completed'
        ).aggregate(total=Sum('direct_payment_amount'))['total'] or Decimal('0.00')

        financial_data.append(('Surgery', surgery_total))

        # 8. Other Payments (Dynamic)
        other_payments = PatientTransactionModel.objects.filter(
            date__range=[from_date, to_date],
            transaction_type='other_payment',
            status='completed'
        ).values('other_service__name').annotate(
            total=Sum('direct_payment_amount')
        ).order_by('other_service__name')

        for item in other_payments:
            if item['total']:
                financial_data.append((item['other_service__name'] or 'Other Payment', item['total']))

        # Calculate grand total
        grand_total = sum(amount for _, amount in financial_data)

        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30,
                                topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            alignment=TA_CENTER,
            spaceAfter=12
        )

        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=11,
            alignment=TA_CENTER,
            spaceAfter=6
        )

        # Header
        elements.append(Paragraph(report_title, title_style))
        elements.append(Paragraph(
            f"Period: {from_date.strftime('%B %d, %Y')} to {to_date.strftime('%B %d, %Y')}",
            subtitle_style
        ))
        elements.append(Spacer(1, 0.4 * inch))

        # Build table data
        table_data = [['Category', 'Amount']]

        for category, amount in financial_data:
            table_data.append([category, f'₦{amount:,.2f}'])

        # Add grand total
        table_data.append(['', ''])
        table_data.append(['GRAND TOTAL', f'₦{grand_total:,.2f}'])

        # Create table
        table = Table(table_data, colWidths=[4 * inch, 2 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('GRID', (0, 0), (-1, -3), 1, colors.grey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#28a745')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 12),
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.black),
        ]))

        elements.append(table)

        # Add summary box
        elements.append(Spacer(1, 0.5 * inch))

        summary_data = [[f'Total Revenue: ₦{grand_total:,.2f}']]
        summary_table = Table(summary_data, colWidths=[6 * inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#28a745')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('GRID', (0, 0), (-1, -1), 2, colors.black),
        ]))

        elements.append(summary_table)

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f"financial_report_{from_date.strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response