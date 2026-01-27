import calendar
import logging
import json
from datetime import datetime, date, timedelta
from decimal import Decimal

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from io import BytesIO
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.models import User
from django.contrib.messages.views import SuccessMessageMixin
from django.db import transaction
from django.db.models import Q, Count, Sum, Value
from django.db.models.functions import Concat
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils.timezone import now
from django.views import View
from django.views.generic import (
    CreateView, ListView, UpdateView, DeleteView, DetailView, TemplateView
)
from openpyxl.styles import Font, Alignment

from admin_site.models import SiteInfoModel
from finance.models import PatientTransactionModel
from insurance.models import InsuranceClaimModel
from patient.models import PatientModel, PatientWalletModel
from .models import *
from .forms import *
from django.db.models import Q, Count, Case, When, IntegerField
logger = logging.getLogger(__name__)


# -------------------------
# Mixins
# -------------------------
class FlashFormErrorsMixin:
    """
    Mixin for CreateView/UpdateView to flash form errors and redirect safely.
    Use before SuccessMessageMixin in MRO so messages appear before redirect.
    """

    def form_invalid(self, form):
        try:
            for field, errors in form.errors.items():
                field_name = form.fields.get(field).label if form.fields.get(field) else field
                for error in errors:
                    messages.error(self.request, f"{field_name}: {error}")
        except Exception:
            logger.exception("Error while processing form_invalid errors.")
            messages.error(self.request, "There was an error processing the form. Please try again.")
        return redirect(self.get_success_url())


# -------------------------
# Lab Test Category Views
# -------------------------
class LabTestCategoryCreateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, SuccessMessageMixin,
    CreateView
):
    model = LabTestCategoryModel
    permission_required = 'laboratory.add_labtestcategorymodel'
    form_class = LabTestCategoryForm
    template_name = 'laboratory/category/index.html'
    success_message = 'Lab Test Category Successfully Created'

    def get_success_url(self):
        return reverse('lab_category_index')

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET':
            return redirect(reverse('lab_category_index'))
        return super().dispatch(request, *args, **kwargs)


class LabTestCategoryListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = LabTestCategoryModel
    permission_required = 'laboratory.view_labtestcategorymodel'
    template_name = 'laboratory/category/index.html'
    context_object_name = "category_list"

    def get_queryset(self):
        return LabTestCategoryModel.objects.all().order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = LabTestCategoryForm()
        return context


class LabTestCategoryUpdateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, SuccessMessageMixin, UpdateView
):
    model = LabTestCategoryModel
    permission_required = 'laboratory.add_labtestcategorymodel'
    form_class = LabTestCategoryForm
    template_name = 'laboratory/category/index.html'
    success_message = 'Lab Test Category Successfully Updated'

    def get_success_url(self):
        return reverse('lab_category_index')

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET':
            return redirect(reverse('lab_category_index'))
        return super().dispatch(request, *args, **kwargs)


class LabTestCategoryDeleteView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, DeleteView):
    model = LabTestCategoryModel
    permission_required = 'laboratory.add_labtestcategorymodel'
    template_name = 'laboratory/category/delete.html'
    context_object_name = "category"
    success_message = 'Lab Test Category Successfully Deleted'

    def get_success_url(self):
        return reverse('lab_category_index')


# -------------------------
# Lab Test Template Views
# -------------------------
class LabTestTemplateCreateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin,
    CreateView
):
    model = LabTestTemplateModel
    permission_required = 'laboratory.add_labtestcategorymodel'
    form_class = LabTestTemplateForm
    template_name = 'laboratory/template/create.html'
    success_message = 'Lab Test Template Successfully Created'

    def get_success_url(self):
        return reverse('lab_template_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['category_list'] = LabTestCategoryModel.objects.all().order_by('name')
        return context


class LabTestTemplateListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = LabTestTemplateModel
    permission_required = 'laboratory.view_labtestcategorymodel'
    template_name = 'laboratory/template/index.html'
    context_object_name = "template_list"

    def get_queryset(self):
        return LabTestTemplateModel.objects.select_related('category').order_by('category__name', 'name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['category_list'] = LabTestCategoryModel.objects.all().order_by('name')
        return context


class LabTestTemplateDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = LabTestTemplateModel
    permission_required = 'laboratory.view_labtestcategorymodel'
    template_name = 'laboratory/template/detail.html'
    context_object_name = "template"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        template = self.object

        # Get orders for this template
        orders = LabTestOrderModel.objects.filter(template=template).select_related('patient').order_by('-ordered_at')[
                 :10]

        # Get equipment that supports this template
        equipment = template.equipment.filter(status='active')

        # Get reagents used for this template
        reagents = template.reagents.filter(is_active=True)

        context.update({
            'recent_orders': orders,
            'equipment_list': equipment,
            'reagent_list': reagents,
            'total_orders': LabTestOrderModel.objects.filter(template=template).count(),
            'lab_setting': LabSettingModel.objects.first()
        })
        return context


class LabTestTemplateUpdateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, UpdateView
):
    model = LabTestTemplateModel
    permission_required = 'laboratory.add_labtestcategorymodel'
    form_class = LabTestTemplateForm
    template_name = 'laboratory/template/edit.html'
    success_message = 'Lab Test Template Successfully Updated'

    def get_success_url(self):
        return reverse('lab_template_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['category_list'] = LabTestCategoryModel.objects.all().order_by('name')
        context['template'] = self.object
        return context


class LabTestTemplateToggleStatusView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'lab.add_labtestcategorymodel'

    def post(self, request, pk):
        template = get_object_or_404(LabTestTemplateModel, pk=pk)
        action = request.POST.get('action')

        if action == 'deactivate':
            reason = request.POST.get('reason', '').strip()
            if not reason:
                messages.error(request, 'Reason for deactivation is required.')
                return redirect('lab_template_detail', pk=pk)

            template.is_active = False
            template.reason_for_deactivate = reason
            template.save()
            messages.success(request, f'Template "{template.name}" has been deactivated.')

        elif action == 'activate':
            template.is_active = True
            template.reason_for_deactivate = ''
            template.save()
            messages.success(request, f'Template "{template.name}" has been activated.')

        else:
            messages.error(request, 'Invalid action.')

        return redirect('lab_template_detail', pk=pk)


class LabTestTemplateDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = LabTestTemplateModel
    permission_required = 'laboratory.add_labtestcategorymodel'
    template_name = 'laboratory/template/delete.html'
    context_object_name = "template"
    success_message = 'Lab Test Template Successfully Deleted'

    def get_success_url(self):
        return reverse('lab_template_index')


# -------------------------
# Lab Entry Point - Patient Verification
# -------------------------
class LabEntryView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Entry point for lab operations - patient verification"""
    permission_required = 'laboratory.add_labtestordermodel'
    template_name = 'laboratory/order/entry.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Laboratory - Patient Verification'
        context['lab_setting'] = LabSettingModel.objects.first()
        return context


@login_required
def verify_lab_patient_ajax(request):
    """AJAX view to verify patient by card number"""
    if not request.user.has_perm('laboratory.view_labtestordermodel'):
        return JsonResponse({'success': False, 'error': 'Permission denied'})

    card_number = request.GET.get('card_number', '').strip()
    if not card_number:
        return JsonResponse({'success': False, 'error': 'Please enter a card number'})

    try:
        patient = PatientModel.objects.get(card_number__iexact=card_number, status='active')

        # Get internal test counts
        test_counts = {
            'total': LabTestOrderModel.objects.filter(patient=patient).count(),
            'pending': LabTestOrderModel.objects.filter(patient=patient, status='pending').count(),
            'paid': LabTestOrderModel.objects.filter(patient=patient, status='paid').count(),
            'processing': LabTestOrderModel.objects.filter(patient=patient,
                                                           status__in=['collected', 'processing']).count(),
            'completed': LabTestOrderModel.objects.filter(patient=patient, status='completed').count(),
        }

        # --- NEW CODE: Count external orders awaiting results ---
        external_pending_count = ExternalLabTestOrder.objects.filter(
            patient=patient
        ).filter(
            Q(result_file__isnull=True) | Q(result_file='')
        ).count()
        test_counts['external_pending_results'] = external_pending_count
        # --- END NEW CODE ---

        return JsonResponse({
            'success': True,
            'patient': {
                'id': patient.id,
                'full_name': patient.__str__(),
                'patient_id': patient.card_number,
                'phone': patient.mobile,
                'email': getattr(patient, 'email', ''),
                'age': patient.age() if hasattr(patient, 'age') and callable(patient.age) else '',
                'gender': getattr(patient, 'gender', ''),
                'address': patient.address or 'Not provided',
            },
            'test_counts': test_counts
        })

    except PatientModel.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': f'No patient found with card number: {card_number}'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'An error occurred while verifying patient {str(e)}'
        })


class PatientExternalLabTestListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """View a patient's external lab tests with filtering and result management."""
    model = ExternalLabTestOrder
    # It's good practice to have a specific permission for external orders
    permission_required = 'laboratory.view_externallabtestorder'
    template_name = 'laboratory/order/patient_external_tests.html'
    context_object_name = 'external_orders'
    paginate_by = 20

    def get_queryset(self):
        patient_id = self.kwargs.get('patient_id')
        self.patient = get_object_or_404(PatientModel, id=patient_id, status='active')

        queryset = ExternalLabTestOrder.objects.filter(
            patient=self.patient
        ).select_related(
            'ordered_by', 'result_uploaded_by'
        ).order_by('-ordered_at')

        # Filter by result status if provided via GET parameter
        result_status = self.request.GET.get('result_status', '')
        if result_status == 'pending':
            queryset = queryset.filter(Q(result_file__isnull=True) | Q(result_file=''))
        elif result_status == 'uploaded':
            queryset = queryset.filter(result_file__isnull=False).exclude(result_file='')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['patient'] = self.patient

        # Group orders by date for a cleaner display, similar to the internal tests view
        orders_by_date = {}
        for order in context['external_orders']:
            order_date = order.ordered_at.date()
            if order_date not in orders_by_date:
                orders_by_date[order_date] = []
            orders_by_date[order_date].append(order)
        context['orders_by_date'] = orders_by_date

        # Get counts for the summary cards
        all_orders = ExternalLabTestOrder.objects.filter(patient=self.patient)
        context['result_counts'] = {
            'total': all_orders.count(),
            'pending': all_orders.filter(Q(result_file__isnull=True) | Q(result_file='')).count(),
            'uploaded': all_orders.filter(result_file__isnull=False).exclude(result_file='').count(),
        }

        # Pass the current filter to the template to highlight the active filter button
        context['current_filter'] = self.request.GET.get('result_status', '')

        return context


class UploadExternalLabResultView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Handles the POST request for uploading or changing a result file."""
    permission_required = 'laboratory.change_externallabtestorder'

    def post(self, request, order_id):
        order = get_object_or_404(ExternalLabTestOrder, id=order_id)
        result_file = request.FILES.get('result_file')

        if not result_file:
            messages.error(request, 'No file was selected for upload.')
            return redirect(reverse('patient_external_lab_tests', kwargs={'patient_id': order.patient.id}))

        # Update the order instance with the file and tracking info
        order.result_file = result_file
        order.result_uploaded_by = request.user
        order.result_uploaded_at = timezone.now()
        order.save()

        messages.success(request, f"Result for order {order.order_number} was uploaded successfully.")
        return redirect(reverse('patient_external_lab_tests', kwargs={'patient_id': order.patient.id}))


# -------------------------
# Patient Lab Tests Management
# -------------------------
class PatientLabTestsView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """View patient's lab tests with grouping by date"""
    model = LabTestOrderModel
    permission_required = 'laboratory.view_labtestordermodel'
    template_name = 'laboratory/order/patient_tests.html'
    context_object_name = 'orders'
    paginate_by = 30

    def get_queryset(self):
        patient_id = self.kwargs.get('patient_id')
        self.patient = get_object_or_404(PatientModel, id=patient_id, status='active')

        queryset = LabTestOrderModel.objects.filter(
            patient=self.patient
        ).select_related(
            'template', 'template__category', 'ordered_by', 'payment_by'
        ).order_by('-ordered_at')

        # Filter by status if provided
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['patient'] = self.patient
        context['status_choices'] = LabTestOrderModel.STATUS_CHOICES

        # Group orders by date
        orders_by_date = {}
        for order in context['orders']:
            order_date = order.ordered_at.date()
            if order_date not in orders_by_date:
                orders_by_date[order_date] = []
            orders_by_date[order_date].append(order)

        context['orders_by_date'] = orders_by_date

        # Get test counts for summary
        context['test_counts'] = {
            'total': LabTestOrderModel.objects.filter(patient=self.patient).count(),
            'pending': LabTestOrderModel.objects.filter(patient=self.patient, status='pending').count(),
            'paid': LabTestOrderModel.objects.filter(patient=self.patient, status='paid').count(),
            'collected': LabTestOrderModel.objects.filter(patient=self.patient, status='collected').count(),
            'processing': LabTestOrderModel.objects.filter(patient=self.patient, status='processing').count(),
            'completed': LabTestOrderModel.objects.filter(patient=self.patient, status='completed').count(),
        }

        # Get available test templates for new order
        context['template_list'] = LabTestTemplateModel.objects.filter(
            is_active=True
        ).select_related('category').order_by('category__name', 'name')

        return context


# -------------------------
# Lab Test Order CRUD Views
# -------------------------

class LabTestOrderCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = LabTestOrderModel
    permission_required = 'laboratory.add_labtestordermodel'
    form_class = LabTestOrderCreateForm
    template_name = 'laboratory/order/create.html'

    def get_initial(self):
        initial = super().get_initial()
        patient_id = self.kwargs.get('patient_id')
        if patient_id:
            initial['patient'] = patient_id
        initial['ordered_by'] = self.request.user
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        patient_id = self.kwargs.get('patient_id')
        if patient_id:
            context['patient'] = get_object_or_404(PatientModel, id=patient_id)
        context['template_list'] = LabTestTemplateModel.objects.filter(is_active=True).select_related('category').order_by('category__name', 'name')
        return context

    def post(self, request, *args, **kwargs):
        form = self.get_form()
        # Prefer an array-style field 'tests' (JS change recommended below)
        test_ids = request.POST.getlist('tests')
        # Fallback: capture any test_0, test_1 style
        if not test_ids:
            for k, v in request.POST.items():
                if k.startswith('test_') and v:
                    test_ids.append(v)

        if not test_ids:
            messages.error(request, "Please select at least one test.")
            return redirect(request.path)

        if not form.is_valid():
            for field, errs in form.errors.items():
                for e in errs:
                    messages.error(request, f"{field}: {e}")
            return redirect(request.path)

        created_orders = []
        try:
            with transaction.atomic():
                patient = get_object_or_404(PatientModel, id=self.kwargs.get('patient_id'))
                for tid in test_ids:
                    try:
                        template = LabTestTemplateModel.objects.get(pk=int(tid))
                    except (ValueError, LabTestTemplateModel.DoesNotExist):
                        messages.warning(request, f"Test id {tid} not found — skipped.")
                        continue

                    order = LabTestOrderModel(
                        patient = patient,
                        template = template,
                        ordered_by = request.user,
                        status = 'pending',
                        amount_charged = template.price,
                        special_instructions = form.cleaned_data.get('special_instructions', ''),
                        expected_completion = form.cleaned_data.get('expected_completion'),
                    )
                    order.save()
                    created_orders.append(order)
        except Exception:
            logger.exception("Error creating lab orders")
            messages.error(request, "An error occurred creating orders. Contact admin.")
            return redirect(request.path)

        if created_orders:
            messages.success(request, f"Created {len(created_orders)} order(s).")
            return redirect(reverse('patient_lab_tests', kwargs={'patient_id': created_orders[0].patient.id}))
        else:
            messages.error(request, "No valid tests selected.")
            return redirect(request.path)

@login_required
def process_lab_payments(request):
    """Process payments for selected lab tests with wallet balance and insurance validation"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})

    if not request.user.has_perm('laboratory.add_labtestordermodel'):
        return JsonResponse({'success': False, 'error': 'Permission denied'})

    selected_orders = request.POST.getlist('selected_orders')
    if not selected_orders:
        return JsonResponse({'success': False, 'error': 'No tests selected'})

    try:
        with transaction.atomic():
            # Get all selected orders
            orders = LabTestOrderModel.objects.filter(
                id__in=selected_orders,
                status='pending'
            ).select_related('template', 'patient')

            if not orders.exists():
                return JsonResponse({'success': False, 'error': 'No valid pending tests found'})

            # Get patient (assuming all orders are for the same patient)
            patient = orders.first().patient

            # Get or create patient wallet
            wallet, created = PatientWalletModel.objects.get_or_create(
                patient=patient,
                defaults={'amount': Decimal('0.00')}
            )

            # Check for active insurance
            active_insurance = None
            if hasattr(patient, 'insurance_policies'):
                active_insurance = patient.insurance_policies.filter(
                    is_active=True,
                    valid_to__gte=date.today()
                ).select_related('hmo', 'coverage_plan').first()

            # Calculate insurance amount helper function
            def calculate_patient_amount(base_amount, coverage_percentage):
                """Calculate patient's portion after insurance"""
                if coverage_percentage and coverage_percentage > 0:
                    covered_amount = base_amount * (coverage_percentage / 100)
                    return base_amount - covered_amount
                return base_amount

            # Calculate total amounts considering insurance
            total_base_amount = Decimal('0.00')
            total_patient_amount = Decimal('0.00')
            total_insurance_covered = Decimal('0.00')
            order_details = []

            for order in orders:
                # Get base amount
                base_amount = order.amount_charged or order.template.price

                # Apply insurance if applicable
                if active_insurance and active_insurance.coverage_plan.is_lab_covered(order.template):
                    patient_amount = calculate_patient_amount(
                        base_amount,
                        active_insurance.coverage_plan.lab_coverage_percentage
                    )
                    insurance_covered = base_amount - patient_amount
                else:
                    patient_amount = base_amount
                    insurance_covered = Decimal('0.00')

                total_base_amount += base_amount
                total_patient_amount += patient_amount
                total_insurance_covered += insurance_covered

                order_details.append({
                    'order': order,
                    'base_amount': base_amount,
                    'patient_amount': patient_amount,
                    'insurance_covered': insurance_covered
                })

            # Check wallet balance
            if wallet.amount < total_patient_amount:
                return JsonResponse({
                    'success': False,
                    'error': f'Insufficient wallet balance. Required: ₦{total_patient_amount:,.2f}, Available: ₦{wallet.amount:,.2f}. Please visit Finance to fund wallet.',
                    'required_amount': float(total_patient_amount),
                    'available_amount': float(wallet.amount),
                    'shortage': float(total_patient_amount - wallet.amount)
                })

            # Create insurance claim if insurance is active and covers any amount
            insurance_claim = None
            if active_insurance and total_insurance_covered > 0:
                # Determine claim type
                claim_type = 'laboratory' if len(orders) == 1 else 'multiple'

                insurance_claim = InsuranceClaimModel.objects.create(
                    patient_insurance=active_insurance,
                    claim_type=claim_type,
                    total_amount=total_base_amount,
                    covered_amount=total_insurance_covered,
                    patient_amount=total_patient_amount,
                    service_date=timezone.now(),
                    created_by=request.user,
                    notes=f'Auto-generated claim for {len(orders)} lab test(s)'
                )

            # Process payments and update orders
            updated_count = 0

            for detail in order_details:
                order = detail['order']
                patient_amount = detail['patient_amount']
                insurance_covered = detail['insurance_covered']

                # Deduct patient amount from wallet
                wallet.amount -= patient_amount

                # Update order
                order.status = 'paid'
                order.payment_status = True
                order.payment_date = timezone.now()
                order.payment_by = request.user

                if not order.amount_charged:
                    order.amount_charged = detail['base_amount']

                # Link to insurance claim if created
                if insurance_claim:
                    order.insurance_claim = insurance_claim

                order.save()
                updated_count += 1

            # Save wallet
            wallet.save()

            # Prepare response message
            message_parts = [
                f'Successfully processed payment for {updated_count} test(s)',
                f'Patient paid: ₦{total_patient_amount:,.2f}'
            ]

            if insurance_claim:
                message_parts.append(
                    f'Insurance claim {insurance_claim.claim_number} created for ₦{total_insurance_covered:,.2f}')

            return JsonResponse({
                'success': True,
                'message': '. '.join(message_parts),
                'details': {
                    'processed_count': updated_count,
                    'patient_amount_paid': float(total_patient_amount),
                    'insurance_amount_covered': float(total_insurance_covered),
                    'new_wallet_balance': float(wallet.amount),
                    'formatted_wallet_balance': f'₦{wallet.amount:,.2f}',
                    'insurance_claim_number': insurance_claim.claim_number if insurance_claim else None
                }
            })

    except PatientWalletModel.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Patient wallet not found. Please contact Finance to create wallet.'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error processing payment: {str(e)}'
        })


# -------------------------
# Sample Collection
# -------------------------
def collect_sample(request, order_id):
    """Mark test sample as collected"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})

    if not request.user.has_perm('laboratory.change_labtestordermodel'):
        return JsonResponse({'success': False, 'error': 'Permission denied'})

    try:
        order = get_object_or_404(LabTestOrderModel, id=order_id, status='paid')

        sample_label = request.POST.get('sample_label', '').strip()
        expected_completion_str = request.POST.get('expected_completion', '')

        if not sample_label:
            return JsonResponse({'success': False, 'error': 'Sample label is required'})

        # Update order
        order.status = 'collected'
        order.sample_collected_at = timezone.now()
        order.sample_collected_by = request.user
        order.sample_label = sample_label

        # Set expected completion if provided
        if expected_completion_str:
            try:
                expected_completion = datetime.strptime(expected_completion_str, '%Y-%m-%dT%H:%M')
                order.expected_completion = expected_completion
            except ValueError:
                pass  # Invalid date format, ignore

        order.save()

        return JsonResponse({
            'success': True,
            'message': f'Sample collected for {order.template.name}',
            'sample_label': sample_label
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error collecting sample: {str(e)}'})


class LabTestOrderListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = LabTestOrderModel
    permission_required = 'laboratory.view_labtestordermodel'
    template_name = 'laboratory/order/index.html'
    context_object_name = "order_list"
    paginate_by = 20

    # In your existing LabTestOrderListView.get_queryset() method
    # REPLACE the existing queryset line with:

    def get_queryset(self):
        queryset = LabTestOrderModel.objects.select_related(
            'patient', 'template', 'ordered_by'
        ).exclude(
            source='walkin'  # ADD THIS LINE - Exclude walk-in orders
        ).exclude(
            status__in=['pending', 'cancelled']
        ).order_by('-ordered_at')

        # Rest of your existing filter logic remains the same...
        status = self.request.GET.get('status')
        template_id = self.request.GET.get('template')
        search_query = self.request.GET.get('search', '').strip()

        if status:
            queryset = queryset.filter(status=status)

        if template_id:
            queryset = queryset.filter(template__id=template_id)

        if search_query:
            queryset = queryset.annotate(
                patient_full_name=Concat(
                    'patient__first_name', Value(' '), 'patient__last_name'
                )
            ).filter(
                Q(patient_full_name__icontains=search_query) |
                Q(patient__card_number__icontains=search_query) |
                Q(order_number__icontains=search_query) |
                Q(sample_label__icontains=search_query)
            )

        return queryset


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Pass the current search query to the template to keep it in the search box
        context['search_query'] = self.request.GET.get('search', '').strip()

        context['status_choices'] = LabTestOrderModel.STATUS_CHOICES
        context['template_list'] = LabTestTemplateModel.objects.filter(is_active=True).order_by('name')

        # Add counts for dashboard
        # These counts are for all orders, not just the filtered ones.
        context['status_counts'] = {
            'pending': LabTestOrderModel.objects.filter(status='pending').count(),
            'paid': LabTestOrderModel.objects.filter(status='paid').count(),
            'collected': LabTestOrderModel.objects.filter(status='collected').count(),
            'processing': LabTestOrderModel.objects.filter(status='processing').count(),
            'completed': LabTestOrderModel.objects.filter(status='completed').count(),
        }

        return context


class LabTestOrderDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = LabTestOrderModel
    permission_required = 'laboratory.view_labtestordermodel'
    template_name = 'laboratory/order/detail.html'
    context_object_name = "order"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        order = self.object

        # Check if results exist
        has_results = hasattr(order, 'result')

        context.update({
            'has_results': has_results,
            'lab_setting': LabSettingModel.objects.first(),
            'can_process_payment': order.status == 'pending' and self.request.user.has_perm(
                'laboratory.change_labtestordermodel'),
            'can_collect_sample': order.status == 'paid' and self.request.user.has_perm(
                'laboratory.change_labtestordermodel'),
            'can_process_test': order.status == 'collected' and self.request.user.has_perm(
                'laboratory.change_labtestordermodel'),
            'can_complete_test': order.status == 'processing' and self.request.user.has_perm(
                'laboratory.change_labtestordermodel'),
        })

        return context


class LabTestOrderUpdateView(
    LoginRequiredMixin, PermissionRequiredMixin, UpdateView
):
    model = LabTestOrderModel
    permission_required = 'laboratory.add_labtestordermodel'
    form_class = LabTestOrderForm
    template_name = 'laboratory/order/edit.html'

    def get_success_url(self):
        return reverse('lab_order_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['order'] = self.object
        context['patient_list'] = PatientModel.objects.filter(is_active=True).order_by('first_name')
        context['template_list'] = LabTestTemplateModel.objects.filter(is_active=True).order_by('name')
        return context


# -------------------------
# Lab Test Result Views
# -------------------------

@login_required
@permission_required('laboratory.view_labtestordermodel', raise_exception=True)
def laboratory_dashboard(request):
    """Main laboratory dashboard with comprehensive statistics"""

    # Get current date and time ranges
    today = timezone.now().date()
    this_week_start = today - timedelta(days=today.weekday())
    this_month_start = today.replace(day=1)
    last_month_start = (this_month_start - timedelta(days=1)).replace(day=1)
    last_month_end = this_month_start - timedelta(days=1)

    # === BASIC STATISTICS ===

    # Total counts
    total_orders = LabTestOrderModel.objects.count()
    total_templates = LabTestTemplateModel.objects.filter(is_active=True).count()
    total_categories = LabTestCategoryModel.objects.count()
    completed_tests = LabTestOrderModel.objects.filter(status='completed').count()

    # Today's statistics
    orders_today = LabTestOrderModel.objects.filter(ordered_at__date=today).count()
    completed_today = LabTestOrderModel.objects.filter(
        status='completed',
        processed_at__date=today
    ).count()
    pending_today = LabTestOrderModel.objects.filter(
        ordered_at__date=today,
        status__in=['pending', 'paid', 'collected', 'processing']
    ).count()

    # This week's statistics
    orders_week = LabTestOrderModel.objects.filter(ordered_at__date__gte=this_week_start).count()
    completed_week = LabTestOrderModel.objects.filter(
        status='completed',
        processed_at__date__gte=this_week_start
    ).count()

    # This month's statistics
    orders_month = LabTestOrderModel.objects.filter(ordered_at__date__gte=this_month_start).count()
    completed_month = LabTestOrderModel.objects.filter(
        status='completed',
        processed_at__date__gte=this_month_start
    ).count()

    # Last month's statistics for growth calculation
    orders_last_month = LabTestOrderModel.objects.filter(
        ordered_at__date__gte=last_month_start,
        ordered_at__date__lte=last_month_end
    ).count()
    completed_last_month = LabTestOrderModel.objects.filter(
        status='completed',
        processed_at__date__gte=last_month_start,
        processed_at__date__lte=last_month_end
    ).count()

    # Calculate growth percentages
    orders_growth = 0
    completed_growth = 0
    if orders_last_month > 0:
        orders_growth = round(((orders_month - orders_last_month) / orders_last_month) * 100, 1)
    if completed_last_month > 0:
        completed_growth = round(((completed_month - completed_last_month) / completed_last_month) * 100, 1)

    # === STATUS DISTRIBUTION ===
    status_distribution = LabTestOrderModel.objects.values('status').annotate(
        count=Count('id')
    ).order_by('status')

    # Convert to format suitable for charts
    status_chart_data = [
        {'name': status['status'].title(), 'value': status['count']}
        for status in status_distribution
    ]

    # === REVENUE STATISTICS ===

    # Total revenue
    total_revenue = LabTestOrderModel.objects.filter(
        payment_status=True
    ).aggregate(total=Sum('amount_charged'))['total'] or Decimal('0.00')

    # Today's revenue
    revenue_today = LabTestOrderModel.objects.filter(
        payment_status=True,
        payment_date__date=today
    ).aggregate(total=Sum('amount_charged'))['total'] or Decimal('0.00')

    # This week's revenue
    revenue_week = LabTestOrderModel.objects.filter(
        payment_status=True,
        payment_date__date__gte=this_week_start
    ).aggregate(total=Sum('amount_charged'))['total'] or Decimal('0.00')

    # This month's revenue
    revenue_month = LabTestOrderModel.objects.filter(
        payment_status=True,
        payment_date__date__gte=this_month_start
    ).aggregate(total=Sum('amount_charged'))['total'] or Decimal('0.00')

    # === CATEGORY WISE STATISTICS ===
    category_stats = LabTestCategoryModel.objects.annotate(
        total_orders=Count('templates__orders'),
        completed_orders=Count('templates__orders', filter=Q(templates__orders__status='completed')),
        revenue=Sum('templates__orders__amount_charged', filter=Q(templates__orders__payment_status=True))
    ).order_by('-total_orders')

    # Format for chart
    category_chart_data = [
        {
            'name': cat.name,
            'value': cat.total_orders,
            'revenue': float(cat.revenue or 0)
        }
        for cat in category_stats[:10]  # Top 10 categories
    ]

    # === POPULAR TESTS ===
    popular_tests = LabTestTemplateModel.objects.annotate(
        order_count=Count('orders')
    ).filter(order_count__gt=0).order_by('-order_count')[:10]

    popular_tests_data = [
        {
            'name': test.name,
            'orders': test.order_count,
            'revenue': float(
                LabTestOrderModel.objects.filter(
                    template=test,
                    payment_status=True
                ).aggregate(total=Sum('amount_charged'))['total'] or 0
            )
        }
        for test in popular_tests
    ]

    # === DAILY TRENDS (LAST 7 DAYS) ===
    daily_trends = []
    for i in range(7):
        date = today - timedelta(days=6 - i)
        orders_count = LabTestOrderModel.objects.filter(ordered_at__date=date).count()
        completed_count = LabTestOrderModel.objects.filter(
            status='completed',
            processed_at__date=date
        ).count()
        revenue = LabTestOrderModel.objects.filter(
            payment_status=True,
            payment_date__date=date
        ).aggregate(total=Sum('amount_charged'))['total'] or Decimal('0.00')

        daily_trends.append({
            'date': date.strftime('%Y-%m-%d'),
            'orders': orders_count,
            'completed': completed_count,
            'revenue': float(revenue)
        })

    # === MONTHLY TRENDS (LAST 12 MONTHS) ===
    monthly_trends = []
    for i in range(12):
        month_start = (this_month_start - timedelta(days=1)).replace(day=1)
        for _ in range(i):
            month_start = (month_start - timedelta(days=1)).replace(day=1)

        month_end = (month_start.replace(month=month_start.month + 1)
                     if month_start.month < 12
                     else month_start.replace(year=month_start.year + 1, month=1)) - timedelta(days=1)

        orders_count = LabTestOrderModel.objects.filter(
            ordered_at__date__gte=month_start,
            ordered_at__date__lte=month_end
        ).count()

        completed_count = LabTestOrderModel.objects.filter(
            status='completed',
            processed_at__date__gte=month_start,
            processed_at__date__lte=month_end
        ).count()

        revenue = LabTestOrderModel.objects.filter(
            payment_status=True,
            payment_date__date__gte=month_start,
            payment_date__date__lte=month_end
        ).aggregate(total=Sum('amount_charged'))['total'] or Decimal('0.00')

        monthly_trends.insert(0, {
            'month': month_start.strftime('%b %Y'),
            'orders': orders_count,
            'completed': completed_count,
            'revenue': float(revenue)
        })

    # === SOURCE DISTRIBUTION ===
    source_distribution = LabTestOrderModel.objects.values('source').annotate(
        count=Count('id')
    )

    source_chart_data = [
        {
            'name': 'Doctor Prescribed' if source['source'] == 'doctor' else 'Lab Direct',
            'value': source['count']
        }
        for source in source_distribution
    ]

    # === RECENT ACTIVITY ===
    recent_orders = LabTestOrderModel.objects.select_related(
        'patient', 'template', 'ordered_by'
    ).order_by('-ordered_at')[:10]

    # === AVERAGE PROCESSING TIME ===
    completed_orders = LabTestOrderModel.objects.filter(
        status='completed',
        processed_at__isnull=False
    )

    avg_processing_hours = 0
    if completed_orders.exists():
        total_processing_time = sum([
            (order.processed_at - order.ordered_at).total_seconds() / 3600
            for order in completed_orders
            if order.processed_at and order.ordered_at
        ])
        avg_processing_hours = round(total_processing_time / completed_orders.count(), 1)

    # === PENDING TASKS ===
    pending_collection = LabTestOrderModel.objects.filter(status='paid').count()
    pending_processing = LabTestOrderModel.objects.filter(status='collected').count()
    pending_verification = LabTestResultModel.objects.filter(is_verified=False).count()

    context = {
        # Basic stats
        'total_orders': total_orders,
        'total_templates': total_templates,
        'total_categories': total_categories,
        'completed_tests': completed_tests,

        # Daily stats
        'orders_today': orders_today,
        'completed_today': completed_today,
        'pending_today': pending_today,

        # Weekly stats
        'orders_week': orders_week,
        'completed_week': completed_week,

        # Monthly stats
        'orders_month': orders_month,
        'completed_month': completed_month,
        'orders_growth': orders_growth,
        'completed_growth': completed_growth,

        # Revenue
        'total_revenue': total_revenue,
        'revenue_today': revenue_today,
        'revenue_week': revenue_week,
        'revenue_month': revenue_month,

        # Charts data
        'status_distribution': json.dumps(status_chart_data),
        'category_distribution': json.dumps(category_chart_data),
        'popular_tests': popular_tests_data,
        'daily_trends': json.dumps(daily_trends),
        'monthly_trends': json.dumps(monthly_trends),
        'source_distribution': json.dumps(source_chart_data),

        # Other stats
        'avg_processing_hours': avg_processing_hours,
        'recent_orders': recent_orders,

        # Pending tasks
        'pending_collection': pending_collection,
        'pending_processing': pending_processing,
        'pending_verification': pending_verification,
    }

    return render(request, 'laboratory/dashboard.html', context)


@login_required
def laboratory_dashboard_print(request):
    """Printable version of laboratory dashboard"""
    # Get the same context as main dashboard but simplified for printing
    context = laboratory_dashboard(request).context_data
    return render(request, 'laboratory/dashboard_print.html', context)


@login_required
def laboratory_analytics_api(request):
    """API endpoint for dynamic chart updates"""
    chart_type = request.GET.get('type')

    if chart_type == 'daily_revenue':
        # Last 30 days revenue
        data = []
        today = timezone.now().date()
        for i in range(30):
            date = today - timedelta(days=29 - i)
            revenue = LabTestOrderModel.objects.filter(
                payment_status=True,
                payment_date__date=date
            ).aggregate(total=Sum('amount_charged'))['total'] or Decimal('0.00')

            data.append({
                'date': date.strftime('%Y-%m-%d'),
                'revenue': float(revenue)
            })

        return JsonResponse({'data': data})

    elif chart_type == 'category_performance':
        # Category wise performance this month
        this_month_start = timezone.now().date().replace(day=1)

        categories = LabTestCategoryModel.objects.annotate(
            orders_this_month=Count(
                'templates__orders',
                filter=Q(templates__orders__ordered_at__date__gte=this_month_start)
            ),
            revenue_this_month=Sum(
                'templates__orders__amount_charged',
                filter=Q(
                    templates__orders__payment_status=True,
                    templates__orders__payment_date__date__gte=this_month_start
                )
            )
        ).order_by('-orders_this_month')[:10]

        data = [
            {
                'name': cat.name,
                'orders': cat.orders_this_month,
                'revenue': float(cat.revenue_this_month or 0)
            }
            for cat in categories
        ]

        return JsonResponse({'data': data})

    return JsonResponse({'error': 'Invalid chart type'}, status=400)


class LabReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    template_name = 'laboratory/reports/index.html'
    permission_required = 'laboratory.view_labtestordermodel'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Date range from request or default to this month
        from_date = self.request.GET.get('from_date')
        to_date = self.request.GET.get('to_date')

        if not from_date:
            from_date = date.today().replace(day=1)  # First day of current month
        else:
            from_date = datetime.strptime(from_date, '%Y-%m-%d').date()

        if not to_date:
            # Last day of current month
            if date.today().month == 12:
                to_date = date.today().replace(day=31)
            else:
                to_date = (date.today().replace(month=date.today().month + 1, day=1) - timedelta(days=1))
        else:
            to_date = datetime.strptime(to_date, '%Y-%m-%d').date()

        # Default title
        month_year = from_date.strftime('%B %Y')
        default_title = f'Laboratory Test Report for the Month of {month_year}'
        report_title = self.request.GET.get('title', default_title)

        # Filter orders - exclude pending and cancelled from total
        all_orders = LabTestOrderModel.objects.filter(
            ordered_at__date__range=[from_date, to_date]
        )

        # Total orders exclude pending and cancelled
        orders = all_orders.exclude(status__in=['pending', 'cancelled'])

        # Get order by parameter
        order_by = self.request.GET.get('order_by', 'name')

        # Test type breakdown
        test_breakdown = orders.values(
            'template__name', 'template__id'
        ).annotate(
            total_orders=Count('id'),
            completed_orders=Count('id', filter=Q(status='completed'))
        )

        # Apply ordering
        if order_by == 'total':
            test_breakdown = test_breakdown.order_by('-total_orders', 'template__name')
        elif order_by == 'completed':
            test_breakdown = test_breakdown.order_by('-completed_orders', 'template__name')
        else:  # default to name
            test_breakdown = test_breakdown.order_by('template__name')

        # Summary statistics
        context.update({
            'from_date': from_date,
            'to_date': to_date,
            'report_title': report_title,
            'test_breakdown': test_breakdown,
            'order_by': order_by,

            # Summary stats
            'total_tests': orders.count(),
            'completed_tests': all_orders.filter(status='completed').count(),
            'processing_tests': all_orders.filter(status='processing').count(),
            'collected_tests': all_orders.filter(status='collected').count(),
            'paid_tests': all_orders.filter(status='paid').count(),
            'pending_tests': all_orders.filter(status='pending').count(),
            'cancelled_tests': all_orders.filter(status='cancelled').count(),
        })

        return context


class LabReportExportExcelView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'laboratory.view_labtestordermodel'

    def get(self, request, *args, **kwargs):
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        report_title = request.GET.get('title', 'Laboratory Test Report')
        show_total = request.GET.get('show_total', 'true') == 'true'
        show_completed = request.GET.get('show_completed', 'true') == 'true'
        order_by = request.GET.get('order_by', 'name')

        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = datetime.strptime(from_date, '%Y-%m-%d').date()

        if not to_date:
            if date.today().month == 12:
                to_date = date.today().replace(day=31)
            else:
                to_date = (date.today().replace(month=date.today().month + 1, day=1) - timedelta(days=1))
        else:
            to_date = datetime.strptime(to_date, '%Y-%m-%d').date()

        # Get lab and site info
        lab_setting = LabSettingModel.objects.first()
        site_info = SiteInfoModel.objects.first()

        # Query data
        orders = LabTestOrderModel.objects.filter(
            ordered_at__date__range=[from_date, to_date]
        ).exclude(status__in=['pending', 'cancelled'])

        test_breakdown = orders.values(
            'template__name', 'template__id'
        ).annotate(
            total_orders=Count('id'),
            completed_orders=Count('id', filter=Q(status='completed'))
        )

        # Apply ordering
        if order_by == 'total':
            test_breakdown = test_breakdown.order_by('-total_orders', 'template__name')
        elif order_by == 'completed':
            test_breakdown = test_breakdown.order_by('-completed_orders', 'template__name')
        else:
            test_breakdown = test_breakdown.order_by('template__name')

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Lab Test Report"

        # Styles
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True, size=12)
        table_header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        row = 1

        # Header section
        if lab_setting and lab_setting.lab_name:
            ws.merge_cells(f'A{row}:D{row}')
            cell = ws[f'A{row}']
            cell.value = lab_setting.lab_name
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            row += 1

            if lab_setting.mobile or lab_setting.email:
                ws.merge_cells(f'A{row}:D{row}')
                cell = ws[f'A{row}']
                contact = []
                if lab_setting.mobile:
                    contact.append(f"Tel: {lab_setting.mobile}")
                if lab_setting.email:
                    contact.append(f"Email: {lab_setting.email}")
                cell.value = " | ".join(contact)
                cell.alignment = Alignment(horizontal='center')
                row += 1
        elif site_info:
            ws.merge_cells(f'A{row}:D{row}')
            cell = ws[f'A{row}']
            cell.value = site_info.name
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            row += 1

            ws.merge_cells(f'A{row}:D{row}')
            cell = ws[f'A{row}']
            contact = []
            if site_info.mobile_1:
                contact.append(f"Tel: {site_info.mobile_1}")
            if site_info.email:
                contact.append(f"Email: {site_info.email}")
            cell.value = " | ".join(contact)
            cell.alignment = Alignment(horizontal='center')
            row += 1

        row += 1

        # Report title
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = report_title
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center')
        row += 1

        # Date range
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = f"Period: {from_date.strftime('%B %d, %Y')} to {to_date.strftime('%B %d, %Y')}"
        cell.alignment = Alignment(horizontal='center')
        row += 2

        # Table headers
        col = 1
        headers = ['S/N', 'Test']
        if show_total:
            headers.append('Total Orders')
        if show_completed:
            headers.append('Completed Orders')

        for header in headers:
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = table_header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            col += 1

        row += 1

        # Data rows
        for idx, test in enumerate(test_breakdown, 1):
            col = 1

            # S/N
            cell = ws.cell(row=row, column=col)
            cell.value = idx
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            col += 1

            # Test name
            cell = ws.cell(row=row, column=col)
            cell.value = test['template__name']
            cell.border = border
            col += 1

            # Total orders
            if show_total:
                cell = ws.cell(row=row, column=col)
                cell.value = test['total_orders']
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                col += 1

            # Completed orders
            if show_completed:
                cell = ws.cell(row=row, column=col)
                cell.value = test['completed_orders']
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                col += 1

            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 40
        if show_total:
            ws.column_dimensions['C'].width = 15
        if show_completed:
            ws.column_dimensions['D'].width = 18

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"lab_test_report_{from_date.strftime('%Y%m%d')}_{to_date.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response


class LabReportExportPDFView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'laboratory.view_labtestordermodel'

    def get(self, request, *args, **kwargs):
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        report_title = request.GET.get('title', 'Laboratory Test Report')
        show_total = request.GET.get('show_total', 'true') == 'true'
        show_completed = request.GET.get('show_completed', 'true') == 'true'
        order_by = request.GET.get('order_by', 'name')
        report_type = request.GET.get('type', 'breakdown')  # breakdown or summary

        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = datetime.strptime(from_date, '%Y-%m-%d').date()

        if not to_date:
            if date.today().month == 12:
                to_date = date.today().replace(day=31)
            else:
                to_date = (date.today().replace(month=date.today().month + 1, day=1) - timedelta(days=1))
        else:
            to_date = datetime.strptime(to_date, '%Y-%m-%d').date()

        # Get lab and site info
        lab_setting = LabSettingModel.objects.first()
        site_info = SiteInfoModel.objects.first()

        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#333333'),
            spaceAfter=6,
            alignment=TA_CENTER
        )

        # Header
        if lab_setting and lab_setting.lab_name:
            elements.append(Paragraph(lab_setting.lab_name, title_style))
            if lab_setting.mobile or lab_setting.email:
                contact = []
                if lab_setting.mobile:
                    contact.append(f"Tel: {lab_setting.mobile}")
                if lab_setting.email:
                    contact.append(f"Email: {lab_setting.email}")
                elements.append(Paragraph(" | ".join(contact), subtitle_style))
        elif site_info:
            elements.append(Paragraph(site_info.name, title_style))
            contact = []
            if site_info.mobile_1:
                contact.append(f"Tel: {site_info.mobile_1}")
            if site_info.email:
                contact.append(f"Email: {site_info.email}")
            elements.append(Paragraph(" | ".join(contact), subtitle_style))

        elements.append(Spacer(1, 0.3 * inch))
        elements.append(Paragraph(report_title, title_style))
        elements.append(Paragraph(
            f"Period: {from_date.strftime('%B %d, %Y')} to {to_date.strftime('%B %d, %Y')}",
            subtitle_style
        ))
        elements.append(Spacer(1, 0.4 * inch))

        if report_type == 'summary':
            # Summary report
            all_orders = LabTestOrderModel.objects.filter(
                ordered_at__date__range=[from_date, to_date]
            )
            orders = all_orders.exclude(status__in=['pending', 'cancelled'])

            summary_data = [
                ['Metric', 'Count'],
                ['Total Tests', str(orders.count())],
                ['Completed Tests', str(all_orders.filter(status='completed').count())],
                ['Processing Tests', str(all_orders.filter(status='processing').count())],
                ['Sample Collected', str(all_orders.filter(status='collected').count())],
                ['Paid (Awaiting Sample)', str(all_orders.filter(status='paid').count())],
                ['Pending Payment', str(all_orders.filter(status='pending').count())],
                ['Cancelled Tests', str(all_orders.filter(status='cancelled').count())],
            ]

            summary_table = Table(summary_data, colWidths=[4 * inch, 2 * inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
            ]))
            elements.append(summary_table)
        else:
            # Breakdown report
            orders = LabTestOrderModel.objects.filter(
                ordered_at__date__range=[from_date, to_date]
            ).exclude(status__in=['pending', 'cancelled'])

            test_breakdown = orders.values(
                'template__name', 'template__id'
            ).annotate(
                total_orders=Count('id'),
                completed_orders=Count('id', filter=Q(status='completed'))
            )

            # Apply ordering
            if order_by == 'total':
                test_breakdown = test_breakdown.order_by('-total_orders', 'template__name')
            elif order_by == 'completed':
                test_breakdown = test_breakdown.order_by('-completed_orders', 'template__name')
            else:
                test_breakdown = test_breakdown.order_by('template__name')

            # Build table data
            headers = ['S/N', 'Test']
            col_widths = [0.5 * inch, 3 * inch]

            if show_total:
                headers.append('Total Orders')
                col_widths.append(1 * inch)
            if show_completed:
                headers.append('Completed Orders')
                col_widths.append(1.2 * inch)

            table_data = [headers]

            for idx, test in enumerate(test_breakdown, 1):
                row = [str(idx), test['template__name']]
                if show_total:
                    row.append(str(test['total_orders']))
                if show_completed:
                    row.append(str(test['completed_orders']))
                table_data.append(row)

            table = Table(table_data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
            ]))
            elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f"lab_test_report_{from_date.strftime('%Y%m%d')}_{to_date.strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class LabTestLogView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    template_name = 'laboratory/reports/log.html'
    permission_required = 'laboratory.view_labtestordermodel'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # === Date range (same logic as your other report) ===
        from_date = self.request.GET.get('from_date')
        to_date = self.request.GET.get('to_date')

        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = datetime.strptime(from_date, '%Y-%m-%d').date()

        if not to_date:
            today = date.today()
            if today.month == 12:
                to_date = today.replace(day=31)
            else:
                to_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        else:
            to_date = datetime.strptime(to_date, '%Y-%m-%d').date()

        default_title = f'Laboratory Test Log ({from_date.strftime("%b %d, %Y")} - {to_date.strftime("%b %d, %Y")})'
        report_title = self.request.GET.get('title', default_title)

        # === Query Data ===
        # Filter: All paid tests (not pending or cancelled)
        orders_qs = LabTestOrderModel.objects.filter(
            ordered_at__date__range=[from_date, to_date]
        ).exclude(
            status__in=['pending', 'cancelled']
        ).select_related(
            'patient',
            'template',
            'result__verified_by'  # Joins patient, template, and the result's verifier
        ).order_by('ordered_at')

        # === Calculate Totals ===
        totals = orders_qs.aggregate(
            total_amount=Sum('amount_charged'),
            total_patients=Count('patient', distinct=True)
        )

        context.update({
            'from_date': from_date,
            'to_date': to_date,
            'report_title': report_title,
            'orders': orders_qs,
            'total_amount': totals['total_amount'] or Decimal('0.00'),
            'total_patients': totals['total_patients'] or 0,
        })

        return context


class LabTestLogExportExcelView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'laboratory.view_labtestordermodel'

    def get(self, request, *args, **kwargs):
        # === Get Parameters ===
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        report_title = request.GET.get('title', 'Laboratory Test Log')

        # Toggles
        show_status = request.GET.get('show_status', 'true') == 'true'
        show_amount = request.GET.get('show_amount', 'true') == 'true'
        show_scientist = request.GET.get('show_scientist', 'true') == 'true'

        # Date filtering
        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        if not to_date:
            today = date.today()
            if today.month == 12:
                to_date = today.replace(day=31)
            else:
                to_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        else:
            to_date = datetime.strptime(to_date, '%Y-%m-%d').date()

        # === Get Data ===
        orders_qs = LabTestOrderModel.objects.filter(
            ordered_at__date__range=[from_date, to_date]
        ).exclude(
            status__in=['pending', 'cancelled']
        ).select_related(
            'patient', 'template', 'result__verified_by'
        ).order_by('ordered_at')

        totals = orders_qs.aggregate(
            total_amount=Sum('amount_charged'),
            total_patients=Count('patient', distinct=True)
        )
        total_amount = totals['total_amount'] or Decimal('0.00')
        total_patients = totals['total_patients'] or 0

        # === Create Workbook ===
        wb = Workbook()
        ws = wb.active
        ws.title = "Lab Test Log"

        # Styles
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True, size=12)
        table_header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        total_font = Font(bold=True, size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))

        # Get lab and site info (using your existing logic)
        lab_setting = LabSettingModel.objects.first()
        site_info = SiteInfoModel.objects.first()
        row = 1

        # Header section
        if lab_setting and lab_setting.lab_name:
            ws.merge_cells(f'A{row}:F{row}')
            cell = ws[f'A{row}']
            cell.value = lab_setting.lab_name
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            row += 1
        elif site_info:
            ws.merge_cells(f'A{row}:F{row}')
            cell = ws[f'A{row}']
            cell.value = site_info.name
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            row += 1
        row += 1

        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = report_title
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center')
        row += 1

        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = f"Period: {from_date.strftime('%B %d, %Y')} to {to_date.strftime('%B %d, %Y')}"
        cell.alignment = Alignment(horizontal='center')
        row += 2

        # === Table Headers ===
        headers = ['S/N', 'Date', 'Patient ID', 'Patient Name', 'Test Name']
        if show_status:
            headers.append('Status')
        if show_amount:
            headers.append('Amount')
        if show_scientist:
            headers.append('Verified By')

        col = 1
        for header in headers:
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = table_header_font
            cell.fill = header_fill
            cell.border = border
            col += 1
        row += 1

        # === Data Rows ===
        for idx, order in enumerate(orders_qs, 1):
            col = 1

            # Static columns
            ws.cell(row=row, column=col, value=idx).border = border;
            col += 1
            ws.cell(row=row, column=col, value=order.ordered_at.strftime('%Y-%m-%d %H:%M')).border = border;
            col += 1
            ws.cell(row=row, column=col, value=order.patient.card_number).border = border;
            col += 1
            ws.cell(row=row, column=col, value=order.patient.__str__()).border = border;
            col += 1
            ws.cell(row=row, column=col, value=order.template.name).border = border;
            col += 1

            # Dynamic columns
            if show_status:
                ws.cell(row=row, column=col, value=order.get_status_display()).border = border;
                col += 1
            if show_amount:
                cell = ws.cell(row=row, column=col, value=order.amount_charged)
                cell.number_format = '#,##0.00'
                cell.border = border
                col += 1

            if show_scientist:
                verifier = 'N/A'

                # Safe check for related object's existence first
                if (hasattr(order, 'result') and
                        order.result and
                        order.result.verified_by):

                    user = order.result.verified_by

                    # Safe check for your nested staff profile
                    if (hasattr(user, 'user_staff_profile') and
                            user.user_staff_profile and
                            user.user_staff_profile.staff):
                        verifier = user.user_staff_profile.staff.__str__()
                    else:
                        verifier = user.get_full_name()  # Fallback to full name

                ws.cell(row=row, column=col, value=verifier).border = border
                col += 1

            row += 1

        # === Total Row ===
        col = 1
        ws.cell(row=row, column=col).border = border;
        col += 1
        ws.cell(row=row, column=col).border = border;
        col += 1
        ws.cell(row=row, column=col).border = border;
        col += 1
        ws.cell(row=row, column=col).border = border;
        col += 1
        ws.cell(row=row, column=col, value="TOTAL").font = total_font;
        cell.border = border;
        col += 1

        if show_status:
            ws.cell(row=row, column=col).border = border;
            col += 1
        if show_amount:
            cell = ws.cell(row=row, column=col, value=total_amount)
            cell.font = total_font
            cell.number_format = '#,##0.00'
            cell.border = border
            col += 1
        if show_scientist:
            cell = ws.cell(row=row, column=col, value=f"{total_patients} Unique Patients")
            cell.font = total_font
            cell.border = border
            col += 1

        # Column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 35
        if show_status: ws.column_dimensions['F'].width = 15
        if show_amount: ws.column_dimensions['G'].width = 15
        if show_scientist: ws.column_dimensions['H'].width = 30

        # === Prepare Response ===
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"lab_test_log_{from_date.strftime('%Y%m%d')}_{to_date.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class LabTestLogExportPDFView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'laboratory.view_labtestordermodel'

    def get(self, request, *args, **kwargs):
        # === Get Parameters ===
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        report_title = request.GET.get('title', 'Laboratory Test Log')

        show_status = request.GET.get('show_status', 'true') == 'true'
        show_amount = request.GET.get('show_amount', 'true') == 'true'
        show_scientist = request.GET.get('show_scientist', 'true') == 'true'

        # Date filtering
        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        if not to_date:
            today = date.today()
            if today.month == 12:
                to_date = today.replace(day=31)
            else:
                to_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        else:
            to_date = datetime.strptime(to_date, '%Y-%m-%d').date()

        # === Get Data ===
        orders_qs = LabTestOrderModel.objects.filter(
            ordered_at__date__range=[from_date, to_date]
        ).exclude(
            status__in=['pending', 'cancelled']
        ).select_related(
            'patient', 'template', 'result__verified_by'
        ).order_by('ordered_at')

        totals = orders_qs.aggregate(
            total_amount=Sum('amount_charged'),
            total_patients=Count('patient', distinct=True)
        )
        total_amount = totals['total_amount'] or Decimal('0.00')
        total_patients = totals['total_patients'] or 0

        # === Create PDF ===
        buffer = BytesIO()
        # Use landscape for more columns
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30,
                                bottomMargin=30)
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, alignment=TA_CENTER,
                                     spaceAfter=12)
        subtitle_style = ParagraphStyle('CustomSubtitle', parent=styles['Normal'], fontSize=11, alignment=TA_CENTER,
                                        spaceAfter=6)

        # Header (using your existing logic)
        lab_setting = LabSettingModel.objects.first()
        site_info = SiteInfoModel.objects.first()
        if lab_setting and lab_setting.lab_name:
            elements.append(Paragraph(lab_setting.lab_name, title_style))
        elif site_info:
            elements.append(Paragraph(site_info.name, title_style))

        elements.append(Spacer(1, 0.1 * inch))
        elements.append(Paragraph(report_title, title_style))
        elements.append(
            Paragraph(f"Period: {from_date.strftime('%B %d, %Y')} to {to_date.strftime('%B %d, %Y')}", subtitle_style))
        elements.append(Spacer(1, 0.2 * inch))

        # === Build Table Data ===
        headers = ['S/N', 'Date', 'Patient ID', 'Patient Name', 'Test Name']
        col_widths = [0.4 * inch, 0.9 * inch, 0.9 * inch, 1.8 * inch, 2.0 * inch,]

        if show_status:
            headers.append('Status')
            col_widths.append(1.5 * inch)
        if show_amount:
            headers.append('Amount')
            col_widths.append(0.8 * inch)
        if show_scientist:
            headers.append('Verified By')
            col_widths.append(1.5 * inch)

        table_data = [headers]

        # Data rows
        for idx, order in enumerate(orders_qs, 1):
            row = [
                str(idx),
                order.ordered_at.strftime('%y-%m-%d %H:%M'),
                order.patient.card_number,
                order.patient.__str__(),
                order.template.name
            ]
            if show_status:
                row.append(order.get_status_display())
            if show_amount:
                row.append(f"{order.amount_charged:,.2f}")


            if show_scientist:
                verifier = 'N/A'
                # Safe check here
                if (hasattr(order, 'result') and
                        order.result and
                        order.result.verified_by):

                    user = order.result.verified_by
                    if (hasattr(user, 'user_staff_profile') and
                            user.user_staff_profile and
                            user.user_staff_profile.staff):
                        verifier = user.user_staff_profile.staff.__str__()
                    else:
                        verifier = user.get_full_name()  # Fallback

                row.append(verifier)
                # === END MODIFIED SECTION ===

            table_data.append(row)

        # Total row
        total_row = ['', '', '', '', 'TOTAL']
        col_span = 5

        if show_status:
            total_row.append('')
            col_span += 1
        if show_amount:
            total_row.append(f"{total_amount:,.2f}")
            col_span += 1
        if show_scientist:
            total_row.append(f"{total_patients} Unique Patients")
            col_span += 1
        table_data.append(total_row)

        # Table Style
        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # S/N center
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f0f0f0')]),

            # Total Row Style
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 9),
            ('GRID', (0, -1), (-1, -1), 1, colors.black),
            ('ALIGN', (4, -1), (-1, -1), 'RIGHT'),  # Align totals to the right
        ]))

        elements.append(table)
        doc.build(elements)
        buffer.seek(0)

        # === Prepare Response ===
        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f"lab_test_log_{from_date.strftime('%Y%m%d')}_{to_date.strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class LabTestResultDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = LabTestResultModel
    permission_required = 'laboratory.view_labtestresultmodel'
    template_name = 'laboratory/result/detail.html'
    context_object_name = 'result'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        result = self.object

        context.update({
            'can_verify': not result.is_verified and self.request.user.has_perm('laboratory.change_labtestresultmodel'),
            'can_edit': not result.is_verified and self.request.user.has_perm('laboratory.change_labtestresultmodel'),
            'order': result.order,
            'patient': result.order.patient,
            'lab_setting': LabSettingModel.objects.first(),
            'template': result.order.template,
            'parameters': result.order.template.test_parameters.get('parameters', []),
            'results': result.results_data.get('results', [])
        })

        return context


class LabTestResultListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """List all lab results with filtering options"""
    model = LabTestResultModel
    permission_required = 'laboratory.view_labtestresultmodel'
    template_name = 'laboratory/result/index.html'
    context_object_name = 'results'
    paginate_by = 20

    def get_queryset(self):
        # Get all filter values from the request
        verified_filter = self.request.GET.get('verified', '')
        search_query = self.request.GET.get('search', '')
        # NEW: Get start and end date strings
        start_date_str = self.request.GET.get('start_date')
        end_date_str = self.request.GET.get('end_date')

        queryset = LabTestResultModel.objects.select_related(
            'order__patient', 'order__template', 'verified_by'
        ).order_by('-created_at')

        # Apply verified filter
        if verified_filter == 'verified':
            queryset = queryset.filter(is_verified=True)
        elif verified_filter == 'unverified':
            queryset = queryset.filter(is_verified=False)

        # NEW: Apply date range filter
        if start_date_str and end_date_str:
            try:
                # Convert string dates to date objects for filtering
                start_date = date.fromisoformat(start_date_str)
                end_date = date.fromisoformat(end_date_str)
                # Use '__date__range' to filter the DateTimeField by date
                queryset = queryset.filter(created_at__date__range=[start_date, end_date])
            except (ValueError, TypeError):
                # Ignore invalid date formats gracefully
                pass

        # Apply search query filter
        if search_query:
            # Annotate the queryset to create a temporary 'full_name' field in the database
            queryset = queryset.annotate(
                search_full_name=Concat(
                    'order__patient__first_name', Value(' '), 'order__patient__last_name'
                )
            ).filter(
                # Now, filter using the new annotated field and your other existing fields
                Q(search_full_name__icontains=search_query) |
                Q(order__order_number__icontains=search_query) |
                Q(order__template__name__icontains=search_query) |
                Q(order__patient__card_number__icontains=search_query)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # MODIFIED: Get the filtered queryset first to make stats dynamic
        # This re-uses the logic from your get_queryset() method
        filtered_queryset = self.get_queryset()

        # MODIFIED: Calculate stats based on the filtered queryset
        context['total_results'] = filtered_queryset.count()
        context['verified_results'] = filtered_queryset.filter(is_verified=True).count()
        context['pending_verification'] = filtered_queryset.filter(is_verified=False).count()

        # --- The rest of the method remains the same ---

        # Get current filter values to maintain state in the template
        context['current_verified'] = self.request.GET.get('verified', '')
        context['search_query'] = self.request.GET.get('search', '')

        # Set default dates and maintain state for date filters
        today_str = date.today().isoformat()
        context['start_date'] = self.request.GET.get('start_date', today_str)
        context['end_date'] = self.request.GET.get('end_date', today_str)

        # Filter choices
        context['verified_choices'] = [
            ('', 'All Statuses'),
            ('verified', 'Verified Only'),
            ('unverified', 'Unverified Only')
        ]

        return context


@login_required
@permission_required('laboratory.can_verify_lab_result', raise_exception=True)
def verify_result(request, pk):
    result_id = pk
    """AJAX endpoint to verify a lab result"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    try:
        result = get_object_or_404(LabTestResultModel, pk=result_id)

        if result.is_verified:
            return JsonResponse({
                'success': False,
                'error': 'Result is already verified.'
            }, status=400)

        # Optional pathologist comments
        pathologist_comments = request.POST.get('pathologist_comments', '').strip()

        with transaction.atomic():
            result.is_verified = True
            result.verified_by = request.user
            result.verified_at = now()
            if pathologist_comments:
                result.pathologist_comments = pathologist_comments
            result.save(update_fields=['is_verified', 'verified_by', 'verified_at', 'pathologist_comments'])

        return JsonResponse({
            'success': True,
            'message': f'Result verified successfully',
            'verified_by': result.verified_by.get_full_name() or result.verified_by.username,
            'verified_at': result.verified_at.strftime('%B %d, %Y at %I:%M %p')
        })

    except LabTestResultModel.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Lab result not found.'
        }, status=404)

    except Exception as e:
        logger.exception("Error verifying result id=%s", result_id)
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while verifying result. Please contact administrator.'
        }, status=500)


@login_required
@permission_required('laboratory.add_labtestresultmodel', raise_exception=True)
def process_to_result_entry(request, order_id):
    """AJAX endpoint to move order from collected to processing status for result entry"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    try:
        order = get_object_or_404(LabTestOrderModel, pk=order_id)

        if order.status != 'collected':
            return JsonResponse({
                'success': False,
                'error': 'Only collected samples can be processed for result entry.'
            }, status=400)

        with transaction.atomic():
            order.status = 'processing'
            order.processed_at = now()
            order.processed_by = request.user
            order.save(update_fields=['status', 'processed_at', 'processed_by'])

        return JsonResponse({
            'success': True,
            'message': f'Order {order.order_number} moved to processing. Ready for result entry.',
            'redirect_url': reverse('lab_result_create_for_order', kwargs={'order_id': order.id})
        })

    except LabTestOrderModel.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Lab order not found.'
        }, status=404)

    except Exception as e:
        logger.exception("Error processing order id=%s", order_id)
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while processing order. Please contact administrator.'
        }, status=500)


class LabTestResultCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create lab test results"""
    model = LabTestResultModel
    permission_required = 'laboratory.add_labtestresultmodel'
    template_name = 'laboratory/result/create.html'
    fields = ['technician_comments']

    def dispatch(self, request, *args, **kwargs):

        # Get the order from URL
        order_id = request.GET.get('order_id')
        if order_id:
            self.order = get_object_or_404(LabTestOrderModel, pk=order_id)
            # Check if result already exists
            if hasattr(self.order, 'result'):
                messages.warning(request, 'Results already exist for this order. Redirecting to edit.')
                return redirect('lab_result_edit', pk=self.order.result.pk)
            # Check if order is ready for results
            if self.order.status not in ['processing', 'collected']:
                messages.error(request, 'This order is not ready for result entry.')
                return redirect('lab_result_dashboard')
        else:
            self.order = None
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.order:
            context.update({
                'order': self.order,
                'patient': self.order.patient,
                'template': self.order.template,
                'parameters': self.order.template.test_parameters.get('parameters', [])
            })
        return context

    def post(self, request, *args, **kwargs):
        if not self.order:
            messages.error(request, 'Invalid order selected.')
            return redirect('lab_result_dashboard')

        # Process the results data from form
        parameters = self.order.template.test_parameters.get('parameters', [])
        results = []

        # Get patient gender for gender-specific ranges
        patient_gender = self.order.patient.gender.lower() if hasattr(self.order.patient,
                                                                      'gender') and self.order.patient.gender else None

        for param in parameters:
            param_code = param.get('code', '')
            value = request.POST.get(f'param_{param_code}', '').strip()

            if value:  # Only add if value is provided
                result_entry = {
                    'parameter_code': param_code,
                    'parameter_name': param.get('name', ''),
                    'value': value,
                    'unit': param.get('unit', ''),
                    'type': param.get('type', 'text')
                }

                # Add normal range if available
                if 'normal_range' in param:
                    normal_range = param['normal_range']
                    result_entry['normal_range'] = normal_range

                    # Determine status (normal/abnormal) for numeric values
                    if param.get('type') == 'numeric':
                        min_val = None
                        max_val = None

                        # Check if it's gender-specific
                        if normal_range.get('gender_specific') and patient_gender:
                            gender_range = normal_range.get(patient_gender, {})
                            min_val = gender_range.get('min')
                            max_val = gender_range.get('max')
                        else:
                            # Use standard range
                            min_val = normal_range.get('min')
                            max_val = normal_range.get('max')

                        # Only evaluate status if we have valid range values
                        if min_val is not None and max_val is not None:
                            try:
                                numeric_value = float(value)
                                min_val = float(min_val)
                                max_val = float(max_val)

                                if numeric_value < min_val:
                                    result_entry['status'] = 'low'
                                elif numeric_value > max_val:
                                    result_entry['status'] = 'high'
                                else:
                                    result_entry['status'] = 'normal'
                            except (ValueError, TypeError):
                                result_entry['status'] = 'normal'
                        else:
                            result_entry['status'] = 'normal'
                    else:
                        result_entry['status'] = 'normal'

                results.append(result_entry)

        # Create the result object
        technician_comments = request.POST.get('technician_comments', '').strip()

        try:
            with transaction.atomic():
                result = LabTestResultModel.objects.create(
                    order=self.order,
                    results_data={'results': results},
                    technician_comments=technician_comments,

                )

                # Update order status
                self.order.status = 'completed'
                self.order.processed_at = now()
                self.order.processed_by = self.request.user
                self.order.save(update_fields=['status', 'processed_at', 'processed_by'])

            messages.success(request, f'Results entered successfully for {self.order.template.name}')
            return redirect('lab_result_detail', pk=result.pk)

        except Exception as e:
            logger.exception("Error creating lab result for order %s", self.order.id)
            messages.error(request, 'An error occurred while saving results. Please try again.')
            return self.get(request, *args, **kwargs)


class LabTestResultUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update lab test results (only if not verified)"""
    model = LabTestResultModel
    permission_required = 'laboratory.add_labtestresultmodel'
    template_name = 'laboratory/result/edit.html'
    fields = ['technician_comments']

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object.is_verified:
            messages.error(request, 'Cannot edit verified results.')
            return redirect('lab_result_detail', pk=self.object.pk)
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'result': self.object,
            'order': self.object.order,
            'patient': self.object.order.patient,
            'template': self.object.order.template,
            'parameters': self.object.order.template.test_parameters.get('parameters', []),
            'existing_results': {r['parameter_code']: r for r in self.object.results_data.get('results', [])}
        })
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()

        # Process updated results
        parameters = self.object.order.template.test_parameters.get('parameters', [])
        results = []

        # Get patient gender for gender-specific ranges
        patient_gender = self.object.order.patient.gender.lower() if hasattr(self.object.order.patient,
                                                                             'gender') and self.object.order.patient.gender else None

        for param in parameters:
            param_code = param.get('code', '')
            value = request.POST.get(f'param_{param_code}', '').strip()

            if value:
                result_entry = {
                    'parameter_code': param_code,
                    'parameter_name': param.get('name', ''),
                    'value': value,
                    'unit': param.get('unit', ''),
                    'type': param.get('type', 'text')
                }

                # Add normal range and status logic
                if 'normal_range' in param:
                    normal_range = param['normal_range']
                    result_entry['normal_range'] = normal_range

                    # Determine status (normal/abnormal) for numeric values
                    if param.get('type') == 'numeric':
                        min_val = None
                        max_val = None

                        # Check if it's gender-specific
                        if normal_range.get('gender_specific') and patient_gender:
                            gender_range = normal_range.get(patient_gender, {})
                            min_val = gender_range.get('min')
                            max_val = gender_range.get('max')
                        else:
                            # Use standard range
                            min_val = normal_range.get('min')
                            max_val = normal_range.get('max')

                        # Only evaluate status if we have valid range values
                        if min_val is not None and max_val is not None:
                            try:
                                numeric_value = float(value)
                                min_val = float(min_val)
                                max_val = float(max_val)

                                if numeric_value < min_val:
                                    result_entry['status'] = 'low'
                                elif numeric_value > max_val:
                                    result_entry['status'] = 'high'
                                else:
                                    result_entry['status'] = 'normal'
                            except (ValueError, TypeError):
                                result_entry['status'] = 'normal'
                        else:
                            result_entry['status'] = 'normal'
                    else:
                        result_entry['status'] = 'normal'

                results.append(result_entry)

        # Update the result
        technician_comments = request.POST.get('technician_comments', '').strip()

        try:
            self.object.results_data = {'results': results}
            self.object.technician_comments = technician_comments
            self.object.save(update_fields=['results_data', 'technician_comments'])

            self.object.order.status = 'completed'
            self.object.order.processed_at = now()
            self.object.order.processed_by = self.request.user
            self.object.order.save(update_fields=['status', 'processed_at', 'processed_by'])

            messages.success(request, 'Lab results updated successfully')
            return redirect('lab_result_detail', pk=self.object.pk)

        except Exception as e:
            logger.exception("Error updating lab result %s", self.object.id)
            messages.error(request, 'An error occurred while updating results. Please try again.')
            return self.get(request, *args, **kwargs)


# -------------------------
# Lab Equipment Views
# -------------------------
class LabEquipmentCreateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin,
    CreateView
):
    model = LabEquipmentModel
    permission_required = 'laboratory.add_labequipmentmodel'
    form_class = LabEquipmentForm
    template_name = 'laboratory/equipment/index.html'
    success_message = 'Lab Equipment Successfully Added'

    def get_success_url(self):
        return reverse('lab_equipment_index')

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET':
            return redirect(reverse('lab_equipment_index'))
        return super().dispatch(request, *args, **kwargs)


class LabEquipmentListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = LabEquipmentModel
    permission_required = 'laboratory.view_labequipmentmodel'
    template_name = 'laboratory/equipment/index.html'
    context_object_name = "equipment_list"

    def get_queryset(self):
        return LabEquipmentModel.objects.all().order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = LabEquipmentForm()
        context['template_list'] = LabTestTemplateModel.objects.filter(is_active=True).order_by('name')
        return context


class LabEquipmentUpdateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, UpdateView
):
    model = LabEquipmentModel
    permission_required = 'laboratory.change_labequipmentmodel'
    form_class = LabEquipmentForm
    template_name = 'laboratory/equipment/index.html'
    success_message = 'Lab Equipment Successfully Updated'

    def get_success_url(self):
        return reverse('lab_equipment_index')

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET':
            return redirect(reverse('lab_equipment_index'))
        return super().dispatch(request, *args, **kwargs)


class LabEquipmentDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = LabEquipmentModel
    permission_required = 'laboratory.delete_labequipmentmodel'
    template_name = 'laboratory/equipment/delete.html'
    context_object_name = "equipment"
    success_message = 'Lab Equipment Successfully Deleted'

    def get_success_url(self):
        return reverse('lab_equipment_index')


# -------------------------
# Lab Reagent Views
# -------------------------
class LabReagentCreateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin,
    CreateView
):
    model = LabReagentModel
    permission_required = 'laboratory.add_labreagentmodel'
    form_class = LabReagentForm
    template_name = 'laboratory/reagent/index.html'
    success_message = 'Lab Reagent Successfully Added'

    def get_success_url(self):
        return reverse('lab_reagent_index')

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET':
            return redirect(reverse('lab_reagent_index'))
        return super().dispatch(request, *args, **kwargs)


class LabReagentListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = LabReagentModel
    permission_required = 'laboratory.view_labreagentmodel'
    template_name = 'laboratory/reagent/index.html'
    context_object_name = "reagent_list"

    def get_queryset(self):
        return LabReagentModel.objects.all().order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = LabReagentForm()
        context['template_list'] = LabTestTemplateModel.objects.filter(is_active=True).order_by('name')

        # Add alerts for low stock and expired reagents
        context['low_stock_reagents'] = LabReagentModel.objects.filter(
            current_stock__lte=models.F('minimum_stock'), is_active=True
        ).count()
        context['expired_reagents'] = LabReagentModel.objects.filter(
            expiry_date__lte=date.today(), is_active=True
        ).count()

        return context


class LabReagentUpdateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, UpdateView
):
    model = LabReagentModel
    permission_required = 'laboratory.change_labreagentmodel'
    form_class = LabReagentForm
    template_name = 'laboratory/reagent/index.html'
    success_message = 'Lab Reagent Successfully Updated'

    def get_success_url(self):
        return reverse('lab_reagent_index')

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET':
            return redirect(reverse('lab_reagent_index'))
        return super().dispatch(request, *args, **kwargs)


class LabReagentDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = LabReagentModel
    permission_required = 'laboratory.delete_labreagentmodel'
    template_name = 'laboratory/reagent/delete.html'
    context_object_name = "reagent"
    success_message = 'Lab Reagent Successfully Deleted'

    def get_success_url(self):
        return reverse('lab_reagent_index')


# -------------------------
# Lab Test Template Builder Views
# -------------------------
class LabTestTemplateBuilderCreateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin,
    CreateView
):
    model = LabTestTemplateBuilderModel
    permission_required = 'laboratory.add_labtesttemplatebuildermodel'
    form_class = LabTestTemplateBuilderForm
    template_name = 'laboratory/template_builder/create.html'
    success_message = 'Template Builder Successfully Created'

    def get_success_url(self):
        return reverse('lab_template_builder_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['category_list'] = LabTestCategoryModel.objects.all().order_by('name')
        return context


class LabTestTemplateBuilderListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = LabTestTemplateBuilderModel
    permission_required = 'laboratory.view_labtesttemplatebuildermodel'
    template_name = 'laboratory/template_builder/index.html'
    context_object_name = "builder_list"

    def get_queryset(self):
        return LabTestTemplateBuilderModel.objects.select_related('category', 'created_template').order_by(
            '-created_at')


class LabTestTemplateBuilderDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = LabTestTemplateBuilderModel
    permission_required = 'laboratory.view_labtesttemplatebuildermodel'
    template_name = 'laboratory/template_builder/detail.html'
    context_object_name = "builder"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        builder = self.object

        context.update({
            'can_build': not builder.is_processed and self.request.user.has_perm('laboratory.add_labtesttemplatemodel'),
            'preset_parameters': builder._get_preset_parameters() if builder.parameter_preset != 'custom' else [],
        })

        return context


# -------------------------
# Lab Settings Views
# -------------------------
class LabSettingCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = LabSettingModel
    form_class = LabSettingForm
    permission_required = 'laboratory.add_labsettingmodel'
    success_message = 'Lab Setting Created Successfully'
    template_name = 'laboratory/setting/create.html'

    def dispatch(self, request, *args, **kwargs):
        setting = LabSettingModel.objects.first()
        if setting:
            return redirect('lab_setting_edit', pk=setting.pk)
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self):
        return reverse('lab_setting_detail', kwargs={'pk': self.object.pk})


class LabSettingDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = LabSettingModel
    permission_required = 'laboratory.view_labsettingmodel'
    template_name = 'laboratory/setting/detail.html'
    context_object_name = "lab_setting"

    def get_object(self):
        return LabSettingModel.objects.first()


class LabSettingUpdateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, UpdateView
):
    model = LabSettingModel
    form_class = LabSettingForm
    permission_required = 'laboratory.add_labsettingmodel'
    success_message = 'Lab Setting Updated Successfully'
    template_name = 'laboratory/setting/create.html'

    def get_object(self):
        return LabSettingModel.objects.first()

    def get_success_url(self):
        return reverse('lab_setting_detail', kwargs={'pk': self.object.pk})


# -------------------------
# Action Views (Status Updates)
# -------------------------


@login_required
@permission_required('laboratory.can_verify_lab_result', raise_exception=True)
def unverify_result(request, pk):
    # 1. We must check that this is a POST request
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)

    result = get_object_or_404(LabTestResultModel, pk=pk)

    # 2. Check if already verified and return a JSON error
    if not result.is_verified:
        return JsonResponse({'success': False, 'error': 'This result has not yet been verified.'})

    try:
        # The core logic remains inside a transaction
        with transaction.atomic():
            result.is_verified = False
            result.verified_by = None
            result.verified_at = None

            # 3. Get pathologist comments from the AJAX request
            result.pathologist_comments = ''

            result.save()

        # 4. Return a success JSON response
        message = f"Result unverified for order {result.order.order_number}"
        return JsonResponse({'success': True, 'message': message})

    except Exception as e:
        # 5. Catch any errors and return a server error JSON response
        logger.exception("Error unverifying result id=%s: %s", pk, e)
        return JsonResponse({
            'success': False,
            'error': 'An unexpected error occurred. Please contact support.'
        }, status=500)


@login_required
@permission_required('laboratory.change_labtesttemplatebuildermodel', raise_exception=True)
def build_template(request, pk):
    builder = get_object_or_404(LabTestTemplateBuilderModel, pk=pk)
    try:
        if builder.is_processed:
            messages.info(request, "Template has already been built.")
            return redirect(reverse('lab_template_builder_detail', kwargs={'pk': pk}))

        result = builder.build_template()
        if result.get('error'):
            messages.error(request, result['error'])
        else:
            messages.success(request, f"Template '{builder.name}' successfully built!")
            return redirect(reverse('lab_template_detail', kwargs={'pk': result['template'].pk}))

    except Exception:
        logger.exception("Error building template from builder id=%s", pk)
        messages.error(request, "An error occurred while building template. Contact admin.")
    return redirect(reverse('lab_template_builder_detail', kwargs={'pk': pk}))


# -------------------------
# Bulk Actions
# -------------------------
@login_required
@permission_required('laboratory.delete_labtestcategorymodel', raise_exception=True)
def multi_category_action(request):
    """Handle bulk actions on categories (e.g., delete)."""
    if request.method == 'POST':
        category_ids = request.POST.getlist('category')
        action = request.POST.get('action')

        if not category_ids:
            messages.error(request, 'No category selected.')
            return redirect(reverse('lab_category_index'))

        try:
            with transaction.atomic():
                categories = LabTestCategoryModel.objects.filter(id__in=category_ids)
                if action == 'delete':
                    count, _ = categories.delete()
                    messages.success(request, f'Successfully deleted {count} category(s).')
                else:
                    messages.error(request, 'Invalid action.')
        except Exception:
            logger.exception("Bulk category action failed for ids=%s action=%s", category_ids, action)
            messages.error(request, "An error occurred performing that action. Try again or contact admin.")
        return redirect(reverse('lab_category_index'))

    # GET - confirm action
    category_ids = request.GET.getlist('category')
    if not category_ids:
        messages.error(request, 'No category selected.')
        return redirect(reverse('lab_category_index'))

    action = request.GET.get('action')
    context = {'category_list': LabTestCategoryModel.objects.filter(id__in=category_ids)}

    if action == 'delete':
        return render(request, 'laboratory/category/multi_delete.html', context)

    messages.error(request, 'Invalid action.')
    return redirect(reverse('lab_category_index'))


@login_required
@permission_required('laboratory.change_labtestordermodel', raise_exception=True)
def multi_order_action(request):
    """Handle bulk actions on orders (e.g., update status)."""
    if request.method == 'POST':
        order_ids = request.POST.getlist('order')
        action = request.POST.get('action')

        if not order_ids:
            messages.error(request, 'No order selected.')
            return redirect(reverse('lab_order_index'))

        try:
            with transaction.atomic():
                orders = LabTestOrderModel.objects.filter(id__in=order_ids)

                if action == 'mark_paid':
                    paid_count = 0
                    for order in orders.filter(status='pending'):
                        order.status = 'paid'
                        order.payment_status = True
                        order.payment_date = now()
                        order.payment_by = request.user
                        order.save(update_fields=['status', 'payment_status', 'payment_date', 'payment_by'])
                        paid_count += 1
                    messages.success(request, f'Marked {paid_count} order(s) as paid.')

                elif action == 'cancel':
                    cancelled_count = 0
                    for order in orders.exclude(status__in=['completed', 'cancelled']):
                        order.status = 'cancelled'
                        order.save(update_fields=['status'])
                        cancelled_count += 1
                    messages.success(request, f'Cancelled {cancelled_count} order(s).')
                else:
                    messages.error(request, 'Invalid action.')
        except Exception:
            logger.exception("Bulk order action failed for ids=%s action=%s", order_ids, action)
            messages.error(request, "An error occurred performing that action. Try again or contact admin.")
        return redirect(reverse('lab_order_index'))

    # GET - confirm action
    order_ids = request.GET.getlist('order')
    if not order_ids:
        messages.error(request, 'No order selected.')
        return redirect(reverse('lab_order_index'))

    action = request.GET.get('action')
    context = {'order_list': LabTestOrderModel.objects.filter(id__in=order_ids).select_related('patient', 'template')}

    if action in ['mark_paid', 'cancel']:
        return render(request, 'laboratory/order/multi_action.html', {
            **context,
            'action': action,
            'action_title': 'Mark as Paid' if action == 'mark_paid' else 'Cancel Orders'
        })

    messages.error(request, 'Invalid action.')
    return redirect(reverse('lab_order_index'))


# -------------------------
# AJAX/API Views
# -------------------------
@login_required
def get_template_details(request):
    """Get template details for AJAX requests."""
    template_id = request.GET.get('template_id')
    if not template_id:
        return JsonResponse({'error': 'template_id is required'}, status=400)

    try:
        template = LabTestTemplateModel.objects.get(id=template_id)
        data = {
            'id': template.id,
            'name': template.name,
            'code': template.code,
            'price': str(template.price),
            'sample_type': template.sample_type,
            'sample_volume': template.sample_volume,
            'parameters': template.parameter_names,
            'test_parameters': template.test_parameters
        }
        return JsonResponse(data)
    except LabTestTemplateModel.DoesNotExist:
        return JsonResponse({'error': 'Template not found'}, status=404)
    except Exception:
        logger.exception("Failed fetching template details for id=%s", template_id)
        return JsonResponse({'error': 'Internal error'}, status=500)


@login_required
def get_patient_orders(request):
    """Get orders for a specific patient."""
    patient_id = request.GET.get('patient_id')
    if not patient_id:
        return JsonResponse({'error': 'patient_id is required'}, status=400)

    try:
        orders = LabTestOrderModel.objects.filter(
            patient_id=patient_id
        ).select_related('template').order_by('-ordered_at')[:10]

        data = {
            'orders': [
                {
                    'id': order.id,
                    'order_number': order.order_number,
                    'template_name': order.template.name,
                    'status': order.status,
                    'ordered_at': order.ordered_at.strftime('%Y-%m-%d %H:%M'),
                }
                for order in orders
            ]
        }
        return JsonResponse(data)
    except Exception:
        logger.exception("Failed fetching patient orders for patient_id=%s", patient_id)
        return JsonResponse({'error': 'Internal error'}, status=500)


@login_required
def lab_dashboard_data(request):
    """Get dashboard statistics for AJAX requests."""
    try:
        today = date.today()

        data = {
            'today_orders': LabTestOrderModel.objects.filter(ordered_at__date=today).count(),
            'pending_payments': LabTestOrderModel.objects.filter(status='pending').count(),
            'samples_to_collect': LabTestOrderModel.objects.filter(status='paid').count(),
            'tests_processing': LabTestOrderModel.objects.filter(status='processing').count(),
            'pending_results': LabTestOrderModel.objects.filter(status='collected').count(),
            'pending_verification': LabTestResultModel.objects.filter(is_verified=False).count(),
            'low_stock_reagents': LabReagentModel.objects.filter(
                current_stock__lte=models.F('minimum_stock'), is_active=True
            ).count(),
            'expired_reagents': LabReagentModel.objects.filter(
                expiry_date__lte=today, is_active=True
            ).count(),
            'inactive_equipment': LabEquipmentModel.objects.filter(status='inactive').count(),
        }
        return JsonResponse(data)
    except Exception:
        logger.exception("Failed fetching lab dashboard data")
        return JsonResponse({'error': 'Internal error'}, status=500)


# -------------------------
# Print Views
# -------------------------
@login_required
@permission_required('laboratory.view_labtestordermodel', raise_exception=True)
def print_order(request, pk):
    order = get_object_or_404(LabTestOrderModel, pk=pk)
    context = {
        'order': order,
        'lab_setting': LabSettingModel.objects.first(),
    }
    return render(request, 'laboratory/print/order.html', context)


@login_required
@permission_required('laboratory.view_labtestresultmodel', raise_exception=True)
def print_result(request, pk):
    result = get_object_or_404(LabTestResultModel, pk=pk)
    context = {
        'result': result,
        'order': result.order,
        'patient': result.order.patient,
        'lab_setting': LabSettingModel.objects.first(),
    }
    return render(request, 'laboratory/print/result.html', context)


@login_required
@permission_required('laboratory.add_labtestordermodel', raise_exception=True)
def process_payment(request, pk):
    order = get_object_or_404(LabTestOrderModel, pk=pk)
    try:
        if order.status != 'pending':
            messages.info(request, "Order payment has already been processed or order is not pending.")
            return redirect(reverse('lab_order_detail', kwargs={'pk': pk}))

        with transaction.atomic():
            order.status = 'paid'
            order.payment_status = True
            order.payment_date = now()
            order.payment_by = request.user
            order.save(update_fields=['status', 'payment_status', 'payment_date', 'payment_by'])

        messages.success(request, f"Payment processed for order {order.order_number}")
    except Exception:
        logger.exception("Error processing payment for order id=%s", pk)
        messages.error(request, "An error occurred while processing payment. Contact admin.")
    return redirect(reverse('lab_order_detail', kwargs={'pk': pk}))


@login_required
@permission_required('laboratory.change_labtestordermodel', raise_exception=True)
def collect_sample(request, order_id):
    """
    Collect sample for a lab test order - AJAX endpoint
    Returns JSON response for frontend consumption
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    try:
        order = get_object_or_404(LabTestOrderModel, pk=order_id)

        # Check if order status allows sample collection
        if order.status != 'paid':
            return JsonResponse({
                'success': False,
                'error': 'Sample can only be collected for paid orders.'
            }, status=400)

        # Get form data
        sample_label = request.POST.get('sample_label', '').strip()
        expected_completion = request.POST.get('expected_completion', '').strip()

        # Validate required fields
        if not sample_label:
            return JsonResponse({
                'success': False,
                'error': 'Sample label is required.'
            }, status=400)

        # Process sample collection
        with transaction.atomic():
            order.status = 'collected'
            order.sample_collected_at = now()
            order.sample_collected_by = request.user
            order.sample_label = sample_label

            # Handle expected completion if provided
            if expected_completion:
                try:
                    # Parse the datetime-local input
                    from datetime import datetime
                    expected_dt = datetime.fromisoformat(expected_completion)
                    order.expected_completion = expected_dt
                except ValueError:
                    # Invalid datetime format - ignore but don't fail
                    pass

            # Save the order
            update_fields = [
                'status',
                'sample_collected_at',
                'sample_collected_by',
                'sample_label'
            ]
            if expected_completion:
                update_fields.append('expected_completion')

            order.save(update_fields=update_fields)

        return JsonResponse({
            'success': True,
            'message': f'Sample collected successfully for order {order.order_number}',
            'order_id': order.id,
            'sample_label': order.sample_label
        })

    except LabTestOrderModel.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Lab test order not found.'
        }, status=404)

    except Exception as e:
        logger.exception("Error collecting sample for order id=%s", order_id)
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while collecting sample. Please contact administrator.'
        }, status=500)


@login_required
@permission_required('laboratory.change_labtestordermodel', raise_exception=True)
def start_processing(request, pk):
    order = get_object_or_404(LabTestOrderModel, pk=pk)
    try:
        if order.status != 'collected':
            messages.info(request, "Test can only be processed for collected samples.")
            return redirect(reverse('lab_order_detail', kwargs={'pk': pk}))

        with transaction.atomic():
            order.status = 'processing'
            order.processed_at = now()
            order.processed_by = request.user
            order.save(update_fields=['status', 'processed_at', 'processed_by'])

        messages.success(request, f"Processing started for order {order.order_number}")
    except Exception:
        logger.exception("Error starting processing for order id=%s", pk)
        messages.error(request, "An error occurred while starting processing. Contact admin.")
    return redirect(reverse('lab_order_detail', kwargs={'pk': pk}))


@login_required
@permission_required('laboratory.change_labtestordermodel', raise_exception=True)
def complete_test(request, pk):
    order = get_object_or_404(LabTestOrderModel, pk=pk)
    try:
        if order.status != 'processing':
            messages.info(request, "Test can only be completed for processing orders.")
            return redirect(reverse('lab_order_detail', kwargs={'pk': pk}))

        # Check if results exist
        if not hasattr(order, 'result'):
            messages.error(request, "Cannot complete test without results. Please add results first.")
            return redirect(reverse('lab_order_detail', kwargs={'pk': pk}))

        with transaction.atomic():
            order.status = 'completed'
            order.save(update_fields=['status'])

        messages.success(request, f"Test completed for order {order.order_number}")
    except Exception:
        logger.exception("Error completing test for order id=%s", pk)
        messages.error(request, "An error occurred while completing test. Contact admin.")
    return redirect(reverse('lab_order_detail', kwargs={'pk': pk}))


@login_required
@permission_required('laboratory.add_labtestordermodel', raise_exception=True)
def walkin_lab_page(request):
    """Main walk-in lab page"""
    return render(request, 'laboratory/order/walkin_index.html')


@login_required
def walkin_lab_list_ajax(request):
    """Get all walk-in transactions with lab orders"""
    try:
        from django.db.models import Q, Count, Sum, F
        from datetime import timedelta
        from django.utils import timezone

        # Get date filters from request
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        # Get parent transactions with walk-in lab orders
        parent_transactions = PatientTransactionModel.objects.filter(
            transaction_type='direct_payment',
            source='walkin',
            parent_transaction__isnull=True,
            child_transactions__lab_structure__isnull=False,
        ).filter(
            Q(child_transactions__lab_structure__status='paid') |
            Q(child_transactions__lab_structure__status='collected') |
            Q(child_transactions__lab_structure__status='processing') |
            Q(child_transactions__lab_structure__status='completed')
        )

        # Apply date filtering
        if start_date and end_date:
            parent_transactions = parent_transactions.filter(created_at__date__range=[start_date, end_date])
        else:
            # Default to last 7 days
            parent_transactions = parent_transactions.filter(created_at__gte=timezone.now() - timedelta(days=7))

        parent_transactions = parent_transactions.annotate(
            lab_count=Count('child_transactions__lab_structure', distinct=True)
        ).filter(
            lab_count__gt=0
        ).distinct().select_related('received_by').order_by('-created_at')

        transactions_data = []

        for parent_txn in parent_transactions:
            # Get all lab orders in this transaction
            lab_orders = LabTestOrderModel.objects.filter(
                transactions__parent_transaction=parent_txn,
                source='walkin',
                status__in=['paid', 'collected', 'processing', 'completed']
            ).select_related('template')

            if not lab_orders.exists():
                continue

            # Calculate total and build summary
            total_amount = Decimal('0.00')
            lab_items = []

            for order in lab_orders:
                total_amount += order.total_amount
                lab_items.append({
                    'id': order.id,
                    'name': order.template.name,
                    'status': order.status,
                    'status_display': order.get_status_display(),
                })

            # Build summary string
            if len(lab_items) == 1:
                lab_summary = f"{lab_items[0]['name']}"
            else:
                lab_summary = f"{len(lab_items)} tests: " + ", ".join([
                    item['name'] for item in lab_items[:2]
                ])
                if len(lab_items) > 2:
                    lab_summary += f", +{len(lab_items) - 2} more"

            # Overall status
            statuses = [item['status'] for item in lab_items]
            if 'paid' in statuses:
                overall_status = 'paid'
            elif 'collected' in statuses:
                overall_status = 'collected'
            elif 'processing' in statuses:
                overall_status = 'processing'
            else:
                overall_status = 'completed'

            customer_name = parent_txn.customer_name or lab_orders.first().customer_display

            transactions_data.append({
                'transaction_id': parent_txn.transaction_id,
                'customer_name': customer_name,
                'date': parent_txn.created_at.strftime('%Y-%m-%d %H:%M'),
                'total_amount': float(total_amount),
                'formatted_amount': f'₦{total_amount:,.2f}',
                'lab_count': len(lab_items),
                'lab_summary': lab_summary,
                'lab_items': lab_items,
                'status': overall_status,
                'status_display': dict(LabTestOrderModel.STATUS_CHOICES).get(overall_status, 'Unknown')
            })

        return JsonResponse({
            'success': True,
            'transactions': transactions_data,
            'total_count': len(transactions_data)
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'error': f'Error fetching walk-in lab orders: {str(e)}'
        }, status=500)


@login_required
def walkin_lab_detail_ajax(request):
    """Get details of walk-in transaction lab orders"""
    transaction_id = request.GET.get('transaction_id', '').strip()

    if not transaction_id:
        return JsonResponse({'error': 'Transaction ID required'}, status=400)

    try:
        # Get parent transaction
        parent_txn = PatientTransactionModel.objects.get(
            transaction_id=transaction_id,
            transaction_type='direct_payment',
            source='walkin',
            parent_transaction__isnull=True
        )

        # Get all lab orders
        lab_orders = LabTestOrderModel.objects.filter(
            transactions__parent_transaction=parent_txn,
            source='walkin'
        ).select_related('template', 'ordered_by', 'sample_collected_by').order_by('id')

        if not lab_orders.exists():
            return JsonResponse({
                'error': 'No lab orders found for this transaction'
            }, status=404)

        customer_name = parent_txn.customer_name or lab_orders.first().customer_display

        # Build orders data
        orders_data = []
        total_amount = Decimal('0.00')

        for order in lab_orders:
            # Check if result exists
            has_result = hasattr(order, 'result')
            is_verified = has_result and order.result.is_verified if has_result else False

            orders_data.append({
                'id': order.id,
                'order_number': order.order_number,
                'test_name': order.template.name,
                'sample_type': order.template.sample_type,
                'amount': float(order.amount_charged or order.template.price),
                'status': order.status,
                'status_display': order.get_status_display(),
                'ordered_date': order.ordered_at.strftime('%Y-%m-%d %H:%M'),
                'ordered_by': str(order.ordered_by) if order.ordered_by else 'N/A',
                'sample_label': order.sample_label or '',
                'sample_collected_at': order.sample_collected_at.strftime('%Y-%m-%d %H:%M') if order.sample_collected_at else '',
                'sample_collected_by': str(order.sample_collected_by) if order.sample_collected_by else '',
                'special_instructions': order.special_instructions or '',
                'has_result': has_result,
                'is_verified': is_verified,
                'result_id': order.result.id if has_result else None,
            })

            total_amount += order.amount_charged or order.template.price

        transaction_data = {
            'transaction_id': parent_txn.transaction_id,
            'customer_name': customer_name,
            'date': parent_txn.created_at.strftime('%Y-%m-%d %H:%M'),
            'payment_method': parent_txn.payment_method,
            'total_amount': float(total_amount),
            'formatted_amount': f'₦{total_amount:,.2f}',
            'lab_orders': orders_data
        }

        return JsonResponse({
            'success': True,
            'transaction': transaction_data
        })

    except PatientTransactionModel.DoesNotExist:
        return JsonResponse({
            'error': 'Transaction not found'
        }, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'error': f'Error fetching transaction details: {str(e)}'
        }, status=500)
