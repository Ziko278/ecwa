import logging
import json
from decimal import Decimal
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from io import BytesIO
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.models import User
from django.contrib.messages.views import SuccessMessageMixin
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.db.models import Q, Count, Sum, Case, When, IntegerField, Value
from django.db.models.functions import Concat
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils.dateparse import parse_datetime
from django.utils.timezone import now
from django.views import View
from django.views.generic import (
    CreateView, ListView, UpdateView, DeleteView, DetailView, TemplateView
)

from admin_site.models import SiteInfoModel
from insurance.models import InsuranceClaimModel
from patient.models import PatientModel, PatientWalletModel
from .models import *
from .forms import *
from django.utils.text import slugify

logger = logging.getLogger(__name__)


# -------------------------
# Mixins
# -------------------------
class FlashFormErrorsMixin:
    """
    Mixin for CreateView/UpdateView to flash form errors and redirect safely.
    Use before SuccessMessageMixin in MRO so messages appear before redirect.
    """

    def form_invalid(self, form):
        try:
            for field, errors in form.errors.items():
                field_name = form.fields.get(field).label if form.fields.get(field) else field
                for error in errors:
                    messages.error(self.request, f"{field_name}: {error}")
        except Exception:
            logger.exception("Error while processing form_invalid errors.")
            messages.error(self.request, "There was an error processing the form. Please try again.")
        return redirect(self.get_success_url())


# -------------------------
# Scan Category Views
# -------------------------
class ScanCategoryCreateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, SuccessMessageMixin, CreateView
):
    model = ScanCategoryModel
    permission_required = 'scan.add_scancategorymodel'
    form_class = ScanCategoryForm
    template_name = 'scan/category/index.html'
    success_message = 'Scan Category Successfully Created'

    def get_success_url(self):
        return reverse('scan_category_index')

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET':
            return redirect(reverse('scan_category_index'))
        return super().dispatch(request, *args, **kwargs)


class ScanCategoryListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = ScanCategoryModel
    permission_required = 'scan.view_scancategorymodel'
    template_name = 'scan/category/index.html'
    context_object_name = "category_list"

    def get_queryset(self):
        return ScanCategoryModel.objects.all().order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = ScanCategoryForm()
        return context


class ScanCategoryUpdateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, SuccessMessageMixin, UpdateView
):
    model = ScanCategoryModel
    permission_required = 'scan.add_scancategorymodel'
    form_class = ScanCategoryForm
    template_name = 'scan/category/index.html'
    success_message = 'Scan Category Successfully Updated'

    def get_success_url(self):
        return reverse('scan_category_index')

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET':
            return redirect(reverse('scan_category_index'))
        return super().dispatch(request, *args, **kwargs)


class ScanCategoryDeleteView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, DeleteView):
    model = ScanCategoryModel
    permission_required = 'scan.add_scancategorymodel'
    template_name = 'scan/category/delete.html'
    context_object_name = "category"
    success_message = 'Scan Category Successfully Deleted'

    def get_success_url(self):
        return reverse('scan_category_index')


# -------------------------
# Scan Template Views
# -------------------------
class ScanTemplateCreateView(
    LoginRequiredMixin, PermissionRequiredMixin,
    CreateView
):
    model = ScanTemplateModel
    permission_required = 'scan.add_scancategorymodel'
    form_class = ScanTemplateForm
    template_name = 'scan/template/create.html'
    success_message = 'Scan Template Successfully Created'

    def get_success_url(self):
        return reverse('scan_template_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['category_list'] = ScanCategoryModel.objects.all().order_by('name')
        return context


class ScanTemplateListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = ScanTemplateModel
    permission_required = 'scan.view_scancategorymodel'
    template_name = 'scan/template/index.html'
    context_object_name = "template_list"

    def get_queryset(self):
        return ScanTemplateModel.objects.select_related('category').order_by('category__name', 'name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['category_list'] = ScanCategoryModel.objects.all().order_by('name')
        return context


class ScanTemplateDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = ScanTemplateModel
    permission_required = 'scan.view_scancategorymodel'
    template_name = 'scan/template/detail.html'
    context_object_name = "template"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        template = self.object

        # Get orders for this template
        orders = ScanOrderModel.objects.filter(template=template).select_related('patient').order_by('-ordered_at')[:10]

        context.update({
            'recent_orders': orders,
            'total_orders': ScanOrderModel.objects.filter(template=template).count(),
            'expected_images': template.expected_images,
            'scan_parameters': template.scan_parameters,
        })
        return context


class ScanTemplateUpdateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, UpdateView
):
    model = ScanTemplateModel
    permission_required = 'scan.add_scancategorymodel'
    form_class = ScanTemplateForm
    template_name = 'scan/template/edit.html'
    success_message = 'Scan Template Successfully Updated'

    def get_success_url(self):
        return reverse('scan_template_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['category_list'] = ScanCategoryModel.objects.all().order_by('name')
        context['template'] = self.object
        return context


class ScanTemplateDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = ScanTemplateModel
    permission_required = 'scan.add_scancategorymodel'
    template_name = 'scan/template/delete.html'
    context_object_name = "template"
    success_message = 'Scan Template Successfully Deleted'

    def get_success_url(self):
        return reverse('scan_template_index')


# -------------------------
# Scan Entry Point - Patient Verification
# -------------------------
class ScanEntryView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Entry point for scan operations - patient verification"""
    permission_required = 'scan.add_scanordermodel'
    template_name = 'scan/order/entry.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Radiology - Patient Verification'
        return context


def verify_scan_patient_ajax(request):
    """AJAX view to verify patient by card number for scans"""
    if not request.user.has_perm('scan.view_scanordermodel'):
        return JsonResponse({'success': False, 'error': 'Permission denied'})

    card_number = request.GET.get('card_number', '').strip()
    if not card_number:
        return JsonResponse({'success': False, 'error': 'Please enter a card number'})

    try:
        patient = PatientModel.objects.get(card_number__iexact=card_number, status='active')

        # Get scan counts
        scan_counts = {
            'total': ScanOrderModel.objects.filter(patient=patient).count(),
            'pending': ScanOrderModel.objects.filter(patient=patient, status='pending').count(),
            'paid': ScanOrderModel.objects.filter(patient=patient, status='paid').count(),
            'in_progress': ScanOrderModel.objects.filter(patient=patient, status='in_progress').count(),
            'completed': ScanOrderModel.objects.filter(patient=patient, status='completed').count(),
            'external_pending_results':  ExternalScanOrder.objects.filter(patient=patient).filter(
                            Q(result_file__isnull=True) | Q(result_file='')).count()
        }

        return JsonResponse({
            'success': True,
            'patient': {
                'id': patient.id,
                'full_name': patient.__str__(),
                'patient_id': patient.card_number,
                'phone': patient.mobile,
                'email': getattr(patient, 'email', ''),
                'age': patient.age() if hasattr(patient, 'age') and callable(patient.age) else '',
                'gender': getattr(patient, 'gender', ''),
                'address': patient.address or 'Not provided',
            },
            'scan_counts': scan_counts
        })

    except PatientModel.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': f'No patient found with card number: {card_number}'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'An error occurred while verifying patient: {str(e)}'
        })


class PatientExternalScanListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """View a patient's external scan orders with filtering and result management."""
    model = ExternalScanOrder
    # Assumes a specific permission for external scans exists
    permission_required = 'radiology.view_externalscanorder'
    template_name = 'scan/order/patient_external_scans.html'
    context_object_name = 'external_scans'
    paginate_by = 20

    def get_queryset(self):
        patient_id = self.kwargs.get('patient_id')
        self.patient = get_object_or_404(PatientModel, id=patient_id, status='active')

        queryset = ExternalScanOrder.objects.filter(
            patient=self.patient
        ).select_related(
            'ordered_by', 'result_uploaded_by'
        ).order_by('-ordered_at')

        # Filter by result status if provided via GET parameter
        result_status = self.request.GET.get('result_status', '')
        if result_status == 'pending':
            queryset = queryset.filter(Q(result_file__isnull=True) | Q(result_file=''))
        elif result_status == 'uploaded':
            queryset = queryset.filter(result_file__isnull=False).exclude(result_file='')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['patient'] = self.patient

        # Group orders by date for a cleaner display
        orders_by_date = {}
        for order in context['external_scans']:
            order_date = order.ordered_at.date()
            if order_date not in orders_by_date:
                orders_by_date[order_date] = []
            orders_by_date[order_date].append(order)
        context['orders_by_date'] = orders_by_date

        # Get counts for the summary cards
        all_orders = ExternalScanOrder.objects.filter(patient=self.patient)
        context['result_counts'] = {
            'total': all_orders.count(),
            'pending': all_orders.filter(Q(result_file__isnull=True) | Q(result_file='')).count(),
            'uploaded': all_orders.filter(result_file__isnull=False).exclude(result_file='').count(),
        }

        # Pass the current filter to the template
        context['current_filter'] = self.request.GET.get('result_status', '')

        return context


class UploadExternalScanResultView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Handles the POST request for uploading or changing a scan result file."""
    permission_required = 'radiology.change_externalscanorder'

    def post(self, request, order_id):
        order = get_object_or_404(ExternalScanOrder, id=order_id)
        result_file = request.FILES.get('result_file')

        if not result_file:
            messages.error(request, 'No file was selected for upload.')
            return redirect(reverse('patient_external_scan_list', kwargs={'patient_id': order.patient.id}))

        # Update the order instance with the file and tracking info
        order.result_file = result_file
        order.result_uploaded_by = request.user
        order.result_uploaded_at = timezone.now()
        order.save()

        messages.success(request, f"Result for scan order {order.order_number} was uploaded successfully.")
        return redirect(reverse('patient_external_scan_list', kwargs={'patient_id': order.patient.id}))

class ScanOrderListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = ScanOrderModel
    permission_required = 'scan.view_scanordermodel'
    template_name = 'scan/order/index.html'
    context_object_name = "order_list"
    paginate_by = 20

    def get_queryset(self):
        # Exclude pending orders to create a list of active scans
        queryset = ScanOrderModel.objects.exclude(status='pending').select_related(
            'patient', 'template', 'ordered_by'
        ).order_by('-ordered_at')

        # Get filter values from the request
        status = self.request.GET.get('status')
        template_id = self.request.GET.get('template')
        search_query = self.request.GET.get('search', '').strip()

        # Filter by status if provided
        if status:
            queryset = queryset.filter(status=status)

        # Filter by template if provided
        if template_id:
            queryset = queryset.filter(template__id=template_id)

        # Apply search query filter across multiple fields
        if search_query:
            # Annotate queryset to create a searchable full_name field
            queryset = queryset.annotate(
                patient_full_name=Concat(
                    'patient__first_name', Value(' '), 'patient__last_name'
                )
            ).filter(
                Q(patient_full_name__icontains=search_query) |
                Q(patient__card_number__icontains=search_query) |
                Q(order_number__icontains=search_query)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Pass the search query back to the template to keep it in the search box
        context['search_query'] = self.request.GET.get('search', '').strip()

        # Use a slice to exclude 'pending' and 'cancelled' from the filter dropdown
        context['status_choices'] = [choice for choice in ScanOrderModel.STATUS_CHOICES if
                                     choice[0] not in ['pending', 'cancelled']]
        context['template_list'] = ScanTemplateModel.objects.filter(is_active=True).order_by('name')

        # Add counts for the dashboard summary cards
        context['status_counts'] = {
            'pending': ScanOrderModel.objects.filter(status='pending').count(),
            'paid': ScanOrderModel.objects.filter(status='paid').count(),
            'scheduled': ScanOrderModel.objects.filter(status='scheduled').count(),
            'in_progress': ScanOrderModel.objects.filter(status='in_progress').count(),
            'completed': ScanOrderModel.objects.filter(status='completed').count(),
        }

        return context


class ScanOrderUpdateView(
    LoginRequiredMixin, PermissionRequiredMixin, UpdateView
):
    model = ScanOrderModel
    permission_required = 'scan.add_scanordermodel'
    form_class = ScanOrderForm
    template_name = 'scan/order/edit.html'

    def get_success_url(self):
        return reverse('scan_order_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['order'] = self.object
        context['patient_list'] = PatientModel.objects.filter(status='active').order_by('first_name')
        context['template_list'] = ScanTemplateModel.objects.filter(is_active=True).order_by('name')
        return context


# -------------------------
# Patient Scan Management
# -------------------------
class PatientScansView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """View patient's scans with grouping by date"""
    model = ScanOrderModel
    permission_required = 'scan.view_scanordermodel'
    template_name = 'scan/order/patient_scans.html'
    context_object_name = 'orders'
    paginate_by = 30

    def get_queryset(self):
        patient_id = self.kwargs.get('patient_id')
        self.patient = get_object_or_404(PatientModel, id=patient_id, status='active')

        queryset = ScanOrderModel.objects.filter(
            patient=self.patient
        ).select_related(
            'template', 'template__category', 'ordered_by', 'payment_by'
        ).order_by('-ordered_at')

        # Filter by status if provided
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['patient'] = self.patient
        context['status_choices'] = ScanOrderModel.STATUS_CHOICES

        # Group orders by date
        orders_by_date = {}
        for order in context['orders']:
            order_date = order.ordered_at.date()
            if order_date not in orders_by_date:
                orders_by_date[order_date] = []
            orders_by_date[order_date].append(order)

        context['orders_by_date'] = orders_by_date

        # Get scan counts for summary
        context['scan_counts'] = {
            'total': ScanOrderModel.objects.filter(patient=self.patient).count(),
            'pending': ScanOrderModel.objects.filter(patient=self.patient, status='pending').count(),
            'paid': ScanOrderModel.objects.filter(patient=self.patient, status='paid').count(),
            'scheduled': ScanOrderModel.objects.filter(patient=self.patient, status='scheduled').count(),
            'in_progress': ScanOrderModel.objects.filter(patient=self.patient, status='in_progress').count(),
            'completed': ScanOrderModel.objects.filter(patient=self.patient, status='completed').count(),
        }

        # Get available scan templates for new order
        context['template_list'] = ScanTemplateModel.objects.filter(
            is_active=True
        ).select_related('category').order_by('category__name', 'name')

        return context


# -------------------------
# Scan Order CRUD Views
# -------------------------
class ScanOrderCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = ScanOrderModel
    permission_required = 'scan.add_scanordermodel'
    form_class = ScanOrderForm
    template_name = 'scan/order/create.html'

    def get_initial(self):
        initial = super().get_initial()
        patient_id = self.kwargs.get('patient_id')
        if patient_id:
            initial['patient'] = patient_id
        initial['ordered_by'] = self.request.user
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        patient_id = self.kwargs.get('patient_id')
        if patient_id:
            context['patient'] = get_object_or_404(PatientModel, id=patient_id)
        context['template_list'] = ScanTemplateModel.objects.filter(is_active=True).select_related('category').order_by(
            'category__name', 'name')
        return context

    def post(self, request, *args, **kwargs):
        form = self.get_form()
        scan_ids = request.POST.getlist('scans')

        if not scan_ids:
            messages.error(request, "Please select at least one scan.")
            return redirect(request.path)

        if not form.is_valid():
            for field, errs in form.errors.items():
                for e in errs:
                    messages.error(request, f"{field}: {e}")
            return redirect(request.path)

        created_orders = []
        try:
            with transaction.atomic():
                patient = get_object_or_404(PatientModel, id=self.kwargs.get('patient_id'))
                for sid in scan_ids:
                    try:
                        template = ScanTemplateModel.objects.get(pk=int(sid))
                    except (ValueError, ScanTemplateModel.DoesNotExist):
                        messages.warning(request, f"Scan id {sid} not found - skipped.")
                        continue

                    order = ScanOrderModel(
                        patient=patient,
                        template=template,
                        ordered_by=request.user,
                        status='pending',
                        amount_charged=template.price,
                        clinical_indication=form.cleaned_data.get('clinical_indication', ''),
                        special_instructions=form.cleaned_data.get('special_instructions', ''),
                    )
                    order.save()
                    created_orders.append(order)
        except Exception:
            logger.exception("Error creating scan orders")
            messages.error(request, "An error occurred creating orders. Contact admin.")
            return redirect(request.path)

        if created_orders:
            messages.success(request, f"Created {len(created_orders)} scan order(s).")
            return redirect(reverse('patient_scan_orders', kwargs={'patient_id': created_orders[0].patient.id}))
        else:
            messages.error(request, "No valid scans selected.")
            return redirect(request.path)


class ScanOrderDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = ScanOrderModel
    permission_required = 'scan.view_scanordermodel'
    template_name = 'scan/order/detail.html'
    context_object_name = "order"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        order = self.object

        # Check if results exist
        has_results = hasattr(order, 'result')

        context.update({
            'scan_setting': ScanSettingModel.objects.first(),
            'has_results': has_results,
            'can_process_payment': order.status == 'pending' and self.request.user.has_perm(
                'scan.change_scanordermodel'),
            'can_schedule': order.status == 'paid' and self.request.user.has_perm('scan.change_scanordermodel'),
            'can_start_scan': order.status == 'scheduled' and self.request.user.has_perm(
                'scan.change_scanordermodel'),
            'can_complete_scan': order.status == 'in_progress' and self.request.user.has_perm(
                'scan.change_scanordermodel'),
            'expected_images': order.template.expected_images,
        })

        # If has results, get images
        if has_results:
            context['scan_images'] = order.result.images.all().order_by('sequence_number')

        return context


# -------------------------
# Scan Payment Processing
# -------------------------
def process_scan_payments(request):
    """Process payments for selected scans with wallet balance and insurance validation"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})

    if not request.user.has_perm('scan.change_scanordermodel'):
        return JsonResponse({'success': False, 'error': 'Permission denied'})

    selected_orders = request.POST.getlist('selected_orders')
    if not selected_orders:
        return JsonResponse({'success': False, 'error': 'No scans selected'})

    try:
        with transaction.atomic():
            # Get all selected orders
            orders = ScanOrderModel.objects.filter(
                id__in=selected_orders,
                status='pending'
            ).select_related('template', 'patient')

            if not orders.exists():
                return JsonResponse({'success': False, 'error': 'No valid pending scans found'})

            # Get patient (assuming all orders are for the same patient)
            patient = orders.first().patient

            # Get or create patient wallet
            wallet, created = PatientWalletModel.objects.get_or_create(
                patient=patient,
                defaults={'amount': Decimal('0.00')}
            )

            # Check for active insurance
            active_insurance = None
            if hasattr(patient, 'insurance_policies'):
                active_insurance = patient.insurance_policies.filter(
                    is_active=True,
                    valid_to__gte=date.today()
                ).select_related('hmo', 'coverage_plan').first()

            # Calculate insurance amount helper function
            def calculate_patient_amount(base_amount, coverage_percentage):
                """Calculate patient's portion after insurance"""
                if coverage_percentage and coverage_percentage > 0:
                    covered_amount = base_amount * (coverage_percentage / 100)
                    return base_amount - covered_amount
                return base_amount

            # Calculate total amounts considering insurance
            total_base_amount = Decimal('0.00')
            total_patient_amount = Decimal('0.00')
            total_insurance_covered = Decimal('0.00')
            order_details = []

            for order in orders:
                # Get base amount
                base_amount = order.amount_charged or order.template.price

                # Apply insurance if applicable (assuming scan coverage)
                if active_insurance and hasattr(active_insurance.coverage_plan,
                                                'is_scan_covered') and active_insurance.coverage_plan.is_scan_covered(
                    order.template):
                    patient_amount = calculate_patient_amount(
                        base_amount,
                        getattr(active_insurance.coverage_plan, 'scan_coverage_percentage', 0)
                    )
                    insurance_covered = base_amount - patient_amount
                else:
                    patient_amount = base_amount
                    insurance_covered = Decimal('0.00')

                total_base_amount += base_amount
                total_patient_amount += patient_amount
                total_insurance_covered += insurance_covered

                order_details.append({
                    'order': order,
                    'base_amount': base_amount,
                    'patient_amount': patient_amount,
                    'insurance_covered': insurance_covered
                })

            # Check wallet balance
            if wallet.amount < total_patient_amount:
                return JsonResponse({
                    'success': False,
                    'error': f'Insufficient wallet balance. Required: ₦{total_patient_amount:,.2f}, Available: ₦{wallet.amount:,.2f}. Please visit Finance to fund wallet.',
                    'required_amount': float(total_patient_amount),
                    'available_amount': float(wallet.amount),
                    'shortage': float(total_patient_amount - wallet.amount)
                })

            # Create insurance claim if insurance is active and covers any amount
            insurance_claim = None
            if active_insurance and total_insurance_covered > 0:
                claim_type = 'radiology' if len(orders) == 1 else 'multiple'

                insurance_claim = InsuranceClaimModel.objects.create(
                    patient_insurance=active_insurance,
                    claim_type=claim_type,
                    total_amount=total_base_amount,
                    covered_amount=total_insurance_covered,
                    patient_amount=total_patient_amount,
                    service_date=now(),
                    created_by=request.user,
                    notes=f'Auto-generated claim for {len(orders)} scan(s)'
                )

            # Process payments and update orders
            updated_count = 0

            for detail in order_details:
                order = detail['order']
                patient_amount = detail['patient_amount']

                # Deduct patient amount from wallet
                wallet.amount -= patient_amount

                # Update order
                order.status = 'paid'
                order.payment_status = True
                order.payment_date = now()
                order.payment_by = request.user

                if not order.amount_charged:
                    order.amount_charged = detail['base_amount']

                # Link to insurance claim if created
                if insurance_claim:
                    order.insurance_claim = insurance_claim

                order.save()
                updated_count += 1

            # Save wallet
            wallet.save()

            # Prepare response message
            message_parts = [
                f'Successfully processed payment for {updated_count} scan(s)',
                f'Patient paid: ₦{total_patient_amount:,.2f}'
            ]

            if insurance_claim:
                message_parts.append(
                    f'Insurance claim {insurance_claim.claim_number} created for ₦{total_insurance_covered:,.2f}')

            return JsonResponse({
                'success': True,
                'message': '. '.join(message_parts),
                'details': {
                    'processed_count': updated_count,
                    'patient_amount_paid': float(total_patient_amount),
                    'insurance_amount_covered': float(total_insurance_covered),
                    'new_wallet_balance': float(wallet.amount),
                    'formatted_wallet_balance': f'₦{wallet.amount:,.2f}',
                    'insurance_claim_number': insurance_claim.claim_number if insurance_claim else None
                }
            })

    except PatientWalletModel.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Patient wallet not found. Please contact Finance to create wallet.'
        })
    except Exception as e:
        logger.exception("Error processing scan payments")
        return JsonResponse({
            'success': False,
            'error': f'Error processing payment: {str(e)}'
        })


# -------------------------
# Scan Scheduling
# -------------------------
@login_required
@permission_required('scan.add_scanordermodel', raise_exception=True)
def schedule_scan(request, order_id):
    """Schedule a paid scan"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})

    try:
        order = get_object_or_404(ScanOrderModel, id=order_id, status='paid')

        scheduled_date_str = request.POST.get('scheduled_date', '').strip()
        if not scheduled_date_str:
            return JsonResponse({'success': False, 'error': 'Scheduled date is required'})

        try:
            scheduled_date = datetime.fromisoformat(scheduled_date_str)
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Invalid date format'})

        # Update order
        order.status = 'scheduled'
        order.scheduled_date = scheduled_date
        order.scheduled_by = request.user
        order.save()

        return JsonResponse({
            'success': True,
            'message': f'Scan scheduled for {order.template.name}',
            'scheduled_date': scheduled_date.strftime('%Y-%m-%d %H:%M')
        })

    except Exception as e:
        logger.exception("Error scheduling scan")
        return JsonResponse({'success': False, 'error': f'Error scheduling scan: {str(e)}'})


# -------------------------
# Scan Result Views
# -------------------------
class ScanResultDashboardView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """Main dashboard for scan results - shows orders ready for scanning"""
    model = ScanOrderModel
    permission_required = 'scan.view_scanordermodel'
    template_name = 'scan/result/dashboard.html'
    context_object_name = 'orders'
    paginate_by = 20

    def get_queryset(self):
        status_filter = self.request.GET.get('status', '')
        search_query = self.request.GET.get('search', '')

        # Base queryset - orders that can have scans performed
        queryset = ScanOrderModel.objects.filter(
            status__in=['scheduled', 'in_progress', 'completed']
        ).select_related('patient', 'template', 'scheduled_by').order_by('-scheduled_date')

        # Apply filters
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        if search_query:
            queryset = queryset.filter(
                Q(patient__first_name__icontains=search_query) |
                Q(patient__last_name__icontains=search_query) |
                Q(order_number__icontains=search_query) |
                Q(template__name__icontains=search_query)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Status counts for dashboard
        status_counts = ScanOrderModel.objects.filter(
            status__in=['scheduled', 'in_progress', 'completed']
        ).aggregate(
            scheduled=Count(Case(When(status='scheduled', then=1), output_field=IntegerField())),
            in_progress=Count(Case(When(status='in_progress', then=1), output_field=IntegerField())),
            completed=Count(Case(When(status='completed', then=1), output_field=IntegerField())),
        )

        # Results needing verification
        unverified_count = ScanResultModel.objects.filter(is_verified=False).count()

        context.update({
            'status_counts': status_counts,
            'unverified_count': unverified_count,
            'status_choices': ScanOrderModel.STATUS_CHOICES,
            'current_status': self.request.GET.get('status', ''),
            'search_query': self.request.GET.get('search', ''),
        })

        return context


class ScanResultCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = ScanResultModel
    form_class = ScanResultForm
    permission_required = 'scan.add_scanresultmodel'
    template_name = 'scan/result/create.html'

    def dispatch(self, request, *args, **kwargs):
        order_id = kwargs.get('order_id')
        self.order = None
        if order_id:
            self.order = get_object_or_404(ScanOrderModel, pk=order_id)
            # Check if results already exist
            if hasattr(self.order, 'scanresultmodel'):
                messages.warning(request, 'Results already exist. Redirecting to edit.')
                return redirect('scan_result_edit', pk=self.order.scanresultmodel.pk)
            # Check if order is ready for result entry
            if self.order.status not in ['in_progress', 'completed']:
                messages.error(request, 'This order is not ready for result entry.')
                return redirect('scan_result_dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.order:
            context.update({
                'order': self.order,
                'patient': self.order.patient,
                'template': self.order.template,
                'measurements': self.order.template.scan_parameters.get('measurements', []),
            })
        else:
            # Provide orders available for result entry
            context['available_orders'] = ScanOrderModel.objects.filter(
                status__in=['in_progress', 'completed']
            ).exclude(
                id__in=ScanResultModel.objects.values_list('order_id', flat=True)
            ).select_related('patient', 'template').order_by('-scan_completed_at')
        return context

    def form_valid(self, form):
        """
        Save ScanResultModel and then create ScanImageModel records for any uploaded expected images.
        """
        try:
            with transaction.atomic():
                result = form.save(commit=False)

                # If view was invoked with order_id, use that; otherwise support order selected in the form
                if self.order:
                    result.order = self.order
                else:
                    # attempt to get order from form.cleaned_data (only if your form allows it)
                    chosen_order = form.cleaned_data.get('order') if hasattr(form, 'cleaned_data') else None
                    if chosen_order:
                        result.order = chosen_order
                    else:
                        raise ValueError("No order specified for scan result")

                # performed_by - be defensive
                performed_by = self.request.user
                result.performed_by = performed_by

                # performed_at: prefer form input if present, else now
                if not result.performed_at:
                    result.performed_at = timezone.now()

                # If amount_charged or other fields are needed populate them -- measured values handled below
                # Process measured values from form data
                measurements = result.order.template.scan_parameters.get('measurements', [])
                measured_values = {'measurements': []}

                for measurement in measurements:
                    # After
                    param_name = measurement.get('name', '')
                    param_code = slugify(param_name)
                    value = self.request.POST.get(f'measurement_{param_code}', '').strip()
                    comment = self.request.POST.get(f'measurement_{param_code}_comment', '').strip()

                    if value:
                        measurement_result = {
                            'code': param_code,
                            'name': measurement.get('name', ''),
                            'value': value,
                            'unit': measurement.get('unit', ''),
                            'type': measurement.get('type', 'text'),
                            'comment': comment
                        }

                        if 'normal_range' in measurement and measurement.get('type') == 'numeric':
                            measurement_result['normal_range'] = measurement['normal_range']
                            try:
                                numeric_value = float(value)
                                min_val = measurement['normal_range'].get('min')
                                max_val = measurement['normal_range'].get('max')

                                if min_val is not None and max_val is not None:
                                    if numeric_value < float(min_val):
                                        measurement_result['status'] = 'low'
                                    elif numeric_value > float(max_val):
                                        measurement_result['status'] = 'high'
                                    else:
                                        measurement_result['status'] = 'normal'
                            except (ValueError, TypeError):
                                measurement_result['status'] = 'normal'

                        measured_values['measurements'].append(measurement_result)

                if measured_values['measurements']:
                    result.measured_values = measured_values
                print("Measured values being saved:", measured_values)
                # Save the scan result first (so we can attach images)
                result.save()

                # Update the order status to completed if needed
                if result.order.status != 'completed':
                    result.order.status = 'completed'
                    result.order.scan_completed_at = timezone.now()
                    result.order.save(update_fields=['status', 'scan_completed_at'])

                # --- Handle expected image uploads submitted with deterministic names ---
                template_expected = result.order.template.expected_images or []
                for idx, expected in enumerate(template_expected):
                    file_field_name = f'expected_image_{idx}'
                    uploaded_file = self.request.FILES.get(file_field_name)
                    if uploaded_file:
                        # retrieve view metadata sent from the template
                        view_field_name = f'expected_image_view_{idx}'
                        view_name = self.request.POST.get(view_field_name, expected.get('view', ''))

                        # determine next sequence_number for this result
                        last_seq = result.images.aggregate(max_seq=models.Max('sequence_number'))['max_seq'] or 0
                        seq_no = last_seq + 1

                        # create ScanImageModel instance
                        ScanImageModel.objects.create(
                            scan_result=result,
                            image=uploaded_file,
                            view_type=view_name,
                            description=expected.get('description', ''),
                            sequence_number=seq_no,
                            image_quality='good',  # default, can be adjusted
                            technical_parameters={}
                        )

                # Also support radiology_report_image file (already in the form)
                report_img = self.request.FILES.get('radiology_report_image')
                if report_img:
                    result.radiology_report_image = report_img
                    result.save(update_fields=['radiology_report_image'])

            messages.success(self.request, f'Results entered successfully for {result.order.template.name}')
            self.object = result
            return super().form_valid(form)

        except PermissionDenied as p:
            messages.error(self.request, str(p))
            return self.form_invalid(form)

        except Exception as e:
            logger.exception("Error creating scan result for order %s", getattr(self.order, 'id', 'unknown'))
            messages.error(self.request, f'An error occurred while saving results: {str(e)}')
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse('scan_result_detail', kwargs={'pk': self.object.pk})


class ScanResultListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """List all scan results with filtering options"""
    model = ScanResultModel
    permission_required = 'scan.view_scanresultmodel'
    template_name = 'scan/result/index.html'
    context_object_name = 'results'
    paginate_by = 20

    def get_queryset(self):
        verified_filter = self.request.GET.get('verified', '')
        search_query = self.request.GET.get('search', '')
        start_date_str = self.request.GET.get('start_date')
        end_date_str = self.request.GET.get('end_date')

        queryset = ScanResultModel.objects.select_related(
            'order__patient', 'order__template', 'verified_by'
        ).order_by('-created_at')

        if verified_filter == 'verified':
            queryset = queryset.filter(is_verified=True)
        elif verified_filter == 'unverified':
            queryset = queryset.filter(is_verified=False)

        if start_date_str and end_date_str:
            try:
                start_date = date.fromisoformat(start_date_str)
                end_date = date.fromisoformat(end_date_str)
                queryset = queryset.filter(created_at__date__range=[start_date, end_date])
            except (ValueError, TypeError):
                pass

        if search_query:
            queryset = queryset.annotate(
                search_full_name=Concat(
                    'order__patient__first_name', Value(' '), 'order__patient__last_name'
                )
            ).filter(
                # Searches metadata
                Q(search_full_name__icontains=search_query) |
                Q(order__order_number__icontains=search_query) |
                Q(order__template__name__icontains=search_query) |
                Q(order__patient__card_number__icontains=search_query) |

                # --- NEW: Search the report content itself ---
                Q(findings__icontains=search_query) |
                Q(impression__icontains=search_query)
            )

        return queryset

    def get_context_data(self, **kwargs):
        # ... your get_context_data method remains the same ...
        context = super().get_context_data(**kwargs)
        filtered_queryset = self.get_queryset()
        context['total_results'] = filtered_queryset.count()
        context['verified_results'] = filtered_queryset.filter(is_verified=True).count()
        context['pending_verification'] = filtered_queryset.filter(is_verified=False).count()
        context['current_verified'] = self.request.GET.get('verified', '')
        context['search_query'] = self.request.GET.get('search', '')
        today_str = date.today().isoformat()
        context['start_date'] = self.request.GET.get('start_date', today_str)
        context['end_date'] = self.request.GET.get('end_date', today_str)
        context['verified_choices'] = [
            ('', 'All Statuses'), ('verified', 'Verified Only'), ('unverified', 'Unverified Only')
        ]
        return context


class ScanResultUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = ScanResultModel
    form_class = ScanResultForm
    permission_required = 'scan.add_scanresultmodel'
    template_name = 'scan/result/edit.html'

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.order = self.object.order

        # Check permissions - only allow editing if not finalized
        if self.object.status == 'finalized' and not request.user.has_perm('scan.finalize_scanresultmodel'):
            messages.error(request, 'Cannot edit finalized results. Contact administrator.')
            return redirect('scan_result_detail', pk=self.object.pk)

        return super().dispatch(request, *args, **kwargs)

    # After
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get the definitions and existing values
        measurement_definitions = self.order.template.scan_parameters.get('measurements', [])
        existing_values_list = self.object.measured_values.get('measurements', [])

        # Create a dictionary for quick lookups of existing values
        existing_values_map = {item['code']: item for item in existing_values_list}

        # Combine the definitions with the existing values
        combined_measurements = []
        for definition in measurement_definitions:
            slugified_name = slugify(definition.get('name', ''))
            existing_data = existing_values_map.get(slugified_name, {})

            # Merge definition with existing data
            combined = definition.copy()
            combined['value'] = existing_data.get('value', '')
            combined['comment'] = existing_data.get('comment', '')
            combined_measurements.append(combined)

        context.update({
            'order': self.order,
            'patient': self.order.patient,
            'template': self.order.template,
            'measurements': combined_measurements,  # Use the new combined list
            'existing_images': self.object.images.all().order_by('sequence_number'),
        })
        return context

    def get_initial(self):
        """Populate form with existing measurement values"""
        initial = super().get_initial()

        # Populate measurement values from existing data
        measurements_data = self.object.measurements_data
        existing_measurements = measurements_data.get('measurements', [])

        # Create a lookup dictionary for existing values
        measurement_values = {}
        # After
        for measurement in existing_measurements:
            # Get the name from the saved data and slugify it
            name = measurement.get('name', '')
            if name:
                slugified_name = slugify(name)
                measurement_values[f'measurement_{slugified_name}'] = measurement.get('value', '')
                measurement_values[f'measurement_{slugified_name}_comment'] = measurement.get('comment', '')

        initial.update(measurement_values)
        return initial

    def form_valid(self, form):
        """
        Update ScanResultModel and handle image uploads.
        """
        try:
            with transaction.atomic():
                result = form.save(commit=False)

                # Update performed_by if changed
                result.performed_by = self.request.user

                # Update performed_at if provided, otherwise keep existing
                if not result.performed_at:
                    result.performed_at = self.object.performed_at

                # Process measured values from form data
                measurements = result.order.template.scan_parameters.get('measurements', [])
                measured_values = {'measurements': []}

                for measurement in measurements:
                    # After
                    param_name = measurement.get('name', '')
                    param_code = slugify(param_name)
                    value = self.request.POST.get(f'measurement_{param_code}', '').strip()
                    comment = self.request.POST.get(f'measurement_{param_code}_comment', '').strip()

                    if value:
                        measurement_result = {
                            'code': param_code,
                            'name': measurement.get('name', ''),
                            'value': value,
                            'unit': measurement.get('unit', ''),
                            'type': measurement.get('type', 'text'),
                            'comment': comment
                        }

                        if 'normal_range' in measurement and measurement.get('type') == 'numeric':
                            measurement_result['normal_range'] = measurement['normal_range']
                            try:
                                numeric_value = float(value)
                                min_val = measurement['normal_range'].get('min')
                                max_val = measurement['normal_range'].get('max')

                                if min_val is not None and max_val is not None:
                                    if numeric_value < float(min_val):
                                        measurement_result['status'] = 'low'
                                    elif numeric_value > float(max_val):
                                        measurement_result['status'] = 'high'
                                    else:
                                        measurement_result['status'] = 'normal'
                            except (ValueError, TypeError):
                                measurement_result['status'] = 'normal'

                        measured_values['measurements'].append(measurement_result)

                if measured_values['measurements']:
                    result.measured_values = measured_values

                # Save the updated result
                result.save()

                # Handle deletion of existing images if requested
                delete_image_ids = self.request.POST.getlist('delete_images')
                if delete_image_ids:
                    result.images.filter(id__in=delete_image_ids).delete()

                # --- Handle new expected image uploads ---
                template_expected = result.order.template.expected_images or []
                for idx, expected in enumerate(template_expected):
                    file_field_name = f'expected_image_{idx}'
                    uploaded_file = self.request.FILES.get(file_field_name)
                    if uploaded_file:
                        # retrieve view metadata sent from the template
                        view_field_name = f'expected_image_view_{idx}'
                        view_name = self.request.POST.get(view_field_name, expected.get('view', ''))

                        # determine next sequence_number for this result
                        last_seq = result.images.aggregate(max_seq=models.Max('sequence_number'))['max_seq'] or 0
                        seq_no = last_seq + 1

                        # create ScanImageModel instance
                        ScanImageModel.objects.create(
                            scan_result=result,
                            image=uploaded_file,
                            view_type=view_name,
                            description=expected.get('description', ''),
                            sequence_number=seq_no,
                            image_quality='good',
                            technical_parameters={}
                        )

                # Handle radiology report image upload/replacement
                report_img = self.request.FILES.get('radiology_report_image')
                if report_img:
                    result.radiology_report_image = report_img
                    result.save(update_fields=['radiology_report_image'])

                # Handle report image deletion if requested
                if self.request.POST.get('delete_radiology_report') == 'true':
                    if result.radiology_report_image:
                        result.radiology_report_image.delete(save=False)
                        result.radiology_report_image = None
                        result.save(update_fields=['radiology_report_image'])

            messages.success(self.request, f'Results updated successfully for {result.order.template.name}')
            self.object = result
            return super().form_valid(form)

        except PermissionDenied as p:
            messages.error(self.request, str(p))
            return self.form_invalid(form)

        except Exception as e:
            logger.exception("Error updating scan result %s", self.object.pk)
            messages.error(self.request, f'An error occurred while updating results: {str(e)}')
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse('scan_result_detail', kwargs={'pk': self.object.pk})


class ScanResultDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = ScanResultModel
    permission_required = 'scan.view_scanresultmodel'
    template_name = 'scan/result/detail.html'
    context_object_name = 'result'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        result = self.object

        context.update({
            'can_verify': not result.is_verified and self.request.user.has_perm('scan.change_scanresultmodel'),
            'can_edit': not result.is_verified and self.request.user.has_perm('scan.change_scanresultmodel'),
            'can_upload_images': self.request.user.has_perm('scan.add_scanimagemodel'),
            'order': result.order,
            'patient': result.order.patient,
            'template': result.order.template,
            'expected_images': result.order.template.expected_images,
            'measurements': result.measurements_data.get('measurements', []),
            'scan_images': result.images.all().order_by('sequence_number')
        })

        return context


# -------------------------
# Scan Image Management
# -------------------------
class ScanImageUploadView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Upload individual scan images"""
    model = ScanImageModel
    permission_required = 'scan.add_scanresultmodel'
    form_class = ScanImageForm
    template_name = 'scan/image/upload.html'

    def get_initial(self):
        initial = super().get_initial()
        result_id = self.kwargs.get('result_id')
        if result_id:
            initial['scan_result'] = result_id
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        result_id = self.kwargs.get('result_id')
        if result_id:
            scan_result = get_object_or_404(ScanResultModel, id=result_id)
            context['scan_result'] = scan_result
            context['expected_images'] = scan_result.order.template.expected_images
            context['existing_images'] = scan_result.images.all().order_by('sequence_number')
        return context

    def form_valid(self, form):
        messages.success(self.request, 'Image uploaded successfully')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('scan_result_detail', kwargs={'pk': self.object.scan_result.pk})


# -------------------------
# Scan Status Update Actions
# -------------------------
@login_required
@permission_required('scan.change_scanordermodel', raise_exception=True)
def start_scan(request, order_id):
    """Start scan procedure - AJAX endpoint"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    try:
        order = get_object_or_404(ScanOrderModel, pk=order_id)

        if order.status != 'scheduled':
            return JsonResponse({
                'success': False,
                'error': 'Scan can only be started for scheduled orders.'
            }, status=400)

        with transaction.atomic():
            order.status = 'in_progress'
            order.scan_started_at = now()
            order.performed_by = request.user
            order.save(update_fields=['status', 'scan_started_at', 'performed_by'])

        return JsonResponse({
            'success': True,
            'message': f'Scan started for order {order.order_number}',
            'redirect_url': reverse('scan_result_create_for_order', kwargs={'order_id': order.id})
        })

    except Exception as e:
        logger.exception("Error starting scan for order id=%s", order_id)
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while starting scan. Please contact administrator.'
        }, status=500)


@login_required
@permission_required('scan.can_verify_scan_result', raise_exception=True)
def verify_scan_result(request, pk):
    result_id = pk
    """AJAX endpoint to verify a scan result"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    try:
        result = get_object_or_404(ScanResultModel, pk=result_id)

        if result.is_verified:
            return JsonResponse({
                'success': False,
                'error': 'Result is already verified.'
            }, status=400)

        # Optional radiologist comments
        radiologist_comments = request.POST.get('radiologist_comments', '').strip()

        with transaction.atomic():
            result.is_verified = True
            result.verified_by = request.user
            result.verified_at = now()

            if radiologist_comments:
                result.radiologist_comments = radiologist_comments
            result.save(update_fields=['is_verified', 'verified_by', 'verified_at', 'radiologist_comments'])

        return JsonResponse({
            'success': True,
            'message': f'Scan result verified successfully',
            'verified_by': result.verified_by.get_full_name() or result.verified_by.username,
            'verified_at': result.verified_at.strftime('%B %d, %Y at %I:%M %p')
        })

    except Exception as e:
        logger.exception("Error verifying scan result id=%s", result_id)
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while verifying result. Please contact administrator.'
        }, status=500)


@login_required
@permission_required('scan.can_verify_scan_result', raise_exception=True)
def unverify_scan_result(request, pk):
    """AJAX endpoint to un-verify a scan result."""
    # 1. We must check that this is a POST request
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    try:
        result = get_object_or_404(ScanResultModel, pk=pk)

        # 2. Check if the result is already unverified and return an error
        if not result.is_verified:
            return JsonResponse({
                'success': False,
                'error': 'This result is already unverified.'
            }, status=400)

        # 3. The core logic is wrapped in a database transaction for safety
        with transaction.atomic():
            result.is_verified = False
            result.verified_by = None
            result.verified_at = None

            # 4. Clear the comments that were added during verification
            result.radiologist_comments = ''  # Clearing the field used for radiologist comments

            result.save(update_fields=['is_verified', 'verified_by', 'verified_at', 'radiologist_comments'])

        # 5. Return a success JSON response
        return JsonResponse({
            'success': True,
            'message': 'Scan result has been unverified successfully.'
        })

    except Exception as e:
        # 6. Catch any errors and return a server error JSON response
        logger.exception("Error unverifying scan result id=%s", pk)
        return JsonResponse({
            'success': False,
            'error': 'An unexpected error occurred. Please contact support.'
        }, status=500)


# -------------------------
# AJAX/API Views
# -------------------------
@login_required
def get_scan_template_details(request):
    """Get scan template details for AJAX requests."""
    template_id = request.GET.get('template_id')
    if not template_id:
        return JsonResponse({'error': 'template_id is required'}, status=400)

    try:
        template = ScanTemplateModel.objects.get(id=template_id)
        data = {
            'id': template.id,
            'name': template.name,
            'code': template.code,
            'price': str(template.price),
            'estimated_duration': template.estimated_duration,
            'preparation_required': template.preparation_required,
            'fasting_required': template.fasting_required,
            'equipment_required': template.equipment_required,
            'expected_images': template.expected_images,
            'scan_parameters': template.scan_parameters
        }
        return JsonResponse(data)
    except ScanTemplateModel.DoesNotExist:
        return JsonResponse({'error': 'Template not found'}, status=404)
    except Exception:
        logger.exception("Failed fetching scan template details for id=%s", template_id)
        return JsonResponse({'error': 'Internal error'}, status=500)


@login_required
def get_patient_scans(request):
    """Get scan orders for a specific patient."""
    patient_id = request.GET.get('patient_id')
    if not patient_id:
        return JsonResponse({'error': 'patient_id is required'}, status=400)

    try:
        orders = ScanOrderModel.objects.filter(
            patient_id=patient_id
        ).select_related('template').order_by('-ordered_at')[:10]

        data = {
            'orders': [
                {
                    'id': order.id,
                    'order_number': order.order_number,
                    'template_name': order.template.name,
                    'status': order.status,
                    'ordered_at': order.ordered_at.strftime('%Y-%m-%d %H:%M'),
                    'scheduled_date': order.scheduled_date.strftime('%Y-%m-%d %H:%M') if order.scheduled_date else None,
                }
                for order in orders
            ]
        }
        return JsonResponse(data)
    except Exception:
        logger.exception("Failed fetching patient scans for patient_id=%s", patient_id)
        return JsonResponse({'error': 'Internal error'}, status=500)


@login_required
def scan_dashboard_data(request):
    """Get scan dashboard statistics for AJAX requests."""
    try:
        today = date.today()

        data = {
            'today_orders': ScanOrderModel.objects.filter(ordered_at__date=today).count(),
            'pending_payments': ScanOrderModel.objects.filter(status='pending').count(),
            'scans_to_schedule': ScanOrderModel.objects.filter(status='paid').count(),
            'scheduled_scans': ScanOrderModel.objects.filter(status='scheduled').count(),
            'scans_in_progress': ScanOrderModel.objects.filter(status='in_progress').count(),
            'pending_verification': ScanResultModel.objects.filter(is_verified=False).count(),
            'completed_today': ScanOrderModel.objects.filter(status='completed', ordered_at__date=today).count(),
        }
        return JsonResponse(data)
    except Exception:
        logger.exception("Failed fetching scan dashboard data")
        return JsonResponse({'error': 'Internal error'}, status=500)


# -------------------------
# Dashboard View
# -------------------------
class ScanDashboardView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    template_name = 'scan/dashboard.html'
    permission_required = 'scan.view_scanordermodel'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()

        # Basic stats
        context.update({
            'total_categories': ScanCategoryModel.objects.count(),
            'total_templates': ScanTemplateModel.objects.filter(is_active=True).count(),
            'today_orders': ScanOrderModel.objects.filter(ordered_at__date=today).count(),
            'pending_payments': ScanOrderModel.objects.filter(status='pending').count(),
            'scans_to_schedule': ScanOrderModel.objects.filter(status='paid').count(),
            'scheduled_scans': ScanOrderModel.objects.filter(status='scheduled').count(),
            'scans_in_progress': ScanOrderModel.objects.filter(status='in_progress').count(),
            'completed_today': ScanOrderModel.objects.filter(status='completed', ordered_at__date=today).count(),
            'pending_verification': ScanResultModel.objects.filter(is_verified=False).count(),
        })

        # Recent orders
        context['recent_orders'] = ScanOrderModel.objects.select_related(
            'patient', 'template'
        ).order_by('-ordered_at')[:10]

        # Today's scheduled scans
        context['todays_scheduled'] = ScanOrderModel.objects.filter(
            scheduled_date__date=today,
            status__in=['scheduled', 'in_progress']
        ).select_related('patient', 'template').order_by('scheduled_date')

        return context


# -------------------------
# Image Management Actions
# -------------------------
@login_required
@permission_required('scan.add_scanresultmodel', raise_exception=True)
def delete_scan_image(request, image_id):
    """Delete a scan image - AJAX endpoint"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    try:
        image = get_object_or_404(ScanImageModel, pk=image_id)
        scan_result = image.scan_result

        # Check if scan is verified (cannot delete images from verified scans)
        if scan_result.is_verified:
            return JsonResponse({
                'success': False,
                'error': 'Cannot delete images from verified scan results.'
            }, status=400)

        # Delete the image
        image.delete()

        return JsonResponse({
            'success': True,
            'message': 'Image deleted successfully'
        })

    except Exception as e:
        logger.exception("Error deleting scan image id=%s", image_id)
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while deleting image. Please contact administrator.'
        }, status=500)


@login_required
@permission_required('scan.add_scanresultmodel', raise_exception=True)
def update_image_details(request, image_id):
    """Update image description and view type - AJAX endpoint"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    try:
        image = get_object_or_404(ScanImageModel, pk=image_id)

        # Check if scan is verified
        if image.scan_result.is_verified:
            return JsonResponse({
                'success': False,
                'error': 'Cannot edit images from verified scan results.'
            }, status=400)

        # Update details
        view_type = request.POST.get('view_type', '').strip()
        description = request.POST.get('description', '').strip()
        image_quality = request.POST.get('image_quality', '')

        image.view_type = view_type
        image.description = description
        if image_quality:
            image.image_quality = image_quality

        image.save(update_fields=['view_type', 'description', 'image_quality'])

        return JsonResponse({
            'success': True,
            'message': 'Image details updated successfully'
        })

    except Exception as e:
        logger.exception("Error updating image details id=%s", image_id)
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while updating image. Please contact administrator.'
        }, status=500)


# -------------------------
# Bulk Actions
# -------------------------
@login_required
@permission_required('scan.change_scanordermodel', raise_exception=True)
def multi_scan_order_action(request):
    """Handle bulk actions on scan orders"""
    if request.method == 'POST':
        order_ids = request.POST.getlist('order')
        action = request.POST.get('action')

        if not order_ids:
            messages.error(request, 'No order selected.')
            return redirect(reverse('scan_order_index'))

        try:
            with transaction.atomic():
                orders = ScanOrderModel.objects.filter(id__in=order_ids)

                if action == 'mark_paid':
                    paid_count = 0
                    for order in orders.filter(status='pending'):
                        order.status = 'paid'
                        order.payment_status = True
                        order.payment_date = now()
                        order.payment_by = request.user
                        order.save(update_fields=['status', 'payment_status', 'payment_date', 'payment_by'])
                        paid_count += 1
                    messages.success(request, f'Marked {paid_count} scan order(s) as paid.')

                elif action == 'cancel':
                    cancelled_count = 0
                    for order in orders.exclude(status__in=['completed', 'cancelled']):
                        order.status = 'cancelled'
                        order.save(update_fields=['status'])
                        cancelled_count += 1
                    messages.success(request, f'Cancelled {cancelled_count} scan order(s).')
                else:
                    messages.error(request, 'Invalid action.')
        except Exception:
            logger.exception("Bulk scan order action failed for ids=%s action=%s", order_ids, action)
            messages.error(request, "An error occurred performing that action. Try again or contact admin.")
        return redirect(reverse('scan_order_index'))


# -------------------------
# Print Views
# -------------------------
@login_required
@permission_required('scan.view_scanordermodel', raise_exception=True)
def print_scan_order(request, pk):
    order = get_object_or_404(ScanOrderModel, pk=pk)
    context = {
        'order': order,

    }
    return render(request, 'scan/print/order.html', context)


@login_required
@permission_required('scan.view_scanresultmodel', raise_exception=True)
def print_scan_result(request, pk):
    result = get_object_or_404(ScanResultModel, pk=pk)
    context = {
        'result': result,
        'order': result.order,
        'patient': result.order.patient,
        'images': result.images.all().order_by('sequence_number'),

    }
    return render(request, 'scan/print/result.html', context)


# -------------------------
# Scan Equipment Views
# -------------------------
class ScanEquipmentCreateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, CreateView
):
    model = ScanEquipmentModel
    permission_required = 'scan.add_scanequipmentmodel'
    form_class = ScanEquipmentForm
    template_name = 'scan/equipment/index.html'
    success_message = 'Scan Equipment Successfully Added'

    def get_success_url(self):
        return reverse('scan_equipment_index')

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET':
            return redirect(reverse('scan_equipment_index'))
        return super().dispatch(request, *args, **kwargs)


class ScanEquipmentListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = ScanEquipmentModel
    permission_required = 'scan.view_scanequipmentmodel'
    template_name = 'scan/equipment/index.html'
    context_object_name = "equipment_list"

    def get_queryset(self):
        return ScanEquipmentModel.objects.all().order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = ScanEquipmentForm()
        context['template_list'] = ScanTemplateModel.objects.filter(is_active=True).order_by('name')
        return context


class ScanEquipmentUpdateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, UpdateView
):
    model = ScanEquipmentModel
    permission_required = 'scan.change_scanequipmentmodel'
    form_class = ScanEquipmentForm
    template_name = 'scan/equipment/index.html'
    success_message = 'Scan Equipment Successfully Updated'

    def get_success_url(self):
        return reverse('scan_equipment_index')

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET':
            return redirect(reverse('scan_equipment_index'))
        return super().dispatch(request, *args, **kwargs)


class ScanEquipmentDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = ScanEquipmentModel
    permission_required = 'scan.delete_scanequipmentmodel'
    template_name = 'scan/equipment/delete.html'
    context_object_name = "equipment"
    success_message = 'Scan Equipment Successfully Deleted'

    def get_success_url(self):
        return reverse('scan_equipment_index')


# -------------------------
# Scan Appointment Views
# -------------------------
class ScanAppointmentCreateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, CreateView
):
    model = ScanAppointmentModel
    permission_required = 'scan.add_scanappointmentmodel'
    form_class = ScanAppointmentForm
    template_name = 'scan/appointment/create.html'
    success_message = 'Scan Appointment Successfully Created'

    def get_success_url(self):
        return reverse('scan_appointment_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['order_list'] = ScanOrderModel.objects.filter(status__in=['paid', 'scheduled']).select_related(
            'patient', 'template')
        context['equipment_list'] = ScanEquipmentModel.objects.filter(status='active')
        return context


class ScanAppointmentListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = ScanAppointmentModel
    permission_required = 'scan.view_scanappointmentmodel'
    template_name = 'scan/appointment/index.html'
    context_object_name = "appointment_list"
    paginate_by = 20

    def get_queryset(self):
        return ScanAppointmentModel.objects.select_related('scan_order', 'equipment', 'technician').order_by(
            'appointment_date')


class ScanAppointmentDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = ScanAppointmentModel
    permission_required = 'scan.view_scanappointmentmodel'
    template_name = 'scan/appointment/detail.html'
    context_object_name = "appointment"


class ScanAppointmentUpdateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, UpdateView
):
    model = ScanAppointmentModel
    permission_required = 'scan.change_scanappointmentmodel'
    form_class = ScanAppointmentForm
    template_name = 'scan/appointment/edit.html'
    success_message = 'Scan Appointment Successfully Updated'

    def get_success_url(self):
        return reverse('scan_appointment_detail', kwargs={'pk': self.object.pk})


class ScanAppointmentDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = ScanAppointmentModel
    permission_required = 'scan.delete_scanappointmentmodel'
    template_name = 'scan/appointment/delete.html'
    context_object_name = "appointment"
    success_message = 'Scan Appointment Successfully Deleted'

    def get_success_url(self):
        return reverse('scan_appointment_index')


# -------------------------
# Scan Template Builder Views
# -------------------------
class ScanTemplateBuilderCreateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, CreateView
):
    model = ScanTemplateBuilderModel
    permission_required = 'scan.add_scantemplatebuildermodel'
    form_class = ScanTemplateBuilderForm
    template_name = 'scan/template_builder/create.html'
    success_message = 'Template Builder Successfully Created'

    def get_success_url(self):
        return reverse('scan_template_builder_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['category_list'] = ScanCategoryModel.objects.all().order_by('name')
        return context


class ScanTemplateBuilderListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = ScanTemplateBuilderModel
    permission_required = 'scan.view_scantemplatebuildermodel'
    template_name = 'scan/template_builder/index.html'
    context_object_name = "builder_list"

    def get_queryset(self):
        return ScanTemplateBuilderModel.objects.select_related('category', 'created_template').order_by('-created_at')


class ScanTemplateBuilderDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = ScanTemplateBuilderModel
    permission_required = 'scan.view_scantemplatebuildermodel'
    template_name = 'scan/template_builder/detail.html'
    context_object_name = "builder"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        builder = self.object

        context.update({
            'can_build': not builder.is_processed and self.request.user.has_perm('scan.add_scantemplatemodel'),
            'preset_parameters': builder._get_preset_parameters() if builder.scan_preset != 'custom' else [],
        })
        return context


# -------------------------
# Scan Settings Views
# -------------------------
class ScanSettingCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = ScanSettingModel
    form_class = ScanSettingForm
    permission_required = 'scan.add_scansettingmodel'
    success_message = 'Scan Setting Created Successfully'
    template_name = 'scan/setting/create.html'

    def dispatch(self, request, *args, **kwargs):
        setting = ScanSettingModel.objects.first()
        if setting:
            return redirect('scan_setting_edit', pk=setting.pk)
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self):
        return reverse('scan_setting_detail', kwargs={'pk': self.object.pk})


class ScanSettingDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = ScanSettingModel
    permission_required = 'scan.view_scansettingmodel'
    template_name = 'scan/setting/detail.html'
    context_object_name = "scan_setting"

    def get_object(self):
        return ScanSettingModel.objects.first()


class ScanSettingUpdateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, UpdateView
):
    model = ScanSettingModel
    form_class = ScanSettingForm
    permission_required = 'scan.add_scansettingmodel'
    success_message = 'Scan Setting Updated Successfully'
    template_name = 'scan/setting/create.html'

    def get_object(self):
        return ScanSettingModel.objects.first()

    def get_success_url(self):
        return reverse('scan_setting_detail', kwargs={'pk': self.object.pk})


# -------------------------
# Action Views (Status Updates)
# -------------------------
@login_required
@permission_required('scan.can_verify_scan_result', raise_exception=True)
def verify_result(request, pk):
    result = get_object_or_404(ScanResultModel, pk=pk)
    try:
        if result.is_verified:
            messages.info(request, "Result is already verified.")
            return redirect(reverse('scan_result_detail', kwargs={'pk': pk}))

        with transaction.atomic():
            result.is_verified = True
            result.verified_by = request.user
            result.verified_at = now()
            result.save(update_fields=['is_verified', 'verified_by', 'verified_at'])

        messages.success(request, f"Result verified for order {result.order.order_number}")
    except Exception:
        logger.exception("Error verifying scan result id=%s", pk)
        messages.error(request, "An error occurred while verifying result. Contact admin.")
    return redirect(reverse('scan_result_detail', kwargs={'pk': pk}))


@login_required
@permission_required('scan.change_scantemplatebuildermodel', raise_exception=True)
def build_template(request, pk):
    builder = get_object_or_404(ScanTemplateBuilderModel, pk=pk)
    try:
        if builder.is_processed:
            messages.info(request, "Template has already been built.")
            return redirect(reverse('scan_template_builder_detail', kwargs={'pk': pk}))

        result = builder.build_template()
        if result.get('error'):
            messages.error(request, result['error'])
        else:
            messages.success(request, f"Template '{builder.name}' successfully built!")
            return redirect(reverse('scan_template_detail', kwargs={'pk': result['template'].pk}))

    except Exception:
        logger.exception("Error building scan template from builder id=%s", pk)
        messages.error(request, "An error occurred while building template. Contact admin.")
    return redirect(reverse('scan_template_builder_detail', kwargs={'pk': pk}))


# -------------------------
# Bulk Actions
# -------------------------
@login_required
@permission_required('scan.delete_scancategorymodel', raise_exception=True)
def multi_category_action(request):
    """Handle bulk actions on categories (e.g., delete)."""
    if request.method == 'POST':
        category_ids = request.POST.getlist('category')
        action = request.POST.get('action')

        if not category_ids:
            messages.error(request, 'No category selected.')
            return redirect(reverse('scan_category_index'))

        try:
            with transaction.atomic():
                categories = ScanCategoryModel.objects.filter(id__in=category_ids)
                if action == 'delete':
                    count, _ = categories.delete()
                    messages.success(request, f'Successfully deleted {count} category(s).')
                else:
                    messages.error(request, 'Invalid action.')
        except Exception:
            logger.exception("Bulk category action failed for ids=%s action=%s", category_ids, action)
            messages.error(request, "An error occurred performing that action. Try again or contact admin.")
        return redirect(reverse('scan_category_index'))

    # GET - confirm action
    category_ids = request.GET.getlist('category')
    if not category_ids:
        messages.error(request, 'No category selected.')
        return redirect(reverse('scan_category_index'))

    action = request.GET.get('action')
    context = {'category_list': ScanCategoryModel.objects.filter(id__in=category_ids)}

    if action == 'delete':
        return render(request, 'scan/category/multi_delete.html', context)

    messages.error(request, 'Invalid action.')
    return redirect(reverse('scan_category_index'))


@login_required
@permission_required('scan.change_scanordermodel', raise_exception=True)
def multi_order_action(request):
    """Handle bulk actions on orders (e.g., update status)."""
    if request.method == 'POST':
        order_ids = request.POST.getlist('order')
        action = request.POST.get('action')

        if not order_ids:
            messages.error(request, 'No order selected.')
            return redirect(reverse('scan_order_index'))

        try:
            with transaction.atomic():
                orders = ScanOrderModel.objects.filter(id__in=order_ids)

                if action == 'mark_paid':
                    paid_count = 0
                    for order in orders.filter(status='pending'):
                        order.status = 'paid'
                        order.payment_status = True
                        order.payment_date = now()
                        order.payment_by = request.user
                        order.save(update_fields=['status', 'payment_status', 'payment_date', 'payment_by'])
                        paid_count += 1
                    messages.success(request, f'Marked {paid_count} order(s) as paid.')

                elif action == 'cancel':
                    cancelled_count = 0
                    for order in orders.exclude(status__in=['completed', 'cancelled']):
                        order.status = 'cancelled'
                        order.save(update_fields=['status'])
                        cancelled_count += 1
                    messages.success(request, f'Cancelled {cancelled_count} order(s).')
                else:
                    messages.error(request, 'Invalid action.')
        except Exception:
            logger.exception("Bulk scan order action failed for ids=%s action=%s", order_ids, action)
            messages.error(request, "An error occurred performing that action. Try again or contact admin.")
        return redirect(reverse('scan_order_index'))

    # GET - confirm action
    order_ids = request.GET.getlist('order')
    if not order_ids:
        messages.error(request, 'No order selected.')
        return redirect(reverse('scan_order_index'))

    action = request.GET.get('action')
    context = {'order_list': ScanOrderModel.objects.filter(id__in=order_ids).select_related('patient', 'template')}

    if action in ['mark_paid', 'cancel']:
        return render(request, 'scan/order/multi_action.html', {
            **context,
            'action': action,
            'action_title': 'Mark as Paid' if action == 'mark_paid' else 'Cancel Orders'
        })

    messages.error(request, 'Invalid action.')
    return redirect(reverse('scan_order_index'))


# -------------------------
# AJAX/API Views
# -------------------------
@login_required
def get_template_details(request):
    """Get scan template details for AJAX requests."""
    template_id = request.GET.get('template_id')
    if not template_id:
        return JsonResponse({'error': 'template_id is required'}, status=400)

    try:
        template = ScanTemplateModel.objects.get(id=template_id)
        data = {
            'id': template.id,
            'name': template.name,
            'code': template.code,
            'price': str(template.price),
            'scan_type': template.scan_type,
            'estimated_duration': template.estimated_duration,
            'parameters': template.measurement_names,
            'scan_parameters': template.scan_parameters
        }
        return JsonResponse(data)
    except ScanTemplateModel.DoesNotExist:
        return JsonResponse({'error': 'Template not found'}, status=404)
    except Exception:
        logger.exception("Failed fetching scan template details for id=%s", template_id)
        return JsonResponse({'error': 'Internal error'}, status=500)


@login_required
def get_patient_orders(request):
    """Get orders for a specific patient."""
    patient_id = request.GET.get('patient_id')
    if not patient_id:
        return JsonResponse({'error': 'patient_id is required'}, status=400)

    try:
        orders = ScanOrderModel.objects.filter(patient_id=patient_id).select_related('template').order_by(
            '-ordered_at')[:10]

        data = {
            'orders': [
                {
                    'id': order.id,
                    'order_number': order.order_number,
                    'template_name': order.template.name,
                    'status': order.status,
                    'ordered_at': order.ordered_at.strftime('%Y-%m-%d %H:%M'),
                }
                for order in orders
            ]
        }
        return JsonResponse(data)
    except Exception:
        logger.exception("Failed fetching scan patient orders for patient_id=%s", patient_id)
        return JsonResponse({'error': 'Internal error'}, status=500)


@login_required
def scan_dashboard_data(request):
    """Get dashboard statistics for AJAX requests."""
    try:
        today = date.today()

        data = {
            'today_orders': ScanOrderModel.objects.filter(ordered_at__date=today).count(),
            'pending_payments': ScanOrderModel.objects.filter(status='pending').count(),
            'scheduled_scans': ScanOrderModel.objects.filter(status='scheduled').count(),
            'scans_in_progress': ScanOrderModel.objects.filter(status='in_progress').count(),
            'pending_results': ScanOrderModel.objects.filter(status='in_progress').count(),
            'pending_verification': ScanResultModel.objects.filter(is_verified=False).count(),
            'inactive_equipment': ScanEquipmentModel.objects.filter(status='inactive').count(),
            'maintenance_due': ScanEquipmentModel.objects.filter(next_maintenance__lte=today, status='active').count(),
        }
        return JsonResponse(data)
    except Exception:
        logger.exception("Failed fetching scan dashboard data")
        return JsonResponse({'error': 'Internal error'}, status=500)



# -------------------------
# Print Views
# -------------------------
@login_required
@permission_required('scan.view_scanordermodel', raise_exception=True)
def print_order(request, pk):
    order = get_object_or_404(ScanOrderModel, pk=pk)
    context = {
        'order': order,
        'scan_setting': ScanSettingModel.objects.first(),
    }
    return render(request, 'scan/print/order.html', context)


@login_required
@permission_required('scan.view_scanresultmodel', raise_exception=True)
def print_result(request, pk):
    result = get_object_or_404(ScanResultModel, pk=pk)
    context = {
        'result': result,
        'order': result.order,
        'patient': result.order.patient,
        'scan_setting': ScanSettingModel.objects.first(),
    }
    return render(request, 'scan/print/result.html', context)


# -------------------------
# Order Action Helpers (Payment / Schedule / Start / Complete)
# -------------------------
@login_required
def process_payment(request, pk):
    order = get_object_or_404(ScanOrderModel, pk=pk)
    try:
        if order.status != 'pending':
            messages.info(request, "Order payment has already been processed or order is not pending.")
            return redirect(reverse('scan_order_detail', kwargs={'pk': pk}))

        with transaction.atomic():
            order.status = 'paid'
            order.payment_status = True
            order.payment_date = now()
            order.payment_by = request.user
            order.save(update_fields=['status', 'payment_status', 'payment_date', 'payment_by'])

        messages.success(request, f"Payment processed for order {order.order_number}")
    except Exception:
        logger.exception("Error processing payment for scan order id=%s", pk)
        messages.error(request, "An error occurred while processing payment. Contact admin.")
    return redirect(reverse('scan_order_detail', kwargs={'pk': pk}))


@login_required
@permission_required('scan.change_scanordermodel', raise_exception=True)
def schedule_scan(request, order_id):
    """
    Schedule a paid scan via an AJAX POST request.
    Returns a JSON response.
    """
    # 1. Ensure the request is a POST
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)

    try:
        order = get_object_or_404(ScanOrderModel, id=order_id)

        # 2. Validate the order's current status
        if order.status != 'paid':
            return JsonResponse({
                'success': False,
                'error': 'This scan can only be scheduled if it has been paid for.'
            }, status=400) # 400 = Bad Request

        # 3. Get and validate the scheduled_date from the POST data
        scheduled_date_str = request.POST.get('scheduled_date')
        if not scheduled_date_str:
            return JsonResponse({'success': False, 'error': 'Scheduled date is required.'}, status=400)

        # Use Django's robust parser
        scheduled_date = parse_datetime(scheduled_date_str)
        if not scheduled_date:
            return JsonResponse({'success': False, 'error': 'Invalid date format provided.'}, status=400)

        # 4. Update the database within a transaction
        with transaction.atomic():
            order.status = 'scheduled'
            order.scheduled_date = scheduled_date
            order.scheduled_by = request.user
            order.save(update_fields=['status', 'scheduled_date', 'scheduled_by'])

        # 5. Return a success JSON response
        return JsonResponse({
            'success': True,
            'message': f'Scan for "{order.template.name}" has been successfully scheduled.',
            'scheduled_date': scheduled_date.strftime('%Y-%m-%d %H:%M')
        })

    except ScanOrderModel.DoesNotExist:
         return JsonResponse({'success': False, 'error': 'Order not found.'}, status=404)
    except Exception as e:
        logger.exception("Error scheduling scan for order id=%s", order_id)
        return JsonResponse({
            'success': False,
            'error': 'An unexpected server error occurred. Please contact an administrator.'
        }, status=500) # 500 = Internal Server Error


@login_required
@permission_required('scan.change_scanordermodel', raise_exception=True)
def start_scan(request, order_id):
    """
    Starts a scheduled scan via an AJAX POST request.
    Returns a JSON response.
    """
    # 1. This action should only be performed via POST
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)

    try:
        order = get_object_or_404(ScanOrderModel, id=order_id)

        # 2. Validate that the order is in the correct state
        if order.status != 'scheduled':
            return JsonResponse({
                'success': False,
                'error': 'Scan can only be started for scheduled orders.'
            }, status=400) # 400 = Bad Request

        # 3. Update the database atomically
        with transaction.atomic():
            order.status = 'in_progress'
            order.scan_started_at = now()
            order.performed_by = request.user
            order.save(update_fields=['status', 'scan_started_at', 'performed_by'])

        # 4. Return a success JSON response, including a redirect URL for the JS
        return JsonResponse({
            'success': True,
            'message': f'Scan started for order {order.order_number}. Redirecting to result entry...',
            'redirect_url': reverse('scan_result_create_for_order', kwargs={'order_id': order.id})
        })

    except ScanOrderModel.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Order not found.'}, status=404)
    except Exception as e:
        logger.exception("Error starting scan for order id=%s", order_id)
        return JsonResponse({
            'success': False,
            'error': 'An unexpected server error occurred. Please contact an administrator.'
        }, status=500) # 500 = Internal Server Error


@login_required
@permission_required('scan.change_scanordermodel', raise_exception=True)
def complete_scan(request, pk):
    order = get_object_or_404(ScanOrderModel, pk=pk)
    try:
        if order.status != 'in_progress':
            messages.info(request, "Scan can only be completed for in-progress orders.")
            return redirect(reverse('scan_order_detail', kwargs={'pk': pk}))

        if not hasattr(order, 'result'):
            messages.error(request, "Cannot complete scan without results. Please add results first.")
            return redirect(reverse('scan_order_detail', kwargs={'pk': pk}))

        with transaction.atomic():
            order.status = 'completed'
            order.scan_completed_at = now()
            order.save(update_fields=['status', 'scan_completed_at'])

        messages.success(request, f"Scan completed for order {order.order_number}")
    except Exception:
        logger.exception("Error completing scan for order id=%s", pk)
        messages.error(request, "An error occurred while completing scan. Contact admin.")
    return redirect(reverse('scan_order_detail', kwargs={'pk': pk}))


def _month_start_for_offset(today: date, months_ago: int) -> date:
    """
    Return the first day of the month that is `months_ago` before `today`'s month.
    months_ago=0 => first day of current month
    months_ago=1 => first day of previous month
    """
    # compute month index relative to 1..12
    m = today.month - months_ago
    y = today.year
    while m <= 0:
        m += 12
        y -= 1
    while m > 12:
        m -= 12
        y += 1
    return date(y, m, 1)


def get_scan_dashboard_context(request):
    """Build and return context dict for scan dashboard (so we can reuse it for print)."""
    today = timezone.now().date()
    this_week_start = today - timedelta(days=today.weekday())
    this_month_start = today.replace(day=1)
    last_month_start = _month_start_for_offset(today, 1)
    last_month_end = this_month_start - timedelta(days=1)

    # Basic counts
    total_orders = ScanOrderModel.objects.count()
    # Templates & categories - if models exist; fallback to 0
    try:
        total_templates = ScanTemplateModel.objects.filter(is_active=True).count()
    except Exception:
        total_templates = 0
    try:
        total_categories = ScanCategoryModel.objects.count()
    except Exception:
        total_categories = 0

    completed_scans = ScanOrderModel.objects.filter(status='completed').count()

    # Today's stats
    orders_today = ScanOrderModel.objects.filter(ordered_at__date=today).count()
    completed_today = ScanOrderModel.objects.filter(status='completed', scan_completed_at__date=today).count()
    # pending: orders created today with status in pending/paid/scheduled/in_progress
    pending_today = ScanOrderModel.objects.filter(
        ordered_at__date=today,
        status__in=['pending', 'paid', 'scheduled', 'in_progress']
    ).count()

    # This week
    orders_week = ScanOrderModel.objects.filter(ordered_at__date__gte=this_week_start).count()
    completed_week = ScanOrderModel.objects.filter(status='completed', scan_completed_at__date__gte=this_week_start).count()

    # This month
    orders_month = ScanOrderModel.objects.filter(ordered_at__date__gte=this_month_start).count()
    completed_month = ScanOrderModel.objects.filter(status='completed', scan_completed_at__date__gte=this_month_start).count()

    # Last month (for growth)
    orders_last_month = ScanOrderModel.objects.filter(
        ordered_at__date__gte=last_month_start,
        ordered_at__date__lte=last_month_end
    ).count()
    completed_last_month = ScanOrderModel.objects.filter(
        status='completed',
        scan_completed_at__date__gte=last_month_start,
        scan_completed_at__date__lte=last_month_end
    ).count()

    orders_growth = 0
    completed_growth = 0
    if orders_last_month > 0:
        orders_growth = round(((orders_month - orders_last_month) / orders_last_month) * 100, 1)
    if completed_last_month > 0:
        completed_growth = round(((completed_month - completed_last_month) / completed_last_month) * 100, 1)

    # Status distribution for charts
    status_distribution = ScanOrderModel.objects.values('status').annotate(count=Count('id')).order_by('status')
    status_chart_data = [{'name': s['status'].replace('_', ' ').title(), 'value': s['count']} for s in status_distribution]

    # Revenue stats (payment_status True)
    total_revenue = ScanOrderModel.objects.filter(payment_status=True).aggregate(total=Sum('amount_charged'))['total'] or Decimal('0.00')
    revenue_today = ScanOrderModel.objects.filter(payment_status=True, payment_date__date=today).aggregate(total=Sum('amount_charged'))['total'] or Decimal('0.00')
    revenue_week = ScanOrderModel.objects.filter(payment_status=True, payment_date__date__gte=this_week_start).aggregate(total=Sum('amount_charged'))['total'] or Decimal('0.00')
    revenue_month = ScanOrderModel.objects.filter(payment_status=True, payment_date__date__gte=this_month_start).aggregate(total=Sum('amount_charged'))['total'] or Decimal('0.00')

    # Category wise statistics (if ScanCategoryModel/ScanTemplateModel exist)
    category_chart_data = []
    try:
        category_stats = ScanCategoryModel.objects.annotate(
            total_orders=Count('templates__orders'),
            completed_orders=Count('templates__orders', filter=Q(templates__orders__status='completed')),
            revenue=Sum('templates__orders__amount_charged', filter=Q(templates__orders__payment_status=True))
        ).order_by('-total_orders')
        category_chart_data = [
            {'name': cat.name, 'value': cat.total_orders, 'revenue': float(cat.revenue or 0)}
            for cat in category_stats[:10]
        ]
    except Exception:
        category_chart_data = []

    # Popular scans
    popular_scans = []
    try:
        popular = ScanTemplateModel.objects.annotate(order_count=Count('orders')).filter(order_count__gt=0).order_by('-order_count')[:10]
        popular_scans = [
            {
                'name': t.name,
                'orders': t.order_count,
                'revenue': float(ScanOrderModel.objects.filter(template=t, payment_status=True).aggregate(total=Sum('amount_charged'))['total'] or 0)
            }
            for t in popular
        ]
    except Exception:
        popular_scans = []

    # Daily trends - last 7 days
    daily_trends = []
    for i in range(7):
        d = today - timedelta(days=6 - i)
        orders_count = ScanOrderModel.objects.filter(ordered_at__date=d).count()
        completed_count = ScanOrderModel.objects.filter(status='completed', scan_completed_at__date=d).count()
        revenue = ScanOrderModel.objects.filter(payment_status=True, payment_date__date=d).aggregate(total=Sum('amount_charged'))['total'] or Decimal('0.00')
        daily_trends.append({'date': d.strftime('%Y-%m-%d'), 'orders': orders_count, 'completed': completed_count, 'revenue': float(revenue)})

    # Monthly trends - last 12 months
    monthly_trends = []
    for months_ago in reversed(range(12)):  # oldest first
        month_start = _month_start_for_offset(today, months_ago)
        # compute next month start
        if month_start.month == 12:
            next_month_start = date(month_start.year + 1, 1, 1)
        else:
            next_month_start = date(month_start.year, month_start.month + 1, 1)
        month_end = next_month_start - timedelta(days=1)

        orders_count = ScanOrderModel.objects.filter(ordered_at__date__gte=month_start, ordered_at__date__lte=month_end).count()
        completed_count = ScanOrderModel.objects.filter(status='completed', scan_completed_at__date__gte=month_start, scan_completed_at__date__lte=month_end).count()
        revenue = ScanOrderModel.objects.filter(payment_status=True, payment_date__date__gte=month_start, payment_date__date__lte=month_end).aggregate(total=Sum('amount_charged'))['total'] or Decimal('0.00')

        monthly_trends.append({'month': month_start.strftime('%b %Y'), 'orders': orders_count, 'completed': completed_count, 'revenue': float(revenue)})

    # Source distribution (if you have a source field)
    try:
        source_distribution = ScanOrderModel.objects.values('source').annotate(count=Count('id'))
        source_chart_data = []
        for s in source_distribution:
            name = s.get('source', '')
            if name == 'doctor':
                label = 'Doctor Prescribed'
            elif name == 'direct' or name == 'lab':
                label = 'Direct'
            else:
                label = name.title() if name else 'Unknown'
            source_chart_data.append({'name': label, 'value': s['count']})
    except Exception:
        source_chart_data = []

    # Recent activity
    recent_orders = ScanOrderModel.objects.select_related('patient', 'template', 'ordered_by').order_by('-ordered_at')[:10]

    # Average processing time for completed scans (hours)
    completed_orders_qs = ScanOrderModel.objects.filter(status='completed', scan_completed_at__isnull=False)
    avg_processing_hours = 0
    if completed_orders_qs.exists():
        total_hours = sum([
            (o.scan_completed_at - o.ordered_at).total_seconds() / 3600
            for o in completed_orders_qs
            if o.scan_completed_at and o.ordered_at
        ])
        avg_processing_hours = round(total_hours / completed_orders_qs.count(), 1)

    # Pending tasks
    pending_payment = ScanOrderModel.objects.filter(status='pending').count()
    pending_scheduling = ScanOrderModel.objects.filter(status='paid').count()
    pending_scheduled = ScanOrderModel.objects.filter(status='scheduled').count()

    context = {
        'total_orders': total_orders,
        'total_templates': total_templates,
        'total_categories': total_categories,
        'completed_scans': completed_scans,

        'orders_today': orders_today,
        'completed_today': completed_today,
        'pending_today': pending_today,

        'orders_week': orders_week,
        'completed_week': completed_week,

        'orders_month': orders_month,
        'completed_month': completed_month,
        'orders_growth': orders_growth,
        'completed_growth': completed_growth,

        'total_revenue': total_revenue,
        'revenue_today': revenue_today,
        'revenue_week': revenue_week,
        'revenue_month': revenue_month,

        'status_distribution': status_chart_data,
        'category_distribution': category_chart_data,
        'popular_scans': popular_scans,
        'daily_trends': daily_trends,
        'monthly_trends': monthly_trends,
        'source_distribution': source_chart_data,

        'avg_processing_hours': avg_processing_hours,
        'recent_orders': recent_orders,

        'pending_payment': pending_payment,
        'pending_scheduling': pending_scheduling,
        'pending_scheduled': pending_scheduled,
    }

    return context


@login_required
@permission_required('scan.view_scanordermodel', raise_exception=True)
def scan_dashboard(request):
    """Main scan dashboard view"""
    ctx = get_scan_dashboard_context(request)
    # For JS charts we JSON-encode some items in the template (like your lab templates do)
    import json
    ctx['status_distribution'] = json.dumps(ctx['status_distribution'])
    ctx['category_distribution'] = json.dumps(ctx['category_distribution'])
    ctx['daily_trends'] = json.dumps(ctx['daily_trends'])
    ctx['monthly_trends'] = json.dumps(ctx['monthly_trends'])
    ctx['source_distribution'] = json.dumps(ctx['source_distribution'])
    return render(request, 'scan/dashboard.html', ctx)


@login_required
def scan_dashboard_print(request):
    """Printable version of scan dashboard (reuses same context builder)"""
    ctx = get_scan_dashboard_context(request)
    return render(request, 'scan/dashboard_print.html', ctx)


@login_required
def scan_analytics_api(request):
    """AJAX API for charts (supports daily_revenue and category_performance)"""
    chart_type = request.GET.get('type')
    if chart_type == 'daily_revenue':
        data = []
        today = timezone.now().date()
        for i in range(30):
            d = today - timedelta(days=29 - i)
            revenue = ScanOrderModel.objects.filter(payment_status=True, payment_date__date=d).aggregate(total=Sum('amount_charged'))['total'] or Decimal('0.00')
            data.append({'date': d.strftime('%Y-%m-%d'), 'revenue': float(revenue)})
        return JsonResponse({'data': data})

    elif chart_type == 'category_performance':
        this_month_start = timezone.now().date().replace(day=1)
        try:
            categories = ScanCategoryModel.objects.annotate(
                orders_this_month=Count('templates__orders', filter=Q(templates__orders__ordered_at__date__gte=this_month_start)),
                revenue_this_month=Sum('templates__orders__amount_charged', filter=Q(templates__orders__payment_status=True, templates__orders__payment_date__date__gte=this_month_start))
            ).order_by('-orders_this_month')[:10]

            data = [{'name': c.name, 'orders': c.orders_this_month, 'revenue': float(c.revenue_this_month or 0)} for c in categories]
        except Exception:
            data = []
        return JsonResponse({'data': data})

    return JsonResponse({'error': 'Invalid chart type'}, status=400)


@login_required
@permission_required('scan.view_scanordermodel', raise_exception=True)
def scan_dashboard(request):
    """Main scan dashboard with comprehensive statistics"""

    # Get current date and time ranges
    today = timezone.now().date()
    this_week_start = today - timedelta(days=today.weekday())
    this_month_start = today.replace(day=1)
    last_month_start = (this_month_start - timedelta(days=1)).replace(day=1)
    last_month_end = this_month_start - timedelta(days=1)

    # === BASIC STATISTICS ===

    # Total counts
    total_orders = ScanOrderModel.objects.count()
    total_templates = ScanTemplateModel.objects.filter(is_active=True).count()
    total_categories = ScanCategoryModel.objects.count()
    completed_tests = ScanOrderModel.objects.filter(status='completed').count()

    # Today's statistics
    orders_today = ScanOrderModel.objects.filter(ordered_at__date=today).count()
    completed_today = ScanOrderModel.objects.filter(
        status='completed',
        processed_at__date=today
    ).count()
    pending_today = ScanOrderModel.objects.filter(
        ordered_at__date=today,
        status__in=['pending', 'paid', 'collected', 'processing']
    ).count()

    # This week's statistics
    orders_week = ScanOrderModel.objects.filter(ordered_at__date__gte=this_week_start).count()
    completed_week = ScanOrderModel.objects.filter(
        status='completed',
        processed_at__date__gte=this_week_start
    ).count()

    # This month's statistics
    orders_month = ScanOrderModel.objects.filter(ordered_at__date__gte=this_month_start).count()
    completed_month = ScanOrderModel.objects.filter(
        status='completed',
        processed_at__date__gte=this_month_start
    ).count()

    # Last month's statistics for growth calculation
    orders_last_month = ScanOrderModel.objects.filter(
        ordered_at__date__gte=last_month_start,
        ordered_at__date__lte=last_month_end
    ).count()
    completed_last_month = ScanOrderModel.objects.filter(
        status='completed',
        processed_at__date__gte=last_month_start,
        processed_at__date__lte=last_month_end
    ).count()

    # Calculate growth percentages
    orders_growth = 0
    completed_growth = 0
    if orders_last_month > 0:
        orders_growth = round(((orders_month - orders_last_month) / orders_last_month) * 100, 1)
    if completed_last_month > 0:
        completed_growth = round(((completed_month - completed_last_month) / completed_last_month) * 100, 1)

    # === STATUS DISTRIBUTION ===
    status_distribution = ScanOrderModel.objects.values('status').annotate(
        count=Count('id')
    ).order_by('status')

    # Convert to format suitable for charts
    status_chart_data = [
        {'name': status['status'].title(), 'value': status['count']}
        for status in status_distribution
    ]

    # === REVENUE STATISTICS ===

    # Total revenue
    total_revenue = ScanOrderModel.objects.filter(
        payment_status=True
    ).aggregate(total=Sum('amount_charged'))['total'] or Decimal('0.00')

    # Today's revenue
    revenue_today = ScanOrderModel.objects.filter(
        payment_status=True,
        payment_date__date=today
    ).aggregate(total=Sum('amount_charged'))['total'] or Decimal('0.00')

    # This week's revenue
    revenue_week = ScanOrderModel.objects.filter(
        payment_status=True,
        payment_date__date__gte=this_week_start
    ).aggregate(total=Sum('amount_charged'))['total'] or Decimal('0.00')

    # This month's revenue
    revenue_month = ScanOrderModel.objects.filter(
        payment_status=True,
        payment_date__date__gte=this_month_start
    ).aggregate(total=Sum('amount_charged'))['total'] or Decimal('0.00')

    # === CATEGORY WISE STATISTICS ===
    category_stats = ScanCategoryModel.objects.annotate(
        total_orders=Count('templates__orders'),
        completed_orders=Count('templates__orders', filter=Q(templates__orders__status='completed')),
        revenue=Sum('templates__orders__amount_charged', filter=Q(templates__orders__payment_status=True))
    ).order_by('-total_orders')

    # Format for chart
    category_chart_data = [
        {
            'name': cat.name,
            'value': cat.total_orders,
            'revenue': float(cat.revenue or 0)
        }
        for cat in category_stats[:10]  # Top 10 categories
    ]

    # === POPULAR TESTS ===
    popular_tests = ScanTemplateModel.objects.annotate(
        order_count=Count('orders')
    ).filter(order_count__gt=0).order_by('-order_count')[:10]

    popular_tests_data = [
        {
            'name': test.name,
            'orders': test.order_count,
            'revenue': float(
                ScanOrderModel.objects.filter(
                    template=test,
                    payment_status=True
                ).aggregate(total=Sum('amount_charged'))['total'] or 0
            )
        }
        for test in popular_tests
    ]

    # === DAILY TRENDS (LAST 7 DAYS) ===
    daily_trends = []
    for i in range(7):
        date = today - timedelta(days=6 - i)
        orders_count = ScanOrderModel.objects.filter(ordered_at__date=date).count()
        completed_count = ScanOrderModel.objects.filter(
            status='completed',
            processed_at__date=date
        ).count()
        revenue = ScanOrderModel.objects.filter(
            payment_status=True,
            payment_date__date=date
        ).aggregate(total=Sum('amount_charged'))['total'] or Decimal('0.00')

        daily_trends.append({
            'date': date.strftime('%Y-%m-%d'),
            'orders': orders_count,
            'completed': completed_count,
            'revenue': float(revenue)
        })

    # === MONTHLY TRENDS (LAST 12 MONTHS) ===
    monthly_trends = []
    for i in range(12):
        month_start = (this_month_start - timedelta(days=1)).replace(day=1)
        for _ in range(i):
            month_start = (month_start - timedelta(days=1)).replace(day=1)

        month_end = (month_start.replace(month=month_start.month + 1)
                     if month_start.month < 12
                     else month_start.replace(year=month_start.year + 1, month=1)) - timedelta(days=1)

        orders_count = ScanOrderModel.objects.filter(
            ordered_at__date__gte=month_start,
            ordered_at__date__lte=month_end
        ).count()

        completed_count = ScanOrderModel.objects.filter(
            status='completed',
            processed_at__date__gte=month_start,
            processed_at__date__lte=month_end
        ).count()

        revenue = ScanOrderModel.objects.filter(
            payment_status=True,
            payment_date__date__gte=month_start,
            payment_date__date__lte=month_end
        ).aggregate(total=Sum('amount_charged'))['total'] or Decimal('0.00')

        monthly_trends.insert(0, {
            'month': month_start.strftime('%b %Y'),
            'orders': orders_count,
            'completed': completed_count,
            'revenue': float(revenue)
        })

    # === SOURCE DISTRIBUTION ===
    source_distribution = ScanOrderModel.objects.values('source').annotate(
        count=Count('id')
    )

    source_chart_data = [
        {
            'name': 'Doctor Prescribed' if source['source'] == 'doctor' else 'Scan Direct',
            'value': source['count']
        }
        for source in source_distribution
    ]

    # === RECENT ACTIVITY ===
    recent_orders = ScanOrderModel.objects.select_related(
        'patient', 'template', 'ordered_by'
    ).order_by('-ordered_at')[:10]

    # === AVERAGE PROCESSING TIME ===
    completed_orders = ScanOrderModel.objects.filter(
        status='completed',
        processed_at__isnull=False
    )

    avg_processing_hours = 0
    if completed_orders.exists():
        total_processing_time = sum([
            (order.processed_at - order.ordered_at).total_seconds() / 3600
            for order in completed_orders
            if order.processed_at and order.ordered_at
        ])
        avg_processing_hours = round(total_processing_time / completed_orders.count(), 1)

    # === PENDING TASKS ===
    pending_collection = ScanOrderModel.objects.filter(status='paid').count()
    pending_processing = ScanOrderModel.objects.filter(status='collected').count()
    pending_verification = ScanResultModel.objects.filter(is_verified=False).count()

    context = {
        # Basic stats
        'total_orders': total_orders,
        'total_templates': total_templates,
        'total_categories': total_categories,
        'completed_tests': completed_tests,

        # Daily stats
        'orders_today': orders_today,
        'completed_today': completed_today,
        'pending_today': pending_today,

        # Weekly stats
        'orders_week': orders_week,
        'completed_week': completed_week,

        # Monthly stats
        'orders_month': orders_month,
        'completed_month': completed_month,
        'orders_growth': orders_growth,
        'completed_growth': completed_growth,

        # Revenue
        'total_revenue': total_revenue,
        'revenue_today': revenue_today,
        'revenue_week': revenue_week,
        'revenue_month': revenue_month,

        # Charts data
        'status_distribution': json.dumps(status_chart_data),
        'category_distribution': json.dumps(category_chart_data),
        'popular_tests': popular_tests_data,
        'daily_trends': json.dumps(daily_trends),
        'monthly_trends': json.dumps(monthly_trends),
        'source_distribution': json.dumps(source_chart_data),

        # Other stats
        'avg_processing_hours': avg_processing_hours,
        'recent_orders': recent_orders,

        # Pending tasks
        'pending_collection': pending_collection,
        'pending_processing': pending_processing,
        'pending_verification': pending_verification,
    }

    return render(request, 'scan/dashboard.html', context)



@login_required
def scan_dashboard_print(request):
    """Printable version of scan dashboard"""
    # Get the same context as main dashboard but simplified for printing
    context = scan_dashboard(request).context_data
    return render(request, 'scan/dashboard_print.html', context)


@login_required
def scan_analytics_api(request):
    """API endpoint for dynamic chart updates"""
    chart_type = request.GET.get('type')

    if chart_type == 'daily_revenue':
        # Last 30 days revenue
        data = []
        today = timezone.now().date()
        for i in range(30):
            date = today - timedelta(days=29 - i)
            revenue = ScanOrderModel.objects.filter(
                payment_status=True,
                payment_date__date=date
            ).aggregate(total=Sum('amount_charged'))['total'] or Decimal('0.00')

            data.append({
                'date': date.strftime('%Y-%m-%d'),
                'revenue': float(revenue)
            })

        return JsonResponse({'data': data})

    elif chart_type == 'category_performance':
        # Category wise performance this month
        this_month_start = timezone.now().date().replace(day=1)

        categories = ScanCategoryModel.objects.annotate(
            orders_this_month=Count(
                'templates__orders',
                filter=Q(templates__orders__ordered_at__date__gte=this_month_start)
            ),
            revenue_this_month=Sum(
                'templates__orders__amount_charged',
                filter=Q(
                    templates__orders__payment_status=True,
                    templates__orders__payment_date__date__gte=this_month_start
                )
            )
        ).order_by('-orders_this_month')[:10]

        data = [
            {
                'name': cat.name,
                'orders': cat.orders_this_month,
                'revenue': float(cat.revenue_this_month or 0)
            }
            for cat in categories
        ]

        return JsonResponse({'data': data})

    return JsonResponse({'error': 'Invalid chart type'}, status=400)


class ScanReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    template_name = 'scan/reports/index.html'
    permission_required = 'scan.view_scanordermodel'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Date range from request or default to this month
        from_date = self.request.GET.get('from_date')
        to_date = self.request.GET.get('to_date')

        if not from_date:
            from_date = date.today().replace(day=1)  # First day of current month
        else:
            from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()

        if not to_date:
            # Last day of current month
            if date.today().month == 12:
                to_date = date.today().replace(day=31)
            else:
                to_date = (date.today().replace(month=date.today().month + 1, day=1) - timedelta(days=1))
        else:
            to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()

        # Default title
        month_year = from_date.strftime('%B %Y')
        default_title = f'Scan Test Report for the Month of {month_year}'
        report_title = self.request.GET.get('title', default_title)

        # Filter orders - exclude pending and cancelled from total
        all_orders = ScanOrderModel.objects.filter(
            ordered_at__date__range=[from_date, to_date]
        )

        # Total orders exclude pending and cancelled
        orders = all_orders.exclude(status__in=['pending', 'cancelled'])

        # Get order by parameter
        order_by = self.request.GET.get('order_by', 'name')

        # Test type breakdown
        test_breakdown = orders.values(
            'template__name', 'template__id'
        ).annotate(
            total_orders=Count('id'),
            completed_orders=Count('id', filter=Q(status='completed'))
        )

        # Apply ordering
        if order_by == 'total':
            test_breakdown = test_breakdown.order_by('-total_orders', 'template__name')
        elif order_by == 'completed':
            test_breakdown = test_breakdown.order_by('-completed_orders', 'template__name')
        else:  # default to name
            test_breakdown = test_breakdown.order_by('template__name')

        # Summary statistics
        context.update({
            'from_date': from_date,
            'to_date': to_date,
            'report_title': report_title,
            'test_breakdown': test_breakdown,
            'order_by': order_by,

            # Summary stats
            'total_tests': orders.count(),
            'completed_tests': all_orders.filter(status='completed').count(),
            'processing_tests': all_orders.filter(status='processing').count(),
            'collected_tests': all_orders.filter(status='collected').count(),
            'paid_tests': all_orders.filter(status='paid').count(),
            'pending_tests': all_orders.filter(status='pending').count(),
            'cancelled_tests': all_orders.filter(status='cancelled').count(),
        })

        return context


class ScanReportExportExcelView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'scan.view_scanordermodel'

    def get(self, request, *args, **kwargs):
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        report_title = request.GET.get('title', 'Scan Test Report')
        show_total = request.GET.get('show_total', 'true') == 'true'
        show_completed = request.GET.get('show_completed', 'true') == 'true'
        order_by = request.GET.get('order_by', 'name')

        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()

        if not to_date:
            if date.today().month == 12:
                to_date = date.today().replace(day=31)
            else:
                to_date = (date.today().replace(month=date.today().month + 1, day=1) - timedelta(days=1))
        else:
            to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()

        # Get scan and site info
        scan_setting = ScanSettingModel.objects.first()
        site_info = SiteInfoModel.objects.first()

        # Query data
        orders = ScanOrderModel.objects.filter(
            ordered_at__date__range=[from_date, to_date]
        ).exclude(status__in=['pending', 'cancelled'])

        test_breakdown = orders.values(
            'template__name', 'template__id'
        ).annotate(
            total_orders=Count('id'),
            completed_orders=Count('id', filter=Q(status='completed'))
        )

        # Apply ordering
        if order_by == 'total':
            test_breakdown = test_breakdown.order_by('-total_orders', 'template__name')
        elif order_by == 'completed':
            test_breakdown = test_breakdown.order_by('-completed_orders', 'template__name')
        else:
            test_breakdown = test_breakdown.order_by('template__name')

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Scan Test Report"

        # Styles
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True, size=12)
        table_header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        row = 1

        # Header section
        if scan_setting and scan_setting.scan_name:
            ws.merge_cells(f'A{row}:D{row}')
            cell = ws[f'A{row}']
            cell.value = scan_setting.scan_name
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            row += 1

            if scan_setting.mobile or scan_setting.email:
                ws.merge_cells(f'A{row}:D{row}')
                cell = ws[f'A{row}']
                contact = []
                if scan_setting.mobile:
                    contact.append(f"Tel: {scan_setting.mobile}")
                if scan_setting.email:
                    contact.append(f"Email: {scan_setting.email}")
                cell.value = " | ".join(contact)
                cell.alignment = Alignment(horizontal='center')
                row += 1
        elif site_info:
            ws.merge_cells(f'A{row}:D{row}')
            cell = ws[f'A{row}']
            cell.value = site_info.name
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            row += 1

            ws.merge_cells(f'A{row}:D{row}')
            cell = ws[f'A{row}']
            contact = []
            if site_info.mobile_1:
                contact.append(f"Tel: {site_info.mobile_1}")
            if site_info.email:
                contact.append(f"Email: {site_info.email}")
            cell.value = " | ".join(contact)
            cell.alignment = Alignment(horizontal='center')
            row += 1

        row += 1

        # Report title
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = report_title
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center')
        row += 1

        # Date range
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = f"Period: {from_date.strftime('%B %d, %Y')} to {to_date.strftime('%B %d, %Y')}"
        cell.alignment = Alignment(horizontal='center')
        row += 2

        # Table headers
        col = 1
        headers = ['S/N', 'Test']
        if show_total:
            headers.append('Total Orders')
        if show_completed:
            headers.append('Completed Orders')

        for header in headers:
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = table_header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            col += 1

        row += 1

        # Data rows
        for idx, test in enumerate(test_breakdown, 1):
            col = 1

            # S/N
            cell = ws.cell(row=row, column=col)
            cell.value = idx
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            col += 1

            # Test name
            cell = ws.cell(row=row, column=col)
            cell.value = test['template__name']
            cell.border = border
            col += 1

            # Total orders
            if show_total:
                cell = ws.cell(row=row, column=col)
                cell.value = test['total_orders']
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                col += 1

            # Completed orders
            if show_completed:
                cell = ws.cell(row=row, column=col)
                cell.value = test['completed_orders']
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                col += 1

            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 40
        if show_total:
            ws.column_dimensions['C'].width = 15
        if show_completed:
            ws.column_dimensions['D'].width = 18

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"scan_test_report_{from_date.strftime('%Y%m%d')}_{to_date.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response


class ScanReportExportPDFView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'scan.view_scanordermodel'

    def get(self, request, *args, **kwargs):
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        report_title = request.GET.get('title', 'Scan Test Report')
        show_total = request.GET.get('show_total', 'true') == 'true'
        show_completed = request.GET.get('show_completed', 'true') == 'true'
        order_by = request.GET.get('order_by', 'name')
        report_type = request.GET.get('type', 'breakdown')  # breakdown or summary

        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()

        if not to_date:
            if date.today().month == 12:
                to_date = date.today().replace(day=31)
            else:
                to_date = (date.today().replace(month=date.today().month + 1, day=1) - timedelta(days=1))
        else:
            to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()

        # Get scan and site info
        scan_setting = ScanSettingModel.objects.first()
        site_info = SiteInfoModel.objects.first()

        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#333333'),
            spaceAfter=6,
            alignment=TA_CENTER
        )

        # Header
        if scan_setting and scan_setting.scan_name:
            elements.append(Paragraph(scan_setting.scan_name, title_style))
            if scan_setting.mobile or scan_setting.email:
                contact = []
                if scan_setting.mobile:
                    contact.append(f"Tel: {scan_setting.mobile}")
                if scan_setting.email:
                    contact.append(f"Email: {scan_setting.email}")
                elements.append(Paragraph(" | ".join(contact), subtitle_style))
        elif site_info:
            elements.append(Paragraph(site_info.name, title_style))
            contact = []
            if site_info.mobile_1:
                contact.append(f"Tel: {site_info.mobile_1}")
            if site_info.email:
                contact.append(f"Email: {site_info.email}")
            elements.append(Paragraph(" | ".join(contact), subtitle_style))

        elements.append(Spacer(1, 0.3 * inch))
        elements.append(Paragraph(report_title, title_style))
        elements.append(Paragraph(
            f"Period: {from_date.strftime('%B %d, %Y')} to {to_date.strftime('%B %d, %Y')}",
            subtitle_style
        ))
        elements.append(Spacer(1, 0.4 * inch))

        if report_type == 'summary':
            # Summary report
            all_orders = ScanOrderModel.objects.filter(
                ordered_at__date__range=[from_date, to_date]
            )
            orders = all_orders.exclude(status__in=['pending', 'cancelled'])

            summary_data = [
                ['Metric', 'Count'],
                ['Total Tests', str(orders.count())],
                ['Completed Tests', str(all_orders.filter(status='completed').count())],
                ['Processing Tests', str(all_orders.filter(status='processing').count())],
                ['Sample Collected', str(all_orders.filter(status='collected').count())],
                ['Paid (Awaiting Sample)', str(all_orders.filter(status='paid').count())],
                ['Pending Payment', str(all_orders.filter(status='pending').count())],
                ['Cancelled Tests', str(all_orders.filter(status='cancelled').count())],
            ]

            summary_table = Table(summary_data, colWidths=[4 * inch, 2 * inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
            ]))
            elements.append(summary_table)
        else:
            # Breakdown report
            orders = ScanOrderModel.objects.filter(
                ordered_at__date__range=[from_date, to_date]
            ).exclude(status__in=['pending', 'cancelled'])

            test_breakdown = orders.values(
                'template__name', 'template__id'
            ).annotate(
                total_orders=Count('id'),
                completed_orders=Count('id', filter=Q(status='completed'))
            )

            # Apply ordering
            if order_by == 'total':
                test_breakdown = test_breakdown.order_by('-total_orders', 'template__name')
            elif order_by == 'completed':
                test_breakdown = test_breakdown.order_by('-completed_orders', 'template__name')
            else:
                test_breakdown = test_breakdown.order_by('template__name')

            # Build table data
            headers = ['S/N', 'Test']
            col_widths = [0.5 * inch, 3 * inch]

            if show_total:
                headers.append('Total Orders')
                col_widths.append(1 * inch)
            if show_completed:
                headers.append('Completed Orders')
                col_widths.append(1.2 * inch)

            table_data = [headers]

            for idx, test in enumerate(test_breakdown, 1):
                row = [str(idx), test['template__name']]
                if show_total:
                    row.append(str(test['total_orders']))
                if show_completed:
                    row.append(str(test['completed_orders']))
                table_data.append(row)

            table = Table(table_data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
            ]))
            elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f"scan_test_report_{from_date.strftime('%Y%m%d')}_{to_date.strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class ScanLogView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    template_name = 'scan/reports/log.html'
    permission_required = 'scan.view_scanordermodel'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        from_date = self.request.GET.get('from_date')
        to_date = self.request.GET.get('to_date')

        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()

        if not to_date:
            today = date.today()
            if today.month == 12:
                to_date = today.replace(day=31)
            else:
                to_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        else:
            to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()

        default_title = f'Scan Log ({from_date.strftime("%b %d, %Y")} - {to_date.strftime("%b %d, %Y")})'
        report_title = self.request.GET.get('title', default_title)

        orders_qs = ScanOrderModel.objects.filter(
            ordered_at__date__range=[from_date, to_date]
        ).exclude(
            status__in=['pending', 'cancelled']
        ).select_related(
            'patient',
            'template',
            'result__verified_by__user_staff_profile__staff'
        ).order_by('ordered_at')

        totals = orders_qs.aggregate(
            total_amount=Sum('amount_charged'),
            total_patients=Count('patient', distinct=True)
        )

        context.update({
            'from_date': from_date,
            'to_date': to_date,
            'report_title': report_title,
            'orders': orders_qs,
            'total_amount': totals['total_amount'] or Decimal('0.00'),
            'total_patients': totals['total_patients'] or 0,
        })
        return context


class ScanLogExportExcelView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'scan.view_scanordermodel'

    def get(self, request, *args, **kwargs):
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        report_title = request.GET.get('title', 'Scan Log')
        show_status = request.GET.get('show_status', 'true') == 'true'
        show_amount = request.GET.get('show_amount', 'true') == 'true'
        show_scientist = request.GET.get('show_scientist', 'true') == 'true'

        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()
        if not to_date:
            today = date.today()
            if today.month == 12:
                to_date = today.replace(day=31)
            else:
                to_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        else:
            to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()

        orders_qs = ScanOrderModel.objects.filter(
            ordered_at__date__range=[from_date, to_date]
        ).exclude(
            status__in=['pending', 'cancelled']
        ).select_related(
            'patient', 'template', 'result__verified_by'
        ).order_by('ordered_at')

        totals = orders_qs.aggregate(
            total_amount=Sum('amount_charged'),
            total_patients=Count('patient', distinct=True)
        )
        total_amount = totals['total_amount'] or Decimal('0.00')
        total_patients = totals['total_patients'] or 0

        wb = Workbook()
        ws = wb.active
        ws.title = "Scan Log"

        # Styles
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True, size=12)
        table_header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        total_font = Font(bold=True, size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))

        # Header section adapted for Scan/Site info
        scan_setting = ScanSettingModel.objects.first()
        site_info = SiteInfoModel.objects.first()
        row = 1
        if scan_setting and scan_setting.scan_name:
            ws.merge_cells(f'A{row}:H{row}');
            cell = ws[f'A{row}'];
            cell.value = scan_setting.scan_name;
            cell.font = header_font;
            cell.alignment = Alignment(horizontal='center');
            row += 1
        elif site_info:
            ws.merge_cells(f'A{row}:H{row}');
            cell = ws[f'A{row}'];
            cell.value = site_info.name;
            cell.font = header_font;
            cell.alignment = Alignment(horizontal='center');
            row += 1
        row += 1

        ws.merge_cells(f'A{row}:H{row}');
        cell = ws[f'A{row}'];
        cell.value = report_title;
        cell.font = title_font;
        cell.alignment = Alignment(horizontal='center');
        row += 1
        ws.merge_cells(f'A{row}:H{row}');
        cell = ws[f'A{row}'];
        cell.value = f"Period: {from_date.strftime('%B %d, %Y')} to {to_date.strftime('%B %d, %Y')}";
        cell.alignment = Alignment(horizontal='center');
        row += 2

        # Table Headers
        headers = ['S/N', 'Date', 'Patient ID', 'Patient Name', 'Scan Name']
        if show_status: headers.append('Status')
        if show_amount: headers.append('Amount')
        if show_scientist: headers.append('Verified By')

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num, value=header)
            cell.font = table_header_font;
            cell.fill = header_fill;
            cell.border = border
        row += 1

        # Data Rows
        for idx, order in enumerate(orders_qs, 1):
            col = 1
            ws.cell(row=row, column=col, value=idx).border = border;
            col += 1
            ws.cell(row=row, column=col, value=order.ordered_at.strftime('%Y-%m-%d %H:%M')).border = border;
            col += 1
            ws.cell(row=row, column=col, value=order.patient.card_number).border = border;
            col += 1
            ws.cell(row=row, column=col, value=order.patient.__str__()).border = border;
            col += 1
            ws.cell(row=row, column=col, value=order.template.name).border = border;
            col += 1

            if show_status:
                ws.cell(row=row, column=col, value=order.get_status_display()).border = border;
                col += 1
            if show_amount:
                cell = ws.cell(row=row, column=col, value=order.amount_charged);
                cell.number_format = '#,##0.00';
                cell.border = border;
                col += 1

            if show_scientist:
                verifier = 'N/A'
                if hasattr(order, 'result') and order.result and order.result.verified_by:
                    user = order.result.verified_by
                    if hasattr(user,
                               'user_staff_profile') and user.user_staff_profile and user.user_staff_profile.staff:
                        verifier = user.user_staff_profile.staff.__str__()
                    else:
                        verifier = user.get_full_name()
                ws.cell(row=row, column=col, value=verifier).border = border;
                col += 1
            row += 1

        # Total Row
        ws.cell(row=row, column=5, value="TOTAL").font = total_font
        ws.cell(row=row, column=5).border = border

        col_offset = 6
        if show_status:
            ws.cell(row=row, column=col_offset).border = border;
            col_offset += 1
        if show_amount:
            cell = ws.cell(row=row, column=col_offset, value=total_amount);
            cell.font = total_font;
            cell.number_format = '#,##0.00';
            cell.border = border;
            col_offset += 1
        if show_scientist:
            cell = ws.cell(row=row, column=col_offset, value=f"{total_patients} Unique Patients");
            cell.font = total_font;
            cell.border = border

        ws.column_dimensions['A'].width = 6;
        ws.column_dimensions['B'].width = 18;
        ws.column_dimensions['C'].width = 15;
        ws.column_dimensions['D'].width = 30;
        ws.column_dimensions['E'].width = 35
        if show_status: ws.column_dimensions['F'].width = 15
        if show_amount: ws.column_dimensions['G'].width = 15
        if show_scientist: ws.column_dimensions['H'].width = 30

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"scan_log_{from_date.strftime('%Y%m%d')}_{to_date.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class ScanLogExportPDFView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'scan.view_scanordermodel'

    def get(self, request, *args, **kwargs):
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        report_title = request.GET.get('title', 'Scan Log')
        show_status = request.GET.get('show_status', 'true') == 'true'
        show_amount = request.GET.get('show_amount', 'true') == 'true'
        show_scientist = request.GET.get('show_scientist', 'true') == 'true'

        if not from_date:
            from_date = date.today().replace(day=1)
        else:
            from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()
        if not to_date:
            today = date.today()
            if today.month == 12:
                to_date = today.replace(day=31)
            else:
                to_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        else:
            to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()

        orders_qs = ScanOrderModel.objects.filter(
            ordered_at__date__range=[from_date, to_date]
        ).exclude(
            status__in=['pending', 'cancelled']
        ).select_related(
            'patient', 'template', 'result__verified_by'
        ).order_by('ordered_at')

        totals = orders_qs.aggregate(Sum('amount_charged'), Count('patient', distinct=True))
        total_amount = totals['amount_charged__sum'] or Decimal('0.00')
        total_patients = totals['patient__count'] or 0

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30,
                                bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, alignment=TA_CENTER,
                                     spaceAfter=12)
        subtitle_style = ParagraphStyle('CustomSubtitle', parent=styles['Normal'], fontSize=11, alignment=TA_CENTER,
                                        spaceAfter=6)

        # Header adapted for Scan/Site info
        scan_setting = ScanSettingModel.objects.first()
        site_info = SiteInfoModel.objects.first()
        if scan_setting and scan_setting.scan_name:
            elements.append(Paragraph(scan_setting.scan_name, title_style))
        elif site_info:
            elements.append(Paragraph(site_info.name, title_style))

        elements.append(Spacer(1, 0.1 * inch))
        elements.append(Paragraph(report_title, title_style))
        elements.append(
            Paragraph(f"Period: {from_date.strftime('%B %d, %Y')} to {to_date.strftime('%B %d, %Y')}", subtitle_style))
        elements.append(Spacer(1, 0.2 * inch))

        headers = ['S/N', 'Date', 'Patient ID', 'Patient Name', 'Scan Name']
        col_widths = [0.4 * inch, 0.9 * inch, 0.9 * inch, 1.8 * inch, 2.0 * inch]
        if show_status: headers.append('Status'); col_widths.append(1.0 * inch)
        if show_amount: headers.append('Amount'); col_widths.append(0.8 * inch)
        if show_scientist: headers.append('Verified By'); col_widths.append(1.5 * inch)
        table_data = [headers]

        for idx, order in enumerate(orders_qs, 1):
            row = [str(idx), order.ordered_at.strftime('%y-%m-%d %H:%M'), order.patient.card_number,
                   order.patient.__str__(), order.template.name]
            if show_status: row.append(order.get_status_display())
            if show_amount: row.append(f"{order.amount_charged:,.2f}")
            if show_scientist:
                verifier = 'N/A'
                if hasattr(order, 'result') and order.result and order.result.verified_by:
                    user = order.result.verified_by
                    if hasattr(user,
                               'user_staff_profile') and user.user_staff_profile and user.user_staff_profile.staff:
                        verifier = user.user_staff_profile.staff.__str__()
                    else:
                        verifier = user.get_full_name()
                row.append(verifier)
            table_data.append(row)

        total_row = ['', '', '', '', 'TOTAL']
        if show_status: total_row.append('')
        if show_amount: total_row.append(f"{total_amount:,.2f}")
        if show_scientist: total_row.append(f"{total_patients} Unique Patients")
        table_data.append(total_row)

        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey), ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f0f0f0')]),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'), ('FONTSIZE', (0, -1), (-1, -1), 9),
            ('GRID', (0, -1), (-1, -1), 1, colors.black), ('ALIGN', (4, -1), (-1, -1), 'RIGHT'),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f"scan_log_{from_date.strftime('%Y%m%d')}_{to_date.strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response